#!/usr/bin/env python3
"""
Дашборд brak_team.brak_data — 4 таблицы ТОП-20 как в отчёте.

Источник: brak_team.brak_data → VIEW brak_team.brak_data_norm → weekly/nm matview.

  python write_offs_dashboard.py
  → http://127.0.0.1:8080/

Фильтр: все WH, корпус (несколько wh_id из wh_buildings.json) или свой набор галочками.
Настройка корпусов: wh_buildings.json
"""

from __future__ import annotations

import json
import os
import re
import secrets
import sys
from io import BytesIO
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from threading import RLock
from time import monotonic
from typing import Any

from dotenv import load_dotenv

from brak_dashboard.analytics import (
    build_growth_alerts as _analytics_build_growth_alerts,
    build_period_compare,
    build_top20_churn,
    concentration_shares,
    parse_year_value,
    stable_etag_payload,
    yoy_pct,
)

ROOT = Path(__file__).resolve().parent.parent
CONFIG_PATH = ROOT / "wh_buildings.json"
BASE_TABLE = "brak_team.brak_data"
NORM_VIEW = "brak_team.brak_data_norm"
NORM_VIEW_SQL_PATH = ROOT / "sql" / "brak_data_norm.sql"

DETAIL_COLUMNS = (
    ("shk_id", "ШК"),
    ("date", "Дата"),
    ("type", "Тип"),
    ("total_cost", "Стоимость"),
    ("amount", "Сумма факт"),
    ("share", "Доля"),
    ("office_id", "Офис"),
    ("wh_id", "Блок"),
    ("nm_id", "nm_id"),
    ("subject_name", "Предмет"),
    ("parent_name", "Родитель"),
    ("title", "Наименование"),
    ("brand_name", "Бренд"),
    ("state_id", "state_id"),
    ("reason_id", "reason_id"),
    ("reason_descr", "Причина"),
    ("seller_id", "seller_id"),
    ("supplier_id", "supplier_id"),
    ("owner_product", "Владелец"),
    ("cnt_org", "cnt_org"),
    ("cnt_ors", "cnt_ors"),
    ("cnt_ocr", "cnt_ocr"),
    ("amount_obsh", "Общая ст-ть"),
    ("summa_obshay", "summa_obshay"),
    ("wh_name", "WH"),
    ("source_file", "Файл"),
)
DETAIL_COL_NAMES = [c[0] for c in DETAIL_COLUMNS]
DETAIL_SORTABLE = {
    "date",
    "amount",
    "amount_obsh",
    "summa_obshay",
    "total_cost",
    "share",
    "office_id",
    "wh_id",
    "nm_id",
    "shk_id",
    "reason_id",
    "cnt_org",
    "cnt_ors",
    "cnt_ocr",
    "type",
    "parent_name",
    "reason_descr",
    "title",
    "brand_name",
    "subject_name",
    "owner_product",
    "state_id",
    "wh_name",
    "source_file",
}
DETAIL_CLICK_FILTERS = {
    "wh_id",
    "office_id",
    "nm_id",
    "shk_id",
    "reason_id",
    "parent_name",
    "type",
    "cnt_org",
    "brand_name",
    "subject_name",
    "owner_product",
    "state_id",
    "wh_name",
    "source_file",
}

_env_file = ROOT / ".env"
if _env_file.exists():
    load_dotenv(_env_file)

CACHE_TTL_SEC = int(os.environ.get("REPORT_CACHE_TTL_SEC", "120"))
WEEKS_CACHE_TTL_SEC = int(os.environ.get("WEEKS_CACHE_TTL_SEC", "300"))
NM_CACHE_TTL_SEC = int(os.environ.get("NM_CACHE_TTL_SEC", "300"))
MATVIEW_BOOTSTRAP_TTL_SEC = int(os.environ.get("MATVIEW_BOOTSTRAP_TTL_SEC", "600"))
_CACHE_LOCK = RLock()
_CACHE: dict[str, tuple[float, Any]] = {}
_MV_LOCK = RLock()
_MV_BOOTSTRAP_OK_AT = 0.0
_MV_BOOTSTRAP_FAIL_AT = 0.0
_MV_BOOTSTRAP_FAIL_MSG = ""
_MV_LAST_REFRESH_AT = 0.0
_MV_LAST_REFRESH_NOTE = ""
_MV_TEMP_SUFFIX = "__new"
_ADMIN_SESSIONS_LOCK = RLock()
_ADMIN_SESSIONS: dict[str, float] = {}
ADMIN_SESSION_TTL_SEC = int(os.environ.get("ADMIN_SESSION_TTL_SEC", "28800"))
_PG_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _cache_get(key: str) -> Any | None:
    now = monotonic()
    with _CACHE_LOCK:
        item = _CACHE.get(key)
        if not item:
            return None
        expires_at, payload = item
        if expires_at < now:
            _CACHE.pop(key, None)
            return None
        return payload


def _cache_set(key: str, payload: Any, ttl_sec: int) -> None:
    with _CACHE_LOCK:
        _CACHE[key] = (monotonic() + max(1, ttl_sec), payload)


def _cache_clear_all() -> None:
    with _CACHE_LOCK:
        _CACHE.clear()


def _use_report_matview() -> bool:
    return os.environ.get("USE_WRITE_OFFS_MATVIEW", "1").strip().lower() not in (
        "",
        "0",
        "false",
        "no",
    )


def _matview_name() -> str:
    return os.environ.get(
        "WRITE_OFFS_MATVIEW_NAME", "brak_team.brak_weekly_mv"
    ).strip() or "brak_team.brak_weekly_mv"


def _nm_matview_name() -> str:
    return os.environ.get(
        "WRITE_OFFS_NM_MATVIEW_NAME", "brak_team.brak_nm_weekly_mv"
    ).strip() or "brak_team.brak_nm_weekly_mv"


def _quote_pg_relation_name(name: str) -> str:
    parts = [p.strip() for p in name.split(".")]
    if len(parts) not in (1, 2) or any(not p for p in parts):
        raise ValueError(f"Некорректное имя PostgreSQL relation: {name!r}")
    for part in parts:
        if not _PG_IDENTIFIER_RE.fullmatch(part):
            raise ValueError(f"Некорректное имя PostgreSQL identifier: {part!r}")
    return ".".join(f'"{part}"' for part in parts)


def _relation_basename(name: str) -> str:
    return name.split(".")[-1].strip()


def _ensure_report_matview() -> None:
    """
    Ensures weekly aggregated materialized view exists.
    Expensive DDL check is throttled by MATVIEW_BOOTSTRAP_TTL_SEC.
    """
    global _MV_BOOTSTRAP_OK_AT, _MV_BOOTSTRAP_FAIL_AT, _MV_BOOTSTRAP_FAIL_MSG
    if not _use_report_matview():
        return
    with _MV_LOCK:
        now = monotonic()
        if _MV_BOOTSTRAP_OK_AT and (now - _MV_BOOTSTRAP_OK_AT) < MATVIEW_BOOTSTRAP_TTL_SEC:
            return
        if _MV_BOOTSTRAP_FAIL_AT and (now - _MV_BOOTSTRAP_FAIL_AT) < MATVIEW_BOOTSTRAP_TTL_SEC:
            raise RuntimeError(_MV_BOOTSTRAP_FAIL_MSG or "matview bootstrap failed")

        mv = _matview_name()
        mv_tmp = f"{mv}{_MV_TEMP_SUFFIX}"
        mv_sql = _quote_pg_relation_name(mv)
        mv_tmp_sql = _quote_pg_relation_name(mv_tmp)
        mv_basename_sql = _quote_pg_relation_name(_relation_basename(mv))
        ddl_tmp = f"""
            CREATE MATERIALIZED VIEW IF NOT EXISTS {mv_tmp_sql} AS
            SELECT
                EXTRACT(ISOYEAR FROM date)::int AS iso_year,
                EXTRACT(WEEK FROM date)::int AS week_no,
                office_id,
                wh_id,
                cnt_org,
                reason_id,
                COALESCE(reason_descr, '—') AS reason_descr,
                COALESCE(parent_name, '—') AS parent_name,
                SUM(amount)::numeric AS amount_sum,
                COUNT(*)::bigint AS rows_cnt,
                MAX(date) AS max_date
            FROM {NORM_VIEW}
            WHERE date IS NOT NULL
            GROUP BY 1, 2, 3, 4, 5, 6, 7, 8
        """
        idx = [
            f"CREATE UNIQUE INDEX IF NOT EXISTS brak_weekly_mv_uq ON {mv_sql} (iso_year, week_no, office_id, wh_id, cnt_org, reason_id, reason_descr, parent_name)",
            f"CREATE INDEX IF NOT EXISTS brak_weekly_mv_filter_idx ON {mv_sql} (iso_year, office_id, wh_id, week_no, cnt_org)",
            f"CREATE INDEX IF NOT EXISTS brak_weekly_mv_week_idx ON {mv_sql} (iso_year, week_no)",
        ]
        try:
            _ensure_brak_data_norm()
            with get_conn() as conn:
                conn.autocommit = True
                cur = conn.cursor()
                cur.execute(f"DROP MATERIALIZED VIEW IF EXISTS {mv_tmp_sql} CASCADE")
                cur.execute("SELECT to_regclass(%s)", (mv,))
                mv_regclass = cur.fetchone()
                if mv_regclass and mv_regclass[0]:
                    cur.execute(
                        """
                        SELECT attname
                        FROM pg_attribute
                        WHERE attrelid = to_regclass(%s)
                          AND attnum > 0
                          AND NOT attisdropped
                        """,
                        (mv,),
                    )
                    cols = {str(r[0]) for r in cur.fetchall()}
                    required = {
                        "iso_year",
                        "week_no",
                        "office_id",
                        "wh_id",
                        "cnt_org",
                        "reason_id",
                        "reason_descr",
                        "parent_name",
                        "amount_sum",
                        "rows_cnt",
                        "max_date",
                    }
                    if ("nm_id" in cols) or (not required.issubset(cols)):
                        cur.execute(f"DROP MATERIALIZED VIEW IF EXISTS {mv_sql} CASCADE")
                cur.execute("SELECT to_regclass(%s)", (mv,))
                mv_regclass = cur.fetchone()
                if not mv_regclass or not mv_regclass[0]:
                    cur.execute(ddl_tmp)
                    cur.execute(
                        f"ALTER MATERIALIZED VIEW {mv_tmp_sql} RENAME TO {mv_basename_sql}"
                    )
                for stmt in idx:
                    cur.execute(stmt)
            _MV_BOOTSTRAP_OK_AT = monotonic()
            _MV_BOOTSTRAP_FAIL_AT = 0.0
            _MV_BOOTSTRAP_FAIL_MSG = ""
        except Exception as exc:
            _MV_BOOTSTRAP_FAIL_AT = monotonic()
            _MV_BOOTSTRAP_FAIL_MSG = f"matview bootstrap failed: {exc}"
            raise


def _ensure_nm_matview() -> None:
    if not _use_report_matview():
        return
    _ensure_brak_data_norm()
    mv = _nm_matview_name()
    mv_sql = _quote_pg_relation_name(mv)
    ddl = f"""
        CREATE MATERIALIZED VIEW IF NOT EXISTS {mv_sql} AS
        SELECT
            EXTRACT(ISOYEAR FROM date)::int AS iso_year,
            EXTRACT(WEEK FROM date)::int AS week_no,
            office_id,
            wh_id,
            nm_id,
            COUNT(*)::bigint AS rows_cnt,
            MAX(date) AS max_date
        FROM {NORM_VIEW}
        WHERE date IS NOT NULL
          AND nm_id IS NOT NULL
        GROUP BY 1, 2, 3, 4, 5
    """
    idx = [
        f"CREATE UNIQUE INDEX IF NOT EXISTS brak_nm_weekly_mv_uq ON {mv_sql} (iso_year, week_no, office_id, wh_id, nm_id)",
        f"CREATE INDEX IF NOT EXISTS brak_nm_weekly_mv_filter_idx ON {mv_sql} (iso_year, office_id, week_no, wh_id)",
    ]
    with get_conn() as conn:
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute(ddl)
        for stmt in idx:
            cur.execute(stmt)


def _refresh_report_matview() -> str | None:
    global _MV_LAST_REFRESH_AT, _MV_LAST_REFRESH_NOTE
    if not _use_report_matview():
        return None
    _ensure_report_matview()
    mv = _matview_name()
    mv_sql = _quote_pg_relation_name(mv)
    nm_mv = _nm_matview_name()
    nm_mv_sql = _quote_pg_relation_name(nm_mv)
    _ensure_nm_matview()
    use_concurrently = os.environ.get(
        "WRITE_OFFS_MATVIEW_REFRESH_CONCURRENTLY", "1"
    ).strip().lower() not in ("0", "false", "no")
    note = "Матпредставления обновлены"
    with get_conn() as conn:
        conn.autocommit = True
        cur = conn.cursor()
        if use_concurrently:
            try:
                cur.execute(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {mv_sql}")
                try:
                    cur.execute(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {nm_mv_sql}")
                    note = "Матпредставления обновлены (CONCURRENTLY)"
                except Exception:
                    cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
            except Exception:
                cur.execute(f"REFRESH MATERIALIZED VIEW {mv_sql}")
                cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
        else:
            cur.execute(f"REFRESH MATERIALIZED VIEW {mv_sql}")
            cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
    with _MV_LOCK:
        _MV_LAST_REFRESH_AT = monotonic()
        _MV_LAST_REFRESH_NOTE = note
    return note


def _report_source() -> dict[str, str]:
    if _use_report_matview():
        try:
            _ensure_report_matview()
            return {
                "table": _matview_name(),
                "week_expr": "week_no",
                "isoyear_expr": "iso_year",
                "amount_expr": "amount_sum",
                "year_clause": "iso_year = %s",
                "base_not_null_clause": "",
            }
        except Exception as exc:
            print(
                f"[brak_data] matview unavailable, fallback to {NORM_VIEW}: {exc}",
                file=sys.stderr,
            )
    _ensure_brak_data_norm()
    return {
        "table": NORM_VIEW,
        "week_expr": "EXTRACT(WEEK FROM date)::int",
        "isoyear_expr": "EXTRACT(ISOYEAR FROM date)::int",
        "amount_expr": "amount",
        "year_clause": "date >= %s AND date < %s",
        "base_not_null_clause": "date IS NOT NULL",
    }


@dataclass
class Row:
    row_id: int | None
    name: str
    amounts: dict[int, float]

    def amount(self, week: int) -> float:
        return self.amounts.get(week, 0.0)

    def dynamics(self, week_prev: int, week_last: int) -> float | None:
        w_prev, w_last = self.amount(week_prev), self.amount(week_last)
        if w_prev == 0:
            return None  # Исправлено: рост с нуля — это None, а не 100%
        return (w_last - w_prev) / w_prev * 100

    def average(self, week_prev: int, week_last: int) -> float:
        return (self.amount(week_prev) + self.amount(week_last)) / 2

    def pct_vs_avg(self, week_prev: int, week_last: int) -> float | None:
        avg = self.average(week_prev, week_last)
        if avg == 0:
            return None
        return self.amount(week_last) / avg * 100


class QueryParamError(ValueError):
    pass


def parse_wh_ids(raw: str) -> list[int]:
    wh_ids: list[int] = []
    seen: set[int] = set()
    for item in raw.split(","):
        part = item.strip()
        if not part:
            continue
        try:
            wh_id = int(part)
        except (TypeError, ValueError) as exc:
            raise QueryParamError(
                f"Некорректный wh_ids: значение {part!r} не является числом"
            ) from exc
        if wh_id not in seen:
            seen.add(wh_id)
            wh_ids.append(wh_id)
    return wh_ids


def normalize_database_url(url: str) -> str:
    from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

    url = url.strip()
    if "://" in url:
        scheme, rest = url.split("://", 1)
        if "+" in scheme:
            scheme = scheme.split("+", 1)[0]
        url = f"{scheme}://{rest}"

    parsed = urlparse(url)
    if not parsed.scheme:
        return url

    query = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k and v]
    if not query and parsed.query:
        pass
    fixed = parsed._replace(query=urlencode(query))
    return urlunparse(fixed)


def db_config() -> dict[str, Any]:
    timeout = int(os.environ.get("DB_CONNECT_TIMEOUT", "15"))
    host = os.environ.get("DB_HOST", "").strip()
    user = os.environ.get("DB_USER", "").strip()
    password = os.environ.get("DB_PASSWORD", "")

    if host and user and password:
        cfg: dict[str, Any] = {
            "host": host,
            "port": int(os.environ.get("DB_PORT", "5432")),
            "dbname": os.environ.get("DB_NAME", "botdb"),
            "user": user,
            "password": password,
            "connect_timeout": timeout,
        }
        sslmode = os.environ.get("DB_SSLMODE")
        if sslmode:
            cfg["sslmode"] = sslmode
        return cfg

    database_url = os.environ.get("DATABASE_URL", "").strip()
    if database_url:
        return {"dsn": normalize_database_url(database_url), "connect_timeout": timeout}

    return {
        "host": host or "localhost",
        "port": int(os.environ.get("DB_PORT", "5432")),
        "dbname": os.environ.get("DB_NAME", "botdb"),
        "user": user,
        "password": password,
        "connect_timeout": timeout,
    }


def check_db_env() -> str | None:
    if os.environ.get("DATABASE_URL", "").strip():
        return None
    missing = [k for k in ("DB_HOST", "DB_USER", "DB_PASSWORD") if not os.environ.get(k)]
    if missing:
        return "Не заданы: DATABASE_URL или " + ", ".join(missing)
    return None


@contextmanager
def get_conn():
    import psycopg2

    conn = psycopg2.connect(**db_config())
    try:
        yield conn
    finally:
        conn.close()


def _ensure_brak_data_norm() -> None:
    """Create/replace typed VIEW over brak_team.brak_data."""
    ddl = NORM_VIEW_SQL_PATH.read_text(encoding="utf-8")
    with get_conn() as conn:
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute(ddl)


def load_config() -> dict:
    if CONFIG_PATH.exists():
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    return {
        "office_id": None,
        "week_year": 2026,
        "week_prev": 20,
        "week_last": 21,
        "wh_catalog": [],
        "buildings": [{"id": "all", "name": "Все WH", "wh_ids": []}],
    }


def catalog_wh_ids(cfg: dict) -> list[int]:
    catalog = cfg.get("wh_catalog") or []
    return [int(w["wh_id"]) for w in catalog]


def to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    return float(v)


def fetch_wh_list(office_id: int | None) -> list[dict]:
    _ensure_brak_data_norm()
    clauses = []
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"""
        SELECT wh_id,
               COUNT(*) AS cnt,
               ROUND(SUM(amount)::numeric, 0) AS total_amount
        FROM {NORM_VIEW}
        {where}
        GROUP BY wh_id
        ORDER BY wh_id
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
    return [
        {"wh_id": r[0], "cnt": r[1], "total_amount": to_float(r[2])}
        for r in rows
    ]


def fetch_db_stats(
    office_id: int | None,
    wh_ids: list[int] | None,
) -> dict[str, Any]:
    _ensure_brak_data_norm()
    clauses: list[str] = []
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"""
        SELECT COUNT(*)::bigint,
               MAX(date)::text,
               COALESCE(ROUND(SUM(amount)::numeric, 0), 0)
        FROM {NORM_VIEW}
        {where}
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
    return {
        "row_count": int(row[0]) if row else 0,
        "max_date": row[1] if row else None,
        "total_amount": to_float(row[2]) if row else 0.0,
    }


def _detail_int_arg(args: Any, name: str) -> int | None:
    raw = args.get(name, "")
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError) as exc:
        raise QueryParamError(f"Некорректный {name}: значение должно быть числом") from exc


def _detail_positive_int_arg(
    args: Any,
    name: str,
    default: int,
    *,
    min_value: int,
    max_value: int,
) -> int:
    value = _detail_int_arg(args, name)
    if value is None:
        value = default
    return max(min_value, min(max_value, value))


def build_detail_where(args: Any) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []

    office_id = _detail_int_arg(args, "office_id")
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)

    wh_id = _detail_int_arg(args, "wh_id")
    wh_raw = str(args.get("wh_ids", "") or "").strip()
    if wh_id is not None and wh_raw:
        raise QueryParamError("Передайте только один фильтр: wh_id или wh_ids")
    if wh_id is not None:
        clauses.append("wh_id = %s")
        params.append(wh_id)
    elif wh_raw:
        wh_ids = parse_wh_ids(wh_raw)
        if not wh_ids:
            raise QueryParamError("Некорректный wh_ids: список пуст")
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)

    row_type = str(args.get("type", "") or "").strip()
    if row_type:
        clauses.append("type = %s")
        params.append(row_type)

    date_from = str(args.get("date_from", "") or "").strip()
    if date_from:
        clauses.append("date >= %s")
        params.append(date_from)

    date_to = str(args.get("date_to", "") or "").strip()
    if date_to:
        clauses.append("date < %s::date + interval '1 day'")
        params.append(date_to)

    reason_id = _detail_int_arg(args, "reason_id")
    if reason_id is not None:
        clauses.append("reason_id = %s")
        params.append(reason_id)

    parent_name = str(args.get("parent_name", "") or "").strip()
    if parent_name:
        clauses.append("COALESCE(parent_name, '—') = %s")
        params.append(parent_name)

    nm_id = _detail_int_arg(args, "nm_id")
    if nm_id is not None:
        clauses.append("nm_id = %s")
        params.append(nm_id)

    shk_id = _detail_int_arg(args, "shk_id")
    if shk_id is not None:
        clauses.append("shk_id = %s")
        params.append(shk_id)

    cnt_org = _detail_int_arg(args, "cnt_org")
    if cnt_org is not None:
        clauses.append("cnt_org = %s")
        params.append(cnt_org)

    brand_name = str(args.get("brand_name", "") or "").strip()
    if brand_name:
        clauses.append("COALESCE(brand_name, '') = %s")
        params.append(brand_name)

    subject_name = str(args.get("subject_name", "") or "").strip()
    if subject_name:
        clauses.append("COALESCE(subject_name, '—') = %s")
        params.append(subject_name)

    owner_product = str(args.get("owner_product", "") or "").strip()
    if owner_product:
        clauses.append("COALESCE(owner_product, '—') = %s")
        params.append(owner_product)

    state_id = str(args.get("state_id", "") or "").strip()
    if state_id:
        clauses.append("COALESCE(state_id, '—') = %s")
        params.append(state_id)

    wh_name = str(args.get("wh_name", "") or "").strip()
    if wh_name:
        clauses.append("COALESCE(wh_name, '') = %s")
        params.append(wh_name)

    source_file = str(args.get("source_file", "") or "").strip()
    if source_file:
        clauses.append("COALESCE(source_file, '') = %s")
        params.append(source_file)

    search = str(args.get("search", "") or "").strip()
    if search:
        like = f"%{search}%"
        clauses.append(
            "("
            "title ILIKE %s OR reason_descr ILIKE %s OR brand_name ILIKE %s OR "
            "subject_name ILIKE %s OR parent_name ILIKE %s OR "
            "owner_product ILIKE %s OR state_id ILIKE %s OR wh_name ILIKE %s OR "
            "source_file ILIKE %s OR "
            "nm_id::text ILIKE %s OR shk_id::text ILIKE %s"
            ")"
        )
        params.extend([like, like, like, like, like, like, like, like, like, like, like])

    if not clauses:
        return "", params
    return " WHERE " + " AND ".join(clauses), params


def _detail_row_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for idx, col in enumerate(DETAIL_COL_NAMES):
        value = row[idx]
        if isinstance(value, Decimal):
            value = float(value)
        elif hasattr(value, "isoformat"):
            value = value.isoformat()
        out[col] = value
    return out


def _detail_order_sql(args: Any) -> str:
    sort_by = str(args.get("sort_by", "") or "").strip()
    sort_dir = str(args.get("sort_dir", "") or "").strip().lower()
    if sort_by not in DETAIL_SORTABLE:
        sort_by = "date"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    nulls = "NULLS LAST" if sort_dir == "desc" else "NULLS FIRST"
    if sort_by == "date":
        return f"ORDER BY date {sort_dir} {nulls}, shk_id"
    return f"ORDER BY {sort_by} {sort_dir} {nulls}, date DESC NULLS LAST, shk_id"


def fetch_detail_page(
    args: Any,
    *,
    page: int,
    per_page: int,
) -> dict[str, Any]:
    where_sql, params = build_detail_where(args)
    cols = ", ".join(DETAIL_COL_NAMES)
    offset = (page - 1) * per_page
    order_sql = _detail_order_sql(args)
    sort_by = str(args.get("sort_by", "") or "").strip() or "date"
    sort_dir = str(args.get("sort_dir", "") or "").strip().lower() or "desc"
    if sort_by not in DETAIL_SORTABLE:
        sort_by = "date"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    _ensure_brak_data_norm()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT COUNT(*)::bigint FROM {NORM_VIEW}{where_sql}", params)
        total_row = cur.fetchone()
        total = int(total_row[0]) if total_row else 0
        cur.execute(
            f"""
            SELECT {cols}
            FROM {NORM_VIEW}
            {where_sql}
            {order_sql}
            LIMIT %s OFFSET %s
            """,
            [*params, per_page, offset],
        )
        rows = cur.fetchall()

    pages = max(1, (total + per_page - 1) // per_page)
    return {
        "columns": [{"key": key, "label": label} for key, label in DETAIL_COLUMNS],
        "rows": [_detail_row_to_dict(tuple(r)) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "click_filters": sorted(DETAIL_CLICK_FILTERS),
    }


def run_db_refresh() -> str | None:
    import urllib.error
    import urllib.request

    notes: list[str] = []
    sql = os.environ.get("DB_REFRESH_SQL", "").strip()
    if sql:
        with get_conn() as conn:
            conn.autocommit = True
            cur = conn.cursor()
            cur.execute(sql)
            try:
                row = cur.fetchone()
                if row and row[0] is not None:
                    notes.append(str(row[0]))
                else:
                    notes.append("SQL обновления выполнен")
            except Exception:
                notes.append("SQL обновления выполнен")

    url = os.environ.get("DB_REFRESH_URL", "").strip()
    if url:
        method = os.environ.get("DB_REFRESH_METHOD", "POST").upper()
        req = urllib.request.Request(url, method=method)
        token = os.environ.get("DB_REFRESH_TOKEN", "").strip()
        if token:
            req.add_header("Authorization", f"Bearer {token}")
        timeout = int(os.environ.get("DB_REFRESH_TIMEOUT", "120"))
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                body = resp.read().decode("utf-8", errors="replace")[:300]
                notes.append(body or f"HTTP {resp.status}")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:300]
            raise RuntimeError(f"DB_REFRESH_URL: HTTP {exc.code} {detail}") from exc

    try:
        mv_note = _refresh_report_matview()
        if mv_note:
            notes.append(mv_note)
    except Exception as exc:
        notes.append(f"matview refresh error: {exc}")
    if notes:
        return " | ".join(notes)
    return None


def check_refresh_access() -> str | None:
    token = os.environ.get("REFRESH_API_TOKEN", "").strip()
    if not token:
        return None

    from flask import request

    got = request.headers.get("X-Refresh-Token", "").strip() or request.args.get(
        "refresh_token", ""
    ).strip()
    if got != token:
        return "Недостаточно прав для обновления данных"
    return None


def _admin_login_password() -> tuple[str, str]:
    login = os.environ.get("ADMIN_LOGIN", "admin").strip() or "admin"
    password = os.environ.get("ADMIN_PASSWORD", "").strip()
    if not password:
        password = os.environ.get("REFRESH_API_TOKEN", "").strip()
    return login, password


def _admin_session_create() -> str:
    sid = secrets.token_urlsafe(32)
    with _ADMIN_SESSIONS_LOCK:
        _ADMIN_SESSIONS[sid] = monotonic() + max(60, ADMIN_SESSION_TTL_SEC)
    return sid


def _admin_session_valid(sid: str) -> bool:
    if not sid:
        return False
    now = monotonic()
    with _ADMIN_SESSIONS_LOCK:
        exp = _ADMIN_SESSIONS.get(sid)
        if exp is None:
            return False
        if exp < now:
            _ADMIN_SESSIONS.pop(sid, None)
            return False
        _ADMIN_SESSIONS[sid] = now + max(60, ADMIN_SESSION_TTL_SEC)
        return True


def _admin_session_delete(sid: str) -> None:
    if not sid:
        return
    with _ADMIN_SESSIONS_LOCK:
        _ADMIN_SESSIONS.pop(sid, None)


def check_admin_session_access() -> str | None:
    from flask import request

    sid = request.headers.get("X-Admin-Session", "").strip() or request.args.get(
        "admin_session", ""
    ).strip()
    if not _admin_session_valid(sid):
        return "Требуется вход администратора"
    return None


def check_admin_access() -> str | None:
    session_err = check_admin_session_access()
    if session_err is None:
        return None
    if not os.environ.get("REFRESH_API_TOKEN", "").strip():
        return session_err
    legacy_err = check_refresh_access()
    if legacy_err is None:
        return None
    return session_err


def fetch_totals_all(
    *,
    wh_ids: list[int] | None,
    office_id: int | None,
    org0_only: bool,
    year: int,
    week_prev: int,
    week_last: int,
) -> dict[str, float]:
    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    if org0_only:
        clauses.append("cnt_org = 0")
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    sql = f"""
        SELECT
            COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS total_prev,
            COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS total_last
        FROM {src["table"]}
        {where}
    """
    qparams = [week_prev, week_last, *params]
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, qparams)
        row = cur.fetchone()
    return {
        "w_prev": to_float(row[0] if row else 0),
        "w_last": to_float(row[1] if row else 0),
    }


def enrich_coverages(report: dict, all_totals: dict[str, dict[str, float]]) -> dict:
    for key, totals_key in (
        ("defects", "defects_total"),
        ("defects_org0", "defects_org0_total"),
        ("categories", "categories_total"),
        ("categories_org0", "categories_org0_total"),
    ):
        top = report[totals_key]
        full = all_totals[key]
        top_last = top["w_last"]
        all_last = full["w_last"]
        cover = (top_last / all_last * 100) if all_last else None
        top["all_w_prev"] = full["w_prev"]
        top["all_w_last"] = full["w_last"]
        top["top20_cover_last"] = cover
    return report


def build_kpi_payload(report: dict) -> dict[str, float]:
    d = report.get("defects_total", {}) or {}
    all_totals = report.get("all_totals") or {}
    all_defects = all_totals.get("defects") or {}
    all_org0 = all_totals.get("defects_org0") or {}
    total_last = to_float(all_defects.get("w_last"))
    org0_last = to_float(all_org0.get("w_last"))
    top20_last = to_float(d.get("w_last"))
    cover = d.get("top20_cover_last")
    if cover is None and total_last:
        cover = top20_last / total_last * 100
    avg4 = to_float(d.get("avg4"))
    vs_avg4 = None
    if avg4:
        vs_avg4 = (top20_last / avg4 - 1.0) * 100
    return {
        "total_last": total_last,
        "top20_last": top20_last,
        "org0_last": org0_last,
        "org0_share": (org0_last / total_last * 100) if total_last else 0.0,
        "top20_cover": to_float(cover) if cover is not None else 0.0,
        "top20_avg4": avg4,
        "top20_vs_avg4": to_float(vs_avg4) if vs_avg4 is not None else 0.0,
    }


def _avg4_weeks(weeks: list[int], week_last: int) -> list[int]:
    prior = sorted(w for w in weeks if w <= week_last)
    if not prior:
        return [week_last]
    return prior[-4:]


def _row_avg4(amounts: dict[int, float], avg_weeks: list[int]) -> float:
    if not avg_weeks:
        return 0.0
    return sum(to_float(amounts.get(w, 0)) for w in avg_weeks) / len(avg_weeks)


def _corpus_wh_map(cfg: dict | None = None) -> dict[int, list[int]]:
    cfg = cfg or load_config()
    out: dict[int, list[int]] = {1: [], 2: [], 3: []}
    for b in cfg.get("buildings") or []:
        bid = str(b.get("id") or "")
        if not bid.startswith("korpus_"):
            continue
        try:
            corpus = int(bid.split("_", 1)[1])
        except (TypeError, ValueError):
            continue
        if corpus not in out:
            continue
        ids: list[int] = []
        for raw in b.get("wh_ids") or []:
            try:
                ids.append(int(raw))
            except (TypeError, ValueError):
                continue
        out[corpus] = sorted(set(ids))
    if any(out.values()):
        return out
    for w in cfg.get("wh_catalog") or []:
        try:
            wid = int(w.get("wh_id"))
            corpus = int(w.get("corpus") or 0)
        except (TypeError, ValueError):
            continue
        if corpus in out:
            out[corpus].append(wid)
    for corpus in out:
        out[corpus] = sorted(set(out[corpus]))
    return out


def fetch_top_wh_amounts(
    *,
    wh_ids: list[int],
    office_id: int | None,
    year: int,
    week_no: int,
    limit: int = 3,
) -> list[dict[str, Any]]:
    if not wh_ids:
        return []
    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    clauses.append("wh_id = ANY(%s)")
    params.append(wh_ids)
    clauses.append(f"{src['week_expr']} = %s")
    params.append(week_no)
    where = " WHERE " + " AND ".join(clauses)
    sql = f"""
        SELECT wh_id, COALESCE(SUM({src['amount_expr']}), 0) AS amount
        FROM {src['table']}
        {where}
        GROUP BY wh_id
        ORDER BY amount DESC
        LIMIT %s
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, [*params, max(1, min(10, limit))])
        raw = cur.fetchall()
    return [
        {"wh_id": int(r[0]), "amount": to_float(r[1])}
        for r in raw
        if r[0] is not None
    ]


_DIM_BREAKDOWN_SPECS: dict[str, tuple[str, str]] = {
    "subject": ("COALESCE(NULLIF(BTRIM(subject_name), ''), '—')", "Предмет"),
    "owner": ("COALESCE(NULLIF(BTRIM(owner_product), ''), '—')", "Владелец"),
    "state": ("COALESCE(NULLIF(BTRIM(state_id), ''), '—')", "Статус"),
}


def fetch_dimension_breakdowns(
    *,
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
    limit: int = 10,
) -> dict[str, Any]:
    """TOP slices by subject / owner / state for the selected week pair (from norm view)."""
    cache_key = (
        f"dim_brk_v1:{year}:{week_prev}:{week_last}:{office_id}:"
        f"{','.join(map(str, wh_ids or []))}:{limit}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    _ensure_brak_data_norm()
    year_start = date.fromisocalendar(year, 1, 1)
    year_end = date.fromisocalendar(year + 1, 1, 1)
    clauses = ["date IS NOT NULL", "date >= %s", "date < %s"]
    params: list[Any] = [year_start, year_end]
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = "EXTRACT(WEEK FROM date)::int"
    lim = max(3, min(20, int(limit)))

    out: dict[str, Any] = {"week_prev": week_prev, "week_last": week_last, "dims": {}}
    with get_conn() as conn:
        cur = conn.cursor()
        for key, (dim_expr, label) in _DIM_BREAKDOWN_SPECS.items():
            sql = f"""
                WITH base AS (
                    SELECT
                        {dim_expr} AS name,
                        {week_expr} AS week_no,
                        amount
                    FROM {NORM_VIEW}
                    {where}
                      AND {week_expr} IN (%s, %s)
                ),
                agg AS (
                    SELECT
                        name,
                        COALESCE(SUM(amount) FILTER (WHERE week_no = %s), 0) AS amount_prev,
                        COALESCE(SUM(amount) FILTER (WHERE week_no = %s), 0) AS amount_last
                    FROM base
                    GROUP BY name
                ),
                tot AS (
                    SELECT COALESCE(SUM(amount_last), 0) AS total_last FROM agg
                )
                SELECT
                    a.name,
                    a.amount_prev,
                    a.amount_last,
                    CASE WHEN a.amount_prev > 0
                         THEN (a.amount_last - a.amount_prev) / a.amount_prev * 100
                         ELSE NULL END AS dynamics,
                    CASE WHEN t.total_last > 0
                         THEN a.amount_last / t.total_last * 100
                         ELSE 0 END AS share
                FROM agg a
                CROSS JOIN tot t
                ORDER BY a.amount_last DESC NULLS LAST, a.name
                LIMIT %s
            """
            cur.execute(sql, [*params, week_prev, week_last, week_prev, week_last, lim])
            rows = [
                {
                    "name": str(r[0]),
                    "amount_prev": to_float(r[1]),
                    "amount_last": to_float(r[2]),
                    "dynamics": None if r[3] is None else to_float(r[3]),
                    "share": to_float(r[4]),
                }
                for r in cur.fetchall()
            ]
            out["dims"][key] = {"label": label, "rows": rows}

    _cache_set(cache_key, out, CACHE_TTL_SEC)
    return out


def fetch_corpus_comparison(
    *,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
    cfg: dict | None = None,
) -> list[dict[str, Any]]:
    cfg = cfg or load_config()
    corpus_map = _corpus_wh_map(cfg)
    corpus_sig = ";".join(
        f"{c}:{','.join(map(str, ids))}" for c, ids in sorted(corpus_map.items())
    )
    cache_key = f"corpus_cmp_v2:{year}:{week_prev}:{week_last}:{office_id}:{corpus_sig}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    wh_names = {
        int(w["wh_id"]): str(w.get("name") or w["wh_id"])
        for w in (cfg.get("wh_catalog") or [])
        if w.get("wh_id") is not None
    }
    rows: list[dict[str, Any]] = []
    for corpus in (1, 2, 3):
        wh_ids = corpus_map.get(corpus) or []
        if not wh_ids:
            rows.append(
                {
                    "corpus": corpus,
                    "name": f"{corpus} корпус",
                    "wh_count": 0,
                    "amount_prev": 0.0,
                    "amount_last": 0.0,
                    "amount_org0_last": 0.0,
                    "dynamics": None,
                    "org0_share": 0.0,
                    "share_of_total": 0.0,
                    "top_wh": [],
                }
            )
            continue
        totals_all = fetch_totals_all(
            wh_ids=wh_ids,
            office_id=office_id,
            org0_only=False,
            year=year,
            week_prev=week_prev,
            week_last=week_last,
        )
        totals_org0 = fetch_totals_all(
            wh_ids=wh_ids,
            office_id=office_id,
            org0_only=True,
            year=year,
            week_prev=week_prev,
            week_last=week_last,
        )
        prev = to_float(totals_all.get("w_prev"))
        last = to_float(totals_all.get("w_last"))
        org0_last = to_float(totals_org0.get("w_last"))
        dyn = ((last - prev) / prev * 100) if prev else None
        top_wh = fetch_top_wh_amounts(
            wh_ids=wh_ids,
            office_id=office_id,
            year=year,
            week_no=week_last,
            limit=3,
        )
        for item in top_wh:
            item["name"] = wh_names.get(item["wh_id"], str(item["wh_id"]))
            item["share"] = (item["amount"] / last * 100) if last else 0.0
        rows.append(
            {
                "corpus": corpus,
                "name": f"{corpus} корпус",
                "wh_count": len(wh_ids),
                "amount_prev": prev,
                "amount_last": last,
                "amount_org0_last": org0_last,
                "dynamics": dyn,
                "org0_share": (org0_last / last * 100) if last else 0.0,
                "share_of_total": 0.0,
                "top_wh": top_wh,
            }
        )

    total_last = sum(r["amount_last"] for r in rows)
    for r in rows:
        r["share_of_total"] = (r["amount_last"] / total_last * 100) if total_last else 0.0

    _cache_set(cache_key, rows, CACHE_TTL_SEC)
    return rows


def build_growth_alerts(
    report: dict,
    *,
    min_dynamics: float = 15.0,
    min_vs_avg4: float = 20.0,
    min_amount: float = 50000.0,
    limit: int = 8,
) -> list[dict[str, Any]]:
    return _analytics_build_growth_alerts(
        report,
        min_dynamics=min_dynamics,
        min_vs_avg4=min_vs_avg4,
        min_amount=min_amount,
        limit=limit,
    )


def _rows_as_week_snapshot(rows: list[Any], week: int) -> list[dict[str, Any]]:
    """Normalize Row/dict list into churn-friendly snapshots for a week."""
    out: list[dict[str, Any]] = []
    for r in rows or []:
        if isinstance(r, dict):
            amounts = r.get("amounts") or {}
            amt = to_float(amounts.get(week, r.get("w_last", 0)))
            out.append(
                {
                    "row_id": r.get("row_id"),
                    "name": r.get("name") or "—",
                    "w_last": amt,
                    "amount": amt,
                }
            )
        else:
            amounts = getattr(r, "amounts", {}) or {}
            amt = to_float(amounts.get(week, 0))
            out.append(
                {
                    "row_id": getattr(r, "row_id", None),
                    "name": getattr(r, "name", None) or "—",
                    "w_last": amt,
                    "amount": amt,
                }
            )
    return out


def build_report_compare_and_churn(
    report: dict,
    *,
    week_prev: int,
    week_last: int,
    week_a: int | None = None,
    week_b: int | None = None,
    compare_limit: int = 12,
) -> dict[str, Any]:
    wa = week_a if week_a is not None else week_prev
    wb = week_b if week_b is not None else week_last
    defects = report.get("defects") or []
    categories = report.get("categories") or []
    compare = {
        "week_a": wa,
        "week_b": wb,
        "defects": build_period_compare(defects, wa, wb, kind="reason", limit=compare_limit),
        "categories": build_period_compare(
            categories, wa, wb, kind="category", limit=compare_limit
        ),
    }
    # Churn: TOP-20 by week_prev amounts vs TOP-20 by week_last (current report rows).
    prev_def = sorted(
        _rows_as_week_snapshot(defects, week_prev),
        key=lambda x: -x["w_last"],
    )[:20]
    last_def = sorted(
        _rows_as_week_snapshot(defects, week_last),
        key=lambda x: -x["w_last"],
    )[:20]
    # Also include rows that may only appear strongly in prev week via amounts on current top.
    # Re-rank from amounts on the same defect set is enough for UI; for fuller prev-top
    # we additionally sort all defect rows by prev week.
    all_by_prev = sorted(
        _rows_as_week_snapshot(defects, week_prev),
        key=lambda x: -x["w_last"],
    )
    # If report only has last-week TOP-20, prev ranking among them is approximate.
    # Prefer a dedicated prev bundle when available via report["_prev_defects"].
    if report.get("_prev_defects") is not None:
        prev_def = _rows_as_week_snapshot(report["_prev_defects"], week_prev)[:20]
        if not any(r.get("w_last") for r in prev_def):
            prev_def = [
                {
                    "row_id": r.get("row_id") if isinstance(r, dict) else getattr(r, "row_id", None),
                    "name": (r.get("name") if isinstance(r, dict) else getattr(r, "name", None))
                    or "—",
                    "w_last": to_float(
                        (r.get("w_last") if isinstance(r, dict) else getattr(r, "w_last", 0))
                        or 0
                    ),
                    "amount": to_float(
                        (r.get("w_last") if isinstance(r, dict) else getattr(r, "w_last", 0))
                        or 0
                    ),
                }
                for r in (report["_prev_defects"] or [])[:20]
            ]
    else:
        prev_def = all_by_prev[:20]

    churn = {
        "defects": build_top20_churn(prev_def, last_def, kind="reason"),
        "week_prev": week_prev,
        "week_last": week_last,
    }
    return {"compare": compare, "top20_churn": churn}


def fetch_yoy_totals(
    *,
    year: int,
    week_last: int,
    office_id: int | None,
    wh_ids: list[int] | None,
) -> dict[str, Any]:
    sorted_wh = sorted(wh_ids) if wh_ids else []
    cache_key = f"yoy:{year}:{week_last}:{office_id}:{','.join(map(str, sorted_wh))}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    def _week_total(y: int, w: int) -> float:
        src = _report_source()
        clauses: list[str] = []
        params: list[Any] = []
        if src["base_not_null_clause"]:
            clauses.append(src["base_not_null_clause"])
        if _use_report_matview() and src["table"] != NORM_VIEW:
            clauses.append(src["year_clause"])
            params.append(y)
        else:
            year_start = date.fromisocalendar(y, 1, 1)
            year_end = date.fromisocalendar(y + 1, 1, 1)
            clauses.append(src["year_clause"])
            params.extend([year_start, year_end])
        if office_id is not None:
            clauses.append("office_id = %s")
            params.append(office_id)
        if wh_ids:
            clauses.append("wh_id = ANY(%s)")
            params.append(wh_ids)
        clauses.append(f"{src['week_expr']} = %s")
        params.append(w)
        where = " WHERE " + " AND ".join(clauses)
        sql = f"SELECT COALESCE(SUM({src['amount_expr']}), 0) FROM {src['table']}{where}"
        with get_conn() as conn:
            cur = conn.cursor()
            cur.execute(sql, params)
            row = cur.fetchone()
        return to_float(row[0] if row else 0)

    current = _week_total(year, week_last)
    previous = _week_total(year - 1, week_last)
    payload = {
        "year": year,
        "prev_year": year - 1,
        "week": week_last,
        "amount": current,
        "amount_prev_year": previous,
        "yoy_pct": yoy_pct(current, previous),
    }
    _cache_set(cache_key, payload, CACHE_TTL_SEC)
    return payload


def fetch_org0_series(
    *,
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
    limit_weeks: int = 12,
) -> list[dict[str, Any]]:
    weekly = fetch_weekly_dynamics(
        year=year, office_id=office_id, wh_ids=wh_ids, top_n=1
    )
    weeks = weekly.get("weeks") or []
    if limit_weeks > 0:
        weeks = weeks[-limit_weeks:]
    return [
        {
            "week": w["week"],
            "amount_all": to_float(w.get("amount_all")),
            "amount_org0": to_float(w.get("amount_org0")),
            "org0_share": to_float(w.get("org0_share")),
        }
        for w in weeks
    ]


def fetch_search(
    *,
    q: str,
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
    limit: int = 15,
) -> dict[str, Any]:
    query = (q or "").strip()
    if len(query) < 2:
        raise QueryParamError("Укажите q (минимум 2 символа)")
    limit = max(1, min(30, limit))
    sorted_wh = sorted(wh_ids) if wh_ids else []
    cache_key = (
        f"search:{year}:{office_id}:{','.join(map(str, sorted_wh))}:"
        f"{query.lower()}:{limit}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    like = f"%{query}%"
    year_start = date.fromisocalendar(year, 1, 1)
    year_end = date.fromisocalendar(year + 1, 1, 1)
    clauses = ["date IS NOT NULL", "date >= %s AND date < %s"]
    params: list[Any] = [year_start, year_end]
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)

    results: list[dict[str, Any]] = []
    _ensure_brak_data_norm()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT reason_id, MAX(COALESCE(reason_descr, '—')) AS name,
                   COALESCE(SUM(amount), 0) AS amount
            FROM {NORM_VIEW}
            {where}
              AND (
                COALESCE(reason_descr, '') ILIKE %s
                OR CAST(reason_id AS text) ILIKE %s
              )
            GROUP BY reason_id
            ORDER BY amount DESC NULLS LAST
            LIMIT %s
            """,
            [*params, like, like, limit],
        )
        for r in cur.fetchall():
            results.append(
                {
                    "kind": "reason",
                    "reason_id": int(r[0]) if r[0] is not None else None,
                    "name": r[1] or "—",
                    "amount": to_float(r[2]),
                }
            )
        cur.execute(
            f"""
            SELECT COALESCE(parent_name, '—') AS name,
                   COALESCE(SUM(amount), 0) AS amount
            FROM {NORM_VIEW}
            {where}
              AND COALESCE(parent_name, '') ILIKE %s
            GROUP BY 1
            ORDER BY amount DESC
            LIMIT %s
            """,
            [*params, like, limit],
        )
        for r in cur.fetchall():
            results.append(
                {
                    "kind": "category",
                    "parent_name": r[0] or "—",
                    "name": r[0] or "—",
                    "amount": to_float(r[1]),
                }
            )
        cur.execute(
            f"""
            SELECT nm_id, MAX(COALESCE(title, '—')) AS title,
                   COALESCE(SUM(amount), 0) AS amount
            FROM {NORM_VIEW}
            {where}
              AND (
                COALESCE(title, '') ILIKE %s
                OR CAST(nm_id AS text) ILIKE %s
              )
            GROUP BY nm_id
            ORDER BY amount DESC NULLS LAST
            LIMIT %s
            """,
            [*params, like, like, limit],
        )
        for r in cur.fetchall():
            results.append(
                {
                    "kind": "nm",
                    "nm_id": int(r[0]) if r[0] is not None else None,
                    "name": r[1] or "—",
                    "amount": to_float(r[2]),
                }
            )

    results.sort(key=lambda x: -to_float(x.get("amount")))
    payload = {"q": query, "year": year, "results": results[:limit]}
    _cache_set(cache_key, payload, min(60, CACHE_TTL_SEC))
    return payload


def fetch_reason_card(
    *,
    reason_id: int | None = None,
    parent_name: str | None = None,
    office_id: int | None = None,
    wh_ids: list[int] | None = None,
    year: int,
    week_last: int | None = None,
    top_n: int = 5,
) -> dict[str, Any]:
    if reason_id is None and not (parent_name or "").strip():
        raise QueryParamError("Укажите reason_id или parent_name")

    sorted_wh = sorted(wh_ids) if wh_ids else []
    week_key = "latest" if week_last is None else str(week_last)
    cache_key = (
        f"reason_card:{year}:{week_key}:{office_id}:{reason_id}:"
        f"{parent_name or ''}:{','.join(map(str, sorted_wh))}:{top_n}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    if reason_id is not None:
        clauses.append("reason_id = %s")
        params.append(reason_id)
    if parent_name:
        clauses.append("COALESCE(parent_name, '—') = %s")
        params.append(parent_name.strip())
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]

    sql_weeks = f"""
        SELECT {week_expr}::int AS week_no,
               COALESCE(SUM({amount_expr}), 0) AS amount_all,
               COALESCE(SUM({amount_expr}) FILTER (WHERE cnt_org = 0), 0) AS amount_org0
        FROM {src['table']}
        {where}
        GROUP BY 1
        ORDER BY 1
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql_weeks, params)
        week_rows = cur.fetchall()
        weeks_payload = [
            {
                "week": int(r[0]),
                "amount_all": to_float(r[1]),
                "amount_org0": to_float(r[2]),
                "org0_share": (to_float(r[2]) / to_float(r[1]) * 100) if to_float(r[1]) else 0.0,
            }
            for r in week_rows
        ]
        resolved_week_last = week_last
        if resolved_week_last is None and weeks_payload:
            resolved_week_last = weeks_payload[-1]["week"]
        if resolved_week_last is None:
            resolved_week_last = 1
        week_last = resolved_week_last

        title = parent_name or f"reason_id={reason_id}"
        if reason_id is not None:
            cur.execute(
                f"""
                SELECT MAX(COALESCE(reason_descr, '—'))
                FROM {src['table']}
                {where}
                """,
                params,
            )
            title_row = cur.fetchone()
            if title_row and title_row[0]:
                title = str(title_row[0])

        # Top WH from matview/source for selected week.
        wh_clauses = [*clauses, f"{week_expr} = %s"]
        wh_params = [*params, week_last]
        cur.execute(
            f"""
            SELECT wh_id, COALESCE(SUM({amount_expr}), 0) AS amount
            FROM {src['table']}
            WHERE {' AND '.join(wh_clauses)}
            GROUP BY wh_id
            ORDER BY amount DESC
            LIMIT %s
            """,
            [*wh_params, max(1, min(20, top_n))],
        )
        top_wh = [
            {"wh_id": int(r[0]), "amount": to_float(r[1])}
            for r in cur.fetchall()
            if r[0] is not None
        ]

        # nm_id / brand from normalized base table (richer dimensions).
        _ensure_brak_data_norm()
        base_clauses = ["date IS NOT NULL"]
        base_params: list[Any] = []
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        base_clauses.append("date >= %s AND date < %s")
        base_params.extend([year_start, year_end])
        if office_id is not None:
            base_clauses.append("office_id = %s")
            base_params.append(office_id)
        if wh_ids:
            base_clauses.append("wh_id = ANY(%s)")
            base_params.append(wh_ids)
        if reason_id is not None:
            base_clauses.append("reason_id = %s")
            base_params.append(reason_id)
        if parent_name:
            base_clauses.append("COALESCE(parent_name, '—') = %s")
            base_params.append(parent_name.strip())
        base_clauses.append("EXTRACT(WEEK FROM date)::int = %s")
        base_params.append(week_last)
        base_where = " WHERE " + " AND ".join(base_clauses)
        cur.execute(
            f"""
            SELECT nm_id, COALESCE(MAX(title), '—') AS title,
                   COALESCE(SUM(amount), 0) AS amount
            FROM {NORM_VIEW}
            {base_where}
            GROUP BY nm_id
            ORDER BY amount DESC NULLS LAST
            LIMIT %s
            """,
            [*base_params, max(1, min(20, top_n))],
        )
        top_nm = [
            {
                "nm_id": int(r[0]) if r[0] is not None else None,
                "title": r[1] or "—",
                "amount": to_float(r[2]),
            }
            for r in cur.fetchall()
        ]
        cur.execute(
            f"""
            SELECT COALESCE(brand_name, '—') AS brand,
                   COALESCE(SUM(amount), 0) AS amount
            FROM {NORM_VIEW}
            {base_where}
            GROUP BY 1
            ORDER BY amount DESC
            LIMIT %s
            """,
            [*base_params, max(1, min(20, top_n))],
        )
        top_brands = [
            {"brand_name": r[0] or "—", "amount": to_float(r[1])}
            for r in cur.fetchall()
        ]

    last_amt = next((w["amount_all"] for w in weeks_payload if w["week"] == week_last), 0.0)
    last_org0 = next((w["amount_org0"] for w in weeks_payload if w["week"] == week_last), 0.0)
    avg_weeks = _avg4_weeks([w["week"] for w in weeks_payload], week_last)
    avg4 = (
        sum(w["amount_all"] for w in weeks_payload if w["week"] in avg_weeks) / len(avg_weeks)
        if avg_weeks
        else 0.0
    )
    cfg = load_config()
    wh_names = {
        int(w["wh_id"]): str(w.get("name") or w["wh_id"])
        for w in (cfg.get("wh_catalog") or [])
        if w.get("wh_id") is not None
    }
    for item in top_wh:
        item["name"] = wh_names.get(item["wh_id"], str(item["wh_id"]))
        item["share"] = (item["amount"] / last_amt * 100) if last_amt else 0.0

    org0_trend = [
        {
            "week": w["week"],
            "org0_share": to_float(w.get("org0_share")),
            "amount_all": to_float(w.get("amount_all")),
            "amount_org0": to_float(w.get("amount_org0")),
        }
        for w in weeks_payload
    ]

    payload = {
        "year": year,
        "week_last": week_last,
        "reason_id": reason_id,
        "parent_name": parent_name,
        "title": title,
        "weeks": weeks_payload,
        "kpis": {
            "amount_last": last_amt,
            "amount_org0_last": last_org0,
            "org0_share": (last_org0 / last_amt * 100) if last_amt else 0.0,
            "avg4": avg4,
            "vs_avg4": ((last_amt / avg4 - 1.0) * 100) if avg4 else None,
        },
        "concentration": {
            "top3_wh_share": concentration_shares(top_wh, total=last_amt, top_n=3),
            "top5_nm_share": concentration_shares(top_nm, total=last_amt, top_n=5),
        },
        "org0_trend": org0_trend,
        "top_wh": top_wh,
        "top_nm": top_nm,
        "top_brands": top_brands,
        "source": src["table"],
    }
    _cache_set(cache_key, payload, CACHE_TTL_SEC)
    return payload


def fetch_freshness_payload() -> dict[str, Any]:
    env_err = check_db_env()
    out: dict[str, Any] = {
        "ok": not bool(env_err),
        "db_env_error": env_err,
        "max_date": None,
        "row_count": None,
        "matview_enabled": _use_report_matview(),
        "matview_available": False,
        "matview_max_year": None,
        "matview_max_week": None,
        "matview_refresh_age_sec": None,
        "matview_refresh_note": "",
        "cache_ttl_sec": CACHE_TTL_SEC,
    }
    now = monotonic()
    with _MV_LOCK:
        if _MV_LAST_REFRESH_AT:
            out["matview_refresh_age_sec"] = round(now - _MV_LAST_REFRESH_AT, 1)
            out["matview_refresh_note"] = _MV_LAST_REFRESH_NOTE
    if env_err:
        return out
    try:
        cfg = load_config()
        stats = fetch_db_stats(cfg.get("office_id"), None)
        out["max_date"] = stats.get("max_date")
        out["row_count"] = stats.get("row_count")
        if _use_report_matview():
            try:
                _ensure_report_matview()
                with get_conn() as conn:
                    cur = conn.cursor()
                    cur.execute(
                        f"""
                        SELECT MAX(iso_year)::int, MAX(week_no)::int, MAX(max_date)::text
                        FROM {_matview_name()}
                        """
                    )
                    row = cur.fetchone()
                out["matview_available"] = True
                out["matview_max_year"] = int(row[0]) if row and row[0] is not None else None
                out["matview_max_week"] = int(row[1]) if row and row[1] is not None else None
                if row and row[2]:
                    out["max_date"] = out["max_date"] or row[2]
            except Exception as exc:
                out["matview_available"] = False
                out["matview_error"] = str(exc)
    except Exception as exc:
        out["ok"] = False
        out["error"] = str(exc)
    return out


def export_report_xlsx(report: dict, year: int, week_prev: int, week_last: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_prev = PatternFill("solid", fgColor="E2EFDA")
    fill_last = PatternFill("solid", fgColor="FFF2CC")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    fill_metric = PatternFill("solid", fgColor="F5F8FC")
    fill_heat_red = PatternFill("solid", fgColor="F8D7DA")
    fill_heat_green = PatternFill("solid", fgColor="D4EDDA")
    fill_heat_yellow = PatternFill("solid", fgColor="FFF3CD")
    fill_heat_share = PatternFill("solid", fgColor="D5EBEA")
    fill_pct_green = PatternFill("solid", fgColor="C6EFCE")
    fill_pct_red = PatternFill("solid", fgColor="FFC7CE")
    fill_pct_yellow = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_bold = Font(bold=True, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    def apply_heat(cell, value: float | None, mode: str) -> None:
        if value is None:
            return
        if mode == "dynamics":
            if value < -2:
                cell.fill = fill_heat_green
            elif value > 2:
                cell.fill = fill_heat_red
            else:
                cell.fill = fill_heat_yellow
        elif mode == "share":
            cell.fill = fill_heat_share
        else:
            v = value - 100
            if v < -2:
                cell.fill = fill_pct_green
            elif v > 2:
                cell.fill = fill_pct_red
            else:
                cell.fill = fill_pct_yellow

    def write_section(
        ws,
        title: str,
        rows: list[dict],
        total: dict,
        all_totals: dict[str, float],
        show_id: bool,
        name_header: str,
        weeks: list[int],
        week_prev: int,
        week_last: int,
    ) -> None:
        ws.append([title])
        row_title = ws.max_row
        ws.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=2 + len(weeks) + 4)
        c = ws.cell(row_title, 1)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center
        c.border = border

        # Исправлено: заголовок "Доля в ТОП-20" соответствует логике расчёта
        headers = [("ИД" if show_id else "№"), name_header, *[str(w) for w in weeks], f"Динамика {week_last} к {week_prev}", "Доля в ТОП-20", "Среднее 2 нед.", "% посл. к средней"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(hrow, col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border
            if col >= 3 + len(weeks):
                cell.fill = fill_header
            if col == 3 + weeks.index(week_prev):
                cell.fill = fill_prev
            if col == 3 + weeks.index(week_last):
                cell.fill = fill_last

        def write_data_row(r: dict, total_style: bool = False, label: str | None = None):
            row_idx = ws.max_row + 1
            c1 = ws.cell(row_idx, 1, label if label is not None else (r.get("row_id") if r.get("row_id") is not None else r.get("num")))
            c2 = ws.cell(row_idx, 2, r.get("name", ""))
            c1.alignment = align_center
            c2.alignment = align_left
            c1.border = c2.border = border
            if total_style:
                c1.font = c2.font = font_bold
                c1.fill = c2.fill = fill_total

            for i, w in enumerate(weeks, start=3):
                val = r.get("amounts", {}).get(w, 0.0)
                cw = ws.cell(row_idx, i, val)
                cw.number_format = "# ##0"
                cw.alignment = align_right
                cw.border = border
                if total_style:
                    cw.font = font_bold
                    cw.fill = fill_total
                elif w == week_prev:
                    cw.fill = fill_prev
                elif w == week_last:
                    cw.fill = fill_last

            base = 3 + len(weeks)
            c_dyn = ws.cell(row_idx, base, r.get("dynamics"))
            c_share = ws.cell(row_idx, base + 1, r.get("share"))
            c_avg = ws.cell(row_idx, base + 2, r.get("average"))
            c_pct = ws.cell(row_idx, base + 3, r.get("pct_vs_avg"))
            for c in (c_dyn, c_share, c_avg, c_pct):
                c.border = border
                c.alignment = align_right
                c.fill = fill_metric
            c_dyn.number_format = "0.00%"
            c_share.number_format = "0.00%"
            c_avg.number_format = "# ##0"
            c_pct.number_format = "0.00%"
            if c_dyn.value is not None:
                c_dyn.value = c_dyn.value / 100
            if c_share.value is not None:
                c_share.value = c_share.value / 100
            if c_pct.value is not None:
                c_pct.value = c_pct.value / 100

            apply_heat(c_dyn, r.get("dynamics"), "dynamics")
            apply_heat(c_share, r.get("share"), "share")
            apply_heat(c_pct, r.get("pct_vs_avg"), "pct_vs_avg")

            if total_style:
                for c in (c_dyn, c_share, c_avg, c_pct):
                    c.font = font_bold
                    c.fill = fill_total

        for r in rows:
            write_data_row(r)

        write_data_row(
            {
                "name": "Итого ТОП-20",
                "amounts": total.get("amounts", {}),
                "dynamics": total.get("dynamics"),
                "share": total.get("share"),
                "average": total.get("average"),
                "pct_vs_avg": total.get("pct_vs_avg"),
            },
            total_style=True,
            label="Итого",
        )

        all_prev = to_float(all_totals.get("w_prev"))
        all_last = to_float(all_totals.get("w_last"))
        all_amounts = {w: 0.0 for w in weeks}
        all_amounts[week_prev] = all_prev
        all_amounts[week_last] = all_last
        all_avg = (all_prev + all_last) / 2 if (all_prev or all_last) else 0.0
        all_dyn = ((all_last - all_prev) / all_prev * 100) if all_prev else None
        all_pct = (all_last / all_avg * 100) if all_avg else None
        cover = (total.get("w_last", 0) / all_last * 100) if all_last else None
        write_data_row(
            {
                "name": f"Итого по всем (покрытие ТОП-20: {fmt_pct(cover)})",
                "amounts": all_amounts,
                "dynamics": all_dyn,
                "share": None,
                "average": all_avg,
                "pct_vs_avg": all_pct,
            },
            total_style=True,
            label="Итого по всем",
        )
        ws.append([])

    weeks = report["weeks"]
    sections = [
        ("Дефект ТОП-20, рубли", report["defects"], report["defects_total"], report["all_totals"]["defects"], True, "Дефект"),
        ("Дефект ТОП-20, ORG 0, рубли", report["defects_org0"], report["defects_org0_total"], report["all_totals"]["defects_org0"], True, "Дефект"),
        ("ТОП-20 категорий, рубли", report["categories"], report["categories_total"], report["all_totals"]["categories"], False, "Категория"),
        ("ТОП-20 категорий, ORG 0, рубли", report["categories_org0"], report["categories_org0_total"], report["all_totals"]["categories_org0"], False, "Категория"),
    ]

    ws = wb.create_sheet("Дашборд")
    ws.append(["Отчет по браку", "", "Год", year, "Пред. неделя", week_prev, "Посл. неделя", week_last])
    for c in range(1, 9):
        cell = ws.cell(1, c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.alignment = align_center
        cell.border = border
    ws.append([])

    for sec in sections:
        write_section(ws, *sec, weeks=weeks, week_prev=week_prev, week_last=week_last)

    widths = {
        1: 12,
        2: 42,
    }
    for i in range(len(weeks)):
        widths[3 + i] = 12
    widths[3 + len(weeks)] = 16
    widths[4 + len(weeks)] = 12
    widths[5 + len(weeks)] = 14
    widths[6 + len(weeks)] = 18
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_nomenclature_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"

    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_bold = Font(bold=True, size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    latest_year = payload.get("latest_year")
    latest_week = payload.get("latest_week")
    ws.append(["Кол-во брака по номенклатуре", "", "Год", latest_year, "Неделя", latest_week])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    for c in range(1, 7):
        cell = ws.cell(1, c)
        cell.fill = fill_total
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border

    headers = ["Номенклатура", "1 корпус", "2 корпус", "3 корпус", "Итог"]
    ws.append(headers)
    hrow = ws.max_row
    for i in range(1, len(headers) + 1):
        cell = ws.cell(hrow, i)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    rows = payload.get("rows") or []
    for r in rows:
        ws.append(
            [
                r.get("nomenclature"),
                r.get("corpus_1", 0),
                r.get("corpus_2", 0),
                r.get("corpus_3", 0),
                r.get("total", 0),
            ]
        )

    totals = payload.get("totals") or {}
    ws.append(
        [
            "Итого",
            totals.get("corpus_1", 0),
            totals.get("corpus_2", 0),
            totals.get("corpus_3", 0),
            totals.get("total", 0),
        ]
    )
    tr = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(tr, c)
        cell.fill = fill_total
        cell.font = font_bold
        cell.border = border
        if c == 1:
            cell.alignment = align_center
        else:
            cell.number_format = "# ##0"
            cell.alignment = align_right

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 14
    ws.column_dimensions[get_column_letter(4)].width = 14
    ws.column_dimensions[get_column_letter(5)].width = 14

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_reason_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    fill_header = PatternFill("solid", fgColor="1F4E79")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)

    def style_header(ws, row: int, cols: int) -> None:
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    meta = wb.active
    meta.title = "KPI"
    k = payload.get("kpis") or {}
    conc = payload.get("concentration") or {}
    meta.append(["Поле", "Значение"])
    style_header(meta, 1, 2)
    for label, val in (
        ("title", payload.get("title")),
        ("year", payload.get("year")),
        ("week_last", payload.get("week_last")),
        ("reason_id", payload.get("reason_id")),
        ("parent_name", payload.get("parent_name")),
        ("amount_last", k.get("amount_last")),
        ("amount_org0_last", k.get("amount_org0_last")),
        ("org0_share", k.get("org0_share")),
        ("avg4", k.get("avg4")),
        ("vs_avg4", k.get("vs_avg4")),
        ("top3_wh_share", conc.get("top3_wh_share")),
        ("top5_nm_share", conc.get("top5_nm_share")),
    ):
        meta.append([label, val])

    def write_table(name: str, rows: list[dict], headers: list[tuple[str, str]]) -> None:
        ws = wb.create_sheet(name[:31])
        ws.append([h[1] for h in headers])
        style_header(ws, 1, len(headers))
        for r in rows or []:
            ws.append([r.get(h[0]) for h in headers])
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    write_table(
        "Weeks",
        payload.get("weeks") or [],
        [("week", "Неделя"), ("amount_all", "Всего"), ("amount_org0", "ORG0"), ("org0_share", "ORG0 %")],
    )
    write_table(
        "Top WH",
        payload.get("top_wh") or [],
        [("wh_id", "WH"), ("name", "Название"), ("amount", "Сумма"), ("share", "Доля %")],
    )
    write_table(
        "Top nm",
        payload.get("top_nm") or [],
        [("nm_id", "nm_id"), ("title", "Наименование"), ("amount", "Сумма")],
    )
    write_table(
        "Top brands",
        payload.get("top_brands") or [],
        [("brand_name", "Бренд"), ("amount", "Сумма")],
    )

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_weekly_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly"
    fill_header = PatternFill("solid", fgColor="1F4E79")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    headers = ["Неделя", "Всего", "ORG0", "ORG0 %", "Строк", "Топ причин"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for w in payload.get("weeks") or []:
        top = "; ".join(
            f"{t.get('reason_descr')} ({to_float(t.get('amount')):.0f})"
            for t in (w.get("top_reasons") or [])
        )
        ws.append(
            [
                w.get("week"),
                w.get("amount_all"),
                w.get("amount_org0"),
                w.get("org0_share"),
                w.get("row_count"),
                top,
            ]
        )
    for i, width in enumerate([10, 14, 14, 12, 10, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_detail_xlsx(args: Any, *, limit: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    where_sql, params = build_detail_where(args)
    cols = ", ".join(DETAIL_COL_NAMES)
    order_sql = _detail_order_sql(args)
    _ensure_brak_data_norm()
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {cols}
            FROM {NORM_VIEW}
            {where_sql}
            {order_sql}
            LIMIT %s
            """,
            [*params, limit],
        )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"

    fill_header = PatternFill("solid", fgColor="1F4E79")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    ws.append([label for _, label in DETAIL_COLUMNS])
    for col in range(1, len(DETAIL_COLUMNS) + 1):
        cell = ws.cell(1, col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = border

    numeric_cols = {
        "shk_id",
        "total_cost",
        "amount",
        "amount_obsh",
        "summa_obshay",
        "share",
        "office_id",
        "wh_id",
        "nm_id",
        "reason_id",
        "seller_id",
        "supplier_id",
        "cnt_org",
        "cnt_ors",
        "cnt_ocr",
    }
    for row in rows:
        ws.append(list(row))
        row_idx = ws.max_row
        for col_idx, col_name in enumerate(DETAIL_COL_NAMES, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            if col_name in numeric_cols:
                cell.alignment = align_right
                cell.number_format = (
                    "# ##0.00"
                    if col_name in ("amount", "amount_obsh", "summa_obshay", "total_cost", "share")
                    else "# ##0"
                )
            else:
                cell.alignment = align_left

    widths = {
        "shk_id": 14,
        "date": 16,
        "type": 12,
        "title": 42,
        "reason_descr": 34,
        "subject_name": 24,
        "parent_name": 24,
        "brand_name": 22,
        "owner_product": 24,
        "wh_name": 22,
        "source_file": 22,
        "amount_obsh": 14,
        "summa_obshay": 14,
    }
    for idx, col_name in enumerate(DETAIL_COL_NAMES, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def fetch_weekly_dynamics(
    *,
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
    top_n: int = 5,
) -> dict[str, Any]:
    sorted_wh = sorted(wh_ids) if wh_ids else []
    cache_key = f"weekly:{year}:{office_id}:{','.join(map(str, sorted_wh))}:{top_n}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    # Base table: raw rows. Matview: aggregated group rows (approx. volume).
    row_count_expr = "COUNT(*)::bigint"

    sql_weeks = f"""
        SELECT {week_expr}::int AS week_no,
               COALESCE(SUM({amount_expr}), 0) AS amount_all,
               COALESCE(SUM({amount_expr}) FILTER (WHERE cnt_org = 0), 0) AS amount_org0,
               {row_count_expr} AS row_count
        FROM {src["table"]}
        {where}
        GROUP BY 1
        ORDER BY 1
    """
    sql_top = f"""
        SELECT week_no, reason_id, reason_descr, amount_sum
        FROM (
            SELECT {week_expr}::int AS week_no,
                   reason_id,
                   MAX(COALESCE(reason_descr, '—')) AS reason_descr,
                   COALESCE(SUM({amount_expr}), 0) AS amount_sum,
                   ROW_NUMBER() OVER (
                       PARTITION BY {week_expr}::int
                       ORDER BY COALESCE(SUM({amount_expr}), 0) DESC, reason_id
                   ) AS rn
            FROM {src["table"]}
            {where}
            GROUP BY {week_expr}::int, reason_id
        ) t
        WHERE rn <= %s
        ORDER BY week_no, rn
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql_weeks, params)
        week_rows = cur.fetchall()
        cur.execute(sql_top, [*params, max(1, min(20, top_n))])
        top_rows = cur.fetchall()

    weeks = []
    for r in week_rows:
        amount_all = to_float(r[1])
        amount_org0 = to_float(r[2])
        weeks.append(
            {
                "week": int(r[0]),
                "amount_all": amount_all,
                "amount_org0": amount_org0,
                "org0_share": (amount_org0 / amount_all * 100) if amount_all else 0.0,
                "row_count": int(r[3] or 0),
            }
        )
    top_by_week: dict[int, list[dict[str, Any]]] = {}
    for r in top_rows:
        week = int(r[0])
        top_by_week.setdefault(week, []).append(
            {
                "reason_id": int(r[1]) if r[1] is not None else None,
                "reason_descr": r[2] or "—",
                "amount": to_float(r[3]),
            }
        )
    for w in weeks:
        w["top_reasons"] = top_by_week.get(w["week"], [])

    payload = {
        "year": year,
        "source": src["table"],
        "weeks": weeks,
        "totals": {
            "amount_all": sum(w["amount_all"] for w in weeks),
            "amount_org0": sum(w["amount_org0"] for w in weeks),
            "row_count": sum(w["row_count"] for w in weeks),
        },
    }
    _cache_set(cache_key, payload, WEEKS_CACHE_TTL_SEC)
    return payload


def fetch_reason_heatmap(
    *,
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
    top_n: int = 12,
    week_limit: int = 12,
) -> dict[str, Any]:
    """Matrix of top reasons × recent weeks for heatmap UI."""
    sorted_wh = sorted(wh_ids) if wh_ids else []
    top_n = max(3, min(20, top_n))
    week_limit = max(4, min(26, week_limit))
    cache_key = (
        f"heatmap:{year}:{office_id}:{','.join(map(str, sorted_wh))}:"
        f"{top_n}:{week_limit}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]

    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {week_expr}::int AS week_no
            FROM {src["table"]}
            {where}
            GROUP BY 1
            ORDER BY 1 DESC
            LIMIT %s
            """,
            [*params, week_limit],
        )
        weeks = sorted(int(r[0]) for r in cur.fetchall())
        if not weeks:
            payload = {"year": year, "weeks": [], "reasons": [], "cells": []}
            _cache_set(cache_key, payload, CACHE_TTL_SEC)
            return payload

        week_clause = f"{week_expr} = ANY(%s)"
        cur.execute(
            f"""
            SELECT reason_id, MAX(COALESCE(reason_descr, '—')) AS name,
                   COALESCE(SUM({amount_expr}), 0) AS amount
            FROM {src["table"]}
            {where} AND {week_clause}
            GROUP BY reason_id
            ORDER BY amount DESC NULLS LAST
            LIMIT %s
            """,
            [*params, weeks, top_n],
        )
        reasons = [
            {
                "reason_id": int(r[0]) if r[0] is not None else None,
                "name": r[1] or "—",
                "total": to_float(r[2]),
            }
            for r in cur.fetchall()
        ]
        reason_ids = [r["reason_id"] for r in reasons if r["reason_id"] is not None]
        cells: list[dict[str, Any]] = []
        if reason_ids:
            cur.execute(
                f"""
                SELECT reason_id, {week_expr}::int AS week_no,
                       COALESCE(SUM({amount_expr}), 0) AS amount
                FROM {src["table"]}
                {where}
                  AND {week_clause}
                  AND reason_id = ANY(%s)
                GROUP BY reason_id, {week_expr}::int
                """,
                [*params, weeks, reason_ids],
            )
            cells = [
                {
                    "reason_id": int(r[0]),
                    "week": int(r[1]),
                    "amount": to_float(r[2]),
                }
                for r in cur.fetchall()
            ]

    max_amount = max((c["amount"] for c in cells), default=0.0)
    for c in cells:
        c["intensity"] = (c["amount"] / max_amount) if max_amount else 0.0
    payload = {
        "year": year,
        "weeks": weeks,
        "reasons": reasons,
        "cells": cells,
        "max_amount": max_amount,
        "source": src["table"],
    }
    _cache_set(cache_key, payload, CACHE_TTL_SEC)
    return payload


def fetch_corpus_reasons(
    *,
    corpus: int,
    year: int,
    week_prev: int,
    week_last: int,
    office_id: int | None,
    cfg: dict | None = None,
    limit: int = 10,
) -> dict[str, Any]:
    cfg = cfg or load_config()
    corpus_map = _corpus_wh_map(cfg)
    wh_ids = corpus_map.get(int(corpus)) or []
    if not wh_ids:
        raise QueryParamError(f"Корпус {corpus}: нет WH в справочнике")
    limit = max(3, min(20, limit))
    cache_key = (
        f"corpus_reasons:{corpus}:{year}:{week_prev}:{week_last}:{office_id}:"
        f"{','.join(map(str, sorted(wh_ids)))}:{limit}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    weeks = sorted({week_prev, week_last})
    bundle = fetch_top_bundle(
        wh_ids=wh_ids,
        office_id=office_id,
        year=year,
        weeks=weeks,
        week_last=week_last,
        limit=limit,
    )
    rows = []
    for r in bundle.get("defects") or []:
        prev = to_float((r.amounts or {}).get(week_prev, 0))
        last = to_float((r.amounts or {}).get(week_last, 0))
        rows.append(
            {
                "reason_id": r.row_id,
                "name": r.name,
                "amount_prev": prev,
                "amount_last": last,
                "delta": last - prev,
                "dynamics": ((last / prev - 1.0) * 100) if prev else (100.0 if last else None),
            }
        )
    payload = {
        "corpus": int(corpus),
        "name": f"{corpus} корпус",
        "year": year,
        "week_prev": week_prev,
        "week_last": week_last,
        "wh_ids": wh_ids,
        "wh_count": len(wh_ids),
        "reasons": rows,
    }
    _cache_set(cache_key, payload, CACHE_TTL_SEC)
    return payload


def fetch_watchlist_status(
    *,
    reason_ids: list[int],
    year: int,
    week_prev: int,
    week_last: int,
    office_id: int | None,
    wh_ids: list[int] | None,
) -> dict[str, Any]:
    ids = sorted({int(x) for x in reason_ids if x is not None})[:20]
    if not ids:
        return {"year": year, "week_prev": week_prev, "week_last": week_last, "items": []}
    sorted_wh = sorted(wh_ids) if wh_ids else []
    cache_key = (
        f"watchlist:{year}:{week_prev}:{week_last}:{office_id}:"
        f"{','.join(map(str, sorted_wh))}:{','.join(map(str, ids))}"
    )
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    clauses.append("reason_id = ANY(%s)")
    params.append(ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    sql = f"""
        SELECT reason_id,
               MAX(COALESCE(reason_descr, '—')) AS name,
               COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS amount_prev,
               COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS amount_last
        FROM {src["table"]}
        {where}
        GROUP BY reason_id
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, [week_prev, week_last, *params])
        raw = cur.fetchall()
    by_id = {
        int(r[0]): {
            "reason_id": int(r[0]),
            "name": r[1] or "—",
            "amount_prev": to_float(r[2]),
            "amount_last": to_float(r[3]),
        }
        for r in raw
        if r[0] is not None
    }
    items = []
    for rid in ids:
        item = by_id.get(
            rid,
            {
                "reason_id": rid,
                "name": f"reason_id={rid}",
                "amount_prev": 0.0,
                "amount_last": 0.0,
            },
        )
        prev = to_float(item["amount_prev"])
        last = to_float(item["amount_last"])
        item["delta"] = last - prev
        item["dynamics"] = ((last / prev - 1.0) * 100) if prev else (100.0 if last else None)
        items.append(item)
    payload = {
        "year": year,
        "week_prev": week_prev,
        "week_last": week_last,
        "items": items,
    }
    _cache_set(cache_key, payload, CACHE_TTL_SEC)
    return payload


def fetch_status_payload() -> dict[str, Any]:
    env_err = check_db_env()
    env_states = {}
    for name in (
        "DATABASE_URL",
        "DB_HOST",
        "DB_PORT",
        "DB_NAME",
        "DB_USER",
        "DB_PASSWORD",
        "DB_SSLMODE",
    ):
        if name not in os.environ:
            env_states[name] = "missing"
        elif os.environ.get(name, "").strip() == "":
            env_states[name] = "empty"
        else:
            env_states[name] = "set"

    db_status = "error"
    db_detail = env_err or ""
    stats: dict[str, Any] = {}
    matview: dict[str, Any] = {
        "enabled": _use_report_matview(),
        "name": _matview_name(),
        "available": False,
        "bootstrap_ok_age_sec": None,
        "bootstrap_fail_age_sec": None,
        "bootstrap_fail_msg": "",
        "last_refresh_age_sec": None,
        "last_refresh_note": "",
    }
    now = monotonic()
    with _MV_LOCK:
        if _MV_BOOTSTRAP_OK_AT:
            matview["bootstrap_ok_age_sec"] = round(now - _MV_BOOTSTRAP_OK_AT, 1)
        if _MV_BOOTSTRAP_FAIL_AT:
            matview["bootstrap_fail_age_sec"] = round(now - _MV_BOOTSTRAP_FAIL_AT, 1)
            matview["bootstrap_fail_msg"] = _MV_BOOTSTRAP_FAIL_MSG
        if _MV_LAST_REFRESH_AT:
            matview["last_refresh_age_sec"] = round(now - _MV_LAST_REFRESH_AT, 1)
            matview["last_refresh_note"] = _MV_LAST_REFRESH_NOTE

    if not env_err:
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
                db_status = "ok"
                db_detail = "connected"
                cfg = load_config()
                stats = fetch_db_stats(cfg.get("office_id"), None)
                if _use_report_matview():
                    try:
                        _ensure_report_matview()
                        cur.execute(
                            f"""
                            SELECT COUNT(*)::bigint,
                                   MAX(iso_year)::int,
                                   MAX(week_no)::int
                            FROM {_matview_name()}
                            """
                        )
                        mv_row = cur.fetchone()
                        matview["available"] = True
                        matview["row_count"] = int(mv_row[0]) if mv_row else 0
                        matview["max_year"] = int(mv_row[1]) if mv_row and mv_row[1] is not None else None
                        matview["max_week"] = int(mv_row[2]) if mv_row and mv_row[2] is not None else None
                    except Exception as exc:
                        matview["available"] = False
                        matview["error"] = str(exc)
        except Exception as exc:
            db_status = "error"
            db_detail = str(exc)

    with _CACHE_LOCK:
        cache_entries = len(_CACHE)

    return {
        "status": "ok" if db_status == "ok" and not env_err else "degraded",
        "vercel_env": os.environ.get("VERCEL_ENV", "local"),
        "db_env_error": env_err,
        "env": env_states,
        "database": {"status": db_status, "detail": db_detail, **stats},
        "matview": matview,
        "cache": {
            "entries": cache_entries,
            "report_ttl_sec": CACHE_TTL_SEC,
            "weeks_ttl_sec": WEEKS_CACHE_TTL_SEC,
        },
        "admin": {
            "refresh_token_required": bool(
                (
                    os.environ.get("ADMIN_PASSWORD")
                    or os.environ.get("DB_REFRESH_TOKEN")
                    or ""
                ).strip()
            ),
            "active_sessions": len(_ADMIN_SESSIONS),
        },
    }


def fetch_available_weeks(
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
) -> list[int]:
    # Исправлено: сортируем wh_ids для корректного кэширования
    sorted_wh = sorted(wh_ids) if wh_ids else []
    key = f"weeks:{year}:{office_id}:{','.join(map(str, sorted_wh))}"
    cached = _cache_get(key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    sql = f"""
        SELECT DISTINCT {week_expr}::int AS w
        FROM {src["table"]}
        {where}
        ORDER BY w
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
    weeks = [int(r[0]) for r in rows]
    _cache_set(key, weeks, WEEKS_CACHE_TTL_SEC)
    return weeks


def fetch_nomenclature_counts_latest_week(
    *,
    office_id: int | None,
    year: int | None,
) -> dict[str, Any]:
    cache_key = f"nm_latest:{office_id}:{year if year is not None else 'all'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    cfg = load_config()
    wh_catalog = cfg.get("wh_catalog") or []
    wh_to_corpus: dict[int, int] = {}
    for w in wh_catalog:
        try:
            wid = int(w.get("wh_id"))
        except Exception:
            continue
        corpus_raw = w.get("corpus")
        try:
            corpus = int(corpus_raw) if corpus_raw is not None else 0
        except Exception:
            corpus = 0
        if corpus in (1, 2, 3):
            wh_to_corpus[wid] = corpus

    use_nm_mv = _use_report_matview()
    nm_src = {
        "table": NORM_VIEW,
        "week_expr": "EXTRACT(WEEK FROM date)::int",
        "isoyear_expr": "EXTRACT(ISOYEAR FROM date)::int",
        "year_clause": "date >= %s AND date < %s",
        "base_not_null_clause": "date IS NOT NULL",
        "rows_expr": "1",
    }
    if use_nm_mv:
        try:
            _ensure_nm_matview()
            nm_src = {
                "table": _nm_matview_name(),
                "week_expr": "week_no",
                "isoyear_expr": "iso_year",
                "year_clause": "iso_year = %s",
                "base_not_null_clause": "",
                "rows_expr": "rows_cnt",
            }
        except Exception as exc:
            print(
                f"[brak_data] nm matview unavailable, fallback to {NORM_VIEW}: {exc}",
                file=sys.stderr,
            )
    clauses: list[str] = []
    if nm_src["base_not_null_clause"]:
        clauses.append(nm_src["base_not_null_clause"])
    clauses.append("nm_id IS NOT NULL")
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if year is not None:
        if use_nm_mv and nm_src["table"] != NORM_VIEW:
            clauses.append(nm_src["year_clause"])
            params.append(year)
        else:
            y_start = date.fromisocalendar(year, 1, 1)
            y_end = date.fromisocalendar(year + 1, 1, 1)
            clauses.append(nm_src["year_clause"])
            params.extend([y_start, y_end])

    where = " WHERE " + " AND ".join(clauses)
    week_expr = nm_src["week_expr"]
    year_expr = nm_src["isoyear_expr"]
    rows_expr = nm_src["rows_expr"]
    sql = f"""
        WITH base AS (
            SELECT
                {year_expr} AS iso_year,
                {week_expr} AS week_no,
                wh_id,
                nm_id,
                {rows_expr}::bigint AS rows_cnt
            FROM {nm_src["table"]}
            {where}
        ),
        lw AS (
            SELECT iso_year, week_no
            FROM base
            GROUP BY iso_year, week_no
            ORDER BY iso_year DESC, week_no DESC
            LIMIT 1
        )
        SELECT b.nm_id, b.wh_id, COALESCE(SUM(b.rows_cnt), 0)::bigint AS cnt
        FROM base b
        JOIN lw ON lw.iso_year = b.iso_year AND lw.week_no = b.week_no
        GROUP BY b.nm_id, b.wh_id
        ORDER BY b.nm_id, b.wh_id
    """
    sql_week = f"""
        WITH base AS (
            SELECT
                {year_expr} AS iso_year,
                {week_expr} AS week_no
            FROM {nm_src["table"]}
            {where}
        )
        SELECT iso_year, week_no
        FROM base
        GROUP BY iso_year, week_no
        ORDER BY iso_year DESC, week_no DESC
        LIMIT 1
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql_week, params)
        wk = cur.fetchone()
        cur.execute(sql, params)
        rows = cur.fetchall()

    by_nm: dict[int, dict[str, int]] = {}
    for nm_id, wh_id, cnt in rows:
        nm = int(nm_id)
        corpus = wh_to_corpus.get(int(wh_id), 0)
        if nm not in by_nm:
            by_nm[nm] = {"c1": 0, "c2": 0, "c3": 0, "total": 0}
        if corpus == 1:
            by_nm[nm]["c1"] += int(cnt)
        elif corpus == 2:
            by_nm[nm]["c2"] += int(cnt)
        elif corpus == 3:
            by_nm[nm]["c3"] += int(cnt)
        by_nm[nm]["total"] += int(cnt)

    matrix_rows = [
        {
            "nomenclature": nm,
            "corpus_1": vals["c1"],
            "corpus_2": vals["c2"],
            "corpus_3": vals["c3"],
            "total": vals["total"],
        }
        for nm, vals in by_nm.items()
        if vals["total"] > 0
    ]
    matrix_rows.sort(key=lambda x: (-x["total"], x["nomenclature"]))

    totals = {
        "corpus_1": sum(r["corpus_1"] for r in matrix_rows),
        "corpus_2": sum(r["corpus_2"] for r in matrix_rows),
        "corpus_3": sum(r["corpus_3"] for r in matrix_rows),
        "total": sum(r["total"] for r in matrix_rows),
    }

    latest_year = int(wk[0]) if wk else None
    latest_week = int(wk[1]) if wk else None
    payload = {
        "latest_year": latest_year,
        "latest_week": latest_week,
        "rows": matrix_rows,
        "totals": totals,
    }
    _cache_set(cache_key, payload, NM_CACHE_TTL_SEC)
    return payload


def fetch_top_bundle(
    *,
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    weeks: list[int],
    week_last: int,
    limit: int = 20,
) -> dict[str, list[Row]]:
    if not weeks:
        weeks = [week_last]

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != NORM_VIEW:
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)

    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    week_cols = ",\n               ".join(
        f"COALESCE(SUM(amount) FILTER (WHERE week_no = %s), 0) AS w_{w}"
        for w in weeks
    )

    base_week_select = "week_no" if week_expr == "week_no" else f"{week_expr} AS week_no"
    sql = f"""
WITH base AS (
    SELECT reason_id,
           COALESCE(reason_descr, '—') AS reason_descr,
           COALESCE(parent_name, '—') AS parent_name,
           {base_week_select},
           cnt_org,
           {amount_expr} AS amount
    FROM {src["table"]}
    {where}
),
defects_all AS (
    SELECT reason_id AS row_id, MAX(reason_descr) AS name, {week_cols}
    FROM base
    GROUP BY reason_id
),
defects_org0 AS (
    SELECT reason_id AS row_id, MAX(reason_descr) AS name, {week_cols}
    FROM base
    WHERE cnt_org = 0
    GROUP BY reason_id
),
cats_all AS (
    SELECT NULL::int AS row_id, parent_name AS name, {week_cols}
    FROM base
    GROUP BY parent_name
),
cats_org0 AS (
    SELECT NULL::int AS row_id, parent_name AS name, {week_cols}
    FROM base
    WHERE cnt_org = 0
    GROUP BY parent_name
),
u AS (
    SELECT 'defects'::text AS bucket, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM defects_all
    UNION ALL
    SELECT 'defects_org0'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM defects_org0
    UNION ALL
    SELECT 'categories'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM cats_all
    UNION ALL
    SELECT 'categories_org0'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM cats_org0
),
r AS (
    SELECT *, ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY w_{week_last} DESC NULLS LAST) AS rn
    FROM u
)
SELECT bucket, row_id, name, {", ".join(f"w_{w}" for w in weeks)}
FROM r
WHERE rn <= %s
ORDER BY bucket, rn
"""
    qparams = [*params, *weeks, *weeks, *weeks, *weeks, limit]

    buckets: dict[str, list[Row]] = {
        "defects": [],
        "defects_org0": [],
        "categories": [],
        "categories_org0": [],
    }
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, qparams)
        raw = cur.fetchall()

    for rec in raw:
        bucket = str(rec[0])
        row_id = int(rec[1]) if rec[1] is not None else None
        name = str(rec[2])
        amounts = {w: to_float(rec[3 + i]) for i, w in enumerate(weeks)}
        buckets[bucket].append(Row(row_id=row_id, name=name, amounts=amounts))
    return buckets


def warm_report_cache_async(
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
) -> None:
    import threading

    def _worker() -> None:
        try:
            build_report_data(wh_ids, office_id, year, week_prev, week_last)
        except Exception:
            pass

    threading.Thread(target=_worker, daemon=True).start()


def add_shares(
    rows: list[Row],
    week_prev: int,
    week_last: int,
    *,
    avg_weeks: list[int] | None = None,
) -> list[dict]:
    """
    Доля считается от суммы ТОП-20. Сумма всех долей в таблице = 100%.
    """
    total_last = sum(r.amount(week_last) for r in rows)
    avg_weeks = avg_weeks or _avg4_weeks(sorted({week_prev, week_last}), week_last)
    out: list[dict] = []
    for i, r in enumerate(rows, start=1):
        share = (r.amount(week_last) / total_last * 100) if total_last else 0
        avg4 = _row_avg4(r.amounts, avg_weeks)
        last = r.amount(week_last)
        vs_avg4 = ((last / avg4 - 1.0) * 100) if avg4 else None
        out.append(
            {
                "num": i,
                "row_id": r.row_id,
                "name": r.name,
                "amounts": dict(r.amounts),
                "w_prev": r.amount(week_prev),
                "w_last": last,
                "dynamics": r.dynamics(week_prev, week_last),
                "share": share,
                "average": r.average(week_prev, week_last),
                "pct_vs_avg": r.pct_vs_avg(week_prev, week_last),
                "avg4": avg4,
                "vs_avg4": vs_avg4,
            }
        )
    return out


def totals(
    rows: list[dict],
    weeks: list[int],
    week_prev: int,
    week_last: int,
    *,
    avg_weeks: list[int] | None = None,
) -> dict:
    """
    Для итоговой строки ТОП-20 доля всегда 100%.
    """
    amounts = {w: sum(x["amounts"].get(w, 0) for x in rows) for w in weeks}
    w_prev = amounts.get(week_prev, 0)
    w_last = amounts.get(week_last, 0)
    avg = (w_prev + w_last) / 2 if rows else 0
    dyn = ((w_last - w_prev) / w_prev * 100) if w_prev else None
    pct_avg = (w_last / avg * 100) if avg else None
    avg_weeks = avg_weeks or _avg4_weeks(weeks, week_last)
    avg4 = _row_avg4(amounts, avg_weeks)
    vs_avg4 = ((w_last / avg4 - 1.0) * 100) if avg4 else None
    return {
        "amounts": amounts,
        "w_prev": w_prev,
        "w_last": w_last,
        "dynamics": dyn,
        "share": 100.0,  # Всегда 100% для итога ТОП-20
        "average": avg,
        "pct_vs_avg": pct_avg,
        "avg4": avg4,
        "vs_avg4": vs_avg4,
    }


def fmt_num(n: float) -> str:
    return f"{n:,.0f}".replace(",", " ")


def fmt_pct(n: float | None) -> str:
    if n is None:
        return "—"
    return f"{n:.2f}%"


def heat_style(value: float | None, mode: str) -> str:
    if value is None:
        return ""
    if mode == "dynamics":
        v = max(-50, min(50, value))
        if v <= 0:
            return (
                "background:rgba(63,143,107,.14);color:#3F8F6B"
                if v < -2
                else "background:rgba(196,138,42,.12);color:#C48A2A"
            )
        return (
            "background:rgba(196,75,90,.14);color:#C44B5A"
            if v > 2
            else "background:rgba(196,138,42,.12);color:#C48A2A"
        )
    if mode == "share":
        v = max(0, min(20, value))
        alpha = v / 20
        return f"background:rgba(47,125,122,{0.08 + alpha * 0.28})"
    v = value - 100
    v = max(-30, min(30, v))
    if v < -2:
        return "background:rgba(63,143,107,.14);color:#3F8F6B"
    if v > 2:
        return "background:rgba(196,75,90,.14);color:#C44B5A"
    return "background:rgba(196,138,42,.12);color:#C48A2A"


def _week_cell_style(week: int, week_prev: int, week_last: int) -> str:
    if week == week_last:
        return "background:rgba(47,125,122,.14);font-weight:600"
    if week == week_prev:
        return "background:rgba(63,143,107,.10)"
    return ""


def render_table(
    title: str,
    rows: list[dict],
    total: dict,
    *,
    show_id: bool,
    weeks: list[int],
    week_prev: int,
    week_last: int,
    name_header: str = "Наименование",
    all_totals: dict[str, float] | None = None,
    drill_kind: str | None = None,
    org0_only: bool = False,
) -> str:
    id_hdr = (
        "<th class='sticky col-id'>ИД</th>"
        if show_id
        else "<th class='sticky col-id'>№</th>"
    )
    week_hdrs = "".join(
        f"<th class='n week-col'>{w}</th>" for w in weeks
    )
    body = []
    for r in rows:
        id_cell = (
            f"<td class='c sticky col-id'>{r['row_id']}</td>"
            if show_id
            else f"<td class='c sticky col-id'>{r['num']}</td>"
        )
        week_cells = "".join(
            f"<td class='n week-col' style='{_e(_week_cell_style(w, week_prev, week_last))}'>"
            f"{fmt_num(r['amounts'].get(w, 0))}</td>"
            for w in weeks
        )
        drill_attrs = ""
        if drill_kind == "reason" and r.get("row_id") is not None:
            drill_attrs = (
                f' class="drill-row" data-drill="reason" data-reason-id="{_e(r["row_id"])}"'
                f' data-org0="{1 if org0_only else 0}" title="Открыть карточку причины"'
            )
        elif drill_kind == "category" and r.get("name"):
            drill_attrs = (
                f' class="drill-row" data-drill="category" data-parent-name="{_e(r["name"])}"'
                f' data-org0="{1 if org0_only else 0}" title="Открыть карточку причины"'
            )
        body.append(
            f"<tr{drill_attrs}>{id_cell}"
            f"<td class='name sticky col-name' title='{_e(r['name'])}'>{_e(r['name'])}</td>"
            f"{week_cells}"
            f"<td class='n metric' style='{_e(heat_style(r['dynamics'], 'dynamics'))}'>{fmt_pct(r['dynamics'])}</td>"
            f"<td class='n metric' style='{_e(heat_style(r['share'], 'share'))}'>{fmt_pct(r['share'])}</td>"
            f"<td class='n metric'>{fmt_num(r['average'])}</td>"
            f"<td class='n metric' style='{_e(heat_style(r['pct_vs_avg'], 'pct_vs_avg'))}'>{fmt_pct(r['pct_vs_avg'])}</td>"
            "</tr>"
        )

    t = total
    week_total_cells = "".join(
        f"<td class='n week-col' style='{_e(_week_cell_style(w, week_prev, week_last))}'>"
        f"<b>{fmt_num(t['amounts'].get(w, 0))}</b></td>"
        for w in weeks
    )
    body.append(
        f"<tr class='total'>"
        f"<td colspan='2' class='sticky col-id'><b>Итого</b></td>"
        f"{week_total_cells}"
        f"<td class='n metric' style='{_e(heat_style(t['dynamics'], 'dynamics'))}'><b>{fmt_pct(t['dynamics'])}</b></td>"
        f"<td class='n metric'><b>{fmt_pct(t['share'])}</b></td>"
        f"<td class='n metric'><b>{fmt_num(t['average'])}</b></td>"
        f"<td class='n metric' style='{_e(heat_style(t['pct_vs_avg'], 'pct_vs_avg'))}'><b>{fmt_pct(t['pct_vs_avg'])}</b></td>"
        f"</tr>"
    )
    if all_totals is not None:
        all_prev = to_float(all_totals.get("w_prev"))
        all_last = to_float(all_totals.get("w_last"))
        cover = (t["w_last"] / all_last * 100) if all_last else None
        all_amounts = {w: 0.0 for w in weeks}
        all_amounts[week_prev] = all_prev
        all_amounts[week_last] = all_last
        all_week_cells = "".join(
            f"<td class='n week-col'><b>{fmt_num(all_amounts.get(w, 0))}</b></td>"
            for w in weeks
        )
        all_avg = (all_prev + all_last) / 2 if (all_prev or all_last) else 0.0
        all_dyn = ((all_last - all_prev) / all_prev * 100) if all_prev else None
        all_pct = (all_last / all_avg * 100) if all_avg else None
        body.append(
            f"<tr class='total'>"
            f"<td colspan='2' class='sticky col-id'><b>Итого по всем</b></td>"
            f"{all_week_cells}"
            f"<td class='n metric'><b>{fmt_pct(all_dyn)}</b></td>"
            f"<td class='n metric'><b>—</b></td>"
            f"<td class='n metric'><b>{fmt_num(all_avg)}</b></td>"
            f"<td class='n metric'><b>{fmt_pct(all_pct)} · покрытие ТОП-20: {fmt_pct(cover)}</b></td>"
            f"</tr>"
        )

    # Исправлено: заголовок "Доля в ТОП-20" соответствует логике расчёта
    return f"""
<section class="panel">
  <h2>{_e(title)}</h2>
  <div class="table-scroll">
  <table>
    <colgroup>
      <col class="col-id">
      <col class="col-name">
    </colgroup>
    <thead>
      <tr>
        {id_hdr}
        <th class="sticky col-name">{_e(name_header)}</th>
        {week_hdrs}
        <th class="metric">Динамика {week_last} к {week_prev}</th>
        <th class="metric">Доля в ТОП-20</th>
        <th class="metric">Среднее за 2 нед.</th>
        <th class="metric">% посл. к средней</th>
      </tr>
    </thead>
    <tbody>{''.join(body)}</tbody>
  </table>
  </div>
</section>
"""


def _e(s: Any) -> str:
    import html

    return html.escape(str(s)) if s is not None else ""


def build_report_data(
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
    weeks: list[int] | None = None,
    show_all_weeks: bool = True,
) -> dict:
    # Исправлено: сортируем wh_ids и weeks для корректного кэширования
    sorted_wh = sorted(wh_ids) if wh_ids else []
    sorted_weeks = sorted(weeks) if weeks else []
    key = (
        f"report:{year}:{week_prev}:{week_last}:{office_id}:{int(show_all_weeks)}:"
        f"{','.join(map(str, sorted_wh))}:{','.join(map(str, sorted_weeks))}"
    )
    cached = _cache_get(key)
    if cached is not None:
        return cached

    if weeks is None:
        weeks = fetch_available_weeks(year, office_id, wh_ids)
    if not weeks:
        weeks = [week_prev, week_last]
    if week_prev not in weeks:
        weeks = sorted(set(weeks) | {week_prev})
    if week_last not in weeks:
        weeks = sorted(set(weeks) | {week_last})
    # Keep full history for 4-week average even when UI shows only 2 weeks.
    available_all = list(weeks)
    avg_weeks = _avg4_weeks(sorted(set(available_all) | {week_prev, week_last}), week_last)
    query_weeks = sorted(set(available_all) | set(avg_weeks))
    if not show_all_weeks:
        weeks = sorted({week_prev, week_last})

    bundle = fetch_top_bundle(
        wh_ids=wh_ids,
        office_id=office_id,
        year=year,
        weeks=query_weeks,
        week_last=week_last,
        limit=20,
    )
    # Separate TOP-20 ranked by previous week for accurate churn.
    prev_bundle = fetch_top_bundle(
        wh_ids=wh_ids,
        office_id=office_id,
        year=year,
        weeks=query_weeks,
        week_last=week_prev,
        limit=20,
    )
    defects = bundle["defects"]
    defects_org0 = bundle["defects_org0"]
    cats = bundle["categories"]
    cats_org0 = bundle["categories_org0"]

    d_rows = add_shares(defects, week_prev, week_last, avg_weeks=avg_weeks)
    d0_rows = add_shares(defects_org0, week_prev, week_last, avg_weeks=avg_weeks)
    c_rows = add_shares(cats, week_prev, week_last, avg_weeks=avg_weeks)
    c0_rows = add_shares(cats_org0, week_prev, week_last, avg_weeks=avg_weeks)

    # Totals/avg4 must use full history before display trimming.
    defects_total = totals(d_rows, query_weeks, week_prev, week_last, avg_weeks=avg_weeks)
    defects_org0_total = totals(d0_rows, query_weeks, week_prev, week_last, avg_weeks=avg_weeks)
    categories_total = totals(c_rows, query_weeks, week_prev, week_last, avg_weeks=avg_weeks)
    categories_org0_total = totals(c0_rows, query_weeks, week_prev, week_last, avg_weeks=avg_weeks)

    # Keep only display weeks in row amounts for HTML tables.
    if not show_all_weeks:
        for rows in (d_rows, d0_rows, c_rows, c0_rows):
            for row in rows:
                row["amounts"] = {w: row["amounts"].get(w, 0.0) for w in weeks}
        for total_row in (
            defects_total,
            defects_org0_total,
            categories_total,
            categories_org0_total,
        ):
            total_row["amounts"] = {w: total_row["amounts"].get(w, 0.0) for w in weeks}

    prev_defects_snap = []
    for r in prev_bundle.get("defects") or []:
        prev_defects_snap.append(
            {
                "row_id": r.row_id,
                "name": r.name,
                "w_last": to_float((r.amounts or {}).get(week_prev, 0)),
                "amount": to_float((r.amounts or {}).get(week_prev, 0)),
            }
        )

    report = {
        "weeks": weeks,
        "avg_weeks": avg_weeks,
        "defects": d_rows,
        "defects_total": defects_total,
        "defects_org0": d0_rows,
        "defects_org0_total": defects_org0_total,
        "categories": c_rows,
        "categories_total": categories_total,
        "categories_org0": c0_rows,
        "categories_org0_total": categories_org0_total,
        "_prev_defects": prev_defects_snap,
    }
    total_all = fetch_totals_all(
        wh_ids=wh_ids,
        office_id=office_id,
        org0_only=False,
        year=year,
        week_prev=week_prev,
        week_last=week_last,
    )
    total_org0 = fetch_totals_all(
        wh_ids=wh_ids,
        office_id=office_id,
        org0_only=True,
        year=year,
        week_prev=week_prev,
        week_last=week_last,
    )
    all_totals = {
        "defects": total_all,
        "defects_org0": total_org0,
        "categories": total_all,
        "categories_org0": total_org0,
    }
    report["all_totals"] = all_totals
    final_report = enrich_coverages(report, all_totals)
    _cache_set(key, final_report, CACHE_TTL_SEC)
    return final_report


SHARED_CSS = r"""
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=Manrope:wght@400;500;600;700&family=Source+Serif+4:opsz,wght@8..60,500;8..60,600&display=swap');
:root {
  --bg: #F3F6F8;
  --bg-soft: #EEF2F5;
  --surface: #FFFFFF;
  --surface-2: #F7FAFC;
  --text: #1E2A32;
  --muted: #6B7A86;
  --line: #D8E0E6;
  --line-soft: #E8EEF2;
  --primary: #2F7D7A;
  --primary-2: #3A9591;
  --primary-soft: rgba(47, 125, 122, 0.12);
  --ok: #3F8F6B;
  --ok-soft: rgba(63, 143, 107, 0.12);
  --warn: #C48A2A;
  --warn-soft: rgba(196, 138, 42, 0.12);
  --danger: #C44B5A;
  --danger-soft: rgba(196, 75, 90, 0.10);
  --shadow: 0 8px 28px rgba(30, 42, 50, 0.06);
  --shadow-sm: 0 2px 10px rgba(30, 42, 50, 0.04);
  --radius: 18px;
  --radius-sm: 10px;
  --pad: 16px;
  --font: "Manrope", sans-serif;
  --display: "Source Serif 4", Georgia, serif;
  --mono: "IBM Plex Mono", ui-monospace, monospace;
  --max: 1600px;
  --ease: .25s ease;
}
* { box-sizing: border-box; }
html, body { margin: 0; }
body {
  font: 13.5px/1.5 var(--font);
  color: var(--text);
  background:
    radial-gradient(900px 420px at 8% -8%, rgba(47,125,122,.10), transparent 55%),
    radial-gradient(720px 380px at 96% 0%, rgba(196,138,42,.07), transparent 50%),
    linear-gradient(180deg, #F7FAFC 0%, var(--bg) 40%, #EEF3F6 100%);
  min-height: 100vh;
}
a { color: var(--primary); text-decoration: none; transition: color var(--ease); }
a:hover { color: var(--primary-2); text-decoration: underline; }
.app-shell { min-height: 100vh; display: flex; flex-direction: column; width: 100%; }
.app-frame,
.app-top-inner,
.page-body,
.page,
.page-shell,
.wrap {
  width: 100%;
  max-width: var(--max);
  margin-left: auto;
  margin-right: auto;
  padding-left: var(--pad);
  padding-right: var(--pad);
  box-sizing: border-box;
}
.page-body .page,
.page-body .wrap,
.page-body .page-shell {
  max-width: none;
  margin: 0;
  padding-left: 0;
  padding-right: 0;
}
.app-top {
  background: rgba(255,255,255,.88);
  border-bottom: 1px solid var(--line-soft);
  position: sticky; top: 0; z-index: 40;
  backdrop-filter: blur(12px);
  width: 100%;
  box-shadow: var(--shadow-sm);
}
.app-top-inner { padding-top: 12px; padding-bottom: 0; }
.brand-row {
  display: flex; align-items: baseline; justify-content: space-between;
  gap: 12px; flex-wrap: wrap; padding-bottom: 10px; width: 100%;
}
.brand-row h1 {
  margin: 0; font-family: var(--display); font-size: 22px; font-weight: 600;
  letter-spacing: .01em; color: var(--text);
}
.brand-row .subtitle, .brand-row .header-subtitle {
  margin: 0; color: var(--muted); font-size: 12.5px; font-weight: 400;
}
.topnav {
  display: flex; gap: 2px; flex-wrap: wrap;
  margin: 0; padding: 0;
  border-top: 1px solid var(--line-soft);
  width: 100%;
}
.topnav a {
  color: var(--muted); text-decoration: none; font-weight: 500; font-size: 13px;
  padding: 10px 12px; border-bottom: 2px solid transparent; border-radius: 0;
  background: transparent; transition: color var(--ease), border-color var(--ease);
}
.topnav a:hover { color: var(--primary); background: transparent; text-decoration: none; }
.topnav a.active {
  color: var(--primary); border-bottom-color: var(--primary); background: transparent;
}
.page-body { padding-top: 16px; padding-bottom: 32px; flex: 1 1 auto; }
header.app-header { background: transparent; border: 0; padding: 0 0 12px; width: 100%; }
header.app-header h1 {
  margin: 0; font-family: var(--display); font-size: 26px; font-weight: 600;
  letter-spacing: .01em; color: var(--text);
}
header.app-header .subtitle, header.app-header .header-subtitle {
  margin-top: 4px; color: var(--muted); font-size: 13px; font-weight: 400;
}
.toolbar, .filters, .panel, .surface, .card, .insight-card, .presets, .table-tools, .freshness {
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: var(--radius);
  box-shadow: var(--shadow-sm);
  width: 100%;
  box-sizing: border-box;
  transition: background var(--ease), border-color var(--ease), transform var(--ease), box-shadow var(--ease);
}
.toolbar, .filters {
  display: flex; flex-wrap: wrap; gap: 12px; align-items: end;
  padding: 14px; margin: 0 0 12px;
}
.toolbar.main-toolbar {
  display: grid;
  grid-template-columns: minmax(280px, 2.2fr) minmax(220px, 1fr) minmax(180px, auto);
  align-items: stretch;
  gap: 14px;
  padding: 16px;
  background:
    linear-gradient(180deg, #FFFFFF 0%, #F9FBFC 100%);
  box-shadow: var(--shadow);
}
.toolbar.main-toolbar .actions { justify-content: flex-start; }
.toolbar > label, .toolbar .actions label, .filters > label, .filters label:not(.wh-grid label) {
  display: flex; flex-direction: column; gap: 5px;
  font-size: 11px; font-weight: 600; color: var(--muted);
  text-transform: uppercase; letter-spacing: .04em;
}
.toolbar fieldset label {
  display: flex; flex-direction: row; align-items: center; gap: 6px;
  font-size: 12px; font-weight: 500; color: var(--text);
  text-transform: none; letter-spacing: 0;
}
.toolbar input, .toolbar select, .filters input, .filters select,
input[type=text], input[type=number], select, textarea {
  height: 36px; border: 1px solid var(--line); border-radius: var(--radius-sm);
  padding: 0 10px; font: 13px var(--font); color: var(--text);
  background: #fff;
  transition: border-color var(--ease), background var(--ease), box-shadow var(--ease);
}
.toolbar input:focus, .toolbar select:focus, .filters input:focus, .filters select:focus,
input[type=text]:focus, input[type=number]:focus, select:focus, textarea:focus {
  outline: none; border-color: var(--primary);
  box-shadow: 0 0 0 3px var(--primary-soft);
}
textarea { height: auto; min-height: 56px; padding: 8px 10px; color: var(--text); background: #fff; }
.btn, .actions button, .presets button, .building-btns button, button {
  height: 36px; border: 1px solid var(--line); background: #fff; color: var(--text);
  border-radius: var(--radius-sm); padding: 0 12px; font: 600 12.5px/1 var(--font);
  cursor: pointer; white-space: nowrap;
  transition: background var(--ease), color var(--ease), border-color var(--ease), box-shadow var(--ease);
}
.btn:hover, .actions button:hover, .presets button:hover, .building-btns button:hover, button:hover {
  background: var(--bg-soft); border-color: #C5D0D8;
}
.btn.primary, .actions button.primary, button.primary {
  background: var(--primary); border-color: var(--primary); color: #fff;
  box-shadow: 0 4px 14px rgba(47,125,122,.22);
}
.btn.primary:hover, .actions button.primary:hover, button.primary:hover {
  background: var(--primary-2); border-color: var(--primary-2); color: #fff;
}
.btn.secondary, .actions button.secondary, button.secondary {
  background: var(--primary-soft); border-color: rgba(47,125,122,.28); color: var(--primary);
}
.btn.secondary:hover, .actions button.secondary:hover, button.secondary:hover {
  background: rgba(47,125,122,.18); border-color: var(--primary); color: var(--primary);
}
.btn.export, .actions button.export, button.export {
  background: #fff; border-color: var(--line); color: var(--muted);
}
.btn.export:hover, .actions button.export:hover, button.export:hover {
  border-color: var(--primary); color: var(--primary); background: var(--primary-soft);
}
.kpis {
  display: grid; grid-template-columns: repeat(6, minmax(0, 1fr));
  gap: 10px; margin: 0 0 12px; width: 100%;
}
.kpi, .kpis .kpi {
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: var(--radius);
  padding: 12px 13px; min-height: 82px;
  display: flex; flex-direction: column; justify-content: space-between;
  transition: transform var(--ease), border-color var(--ease);
}
.kpi:hover, .kpis .kpi:hover { border-color: rgba(47,125,122,.35); transform: translateY(-1px); }
.kpi .k {
  font-size: 10.5px; color: var(--muted); font-weight: 600;
  text-transform: uppercase; letter-spacing: .05em; line-height: 1.25;
}
.kpi .v {
  margin-top: 8px; font-size: 22px; font-weight: 600; font-family: var(--mono);
  letter-spacing: -0.02em; line-height: 1.1; word-break: break-word; color: var(--text);
}
.kpi .v.up, .dyn, .alert-item .dyn, .board .card .dyn { color: var(--danger); }
.kpi .v.down { color: var(--ok); }
.insight-grid, .grid {
  display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 0 0 12px; width: 100%;
}
#reportGrid.grid, .report-grid {
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 12px;
  width: 100%;
  margin: 0 0 12px;
  align-items: stretch;
}
#reportGrid .panel, .report-grid .panel {
  min-width: 0;
  min-height: 0;
  height: 420px;
  max-height: 420px;
  display: flex;
  flex-direction: column;
  overflow: hidden;
}
#reportGrid .panel > h2, .report-grid .panel > h2 {
  flex: 0 0 auto;
  margin: 0 0 10px;
  font-family: var(--display);
  font-size: 16px;
  font-weight: 600;
  letter-spacing: .01em;
  padding-bottom: 8px;
  border-bottom: 1px solid var(--line-soft);
  color: var(--text);
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}
#reportGrid .table-scroll, .report-grid .table-scroll {
  flex: 1 1 auto;
  min-height: 0;
  width: 100%;
  overflow: auto;
  overscroll-behavior: contain;
  border: 1px solid var(--line-soft);
  border-radius: var(--radius-sm);
  background: #fff;
}
#reportGrid .table-scroll table,
.report-grid .table-scroll table {
  width: max-content;
  min-width: 100%;
  table-layout: fixed;
  font-size: 12px;
}
#reportGrid .table-scroll th,
#reportGrid .table-scroll td,
.report-grid .table-scroll th,
.report-grid .table-scroll td {
  padding: 7px 8px;
  vertical-align: middle;
}
#reportGrid .table-scroll .col-id,
.report-grid .table-scroll .col-id {
  width: 52px;
  min-width: 52px;
  text-align: center;
}
#reportGrid .table-scroll .col-name,
.report-grid .table-scroll .col-name,
#reportGrid .table-scroll td.name,
.report-grid .table-scroll td.name {
  width: 168px;
  min-width: 168px;
  max-width: 168px;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}
#reportGrid .table-scroll .week-col,
.report-grid .table-scroll .week-col {
  width: 92px;
  min-width: 92px;
  white-space: nowrap;
}
#reportGrid .table-scroll .metric,
.report-grid .table-scroll .metric {
  width: 88px;
  min-width: 88px;
  white-space: nowrap;
}
#reportGrid .table-scroll .n,
.report-grid .table-scroll .n {
  white-space: nowrap;
  font-variant-numeric: tabular-nums;
}
#reportGrid .table-scroll table thead th,
.report-grid .table-scroll table thead th {
  position: sticky;
  top: 0;
  z-index: 2;
  background: var(--bg-soft);
  box-shadow: 0 1px 0 var(--line-soft);
  white-space: normal;
  line-height: 1.2;
}
.insight-card, .card, .panel { padding: 14px; margin: 0; overflow: hidden; }
.table-wrap, .heatmap-wrap { width: 100%; overflow-x: auto; }
.chips, .pager {
  display: flex; flex-wrap: wrap; gap: 8px; align-items: center;
  margin: 0 0 12px; width: 100%;
}
.pager .spacer { flex: 1 1 auto; }
.row .val { font-family: var(--mono); font-weight: 700; white-space: nowrap; }
.insight-card h3, .card h3, .panel h3, .card h2, .panel h2 {
  margin: 0 0 10px; font-family: var(--display); font-size: 17px; font-weight: 600;
  letter-spacing: .01em;
  padding-bottom: 8px; border-bottom: 1px solid var(--line-soft); color: var(--text);
}
.insight-card .body, .card .body { min-width: 0; }
.muted, .muted-box, .empty, .hint, .note { color: var(--muted); font-size: 12.5px; }
.muted-box, .empty, .note {
  padding: 10px; background: var(--bg-soft); border: 1px dashed var(--line);
  border-radius: var(--radius-sm);
}
.freshness {
  display: flex; flex-wrap: wrap; gap: 10px 14px; align-items: center;
  padding: 9px 12px; margin: 0 0 12px; font-size: 12px; color: var(--muted);
}
.freshness b { color: var(--text); font-family: var(--mono); font-weight: 600; }
.freshness .ok { color: var(--ok); font-weight: 700; }
.freshness .warn { color: var(--warn); font-weight: 700; }
.freshness .err { color: var(--danger); font-weight: 700; }
.presets, .table-tools, .meta {
  display: flex; flex-wrap: wrap; gap: 8px; align-items: center;
  padding: 10px 12px; margin: 0 0 12px; width: 100%; box-sizing: border-box;
  background: var(--surface); border: 1px solid var(--line); border-radius: var(--radius);
}
.presets .label { font-size: 11px; font-weight: 700; color: var(--muted); text-transform: uppercase; }
.alert-item .meta, .item .meta, .watch-item .meta, .row .meta {
  display: block; width: auto; padding: 0; margin: 2px 0 0;
  background: transparent; border: 0; border-radius: 0;
}
table {
  width: 100%; border-collapse: collapse; font-size: 12.5px;
  background: transparent; color: var(--text);
}
th, td {
  padding: 8px 8px; border-bottom: 1px solid var(--line-soft);
  text-align: left; vertical-align: top;
}
th {
  color: var(--muted); font-size: 11px; font-weight: 600;
  text-transform: uppercase; letter-spacing: .03em; background: var(--bg-soft);
}
td.num, th.num, .num { text-align: right; font-family: var(--mono); font-variant-numeric: tabular-nums; }
tr.drill-row { cursor: pointer; }
tr.drill-row:hover, .delta-table tr:hover, .watch-item:hover { background: var(--primary-soft); }
.delta-table { width: 100%; border-collapse: collapse; font-size: 12px; }
.delta-table th, .delta-table td { padding: 6px 7px; border-bottom: 1px solid var(--line-soft); }
.alert-list, .watch-list, .list { display: grid; gap: 0; width: 100%; }
.alert-item, .item, .watch-item, .churn-item {
  border: 0; border-bottom: 1px solid var(--line-soft); border-radius: 0;
  background: transparent; padding: 9px 0; cursor: default; width: 100%;
}
.row {
  display: flex; justify-content: space-between; align-items: flex-start;
  gap: 12px; width: 100%;
  border: 0; border-bottom: 1px solid var(--line-soft); border-radius: 0;
  background: transparent; padding: 9px 0; cursor: default;
}
.alert-item { cursor: pointer; transition: background var(--ease); }
.alert-item:hover { background: var(--danger-soft); }
.alert-item .title, .row .name, .watch-item b { font-weight: 600; font-size: 13px; color: var(--text); }
.alert-item .meta, .item .meta, .row .meta, .row .hint { margin-top: 2px; color: var(--muted); font-size: 12px; }
.alert-item .note-row select, .alert-item .note-row input {
  background: var(--bg-soft); color: var(--text); border-color: var(--line);
}
.thresholds { display:flex; flex-wrap:wrap; gap:8px; align-items:center; font-size:12px; color:var(--muted); margin-bottom:8px; }
.thresholds input { width:72px; }
.spark { display:flex; align-items:flex-end; gap:2px; height:22px; margin-top:6px; }
.spark i { display:block; width:5px; background: var(--primary); border-radius:1px; min-height:2px; }
.churn-cols, .churn-grid { display:grid; grid-template-columns:1fr 1fr 1fr; gap:10px; }
.churn-cols h4, .churn-col h4 {
  margin:0 0 6px; font-family: var(--display); font-size:12px; color:var(--muted);
  text-transform:uppercase; letter-spacing:.04em; font-weight: 500;
}
.corpus-bars { display:grid; gap:8px; }
.corpus-row {
  display:grid; grid-template-columns:110px 1fr 110px; gap:10px; align-items:center;
  padding:7px 4px; cursor:pointer; border-radius: 8px; transition: background var(--ease);
}
.corpus-row:hover { background: var(--primary-soft); }
.corpus-track { height: 8px; background: var(--line-soft); border-radius: 99px; overflow:hidden; }
.corpus-fill { height:100%; background: linear-gradient(90deg, #5AA8A4, var(--primary)); border-radius:99px; }
.corpus-meta { text-align:right; font-family:var(--mono); font-size:12px; font-weight:600; }
.corpus-stats { display:grid; grid-template-columns:repeat(3,1fr); gap:8px; margin-top:10px; }
.corpus-stat { border:1px solid var(--line); border-radius:var(--radius-sm); padding:8px; background:var(--bg-soft); }
.corpus-stat .k { font-size:10px; color:var(--muted); font-weight:700; text-transform:uppercase; }
.corpus-stat .v { margin-top:2px; font-size:14px; font-weight:700; font-family:var(--mono); color:var(--text); }
.dim-list { display:grid; gap:0; }
.dim-row {
  display:grid; grid-template-columns:1fr auto auto; gap:10px; align-items:center;
  padding:7px 0; border-bottom:1px solid var(--line-soft); font-size:13px;
}
.dim-row .name { font-weight:600; min-width:0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.dim-row .amt { font-family:var(--mono); font-weight:700; text-align:right; white-space:nowrap; }
.dim-row .dyn { font-family:var(--mono); font-size:12px; text-align:right; min-width:64px; }
.dim-row .share { color:var(--muted); font-size:11px; font-weight:600; }
.top-wh { margin-top:4px; font-size:11px; color:var(--muted); line-height:1.35; }
.search-box { position:relative; }
.search-drop {
  position:absolute; z-index:30; left:0; right:0; top:100%; background:var(--surface-2);
  border:1px solid var(--line); border-radius:var(--radius-sm); max-height:260px; overflow:auto; display:none;
}
.search-drop.open { display:block; }
.search-drop button {
  display:block; width:100%; text-align:left; border:0; background:transparent;
  height:auto; padding:8px 10px; border-radius:0; color: var(--text);
}
.search-drop button:hover { background: var(--primary-soft); }
.building-btns { display:flex; flex-wrap:wrap; gap:6px; margin-bottom:8px; }
.building-btns button.active, .presets button.active {
  background: var(--primary); color:#fff; border-color:var(--primary);
  box-shadow: 0 3px 10px rgba(47,125,122,.2);
}
.building-btns button.active:hover, .presets button.active:hover {
  background: var(--primary-2); color: #fff; border-color: var(--primary-2);
}
.wh-grid {
  display:grid; grid-template-columns:repeat(auto-fill,minmax(200px,1fr));
  gap:6px 10px; max-height:180px; overflow:auto; padding:6px 0;
  align-content: start;
}
.wh-grid .corpus-hdr {
  grid-column: 1 / -1; margin: 6px 0 2px; font-family: var(--display);
  font-size: 12px; font-weight: 500;
  color: var(--primary); text-transform: uppercase; letter-spacing: .04em;
}
.wh-grid label,
.wh-grid label.wh-item {
  display: flex !important;
  flex-direction: row !important;
  align-items: center !important;
  gap: 8px !important;
  min-height: 28px;
  margin: 0;
  padding: 2px 4px;
  font-size: 12px !important;
  font-weight: 500 !important;
  text-transform: none !important;
  letter-spacing: 0 !important;
  color: var(--text) !important;
  white-space: nowrap;
}
.wh-grid input[type=checkbox] {
  width: 15px; height: 15px; margin: 0; flex: 0 0 auto;
  accent-color: var(--primary);
}
.group {
  border:1px solid var(--line); border-radius:14px; padding:12px; margin:0;
  min-width:0; flex:1 1 280px; background: rgba(255,255,255,.72); width:100%;
  box-shadow: inset 0 1px 0 rgba(255,255,255,.8);
}
.toolbar.main-toolbar .group { flex: none; }
.group legend {
  padding:0 6px; font-size:11px; font-weight:700; color: var(--primary);
  text-transform:uppercase; letter-spacing: .04em;
}
.group.weeks label {
  display: flex; flex-direction: column; gap: 5px;
  font-size: 11px; font-weight: 700; color: var(--muted);
  text-transform: uppercase; letter-spacing: .04em;
}
.weeks-fields {
  display: grid;
  grid-template-columns: 1fr;
  gap: 10px;
  margin-bottom: 10px;
}
.group.weeks .hint {
  display: block;
  padding: 8px 10px;
  background: var(--bg-soft);
  border-radius: 8px;
  border: 1px solid var(--line-soft);
  line-height: 1.4;
}
.actions {
  display:flex; flex-direction:column; gap:10px; min-width:0;
  padding: 4px 2px;
  justify-content: space-between;
}
.actions-row { display:flex; flex-wrap:wrap; gap:6px; align-items:center; }
.actions-row.admin {
  margin-top: auto;
  padding-top: 10px;
  border-top: 1px dashed var(--line);
}
.actions-row .label { font-size:11px; color:var(--muted); font-weight:700; text-transform:uppercase; }
.modal-overlay {
  position:fixed; inset:0; background:rgba(30,42,50,.35); display:none;
  align-items:center; justify-content:center; z-index:80;
  backdrop-filter: blur(4px);
}
.modal-overlay.open { display:flex; }
.modal {
  width:min(420px,92vw); background:var(--surface); border:1px solid var(--line);
  border-radius:var(--radius); padding:16px; color: var(--text);
  box-shadow: var(--shadow);
}
.modal h3 { margin:0 0 10px; font-family: var(--display); font-size:18px; font-weight:600; color:var(--text); }
.modal .body { display:grid; gap:8px; }
.modal .row { border: 0; padding: 0; justify-content: flex-end; }
.bars { display:grid; grid-auto-flow:column; grid-auto-columns:minmax(18px,1fr); gap:3px; align-items:end; height:140px; }
.bar { background: var(--primary); border-radius: 3px 3px 1px 1px; min-height:2px; width:100%; }
.bar.org0 { background: var(--warn); }
.wlabel { font-size:10px; color:var(--muted); text-align:center; font-family:var(--mono); }
.heatmap { border-collapse: separate; border-spacing: 2px; font-size: 11px; }
.heatmap th, .heatmap td { padding: 0; text-align: center; min-width: 26px; height: 24px; }
.heatmap th { color: var(--muted); font-weight: 600; font-family: var(--mono); background: transparent; border:0; text-transform:none; }
.heatmap .rname {
  text-align:left; padding:0 8px 0 0; min-width:150px; max-width:210px;
  white-space:nowrap; overflow:hidden; text-overflow:ellipsis; color:var(--text);
  font-weight:600; cursor:pointer; border:0; background:transparent; text-transform:none;
}
.heatmap td.cell { border-radius: 3px; cursor: pointer; border:0; }
.board { display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:10px; width:100%; }
.col { background:var(--surface); border:1px solid var(--line); border-radius:var(--radius); padding:12px; min-height:280px; }
.col h3 {
  margin:0 0 10px; font-family: var(--display); font-size:15px; font-weight:600;
  display:flex; justify-content:space-between; border:0; padding:0; color:var(--text);
}
.col h3 .n { font-family:var(--mono); color:var(--muted); }
.board .card {
  border:1px solid var(--line); border-radius:var(--radius-sm);
  padding:10px; margin-bottom:8px; cursor:pointer; background:var(--bg-soft);
  transition: background var(--ease), border-color var(--ease);
}
.board .card:hover { background: #fff; border-color: rgba(47,125,122,.35); }
.board .card .t { font-weight:700; font-size:13px; color: var(--text); }
.board .card .m { margin-top:4px; font-size:12px; color:var(--muted); }
.board .card select, .board .card textarea {
  background: var(--surface-2); color: var(--text); border-color: var(--line);
}
.status, #status { margin: 0 0 12px; color: var(--muted); font-size: 12.5px; }
#reportGrid.loading, .loading { opacity: .55; pointer-events: none; }
#reportGrid table, .report-grid table {
  background: transparent; border: 0; border-radius: 0; width: 100%;
}
@media (max-width: 1100px) {
  .kpis { grid-template-columns: repeat(3, minmax(0, 1fr)); }
  .churn-cols, .churn-grid { grid-template-columns: 1fr; }
  .board { grid-template-columns: 1fr 1fr; }
  .toolbar.main-toolbar { grid-template-columns: 1fr 1fr; }
  .toolbar.main-toolbar .actions { grid-column: 1 / -1; }
  #reportGrid.grid, .report-grid { grid-template-columns: 1fr; }
}
@media (max-width: 800px) {
  .insight-grid, .grid, .kpis, .board, .corpus-stats { grid-template-columns: 1fr; }
  #reportGrid.grid, .report-grid { grid-template-columns: 1fr; }
  .corpus-row { grid-template-columns: 1fr; }
  .toolbar.main-toolbar { grid-template-columns: 1fr; }
}
@media print {
  .topnav, .toolbar, .no-print, .app-top, .scrolldown, .hero-band { display: none !important; }
  body { background: #fff; color: #111; }
  .card, .kpi, .panel, .toolbar, .insight-card { background: #fff; color: #111; border-color: #ccc; }
  .card, .kpi, .panel { break-inside: avoid; }
  [data-reveal] { opacity: 1 !important; transform: none !important; filter: none !important; }
}

/* —— Cocos-like motion / reveal —— */
.text-accent { color: var(--primary); }
.hero-band {
  position: relative;
  padding: 28px 0 36px;
  margin: 0 0 18px;
  border-bottom: 1px solid var(--line-soft);
  overflow: hidden;
}
.hero-band::before {
  content: "";
  position: absolute; inset: -20% -10% auto auto;
  width: 420px; height: 420px;
  background: radial-gradient(circle, rgba(47,125,122,.16), transparent 65%);
  pointer-events: none;
  animation: heroGlow 6s ease-in-out infinite alternate;
}
.hero-kicker {
  font-family: var(--font); font-size: 11px; font-weight: 700;
  letter-spacing: .14em; text-transform: uppercase; color: var(--primary);
  margin-bottom: 8px;
}
.hero-band h2 {
  margin: 0; font-family: var(--display); font-weight: 600;
  font-size: clamp(28px, 4vw, 42px); line-height: 1.08;
  letter-spacing: .01em; color: var(--text);
}
.hero-band .hero-lead {
  margin: 10px 0 0; max-width: 52ch; color: var(--muted); font-size: 14px; font-weight: 400;
}
.scrolldown {
  --color: var(--primary);
  display: inline-flex; flex-direction: column; align-items: center; gap: 6px;
  margin-top: 22px; color: var(--muted); text-decoration: none; font-size: 11px;
  text-transform: uppercase; letter-spacing: .08em;
}
.scrolldown .chev {
  width: 10px; height: 10px; border: solid var(--color);
  border-width: 0 2px 2px 0; transform: rotate(45deg);
  animation: chevronBounce 1.6s infinite;
}
.mod-section {
  position: relative;
  padding: 8px 0 22px;
  margin: 0 0 8px;
}
.mod-section::after {
  content: "";
  display: block; width: min(50%, 520px); height: 2px;
  margin: 22px auto 0; background: linear-gradient(90deg, transparent, var(--line), transparent);
}
.mod-section.no-rule::after { display: none; }
.section-title {
  margin: 0 0 14px; font-family: var(--display); font-size: clamp(22px, 3vw, 32px);
  font-weight: 600; line-height: 1.1; letter-spacing: .01em; color: var(--text);
}
.section-title .text-accent { color: var(--primary); }
.section-sub {
  margin: -6px 0 14px; color: var(--muted); font-size: 13px;
}
.insight-card, .card, .panel, .kpi, .col, .board .card {
  will-change: transform, opacity;
}
.insight-card:hover, .card:hover, .panel:hover {
  border-color: rgba(47,125,122,.28);
  box-shadow: 0 12px 32px rgba(30,42,50,.08);
  transform: translateY(-2px);
}
.kpi:hover, .kpis .kpi:hover {
  box-shadow: 0 10px 24px rgba(30,42,50,.08);
}
[data-reveal] {
  opacity: 0;
  transform: translateY(28px);
  filter: blur(2px);
  transition: opacity .7s ease, transform .7s ease, filter .7s ease;
  transition-delay: var(--reveal-delay, 0ms);
}
[data-reveal="fade-down"] { transform: translateY(-28px); }
[data-reveal="fade-left"] { transform: translateX(36px); }
[data-reveal="fade-right"] { transform: translateX(-36px); }
[data-reveal="flip-left"] {
  transform: perspective(900px) rotateY(18deg) translateY(16px);
  transform-origin: left center;
}
[data-reveal="zoom-in"] { transform: scale(.92); }
[data-reveal].is-in {
  opacity: 1;
  transform: none;
  filter: none;
}
.stagger > [data-reveal]:nth-child(1) { --reveal-delay: 0ms; }
.stagger > [data-reveal]:nth-child(2) { --reveal-delay: 80ms; }
.stagger > [data-reveal]:nth-child(3) { --reveal-delay: 160ms; }
.stagger > [data-reveal]:nth-child(4) { --reveal-delay: 240ms; }
.stagger > [data-reveal]:nth-child(5) { --reveal-delay: 320ms; }
.stagger > [data-reveal]:nth-child(6) { --reveal-delay: 400ms; }
.stagger > [data-reveal]:nth-child(7) { --reveal-delay: 480ms; }
.stagger > [data-reveal]:nth-child(8) { --reveal-delay: 560ms; }
@keyframes heroGlow {
  from { opacity: .55; transform: translate(0,0) scale(1); }
  to { opacity: 1; transform: translate(-24px, 18px) scale(1.08); }
}
@keyframes chevronBounce {
  0%, 100% { transform: rotate(45deg) translate(0,0); opacity: .4; }
  50% { transform: rotate(45deg) translate(4px, 4px); opacity: 1; }
}
@keyframes pulseAccent {
  0%, 100% { box-shadow: 0 0 0 0 rgba(47,125,122,.28); }
  50% { box-shadow: 0 0 0 8px rgba(47,125,122,0); }
}
.btn.primary { animation: pulseAccent 2.8s ease-in-out infinite; }
.btn.primary:hover { animation: none; }
.app-top { animation: fadeDownSoft .55s ease both; }
@keyframes fadeDownSoft {
  from { opacity: 0; transform: translateY(-12px); }
  to { opacity: 1; transform: none; }
}
@media (prefers-reduced-motion: reduce) {
  *, *::before, *::after {
    animation: none !important; transition: none !important;
  }
  [data-reveal] { opacity: 1 !important; transform: none !important; filter: none !important; }
}
"""


SHARED_JS = r"""
<script>
(function () {
  const reduced = window.matchMedia && window.matchMedia('(prefers-reduced-motion: reduce)').matches;
  let io = null;
  function observeAll(root) {
    const scope = root || document;
    const nodes = scope.querySelectorAll('[data-reveal]:not(.is-in)');
    if (reduced) {
      nodes.forEach((el) => el.classList.add('is-in'));
      return;
    }
    if (!io) {
      io = new IntersectionObserver((entries) => {
        entries.forEach((entry) => {
          if (!entry.isIntersecting) return;
          entry.target.classList.add('is-in');
          io.unobserve(entry.target);
        });
      }, { threshold: 0.12, rootMargin: '0px 0px -6% 0px' });
    }
    nodes.forEach((el) => io.observe(el));
  }
  window.revealRefresh = function (root) { observeAll(root); };
  if (document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', () => observeAll());
  } else {
    observeAll();
  }
  document.addEventListener('click', (e) => {
    const a = e.target.closest('a.scrolldown[href^="#"]');
    if (!a) return;
    const id = a.getAttribute('href').slice(1);
    const t = document.getElementById(id);
    if (!t) return;
    e.preventDefault();
    t.scrollIntoView({ behavior: reduced ? 'auto' : 'smooth', block: 'start' });
  });
})();
</script>
"""


DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Брак — ТОП-20</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Брак</h1>
  <div class="subtitle">ТОП-20 дефектов и категорий · корпуса · динамика</div>
</div>
<nav class="topnav">
  <a href="/" class="active">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Write-offs · live</div>
  <h2>ТОП-20 <span class="text-accent">брака</span></h2>
  <p class="hero-lead">Дефекты и категории по корпусам, динамика недель, алерты роста и покрытие — тот же отчёт, новый ритм интерфейса.</p>
  <a class="scrolldown" href="#filters">К фильтрам <span class="chev" aria-hidden="true"></span></a>
</section>

<section class="mod-section" id="filters">
  <h2 class="section-title" data-reveal="fade-right">Фильтры <span class="text-accent">и недели</span></h2>
  <p class="section-sub" data-reveal="fade-up">Выберите корпуса и две недели для расчёта динамики — остальное подтянется автоматически.</p>
  <div class="toolbar main-toolbar" data-reveal="fade-up">
  <fieldset class="group">
    <legend>Корпус / WH</legend>
    <div class="building-btns" id="buildingBtns"></div>
    <div class="wh-grid" id="whGrid"></div>
  </fieldset>
  <fieldset class="group weeks">
    <legend>Недели (ISO)</legend>
    <div class="weeks-fields">
      <label>Год <input type="number" id="year" value="2026"></label>
      <label>Пред. неделя <select id="weekPrev"></select></label>
      <label>Посл. неделя <select id="weekLast"></select></label>
    </div>
    <span class="hint">В таблице — все недели года. Динамика, доля и среднее считаются только по двум выбранным.</span>
  </fieldset>
  <div class="actions">
    <div class="actions-row">
      <button type="button" id="btnApply" class="primary">Применить</button>
      <button type="button" id="btnToggleWeeks" class="secondary">Все недели</button>
      <button type="button" id="btnExportXlsx" class="export">Экспорт XLSX</button>
    </div>
    <div class="actions-row">
      <button type="button" id="btnAllWh" class="secondary">Все WH</button>
      <button type="button" id="btnClearWh" class="export">Сбросить WH</button>
    </div>
    <div class="actions-row admin">
      <span class="label">Админ</span>
      <button type="button" id="btnAdminLogin" class="export">Вход</button>
      <button type="button" id="btnAdminLogout" class="export">Выход</button>
      <button type="button" id="btnRefreshData" class="secondary">Обновить данные</button>
    </div>
  </div>
  </div>
</section>

<section class="mod-section" id="kpisSection">
  <h2 class="section-title" data-reveal="fade-right">Ключевые <span class="text-accent">показатели</span></h2>
  <div class="kpis stagger" id="kpis">
  <div class="kpi" data-reveal="flip-left"><div class="k">Всего брак (посл. нед.)</div><div class="v" id="kpiTotal">—</div></div>
  <div class="kpi" data-reveal="flip-left"><div class="k">ТОП-20 (посл. нед.)</div><div class="v" id="kpiTop20">—</div></div>
  <div class="kpi" data-reveal="flip-left"><div class="k">ORG0 от общего</div><div class="v" id="kpiOrg0Share">—</div><div class="spark" id="org0Spark"></div></div>
  <div class="kpi" data-reveal="flip-left"><div class="k">Покрытие ТОП-20</div><div class="v" id="kpiCover">—</div></div>
  <div class="kpi" data-reveal="flip-left"><div class="k">ТОП-20 vs ср. 4 нед.</div><div class="v" id="kpiVsAvg4">—</div></div>
  <div class="kpi" data-reveal="flip-left"><div class="k">YoY (та же неделя)</div><div class="v" id="kpiYoy">—</div></div>
  </div>
  <div class="freshness" id="freshnessBar" data-reveal="fade-up">Свежесть данных: загрузка…</div>
</section>

<section class="mod-section" id="insights">
  <h2 class="section-title" data-reveal="fade-right">Аналитика <span class="text-accent">среза</span></h2>
  <div class="insight-grid stagger">
  <section class="insight-card" data-reveal="flip-left">
    <h3>Сравнение корпусов</h3>
    <div class="body" id="corpusCompare"><div class="muted-box">Загрузка…</div></div>
  </section>
  <section class="insight-card" data-reveal="flip-left">
    <h3>Алерты роста</h3>
    <div class="body">
      <div class="thresholds">
        <label>WoW % <input id="alertWow" type="number" value="15"></label>
        <label>vs ср.4 % <input id="alertVsAvg4" type="number" value="20"></label>
        <label>мин ₽ <input id="alertMinAmount" type="number" value="50000"></label>
        <label><input id="hideClosedAlerts" type="checkbox" checked> скрыть closed</label>
      </div>
      <div id="growthAlerts"><div class="muted-box">Загрузка…</div></div>
    </div>
  </section>
  </div>
  <div class="insight-grid stagger">
  <section class="insight-card" data-reveal="fade-left">
    <h3>Сравнение недель</h3>
    <div class="body" id="periodCompare"><div class="muted-box">Загрузка…</div></div>
  </section>
  <section class="insight-card" data-reveal="fade-left">
    <h3>Что изменило ТОП-20</h3>
    <div class="body" id="top20Churn"><div class="muted-box">Загрузка…</div></div>
  </section>
  </div>
  <div class="insight-grid stagger">
  <section class="insight-card" data-reveal="fade-up">
    <h3>Watchlist причин</h3>
    <div class="body" id="watchlistBox"><div class="muted-box">Пусто — добавьте из карточки причины</div></div>
  </section>
  <section class="insight-card" data-reveal="fade-up">
    <h3>Корпус: ТОП причин</h3>
    <div class="body" id="corpusDrill"><div class="muted-box">Кликните по корпусу слева сверху</div></div>
  </section>
  </div>
  <div class="insight-grid stagger">
  <section class="insight-card" data-reveal="fade-up">
    <h3>ТОП предметов</h3>
    <div class="body" id="dimSubject"><div class="muted-box">Загрузка…</div></div>
  </section>
  <section class="insight-card" data-reveal="fade-up">
    <h3>Владелец / статус</h3>
    <div class="body" id="dimOwnerState"><div class="muted-box">Загрузка…</div></div>
  </section>
  </div>
</section>

<section class="mod-section no-rule" id="reportSection">
  <h2 class="section-title" data-reveal="fade-right">Таблица <span class="text-accent">ТОП-20</span></h2>
  <div class="presets" id="presetsBar" data-reveal="fade-up">
  <span class="label">Пресеты:</span>
  <button type="button" data-preset="all">Все корпуса</button>
  <button type="button" data-preset="korpus_1">1 корпус</button>
  <button type="button" data-preset="korpus_2">2 корпус</button>
  <button type="button" data-preset="korpus_3">3 корпус</button>
  <button type="button" data-preset="latest">Последние 2 недели</button>
  <button type="button" id="btnSavePreset">Сохранить текущий</button>
  <button type="button" id="btnCopyLink">Копировать ссылку</button>
  <select id="reasonBookmarks" style="margin-left:8px;font:12px var(--font);padding:4px 8px;border:1px solid var(--line);border-radius:8px">
    <option value="">Закладки / недавние…</option>
  </select>
  </div>
  <div class="table-tools" data-reveal="fade-up">
  <label>Фильтр таблицы: <input type="text" id="tableSearch" placeholder="Дефект / категория / ID"></label>
  <div class="search-box" style="margin-left:12px">
    <label>Поиск в БД: <input type="text" id="globalSearch" placeholder="причина / категория / nm" autocomplete="off"></label>
    <div class="search-drop" id="searchDrop"></div>
  </div>
  </div>
  <div id="status" data-reveal="fade-up">Загрузка…</div>
  <div id="adminModal" class="modal-overlay" aria-hidden="true">
  <div class="modal" role="dialog" aria-modal="true" aria-labelledby="adminModalTitle">
    <h3 id="adminModalTitle">Вход администратора</h3>
    <div class="body">
      <label>Логин
        <input id="adminLoginInput" type="text" autocomplete="username" placeholder="admin">
      </label>
      <label>Пароль
        <input id="adminPasswordInput" type="password" autocomplete="current-password" placeholder="Введите пароль">
      </label>
      <div id="adminAuthError" class="hint"></div>
      <div class="row">
        <button type="button" id="btnAdminCancel" class="export">Отмена</button>
        <button type="button" id="btnAdminSubmit" class="primary">Войти</button>
      </div>
    </div>
  </div>
  </div>
  <div class="report-grid" id="reportGrid" data-reveal="zoom-in"></div>
</section>
<script>
const CONFIG = __CONFIG_JSON__;
const CATALOG = CONFIG.wh_catalog && CONFIG.wh_catalog.length
  ? CONFIG.wh_catalog
  : (CONFIG.wh_list || []).map(w => ({ wh_id: w.wh_id, name: String(w.wh_id), corpus: 0 }));
const ALL_WH_IDS = CATALOG.map(w => w.wh_id);

let selectedWh = new Set();
let activeBuilding = 'custom';
let availableWeeks = [];
let showAllWeeks = false;
let adminSessionId = '';
let isAdminSession = false;

function whLabel(id) {
  const w = CATALOG.find(x => x.wh_id === id);
  return w ? (id + ' — ' + w.name) : String(id);
}

function fmtInt(v) {
  return Number(v || 0).toLocaleString('ru-RU');
}

function fmtPct(v) {
  return Number(v || 0).toLocaleString('ru-RU', { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + '%';
}

function loadAlertThresholds() {
  try {
    const raw = localStorage.getItem('kurkuma_alert_thresholds');
    if (!raw) return;
    const t = JSON.parse(raw);
    if (t.wow != null) document.getElementById('alertWow').value = t.wow;
    if (t.vs_avg4 != null) document.getElementById('alertVsAvg4').value = t.vs_avg4;
    if (t.min_amount != null) document.getElementById('alertMinAmount').value = t.min_amount;
  } catch (_) {}
}
function saveAlertThresholds() {
  const t = {
    wow: Number(document.getElementById('alertWow').value) || 15,
    vs_avg4: Number(document.getElementById('alertVsAvg4').value) || 20,
    min_amount: Number(document.getElementById('alertMinAmount').value) || 50000,
  };
  localStorage.setItem('kurkuma_alert_thresholds', JSON.stringify(t));
  return t;
}
function alertNotesKey() {
  const year = document.getElementById('year').value;
  const { wl } = selectedWeeks();
  const wh = selectedWh.size ? Array.from(selectedWh).sort((a,b)=>a-b).join(',') : 'all';
  return `kurkuma_alert_notes:${year}:${wl}:${wh}`;
}
function loadAlertNotes() {
  try { return JSON.parse(localStorage.getItem(alertNotesKey()) || '{}') || {}; } catch (_) { return {}; }
}
function saveAlertNotes(notes) {
  localStorage.setItem(alertNotesKey(), JSON.stringify(notes || {}));
}
function reasonStorageGet(key) {
  try { return JSON.parse(localStorage.getItem(key) || '[]') || []; } catch (_) { return []; }
}
function reasonStorageSet(key, arr) {
  localStorage.setItem(key, JSON.stringify((arr || []).slice(0, 10)));
}
function pushReasonRecent(item) {
  const key = 'kurkuma_reason_recent';
  const list = reasonStorageGet(key).filter(x => x.href !== item.href);
  list.unshift(item);
  reasonStorageSet(key, list);
  renderReasonShortcuts();
}
function renderReasonShortcuts() {
  const sel = document.getElementById('reasonBookmarks');
  if (!sel) return;
  const bookmarks = reasonStorageGet('kurkuma_reason_bookmarks');
  const recent = reasonStorageGet('kurkuma_reason_recent');
  const opts = ['<option value="">Закладки / недавние…</option>'];
  if (bookmarks.length) {
    opts.push('<option disabled>—— Закладки ——</option>');
    bookmarks.forEach((b,i) => opts.push(`<option value="b:${i}">★ ${b.title}</option>`));
  }
  if (recent.length) {
    opts.push('<option disabled>—— Недавние ——</option>');
    recent.forEach((b,i) => opts.push(`<option value="r:${i}">${b.title}</option>`));
  }
  sel.innerHTML = opts.join('');
}

function setText(id, value) {
  const el = document.getElementById(id);
  if (el) el.textContent = value;
}
function updateKpis(kpis, yoy) {
  const d = kpis || {};
  setText('kpiTotal', fmtInt(d.total_last || 0));
  setText('kpiTop20', fmtInt(d.top20_last || 0));
  setText('kpiOrg0Share', fmtPct(d.org0_share || 0));
  setText('kpiCover', fmtPct(d.top20_cover || 0));
  setText('kpiVsAvg4', dynText(d.top20_vs_avg4));
  const y = yoy || {};
  setText('kpiYoy', y.yoy_pct == null ? 'нет данных' : dynText(y.yoy_pct));
}

function dynText(v) {
  if (v === null || v === undefined || Number.isNaN(Number(v))) return '—';
  const n = Number(v);
  const sign = n > 0 ? '+' : '';
  return sign + n.toLocaleString('ru-RU', { maximumFractionDigits: 1 }) + '%';
}

function ageText(sec) {
  if (sec == null) return '—';
  const s = Number(sec);
  if (s < 60) return Math.round(s) + ' с';
  if (s < 3600) return Math.round(s / 60) + ' мин';
  return (s / 3600).toFixed(1) + ' ч';
}

function updateFreshness(f) {
  const box = document.getElementById('freshnessBar');
  if (!box) return;
  if (!f) {
    box.innerHTML = 'Свежесть данных: нет данных';
    return;
  }
  const okCls = f.ok ? 'ok' : 'err';
  const mvCls = f.matview_available ? 'ok' : (f.matview_enabled ? 'warn' : '');
  box.innerHTML = `
    <span>Данные на: <b>${f.max_date || '—'}</b></span>
    <span>Строк: <b>${fmtInt(f.row_count || 0)}</b></span>
    <span class="${okCls}">DB ${f.ok ? 'ok' : 'error'}</span>
    <span class="${mvCls}">Matview ${f.matview_available ? `W${f.matview_max_week || '—'} / ${f.matview_max_year || '—'}` : (f.matview_enabled ? 'нет' : 'выкл')}</span>
    <span>Refresh: <b>${ageText(f.matview_refresh_age_sec)}</b></span>
    <a href="/status" style="color:var(--primary-2);font-weight:700;text-decoration:none">Статус →</a>
  `;
}

function renderDimRows(rows) {
  const list = Array.isArray(rows) ? rows : [];
  if (!list.length) return '<div class="muted-box">Нет данных</div>';
  return `<div class="dim-list">${list.map(r => {
    const dyn = r.dynamics;
    const dynCls = dyn == null ? '' : (dyn > 2 ? 'color:var(--danger)' : (dyn < -2 ? 'color:var(--ok)' : ''));
    return `<div class="dim-row">
      <div class="name" title="${String(r.name||'').replaceAll('"','&quot;')}">${r.name || '—'}
        <div class="share">${fmtPct(r.share || 0)} доли</div></div>
      <div class="amt">${fmtInt(r.amount_last)}</div>
      <div class="dyn" style="${dynCls}">${dynText(dyn)}</div>
    </div>`;
  }).join('')}</div>`;
}
function updateDimBreakdowns(payload) {
  const sub = document.getElementById('dimSubject');
  const own = document.getElementById('dimOwnerState');
  if (!sub || !own) return;
  const dims = (payload && payload.dims) || {};
  sub.innerHTML = renderDimRows((dims.subject || {}).rows);
  const ownerHtml = renderDimRows((dims.owner || {}).rows);
  const stateHtml = renderDimRows((dims.state || {}).rows);
  own.innerHTML = `<div class="muted" style="margin-bottom:6px;font-weight:700">Владелец (FBO/FBS/1P)</div>${ownerHtml}
    <div class="muted" style="margin:12px 0 6px;font-weight:700">Статус (state_id)</div>${stateHtml}`;
}
function updateCorpusCompare(rows) {
  const box = document.getElementById('corpusCompare');
  if (!box) return;
  const list = Array.isArray(rows) ? rows : [];
  if (!list.length) {
    box.innerHTML = '<div class="muted-box">Нет данных по корпусам</div>';
    return;
  }
  const maxLast = Math.max(1, ...list.map(r => Number(r.amount_last || 0)));
  const bars = list.map(r => {
    const pct = Math.max(2, Math.round(Number(r.amount_last || 0) / maxLast * 100));
    const dyn = r.dynamics;
    const dynCls = dyn == null ? '' : (dyn > 2 ? ' style="color:var(--danger)"' : (dyn < -2 ? ' style="color:var(--ok)"' : ''));
    const topWh = (r.top_wh || []).map(w => `${w.wh_id} ${w.name}: ${fmtInt(w.amount)}`).join('<br>');
    return `<div class="corpus-row" data-corpus="${r.corpus}" title="Открыть ТОП причин корпуса">
      <div class="name">${r.name}<div class="muted" style="font-weight:600">${r.wh_count || 0} WH · клик → ТОП</div></div>
      <div>
        <div class="corpus-track"><div class="corpus-fill" style="width:${pct}%"></div></div>
        <div class="top-wh">${topWh || 'нет WH'}</div>
      </div>
      <div class="corpus-meta">${fmtInt(r.amount_last)}<div${dynCls}>${dynText(dyn)}</div></div>
    </div>`;
  }).join('');
  const stats = list.map(r => `<div class="corpus-stat">
    <div class="k">${r.name}</div>
    <div class="v">${fmtPct(r.share_of_total || 0)}</div>
    <div class="muted">ORG0 ${fmtPct(r.org0_share || 0)}</div>
  </div>`).join('');
  box.innerHTML = `<div class="corpus-bars">${bars}</div><div class="corpus-stats">${stats}</div>`;
  box.querySelectorAll('.corpus-row[data-corpus]').forEach(el => {
    el.addEventListener('click', () => loadCorpusDrill(Number(el.dataset.corpus)));
  });
}

function watchlistGet() {
  try { return JSON.parse(localStorage.getItem('kurkuma_reason_watchlist') || '[]') || []; } catch (_) { return []; }
}
function watchlistSet(arr) {
  localStorage.setItem('kurkuma_reason_watchlist', JSON.stringify((arr || []).slice(0, 12)));
}
async function loadWatchlist() {
  const box = document.getElementById('watchlistBox');
  if (!box) return;
  const list = watchlistGet().filter(x => x.reason_id != null);
  if (!list.length) {
    box.innerHTML = '<div class="muted-box">Пусто — откройте карточку причины и нажмите «В watchlist»</div>';
    return;
  }
  const { wp, wl } = selectedWeeks();
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
    reason_ids: list.map(x => x.reason_id).join(','),
  });
  if (wh) q.set('wh_ids', wh);
  try {
    const r = await fetch('/api/watchlist?' + q);
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || r.statusText);
    const items = d.items || [];
    box.innerHTML = `<div class="watch-list">${items.map(it => {
      const title = it.name || ('reason ' + it.reason_id);
      return `<div class="watch-item">
        <div><b style="cursor:pointer" data-reason-id="${it.reason_id}">${title}</b>
          <div class="muted">${fmtInt(it.amount_prev)} → ${fmtInt(it.amount_last)}</div></div>
        <div style="font-family:var(--mono);font-weight:700;color:${Number(it.dynamics)>0?'var(--danger)':'var(--ok)'}">${dynText(it.dynamics)}</div>
        <button type="button" data-remove="${it.reason_id}">✕</button>
      </div>`;
    }).join('')}</div>`;
    box.querySelectorAll('[data-reason-id]').forEach(el => {
      el.addEventListener('click', () => openReasonCard({ reason_id: el.dataset.reasonId }));
    });
    box.querySelectorAll('[data-remove]').forEach(btn => {
      btn.addEventListener('click', () => {
        watchlistSet(watchlistGet().filter(x => String(x.reason_id) !== String(btn.dataset.remove)));
        loadWatchlist();
      });
    });
  } catch (e) {
    box.innerHTML = `<div class="muted-box">Ошибка watchlist: ${e.message || e}</div>`;
  }
}

async function loadCorpusDrill(corpus) {
  const box = document.getElementById('corpusDrill');
  if (!box || !corpus) return;
  const { wp, wl } = selectedWeeks();
  const q = new URLSearchParams({
    corpus: String(corpus),
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
  });
  box.innerHTML = '<div class="muted-box">Загрузка корпуса…</div>';
  try {
    const r = await fetch('/api/corpus/reasons?' + q);
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || r.statusText);
    const rows = d.reasons || [];
    if (!rows.length) {
      box.innerHTML = `<div class="muted-box">${d.name || ('Корпус '+corpus)}: нет данных</div>`;
      return;
    }
    box.innerHTML = `<div class="muted" style="margin-bottom:6px">${d.name} · W${d.week_prev}→${d.week_last} · ${d.wh_count} WH</div>
      <table class="delta-table"><thead><tr><th>Причина</th><th class="num">Δ ₽</th><th class="num">%</th></tr></thead>
      <tbody>${rows.map(r => `<tr data-reason-id="${r.reason_id}">
        <td>${r.name}</td><td class="num">${fmtInt(r.delta)}</td><td class="num">${dynText(r.dynamics)}</td>
      </tr>`).join('')}</tbody></table>`;
    box.querySelectorAll('tr[data-reason-id]').forEach(tr => {
      tr.addEventListener('click', () => openReasonCard({ reason_id: tr.dataset.reasonId }));
    });
  } catch (e) {
    box.innerHTML = `<div class="muted-box">Ошибка: ${e.message || e}</div>`;
  }
}

function updateOrg0Spark(series) {
  const box = document.getElementById('org0Spark');
  if (!box) return;
  const list = Array.isArray(series) ? series : [];
  if (!list.length) { box.innerHTML = ''; return; }
  const max = Math.max(1, ...list.map(x => Number(x.org0_share || 0)));
  box.innerHTML = list.map(x => {
    const h = Math.max(2, Math.round(Number(x.org0_share || 0) / max * 26));
    return `<i style="height:${h}px" title="W${x.week}: ${fmtPct(x.org0_share)}"></i>`;
  }).join('');
}

function updatePeriodCompare(compare) {
  const box = document.getElementById('periodCompare');
  if (!box) return;
  const c = compare || {};
  const rows = [...(c.defects || []), ...(c.categories || [])].slice(0, 12);
  if (!rows.length) {
    box.innerHTML = '<div class="muted-box">Нет данных для сравнения</div>';
    return;
  }
  box.innerHTML = `<div class="muted" style="margin-bottom:6px">Неделя ${c.week_a} → ${c.week_b}</div>
    <table class="delta-table"><thead><tr><th>Тип</th><th>Название</th><th class="num">Δ ₽</th><th class="num">%</th></tr></thead>
    <tbody>${rows.map(r => {
      const kind = r.kind === 'category' ? 'кат.' : 'деф.';
      const attrs = r.kind === 'reason' && r.row_id != null
        ? ` data-kind="reason" data-reason-id="${r.row_id}"`
        : (r.kind === 'category' ? ` data-kind="category" data-parent-name="${String(r.name).replaceAll('"','&quot;')}"` : '');
      return `<tr${attrs}><td>${kind}</td><td>${r.name}</td><td class="num">${fmtInt(r.delta)}</td><td class="num">${dynText(r.pct)}</td></tr>`;
    }).join('')}</tbody></table>`;
  box.querySelectorAll('tr[data-kind]').forEach(el => {
    el.addEventListener('click', () => {
      const params = {};
      if (el.dataset.kind === 'reason' && el.dataset.reasonId) params.reason_id = el.dataset.reasonId;
      if (el.dataset.kind === 'category' && el.dataset.parentName) params.parent_name = el.dataset.parentName;
      openReasonCard(params);
    });
  });
}

function updateTop20Churn(churn) {
  const box = document.getElementById('top20Churn');
  if (!box) return;
  const d = (churn && churn.defects) || {};
  const membershipChanged = Boolean(d.membership_changed);
  const leftTitle = membershipChanged ? 'Вошли' : '↑ в рейтинге';
  const midTitle = membershipChanged ? 'Вышли' : '↓ в рейтинге';
  const leftItems = membershipChanged ? (d.entered || []) : (d.rank_up || []);
  const midItems = membershipChanged ? (d.exited || []) : (d.rank_down || []);
  const leftMode = membershipChanged ? 'entered' : 'rank_up';
  const midMode = membershipChanged ? 'exited' : 'rank_down';
  const note = membershipChanged
    ? `<div class="muted" style="margin-bottom:6px">Состав ТОП-20: +${(d.entered||[]).length} / −${(d.exited||[]).length}</div>`
    : `<div class="muted" style="margin-bottom:6px">Состав ТОП-20 не изменился — показаны сдвиги ранга</div>`;
  const col = (title, items, mode) => {
    const list = items || [];
    if (!list.length) return `<div><h4>${title}</h4><div class="muted">нет</div></div>`;
    return `<div><h4>${title}</h4>${list.slice(0,8).map(r => {
      const attrs = r.row_id != null ? ` data-kind="reason" data-reason-id="${r.row_id}"` : '';
      let meta = '';
      if (mode === 'exited') meta = `#${r.rank_prev || '—'} · ${fmtInt(r.amount_prev)}`;
      else if (mode === 'entered') meta = `#${r.rank_last || '—'} · ${fmtInt(r.amount_last)}`;
      else if (mode === 'rank_up' || mode === 'rank_down') {
        meta = `#${r.rank_prev}→#${r.rank_last} · ${fmtInt(r.delta)}`;
      } else {
        const pct = r.amount_prev ? (r.delta / r.amount_prev * 100) : null;
        const rank = (r.rank_prev != null && r.rank_last != null) ? ` · #${r.rank_prev}→#${r.rank_last}` : '';
        meta = `${dynText(pct)} · ${fmtInt(r.delta)}${rank}`;
      }
      return `<div class="churn-item"${attrs}><b>${r.name}</b><div class="muted">${meta}</div></div>`;
    }).join('')}</div>`;
  };
  box.innerHTML = `${note}<div class="churn-cols">${col(leftTitle, leftItems, leftMode)}${col(midTitle, midItems, midMode)}${col('Остались (Δ)', d.stayed, 'stayed')}</div>`;
  box.querySelectorAll('[data-kind]').forEach(el => {
    el.addEventListener('click', () => openReasonCard({ reason_id: el.dataset.reasonId }));
  });
}

function updateGrowthAlerts(alerts, thresholds) {
  const box = document.getElementById('growthAlerts');
  if (!box) return;
  const t = thresholds || saveAlertThresholds();
  let list = Array.isArray(alerts) ? alerts : [];
  const notes = loadAlertNotes();
  const hideClosed = document.getElementById('hideClosedAlerts')?.checked;
  if (hideClosed) list = list.filter(a => (notes[a.alert_key || ''] || {}).status !== 'closed');
  if (!list.length) {
    box.innerHTML = `<div class="muted-box">Сильного роста нет (WoW ≥ +${t.wow}% или vs ср.4 ≥ +${t.vs_avg4}%, сумма ≥ ${fmtInt(t.min_amount)})</div>`;
    return;
  }
  box.innerHTML = `<div class="alert-list">${list.map(a => {
    const key = a.alert_key || ((a.kind==='reason'?'r:':'c:') + (a.row_id ?? a.name));
    const note = notes[key] || { status: 'new', comment: '' };
    const title = `${a.label}: ${a.name}`;
    const trigger = a.trigger === 'vs_avg4' ? 'vs ср.4' : 'WoW';
    const meta = `${fmtInt(a.w_prev)} → ${fmtInt(a.w_last)} · ${trigger} ${dynText(a.trigger === 'vs_avg4' ? a.vs_avg4 : a.dynamics)} · доля ${fmtPct(a.share || 0)}`;
    const attrs = a.kind === 'reason' && a.row_id != null
      ? ` data-kind="reason" data-reason-id="${a.row_id}"`
      : (a.kind === 'category' ? ` data-kind="category" data-parent-name="${String(a.name).replaceAll('"','&quot;')}"` : '');
    return `<div class="alert-item" data-alert-key="${String(key).replaceAll('"','&quot;')}"${attrs}>
      <div class="title">${title}</div>
      <div class="meta"><span class="dyn">${dynText(a.vs_avg4 != null ? a.vs_avg4 : a.dynamics)}</span> · ${meta}</div>
      <div class="note-row" onclick="event.stopPropagation()">
        <select data-role="status">
          <option value="new"${note.status==='new'?' selected':''}>new</option>
          <option value="watching"${note.status==='watching'?' selected':''}>watching</option>
          <option value="escalated"${note.status==='escalated'?' selected':''}>escalated</option>
          <option value="closed"${note.status==='closed'?' selected':''}>closed</option>
        </select>
        <input type="text" data-role="comment" maxlength="120" placeholder="комментарий" value="${String(note.comment||'').replaceAll('"','&quot;')}">
      </div>
    </div>`;
  }).join('')}</div>`;
  box.querySelectorAll('.alert-item').forEach(el => {
    const key = el.dataset.alertKey;
    const persist = () => {
      const n = loadAlertNotes();
      n[key] = {
        status: el.querySelector('[data-role=status]').value,
        comment: el.querySelector('[data-role=comment]').value || '',
      };
      saveAlertNotes(n);
    };
    el.querySelector('[data-role=status]').addEventListener('change', () => { persist(); if (hideClosed) updateGrowthAlerts(alerts, t); });
    el.querySelector('[data-role=comment]').addEventListener('change', persist);
    el.addEventListener('click', (ev) => {
      if (ev.target.closest('.note-row')) return;
      const params = {};
      if (el.dataset.kind === 'reason' && el.dataset.reasonId) params.reason_id = el.dataset.reasonId;
      if (el.dataset.kind === 'category' && el.dataset.parentName) params.parent_name = el.dataset.parentName;
      openReasonCard(params);
    });
  });
}

function currentFilterState() {
  const { wp, wl } = selectedWeeks();
  return {
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
    show_all_weeks: showAllWeeks,
    building: activeBuilding,
    wh_ids: Array.from(selectedWh),
  };
}

function applyFilterState(state) {
  if (!state) return;
  if (state.year) document.getElementById('year').value = state.year;
  if (Array.isArray(state.wh_ids)) {
    selectedWh = new Set(state.wh_ids.map(Number));
    document.querySelectorAll('#whGrid input').forEach(cb => {
      cb.checked = selectedWh.has(parseInt(cb.value, 10));
    });
  }
  if (state.building) {
    activeBuilding = state.building;
    syncBuildingButtons();
  }
  if (typeof state.show_all_weeks === 'boolean') {
    showAllWeeks = state.show_all_weeks;
    updateWeeksToggleLabel();
  }
  fillWeekSelects(
    availableWeeks.length ? availableWeeks : [Number(state.week_prev), Number(state.week_last)].filter(Boolean),
    Number(state.week_prev),
    Number(state.week_last)
  );
}

function loadSavedPresets() {
  try {
    return JSON.parse(localStorage.getItem('dashboardPresets') || '[]');
  } catch (_) {
    return [];
  }
}

function renderSavedPresets() {
  const bar = document.getElementById('presetsBar');
  if (!bar) return;
  bar.querySelectorAll('[data-saved-preset]').forEach(el => el.remove());
  loadSavedPresets().forEach((p, idx) => {
    const btn = document.createElement('button');
    btn.type = 'button';
    btn.dataset.savedPreset = String(idx);
    btn.textContent = p.name || ('Пресет ' + (idx + 1));
    btn.addEventListener('click', async () => {
      applyFilterState(p.state);
      await refreshWeeks();
      applyFilterState(p.state);
      loadReport();
    });
    bar.appendChild(btn);
  });
}

function openReasonCard(params) {
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wl } = selectedWeeks();
  const year = Number(document.getElementById('year').value) || new Date().getFullYear();
  const q = new URLSearchParams(params || {});
  if (wh) q.set('wh_ids', wh);
  q.set('year', String(year));
  q.set('week_last', String(wl));
  const href = '/reason?' + q.toString();
  const title = params.parent_name || (params.reason_id != null ? ('reason ' + params.reason_id) : 'карточка');
  pushReasonRecent({ title, href });
  window.location.href = href;
}

function dashboardQueryParams() {
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
    show_all_weeks: showAllWeeks ? '1' : '0',
  });
  if (wh) q.set('wh_ids', wh);
  if (activeBuilding && activeBuilding !== 'custom') q.set('building', activeBuilding);
  return q;
}

function syncDashboardUrl() {
  const q = dashboardQueryParams();
  history.replaceState(null, '', '/?' + q.toString());
}

function hydrateDashboardFromUrl() {
  const q = new URLSearchParams(window.location.search);
  if (!q.toString()) return false;
  if (q.get('year')) document.getElementById('year').value = q.get('year');
  if (q.get('show_all_weeks') === '1') showAllWeeks = true;
  if (q.get('show_all_weeks') === '0') showAllWeeks = false;
  updateWeeksToggleLabel();
  const building = q.get('building');
  if (building) {
    const b = (CONFIG.buildings || []).find(x => x.id === building);
    if (b) {
      selectedWh = new Set((b.wh_ids && b.wh_ids.length) ? b.wh_ids : ALL_WH_IDS);
      activeBuilding = b.id;
    }
  }
  if (q.get('wh_ids')) {
    const parsed = q.get('wh_ids').split(',').map(x => parseInt(String(x).trim(), 10)).filter(Number.isFinite);
    if (parsed.length) {
      selectedWh = new Set(parsed);
      activeBuilding = building || 'custom';
    }
  }
  return {
    week_prev: q.get('week_prev'),
    week_last: q.get('week_last'),
  };
}

function applyTableSearch() {
  const q = (document.getElementById('tableSearch').value || '').trim().toLowerCase();
  document.querySelectorAll('#reportGrid tbody tr').forEach(tr => {
    if (tr.classList.contains('total')) {
      tr.style.display = '';
      return;
    }
    if (!q) {
      tr.style.display = '';
      return;
    }
    tr.style.display = tr.textContent.toLowerCase().includes(q) ? '' : 'none';
  });
}

function init() {
  if (CONFIG.week_year) document.getElementById('year').value = CONFIG.week_year;
  adminSessionId = (localStorage.getItem('adminSessionId') || '').trim();
  isAdminSession = Boolean(adminSessionId);
  applyAdminUi();
  showAllWeeks = Boolean(CONFIG.show_all_weeks_default);
  updateWeeksToggleLabel();
  const grid = document.getElementById('whGrid');
  let lastCorpus = null;
  CATALOG.forEach(w => {
    if (w.corpus && w.corpus !== lastCorpus) {
      const hdr = document.createElement('div');
      hdr.className = 'corpus-hdr';
      hdr.textContent = w.corpus + ' корпус';
      grid.appendChild(hdr);
      lastCorpus = w.corpus;
    }
    const lab = document.createElement('label');
    const cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.value = w.wh_id;
    cb.checked = true;
    cb.addEventListener('change', () => {
      activeBuilding = 'custom';
      syncBuildingButtons();
      const id = parseInt(cb.value, 10);
      if (cb.checked) selectedWh.add(id); else selectedWh.delete(id);
    });
    selectedWh.add(w.wh_id);
    lab.append(cb, document.createTextNode(' ' + whLabel(w.wh_id)));
    grid.appendChild(lab);
  });

  const btns = document.getElementById('buildingBtns');
  CONFIG.buildings.forEach(b => {
    const btn = document.createElement('button');
    btn.type = 'button';
    btn.textContent = b.name;
    btn.dataset.id = b.id;
    btn.addEventListener('click', () => selectBuilding(b));
    btns.appendChild(btn);
  });

  document.getElementById('year').addEventListener('change', () => refreshWeeks().then(loadReport));
  document.getElementById('weekPrev').addEventListener('change', loadReport);
  document.getElementById('weekLast').addEventListener('change', loadReport);
  document.getElementById('btnAdminLogin').onclick = adminLogin;
  document.getElementById('btnAdminLogout').onclick = adminLogout;
  document.getElementById('btnAdminCancel').onclick = closeAdminModal;
  document.getElementById('btnAdminSubmit').onclick = submitAdminLogin;
  document.getElementById('adminModal').addEventListener('click', (e) => {
    if (e.target && e.target.id === 'adminModal') closeAdminModal();
  });
  document.getElementById('adminPasswordInput').addEventListener('keydown', (e) => {
    if (e.key === 'Enter') submitAdminLogin();
  });
  document.getElementById('btnToggleWeeks').onclick = () => {
    showAllWeeks = !showAllWeeks;
    updateWeeksToggleLabel();
    loadReport();
  };
  document.getElementById('btnClearWh').onclick = () => {
    selectedWh = new Set();
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = false);
    activeBuilding = 'custom';
    syncBuildingButtons();
    document.getElementById('status').textContent = 'Фильтр WH сброшен. Выберите нужный блок/склады и нажмите "Применить".';
  };
  document.getElementById('tableSearch').addEventListener('input', applyTableSearch);
  loadAlertThresholds();
  renderReasonShortcuts();
  ['alertWow','alertVsAvg4','alertMinAmount'].forEach(id => {
    document.getElementById(id).addEventListener('change', () => { saveAlertThresholds(); loadReport(); });
  });
  document.getElementById('hideClosedAlerts').addEventListener('change', () => loadReport());
  document.getElementById('reasonBookmarks').addEventListener('change', (e) => {
    const v = e.target.value || '';
    if (!v) return;
    const [kind, idx] = v.split(':');
    const list = reasonStorageGet(kind === 'b' ? 'kurkuma_reason_bookmarks' : 'kurkuma_reason_recent');
    const item = list[Number(idx)];
    e.target.value = '';
    if (item && item.href) window.location.href = item.href;
  });
  let searchTimer = null;
  const drop = document.getElementById('searchDrop');
  document.getElementById('globalSearch').addEventListener('input', (e) => {
    clearTimeout(searchTimer);
    const q = (e.target.value || '').trim();
    if (q.length < 2) { drop.classList.remove('open'); drop.innerHTML = ''; return; }
    searchTimer = setTimeout(async () => {
      try {
        const params = new URLSearchParams({ q, year: document.getElementById('year').value });
        const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
        if (wh) params.set('wh_ids', wh);
        const r = await fetch('/api/search?' + params);
        const data = await r.json();
        const rows = data.results || [];
        if (!rows.length) { drop.innerHTML = '<button type="button" disabled>Ничего не найдено</button>'; drop.classList.add('open'); return; }
        drop.innerHTML = rows.map(item => {
          const label = item.kind === 'nm'
            ? `nm ${item.nm_id}: ${item.name}`
            : (item.kind === 'category' ? `кат.: ${item.name}` : `деф. ${item.reason_id}: ${item.name}`);
          return `<button type="button" data-kind="${item.kind}" data-reason-id="${item.reason_id||''}" data-parent-name="${String(item.parent_name||'').replaceAll('"','&quot;')}" data-nm-id="${item.nm_id||''}">${label} · ${fmtInt(item.amount)}</button>`;
        }).join('');
        drop.classList.add('open');
        drop.querySelectorAll('button').forEach(btn => {
          btn.onclick = () => {
            if (btn.dataset.kind === 'nm' && btn.dataset.nmId) {
              window.location.href = '/details?nm_id=' + encodeURIComponent(btn.dataset.nmId);
              return;
            }
            const params = {};
            if (btn.dataset.kind === 'reason' && btn.dataset.reasonId) params.reason_id = btn.dataset.reasonId;
            if (btn.dataset.kind === 'category' && btn.dataset.parentName) params.parent_name = btn.dataset.parentName;
            openReasonCard(params);
          };
        });
      } catch (_) {
        drop.classList.remove('open');
      }
    }, 280);
  });
  document.addEventListener('click', (e) => {
    if (!e.target.closest('.search-box')) drop.classList.remove('open');
  });
  document.getElementById('btnApply').onclick = loadReport;
  document.getElementById('btnRefreshData').onclick = refreshData;
  document.getElementById('btnExportXlsx').onclick = exportXlsx;
  document.getElementById('btnAllWh').onclick = async () => {
    selectedWh = new Set(ALL_WH_IDS);
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = true);
    activeBuilding = 'all';
    syncBuildingButtons();
    await refreshWeeks();
    loadReport();
  };
  document.querySelectorAll('#presetsBar [data-preset]').forEach(btn => {
    btn.addEventListener('click', async () => {
      const key = btn.dataset.preset;
      if (key === 'all') {
        selectedWh = new Set(ALL_WH_IDS);
        activeBuilding = 'all';
      } else if (key.startsWith('korpus_')) {
        const b = (CONFIG.buildings || []).find(x => x.id === key);
        if (b) {
          selectedWh = new Set(b.wh_ids || []);
          activeBuilding = b.id;
        }
      } else if (key === 'latest') {
        await refreshWeeks(true);
      }
      document.querySelectorAll('#whGrid input').forEach(cb => {
        cb.checked = selectedWh.has(parseInt(cb.value, 10));
      });
      syncBuildingButtons();
      await refreshWeeks(key === 'latest');
      loadReport();
    });
  });
  document.getElementById('btnSavePreset').onclick = () => {
    const name = prompt('Название пресета', 'Мой фильтр');
    if (!name) return;
    const list = loadSavedPresets();
    list.push({ name: name.trim(), state: currentFilterState() });
    localStorage.setItem('dashboardPresets', JSON.stringify(list.slice(-8)));
    renderSavedPresets();
  };
  document.getElementById('btnCopyLink').onclick = async () => {
    syncDashboardUrl();
    try {
      await navigator.clipboard.writeText(window.location.href);
      document.getElementById('status').textContent = 'Ссылка на текущий срез скопирована.';
    } catch (_) {
      prompt('Скопируйте ссылку среза', window.location.href);
    }
  };
  renderSavedPresets();

  const urlState = hydrateDashboardFromUrl();
  document.querySelectorAll('#whGrid input').forEach(cb => {
    cb.checked = selectedWh.has(parseInt(cb.value, 10));
  });
  syncBuildingButtons();
  loadFreshness();
  refreshWeeks(!(urlState && (urlState.week_prev || urlState.week_last))).then(() => {
    if (urlState && urlState.week_prev) document.getElementById('weekPrev').value = urlState.week_prev;
    if (urlState && urlState.week_last) document.getElementById('weekLast').value = urlState.week_last;
    if (urlState) {
      loadReport();
    } else if (CONFIG.buildings.length) {
      selectBuilding(CONFIG.buildings[0]);
    } else {
      loadReport();
    }
  });
}

function applyAdminUi() {
  const btnRefresh = document.getElementById('btnRefreshData');
  const btnLogin = document.getElementById('btnAdminLogin');
  const btnLogout = document.getElementById('btnAdminLogout');
  const needAuth = Boolean(CONFIG.refresh_token_required);
  if (btnRefresh) btnRefresh.classList.toggle('hidden', !isAdminSession);
  if (btnLogin) btnLogin.disabled = isAdminSession;
  if (btnLogout) btnLogout.disabled = !isAdminSession;
  void needAuth;
}

function adminLogin() {
  openAdminModal();
}

function adminLogout() {
  const status = document.getElementById('status');
  const sid = adminSessionId;
  adminSessionId = '';
  localStorage.removeItem('adminSessionId');
  isAdminSession = false;
  applyAdminUi();
  fetch('/api/admin/logout', {
    method: 'POST',
    headers: sid ? { 'X-Admin-Session': sid } : {},
  }).catch(() => {});
  status.textContent = 'Режим администратора выключен.';
}

function openAdminModal() {
  const modal = document.getElementById('adminModal');
  if (!modal) return;
  const err = document.getElementById('adminAuthError');
  const login = document.getElementById('adminLoginInput');
  const password = document.getElementById('adminPasswordInput');
  if (err) err.textContent = '';
  if (login) login.value = '';
  if (password) password.value = '';
  modal.classList.add('open');
  modal.setAttribute('aria-hidden', 'false');
  setTimeout(() => { if (login) login.focus(); }, 0);
}

function closeAdminModal() {
  const modal = document.getElementById('adminModal');
  if (!modal) return;
  modal.classList.remove('open');
  modal.setAttribute('aria-hidden', 'true');
}

async function submitAdminLogin() {
  const login = document.getElementById('adminLoginInput').value.trim();
  const password = document.getElementById('adminPasswordInput').value;
  const err = document.getElementById('adminAuthError');
  if (!login || !password) {
    err.textContent = 'Введите логин и пароль.';
    return;
  }
  try {
    const r = await fetch('/api/admin/login', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ login, password }),
    });
    const data = await parseApiResponse(r);
    adminSessionId = String(data.session_id || '');
    if (!adminSessionId) throw new Error('Сессия не создана');
    localStorage.setItem('adminSessionId', adminSessionId);
    isAdminSession = true;
    applyAdminUi();
    closeAdminModal();
    document.getElementById('status').textContent = 'Режим администратора включён.';
  } catch (e) {
    err.textContent = e.message || 'Ошибка авторизации';
  }
}

function updateWeeksToggleLabel() {
  const btn = document.getElementById('btnToggleWeeks');
  if (!btn) return;
  btn.textContent = showAllWeeks ? 'Только 2 недели' : 'Все недели';
}

function fillWeekSelects(weeks, prev, last) {
  const selPrev = document.getElementById('weekPrev');
  const selLast = document.getElementById('weekLast');
  selPrev.innerHTML = '';
  selLast.innerHTML = '';
  weeks.forEach(w => {
    const o1 = document.createElement('option');
    o1.value = w; o1.textContent = w;
    const o2 = o1.cloneNode(true);
    selPrev.appendChild(o1);
    selLast.appendChild(o2);
  });
  if (weeks.length >= 2) {
    selPrev.value = String(prev ?? weeks[weeks.length - 2]);
    selLast.value = String(last ?? weeks[weeks.length - 1]);
  } else if (weeks.length === 1) {
    selPrev.value = selLast.value = String(weeks[0]);
  }
}

async function refreshWeeks(forceLatest = false) {
  const year = document.getElementById('year').value;
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const q = new URLSearchParams({ year });
  if (wh) q.set('wh_ids', wh);
  try {
    const r = await fetch('/api/weeks?' + q);
    if (!r.ok) return;
    const data = await r.json();
    availableWeeks = data.weeks || [];
    const curPrev = parseInt(document.getElementById('weekPrev').value, 10);
    const curLast = parseInt(document.getElementById('weekLast').value, 10);
    const hasCurPrev = Number.isFinite(curPrev) && availableWeeks.includes(curPrev);
    const hasCurLast = Number.isFinite(curLast) && availableWeeks.includes(curLast);
    if (forceLatest || !hasCurPrev || !hasCurLast) {
      fillWeekSelects(availableWeeks, data.week_prev, data.week_last);
    } else {
      fillWeekSelects(availableWeeks, curPrev, curLast);
    }
  } catch (_) {}
}

async function selectBuilding(b) {
  activeBuilding = b.id;
  syncBuildingButtons();
  if (!b.wh_ids || b.wh_ids.length === 0) {
    selectedWh = new Set(ALL_WH_IDS);
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = true);
  } else {
    selectedWh = new Set(b.wh_ids);
    document.querySelectorAll('#whGrid input').forEach(cb => {
      cb.checked = selectedWh.has(parseInt(cb.value, 10));
    });
  }
  await refreshWeeks();
  loadReport();
}

function syncBuildingButtons() {
  document.querySelectorAll('#buildingBtns button').forEach(btn => {
    btn.classList.toggle('active', btn.dataset.id === activeBuilding);
  });
}

function selectedWeeks() {
  let wp = document.getElementById('weekPrev').value;
  let wl = document.getElementById('weekLast').value;
  if (!wp || !wl) {
    if (availableWeeks.length >= 2) {
      wp = String(availableWeeks[availableWeeks.length - 2]);
      wl = String(availableWeeks[availableWeeks.length - 1]);
    } else {
      wp = String(CONFIG.week_prev || 20);
      wl = String(CONFIG.week_last || 21);
    }
    fillWeekSelects(availableWeeks.length ? availableWeeks : [parseInt(wp,10)], parseInt(wp,10), parseInt(wl,10));
  }
  return { wp, wl };
}

async function parseApiResponse(r) {
  const raw = await r.text();
  if (!r.ok) {
    try {
      const err = JSON.parse(raw);
      throw new Error(err.error || raw);
    } catch (e) {
      if (e instanceof Error && e.message !== raw) throw e;
      throw new Error(raw || r.statusText);
    }
  }
  return JSON.parse(raw);
}

async function refreshData() {
  const status = document.getElementById('status');
  const grid = document.getElementById('reportGrid');
  const btn = document.getElementById('btnRefreshData');
  if (CONFIG.refresh_token_required && !isAdminSession) {
    status.textContent = 'Сначала выполните вход администратора.';
    return;
  }
  btn.disabled = true;
  grid.classList.add('loading');
  status.textContent = 'Обновление базы и загрузка отчёта…';
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const year = document.getElementById('year').value;
  const q = new URLSearchParams({ year });
  if (wh) q.set('wh_ids', wh);
  const headers = adminSessionId ? { 'X-Admin-Session': adminSessionId } : {};
  try {
    const data = await parseApiResponse(await fetch('/api/refresh?' + q, { method: 'POST', headers }));
    if (data.weeks && data.weeks.length) {
      availableWeeks = data.weeks;
      const wp = data.week_prev ?? availableWeeks[availableWeeks.length - 2];
      const wl = data.week_last ?? availableWeeks[availableWeeks.length - 1];
      fillWeekSelects(availableWeeks, wp, wl);
    } else {
      await refreshWeeks();
    }
    await loadReport();
    const parts = [
      'Данные обновлены',
      data.row_count != null ? data.row_count.toLocaleString('ru') + ' строк' : '',
      data.max_date ? 'посл. дата ' + data.max_date : '',
    ].filter(Boolean);
    if (data.refresh_note) parts.push(data.refresh_note);
    status.textContent = parts.join(' · ');
  } catch (e) {
    if (String(e.message || '').includes('Недостаточно прав')) {
      isAdminSession = false;
      adminSessionId = '';
      localStorage.removeItem('adminSessionId');
      applyAdminUi();
      status.textContent = 'Недостаточно прав: выполните вход администратора с корректным токеном.';
    } else {
    status.textContent = 'Ошибка обновления: ' + e.message;
    }
  } finally {
    btn.disabled = false;
    grid.classList.remove('loading');
  }
}

async function exportXlsx() {
  const status = document.getElementById('status');
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
  });
  q.set('show_all_weeks', showAllWeeks ? '1' : '0');
  if (wh) q.set('wh_ids', wh);
  try {
    status.textContent = 'Готовим XLSX...';
    const r = await fetch('/api/export/xlsx?' + q);
    if (!r.ok) throw new Error(await r.text());
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const building = activeBuilding && activeBuilding !== 'custom' ? activeBuilding : 'wh';
    a.download = `write_offs_${building}_${q.get('year')}_w${wp}-${wl}.xlsx`;
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    status.textContent = 'XLSX сформирован';
  } catch (e) {
    status.textContent = 'Ошибка экспорта: ' + e.message;
  }
}

async function loadFreshness() {
  try {
    const r = await fetch('/api/freshness');
    if (!r.ok) return;
    updateFreshness(await r.json());
  } catch (_) {}
}

async function loadReport() {
  const status = document.getElementById('status');
  const grid = document.getElementById('reportGrid');
  if (!grid) {
    setText('status', 'Ошибка: не найден контейнер отчёта');
    return;
  }
  grid.classList.add('loading');
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const thr = saveAlertThresholds();
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
    alert_wow: String(thr.wow),
    alert_vs_avg4: String(thr.vs_avg4),
    alert_min_amount: String(thr.min_amount),
  });
  q.set('show_all_weeks', showAllWeeks ? '1' : '0');
  if (wh) q.set('wh_ids', wh);
  setText('status', 'Загрузка…');
  try {
    const data = await parseApiResponse(await fetch('/api/report?' + q));
    grid.innerHTML = data.html;
    if (typeof revealRefresh === 'function') revealRefresh(grid);
    updateKpis(data.kpis || null, data.yoy || null);
    updateOrg0Spark(data.org0_series || []);
    updateCorpusCompare(data.corpus_compare || []);
    updateDimBreakdowns(data.dim_breakdowns || null);
    updatePeriodCompare(data.compare || null);
    updateTop20Churn(data.top20_churn || null);
    updateGrowthAlerts(data.growth_alerts || [], data.alert_thresholds || thr);
    loadWatchlist();
    if (data.freshness) updateFreshness(data.freshness);
    else loadFreshness();
    syncDashboardUrl();
    applyTableSearch();
    grid.querySelectorAll('tr.drill-row').forEach(tr => {
      tr.addEventListener('click', () => {
        const kind = tr.dataset.drill;
        const params = {};
        if (kind === 'reason' && tr.dataset.reasonId) params.reason_id = tr.dataset.reasonId;
        if (kind === 'category' && tr.dataset.parentName) params.parent_name = tr.dataset.parentName;
        openReasonCard(params);
      });
    });
    const sorted = Array.from(selectedWh).sort((a,b)=>a-b);
    const label = selectedWh.size === ALL_WH_IDS.length
      ? 'все корпуса (' + ALL_WH_IDS.length + ' WH)'
      : sorted.map(whLabel).join('; ');
    const alertN = (data.growth_alerts || []).length;
    setText('status', `WH: ${label} · расчёт нед. ${data.week_prev}→${data.week_last}, в таблице: ${(data.weeks || []).join(', ')} (${data.year}). Алертов: ${alertN}. Клик по строке/алерту → карточка причины.`);
  } catch (e) {
    setText('status', 'Ошибка: ' + e.message);
    const setBox = (id, html) => { const el = document.getElementById(id); if (el) el.innerHTML = html; };
    setBox('corpusCompare', '<div class="muted-box">Не удалось загрузить сравнение</div>');
    setBox('dimSubject', '<div class="muted-box">Не удалось загрузить предметы</div>');
    setBox('dimOwnerState', '<div class="muted-box">Не удалось загрузить владельца/статус</div>');
    setBox('growthAlerts', '<div class="muted-box">Не удалось загрузить алерты</div>');
    setBox('periodCompare', '<div class="muted-box">Не удалось загрузить сравнение недель</div>');
    setBox('top20Churn', '<div class="muted-box">Не удалось загрузить сдвиг ТОП-20</div>');
  } finally {
    grid.classList.remove('loading');
  }
}

init();
</script>

</div>
</div>
</body>
</html>
"""


DETAILS_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Детализация write_offs</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Детализация</h1>
  <div class="subtitle">Сырые строки · клик по ячейке фильтрует · клик по заголовку сортирует</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details" class="active">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Details · raw</div>
  <h2>Детализация <span class="text-accent">write_offs</span></h2>
  <p class="hero-lead">Сырые строки: клик по ячейке фильтрует, клик по заголовку сортирует.</p>
</section>

<form class="filters" id="filters" data-reveal="fade-up">
  <label>Дата от <input name="date_from" type="date"></label>
  <label>Дата до <input name="date_to" type="date"></label>
  <label>Офис <input name="office_id" type="number" placeholder="office_id"></label>
  <label>WH <input name="wh_id" type="number" placeholder="wh_id"></label>
  <label>WH list <input name="wh_ids" placeholder="1,2,3"></label>
  <label>Тип <input name="type" placeholder="type"></label>
  <label>reason_id <input name="reason_id" type="number"></label>
  <label>Категория <input name="parent_name" placeholder="parent_name"></label>
  <label>Предмет <input name="subject_name" placeholder="subject_name"></label>
  <label>Владелец <input name="owner_product" placeholder="FBO / FBS / 1P"></label>
  <label>Статус <input name="state_id" placeholder="SDO / UDG"></label>
  <label>WH name <input name="wh_name" placeholder="wh_name"></label>
  <label>Файл <input name="source_file" placeholder="source_file"></label>
  <label>nm_id <input name="nm_id" type="number"></label>
  <label>shk_id <input name="shk_id" type="number"></label>
  <label>cnt_org <input name="cnt_org" type="number"></label>
  <label>Бренд <input name="brand_name" placeholder="brand_name"></label>
  <label>Поиск <input name="search" placeholder="title, причина, бренд, nm_id, shk_id"></label>
  <label>На стр.
    <select name="per_page">
      <option>25</option>
      <option selected>50</option>
      <option>100</option>
      <option>250</option>
      <option>500</option>
    </select>
  </label>
  <input type="hidden" name="sort_by" value="date">
  <input type="hidden" name="sort_dir" value="desc">
  <button class="btn primary" type="submit">Применить</button>
  <button class="btn export" type="button" id="btnReset">Сбросить</button>
  <button class="btn secondary" type="button" id="btnExport">Экспорт XLSX</button>
</form>
<div class="chips" id="activeFilters" data-reveal="fade-up"></div>
<div class="meta" id="meta" data-reveal="fade-up">Загрузка…</div>
<section class="panel" data-reveal="zoom-in">
  <div class="table-wrap">
    <table>
      <thead><tr id="thead"></tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
  <div class="pager">
    <button class="btn export" type="button" id="btnPrev">Назад</button>
    <button class="btn export" type="button" id="btnNext">Вперёд</button>
    <span class="muted" id="pageInfo"></span>
    <span class="spacer"></span>
    <span class="muted">Клик по ячейке = фильтр. Экспорт по умолчанию 5000 строк.</span>
  </div>
</section>
<script>
let currentPage = 1;
let columns = [];
let clickFilters = new Set(['wh_id','office_id','nm_id','shk_id','reason_id','parent_name','type','cnt_org','brand_name','subject_name','owner_product','state_id','wh_name','source_file']);
const SORTABLE = new Set(['date','amount','amount_obsh','summa_obshay','total_cost','share','office_id','wh_id','nm_id','shk_id','reason_id','cnt_org','cnt_ors','cnt_ocr','type','parent_name','reason_descr','title','brand_name','subject_name','owner_product','state_id','wh_name','source_file']);

function fmtValue(v) {
  if (v === null || v === undefined || v === '') return '—';
  if (typeof v === 'number') return v.toLocaleString('ru-RU');
  return String(v);
}

function formEl(name) {
  return document.getElementById('filters').elements[name];
}

function paramsFor(page) {
  const fd = new FormData(document.getElementById('filters'));
  const q = new URLSearchParams();
  for (const [k, v] of fd.entries()) {
    const value = String(v || '').trim();
    if (value) q.set(k, value);
  }
  q.set('page', String(page));
  return q;
}

function hydrateFiltersFromUrl() {
  const q = new URLSearchParams(window.location.search);
  const form = document.getElementById('filters');
  for (const [k, v] of q.entries()) {
    const el = form.elements[k];
    if (el) el.value = v;
  }
  if (!form.elements.sort_by.value) form.elements.sort_by.value = 'date';
  if (!form.elements.sort_dir.value) form.elements.sort_dir.value = 'desc';
  currentPage = parseInt(q.get('page') || '1', 10) || 1;
}

function renderActiveFilters() {
  const box = document.getElementById('activeFilters');
  const keys = ['date_from','date_to','office_id','wh_id','wh_ids','type','reason_id','parent_name','nm_id','shk_id','cnt_org','brand_name','search'];
  const chips = [];
  keys.forEach(k => {
    const el = formEl(k);
    const v = (el && el.value || '').trim();
    if (!v) return;
    chips.push(`<span class="chip">${k}=${v}<button type="button" data-clear="${k}" title="Сбросить">×</button></span>`);
  });
  const sortBy = formEl('sort_by').value || 'date';
  const sortDir = formEl('sort_dir').value || 'desc';
  chips.push(`<span class="chip">sort=${sortBy} ${sortDir}</span>`);
  box.innerHTML = chips.length ? chips.join('') : '<span class="muted">Активных фильтров нет. Кликните по ячейке таблицы, чтобы отфильтровать.</span>';
  box.querySelectorAll('[data-clear]').forEach(btn => {
    btn.addEventListener('click', () => {
      const el = formEl(btn.dataset.clear);
      if (el) el.value = '';
      loadDetails(1);
    });
  });
}

function setSort(key) {
  const by = formEl('sort_by');
  const dir = formEl('sort_dir');
  if (by.value === key) {
    dir.value = dir.value === 'asc' ? 'desc' : 'asc';
  } else {
    by.value = key;
    dir.value = (key === 'date' || key === 'amount' || key === 'total_cost') ? 'desc' : 'asc';
  }
  loadDetails(1);
}

function applyCellFilter(key, value) {
  if (value === null || value === undefined || value === '') return;
  const el = formEl(key);
  if (!el) return;
  el.value = String(value);
  if (key === 'wh_id') {
    const list = formEl('wh_ids');
    if (list) list.value = '';
  }
  loadDetails(1);
}

async function loadDetails(page = 1, pushState = true) {
  const meta = document.getElementById('meta');
  const tbody = document.getElementById('tbody');
  const thead = document.getElementById('thead');
  meta.textContent = 'Загрузка…';
  tbody.innerHTML = '';
  const q = paramsFor(page);
  try {
    const r = await fetch('/api/details?' + q);
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    columns = data.columns || [];
    if (Array.isArray(data.click_filters)) clickFilters = new Set(data.click_filters);
    currentPage = data.page || page;
    if (data.sort_by) formEl('sort_by').value = data.sort_by;
    if (data.sort_dir) formEl('sort_dir').value = data.sort_dir;
    if (pushState) history.replaceState(null, '', '/details?' + q);
    const sortBy = formEl('sort_by').value;
    const sortDir = formEl('sort_dir').value;
    thead.innerHTML = columns.map(c => {
      const sortable = SORTABLE.has(c.key);
      const active = sortable && c.key === sortBy;
      const mark = active ? (sortDir === 'asc' ? ' ▲' : ' ▼') : '';
      const cls = sortable ? ` class="sortable${active ? ' active-sort' : ''}" data-sort="${c.key}"` : '';
      return `<th${cls}>${c.label}${mark}</th>`;
    }).join('');
    thead.querySelectorAll('[data-sort]').forEach(th => {
      th.addEventListener('click', () => setSort(th.dataset.sort));
    });
    tbody.innerHTML = (data.rows || []).map(row => {
      return '<tr>' + columns.map(c => {
        const rawVal = row[c.key];
        const value = fmtValue(rawVal);
        const isNum = typeof rawVal === 'number';
        const canClick = clickFilters.has(c.key) && rawVal !== null && rawVal !== undefined && rawVal !== '';
        const cls = [isNum ? 'num' : '', canClick ? 'clickable' : ''].filter(Boolean).join(' ');
        const attrs = canClick
          ? ` class="${cls}" data-filter-key="${c.key}" data-filter-value="${String(rawVal).replaceAll('"', '&quot;')}" title="Фильтр: ${c.key}=${String(value).replaceAll('"', '&quot;')}"`
          : ` class="${cls}" title="${String(value).replaceAll('"', '&quot;')}"`;
        return `<td${attrs}>${value}</td>`;
      }).join('') + '</tr>';
    }).join('');
    tbody.querySelectorAll('[data-filter-key]').forEach(td => {
      td.addEventListener('click', () => applyCellFilter(td.dataset.filterKey, td.dataset.filterValue));
    });
    meta.textContent = `Всего по фильтру: ${Number(data.total || 0).toLocaleString('ru-RU')} · показано: ${(data.rows || []).length}`;
    document.getElementById('pageInfo').textContent = `Страница ${data.page} / ${data.pages}`;
    document.getElementById('btnPrev').disabled = data.page <= 1;
    document.getElementById('btnNext').disabled = data.page >= data.pages;
    renderActiveFilters();
  } catch (e) {
    meta.textContent = 'Ошибка загрузки: ' + (e.message || e);
  }
}

function exportXlsx() {
  const q = paramsFor(currentPage);
  q.delete('page');
  q.set('limit', '5000');
  const reason = q.get('reason_id') || 'all';
  const parent = (q.get('parent_name') || 'all').replaceAll(/[^\\w\\-]+/g, '_').slice(0, 24);
  window.location.href = '/api/details/export/xlsx?' + q;
  // filename is set by server; keep query rich for filters
  void reason; void parent;
}

document.getElementById('filters').addEventListener('submit', (e) => {
  e.preventDefault();
  loadDetails(1);
});
document.getElementById('btnReset').addEventListener('click', () => {
  document.getElementById('filters').reset();
  formEl('sort_by').value = 'date';
  formEl('sort_dir').value = 'desc';
  loadDetails(1);
});
document.getElementById('btnExport').addEventListener('click', exportXlsx);
document.getElementById('btnPrev').addEventListener('click', () => loadDetails(Math.max(1, currentPage - 1)));
document.getElementById('btnNext').addEventListener('click', () => loadDetails(currentPage + 1));

hydrateFiltersFromUrl();
loadDetails(currentPage, false);
</script>

</div>
</div>
</body>
</html>
"""


WEEKLY_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Динамика по неделям</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Динамика</h1>
  <div class="subtitle">Сумма брака, ORG0, топ причин и heatmap</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly" class="active">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Weekly · trend</div>
  <h2>Динамика <span class="text-accent">по неделям</span></h2>
  <p class="hero-lead">Сумма брака, доля ORG0, топ причин и heatmap — тот же срез, с появлением блоков при скролле.</p>
  <a class="scrolldown" href="#weeklyFilters">К фильтрам <span class="chev" aria-hidden="true"></span></a>
</section>

<section class="mod-section" id="weeklyFilters">
  <h2 class="section-title" data-reveal="fade-right">Фильтры</h2>
  <div class="toolbar" data-reveal="fade-up">
  <label>Год <input id="year" type="number" value="2026"></label>
  <label>WH ids <input id="wh_ids" placeholder="пусто = все"></label>
  <label>Топ причин
    <select id="top_n">
      <option>3</option>
      <option selected>5</option>
      <option>10</option>
    </select>
  </label>
  <button class="btn primary" id="btnLoad" type="button">Загрузить</button>
  <button class="btn secondary" id="btnExport" type="button">Экспорт XLSX</button>
  </div>
  <div class="meta" id="meta" data-reveal="fade-up">Загрузка…</div>
</section>

<section class="mod-section">
  <h2 class="section-title" data-reveal="fade-right">График <span class="text-accent">и таблица</span></h2>
<section class="panel" data-reveal="zoom-in">
  <div class="chart">
    <div class="bars" id="bars"></div>
  </div>
  <div style="overflow:auto">
    <table>
      <thead>
        <tr>
          <th>Неделя</th>
          <th class="num">Всего, ₽</th>
          <th class="num">ORG0, ₽</th>
          <th class="num">ORG0 %</th>
          <th class="num">Строк</th>
          <th>Топ причин</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</section>
<section class="panel" style="margin-top:12px" data-reveal="fade-up">
  <div style="padding:12px 14px 4px;font-weight:700">Heatmap: причины × недели</div>
  <div class="heatmap-wrap" style="padding:0 14px 14px" id="heatmapBox"><div class="muted">Загрузка…</div></div>
</section>
</section>
<script>
function fmt(n) { return Number(n || 0).toLocaleString('ru-RU', { maximumFractionDigits: 0 }); }
function pct(n) { return (Number(n || 0)).toLocaleString('ru-RU', { maximumFractionDigits: 1 }) + '%'; }

function hydrate() {
  const q = new URLSearchParams(location.search);
  if (q.get('year')) document.getElementById('year').value = q.get('year');
  if (q.get('wh_ids')) document.getElementById('wh_ids').value = q.get('wh_ids');
  if (q.get('top_n')) document.getElementById('top_n').value = q.get('top_n');
}

function heatColor(intensity) {
  const t = Math.max(0, Math.min(1, Number(intensity || 0)));
  const a = 0.10 + t * 0.55;
  return `rgba(47, 125, 122, ${a.toFixed(3)})`;
}

async function loadHeatmap() {
  const box = document.getElementById('heatmapBox');
  const year = document.getElementById('year').value || new Date().getFullYear();
  const wh = (document.getElementById('wh_ids').value || '').trim();
  const q = new URLSearchParams({ year, top_n: '12', week_limit: '12' });
  if (wh) q.set('wh_ids', wh);
  try {
    const r = await fetch('/api/heatmap?' + q);
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || r.statusText);
    const weeks = d.weeks || [];
    const reasons = d.reasons || [];
    const map = {};
    (d.cells || []).forEach(c => { map[c.reason_id + ':' + c.week] = c; });
    if (!weeks.length || !reasons.length) {
      box.innerHTML = '<div class="muted">Нет данных для heatmap</div>';
      return;
    }
    box.innerHTML = `<table class="heatmap"><thead><tr><th></th>${weeks.map(w=>`<th>W${w}</th>`).join('')}</tr></thead>
      <tbody>${reasons.map(reason => `<tr>
        <td class="rname" data-reason-id="${reason.reason_id}" title="${reason.name}">${reason.name}</td>
        ${weeks.map(w => {
          const cell = map[reason.reason_id + ':' + w];
          const amt = cell ? cell.amount : 0;
          const inten = cell ? cell.intensity : 0;
          return `<td class="cell" style="background:${heatColor(inten)}" title="W${w}: ${fmt(amt)}" data-reason-id="${reason.reason_id}"></td>`;
        }).join('')}
      </tr>`).join('')}</tbody></table>`;
    box.querySelectorAll('[data-reason-id]').forEach(el => {
      el.addEventListener('click', () => {
        const qq = new URLSearchParams({ year, reason_id: el.dataset.reasonId });
        if (wh) qq.set('wh_ids', wh);
        location.href = '/reason?' + qq;
      });
    });
  } catch (e) {
    box.innerHTML = '<div class="muted">Ошибка heatmap: ' + (e.message || e) + '</div>';
  }
}

async function loadWeekly() {
  const meta = document.getElementById('meta');
  const year = document.getElementById('year').value || new Date().getFullYear();
  const wh = (document.getElementById('wh_ids').value || '').trim();
  const topN = document.getElementById('top_n').value || '5';
  const q = new URLSearchParams({ year, top_n: topN });
  if (wh) q.set('wh_ids', wh);
  meta.textContent = 'Загрузка…';
  try {
    const r = await fetch('/api/weekly?' + q);
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    history.replaceState(null, '', '/weekly?' + q);
    const weeks = data.weeks || [];
    const maxAmt = Math.max(1, ...weeks.map(w => w.amount_all || 0));
    document.getElementById('bars').innerHTML = weeks.map(w => {
      const h = Math.max(2, Math.round((w.amount_all || 0) / maxAmt * 180));
      const h0 = Math.max(0, Math.round((w.amount_org0 || 0) / maxAmt * 180));
      return `<div class="bar-wrap" title="W${w.week}: ${fmt(w.amount_all)}">
        <div class="bar-stack" style="--h:${h}px">
          <div class="bar" style="height:${h}px"></div>
          <div class="bar org0" style="height:${h0}px"></div>
        </div>
        <div class="wlabel">${w.week}</div>
      </div>`;
    }).join('');
    document.getElementById('tbody').innerHTML = weeks.map(w => {
      const top = (w.top_reasons || []).map(t => `${t.reason_descr} (${fmt(t.amount)})`).join('; ');
      return `<tr>
        <td>${w.week}</td>
        <td class="num">${fmt(w.amount_all)}</td>
        <td class="num">${fmt(w.amount_org0)}</td>
        <td class="num">${pct(w.org0_share)}</td>
        <td class="num">${fmt(w.row_count)}</td>
        <td class="top">${top || '—'}</td>
      </tr>`;
    }).join('');
    const t = data.totals || {};
    meta.textContent = `Год ${data.year} · источник ${data.source} · всего ${fmt(t.amount_all)} ₽ · ORG0 ${fmt(t.amount_org0)} ₽ · недель ${weeks.length}`;
    if (typeof revealRefresh === 'function') revealRefresh();
    loadHeatmap();
  } catch (e) {
    meta.textContent = 'Ошибка: ' + (e.message || e);
  }
}

document.getElementById('btnLoad').onclick = loadWeekly;
document.getElementById('btnExport').onclick = () => {
  const year = document.getElementById('year').value || new Date().getFullYear();
  const wh = (document.getElementById('wh_ids').value || '').trim();
  const topN = document.getElementById('top_n').value || '5';
  const q = new URLSearchParams({ year, top_n: topN });
  if (wh) q.set('wh_ids', wh);
  window.location.href = '/api/weekly/export/xlsx?' + q;
};
hydrate();
loadWeekly();
</script>

</div>
</div>
</body>
</html>
"""


REASON_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Карточка причины</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1 id="title">Карточка причины</h1>
  <div class="subtitle" id="subtitle">Тренд, топ WH / nm_id / бренды, ORG0</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason" class="active">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Reason · card</div>
  <h2>Карточка <span class="text-accent">причины</span></h2>
  <p class="hero-lead">Тренд, топ WH / nm_id / бренды и ORG0 — данные те же, появление блоков при скролле.</p>
</section>

<div class="meta" id="meta" data-reveal="fade-up">Загрузка…</div>
<div class="kpis stagger" id="kpis"></div>
<div class="grid stagger">
  <section class="card" data-reveal="flip-left">
    <h3>Динамика по неделям</h3>
    <div class="body">
      <div class="bars" id="bars"></div>
      <div id="weekLabels" style="display:grid;grid-auto-flow:column;grid-auto-columns:minmax(22px,1fr);gap:4px;margin-top:4px"></div>
      <div style="margin-top:10px">
        <div class="muted" style="font-size:11px;margin-bottom:4px">Доля ORG0 по неделям</div>
        <div class="bars" id="org0Bars" style="height:72px"></div>
      </div>
    </div>
  </section>
  <section class="card" data-reveal="flip-left">
    <h3>Топ WH (посл. нед.)</h3>
    <div class="body" id="topWh"></div>
  </section>
  <section class="card" data-reveal="fade-left">
    <h3>Топ nm_id</h3>
    <div class="body" id="topNm"></div>
  </section>
  <section class="card" data-reveal="fade-left">
    <h3>Топ бренды</h3>
    <div class="body" id="topBrands"></div>
  </section>
</div>
<script>
function fmt(n){ return Number(n||0).toLocaleString('ru-RU',{maximumFractionDigits:0}); }
function pct(n){ if(n==null) return '—'; return Number(n).toLocaleString('ru-RU',{maximumFractionDigits:1})+'%'; }
function dyn(n){ if(n==null) return '—'; const s=n>0?'+':''; return s+pct(n); }
function table(rows, cols){
  if(!rows.length) return '<div class="muted">Нет данных</div>';
  return '<table><thead><tr>'+cols.map(c=>`<th>${c.label}</th>`).join('')+'</tr></thead><tbody>'+
    rows.map(r=>'<tr>'+cols.map(c=>{
      const v=r[c.key]; const cls=c.num?' class="num"':'';
      return `<td${cls}>${c.num?fmt(v):(v??'—')}</td>`;
    }).join('')+'</tr>').join('')+'</tbody></table>';
}
function bookmarkKey(d){
  return d.reason_id != null ? ('r:'+d.reason_id) : ('c:'+(d.parent_name||''));
}
function toggleBookmark(d){
  const key = 'kurkuma_reason_bookmarks';
  let list = [];
  try { list = JSON.parse(localStorage.getItem(key)||'[]')||[]; } catch(_){}
  const href = location.pathname + location.search;
  const title = d.title || bookmarkKey(d);
  const exists = list.findIndex(x => x.href === href);
  if (exists >= 0) list.splice(exists, 1);
  else list.unshift({ title, href, id: bookmarkKey(d) });
  localStorage.setItem(key, JSON.stringify(list.slice(0,10)));
  return exists < 0;
}
function toggleWatchlist(d){
  if(d.reason_id == null) return false;
  let list = [];
  try { list = JSON.parse(localStorage.getItem('kurkuma_reason_watchlist')||'[]')||[]; } catch(_){}
  const exists = list.findIndex(x => String(x.reason_id) === String(d.reason_id));
  if (exists >= 0) list.splice(exists, 1);
  else list.unshift({ reason_id: d.reason_id, title: d.title || ('reason '+d.reason_id) });
  localStorage.setItem('kurkuma_reason_watchlist', JSON.stringify(list.slice(0,12)));
  return exists < 0;
}
function inWatchlist(d){
  try {
    const list = JSON.parse(localStorage.getItem('kurkuma_reason_watchlist')||'[]')||[];
    return list.some(x => String(x.reason_id) === String(d.reason_id));
  } catch(_){ return false; }
}
async function load(){
  const q = new URLSearchParams(location.search);
  const meta = document.getElementById('meta');
  if(!q.get('reason_id') && !q.get('parent_name')){
    meta.textContent = 'Укажите reason_id или parent_name в URL';
    return;
  }
  try{
    const r = await fetch('/api/reason?' + q);
    const raw = await r.text();
    if(!r.ok) throw new Error(raw||r.statusText);
    const d = JSON.parse(raw);
    const titleEl = document.getElementById('title');
    const subtitleEl = document.getElementById('subtitle');
    if (titleEl) titleEl.textContent = d.title || 'Карточка причины';
    if (subtitleEl) subtitleEl.textContent =
      (d.reason_id!=null ? ('reason_id='+d.reason_id+' · ') : '') +
      (d.parent_name ? ('категория: '+d.parent_name+' · ') : '') +
      `год ${d.year}, неделя ${d.week_last}`;
    try {
      const recent = JSON.parse(localStorage.getItem('kurkuma_reason_recent')||'[]')||[];
      const href = location.pathname + location.search;
      const item = { title: d.title || 'карточка', href };
      localStorage.setItem('kurkuma_reason_recent', JSON.stringify([item, ...recent.filter(x=>x.href!==href)].slice(0,10)));
    } catch(_){}
    const k = d.kpis || {};
    const c = d.concentration || {};
    document.getElementById('kpis').innerHTML = [
      ['Сумма (посл. нед.)', fmt(k.amount_last)],
      ['ORG0', fmt(k.amount_org0_last)],
      ['ORG0 %', pct(k.org0_share)],
      ['Ср. 4 нед.', fmt(k.avg4)],
      ['vs ср. 4', dyn(k.vs_avg4)],
      ['Топ-3 WH', pct(c.top3_wh_share)],
      ['Топ-5 nm', pct(c.top5_nm_share)],
    ].map(([a,b])=>`<div class="kpi" data-reveal="flip-left"><div class="k">${a}</div><div class="v">${b}</div></div>`).join('');
    const weeks = d.weeks || [];
    const maxA = Math.max(1, ...weeks.map(w=>w.amount_all||0));
    document.getElementById('bars').innerHTML = weeks.map(w=>{
      const h = Math.max(2, Math.round((w.amount_all||0)/maxA*140));
      return `<div class="bar" style="height:${h}px" title="W${w.week}: ${fmt(w.amount_all)}"></div>`;
    }).join('');
    document.getElementById('weekLabels').innerHTML = weeks.map(w=>`<div class="wlabel">${w.week}</div>`).join('');
    const org0 = d.org0_trend || weeks;
    const maxShare = Math.max(1, ...org0.map(w=>w.org0_share||0));
    document.getElementById('org0Bars').innerHTML = org0.map(w=>{
      const h = Math.max(2, Math.round((w.org0_share||0)/maxShare*60));
      return `<div class="bar" style="height:${h}px;background:var(--warn)" title="W${w.week}: ${pct(w.org0_share)}"></div>`;
    }).join('');
    document.getElementById('topWh').innerHTML = table(d.top_wh||[], [
      {key:'wh_id', label:'WH'}, {key:'name', label:'Название'}, {key:'amount', label:'Сумма', num:true}
    ]);
    document.getElementById('topNm').innerHTML = table(d.top_nm||[], [
      {key:'nm_id', label:'nm_id'}, {key:'title', label:'Наименование'}, {key:'amount', label:'Сумма', num:true}
    ]);
    document.getElementById('topBrands').innerHTML = table(d.top_brands||[], [
      {key:'brand_name', label:'Бренд'}, {key:'amount', label:'Сумма', num:true}
    ]);
    const detailsQ = new URLSearchParams(q);
    const watchLabel = inWatchlist(d) ? 'В watchlist ✓' : 'В watchlist';
    meta.innerHTML = `Источник: ${d.source} · <a class="btn secondary" href="/details?${detailsQ}">Сырые строки</a> <a class="btn secondary" href="/api/reason/export/xlsx?${q}">XLSX</a> <button class="btn secondary" id="btnBookmark" type="button">В закладки</button> <button class="btn secondary" id="btnWatch" type="button">${watchLabel}</button> <a class="btn secondary" href="/">На дашборд</a>`;
    document.getElementById('btnBookmark').onclick = () => {
      const added = toggleBookmark(d);
      document.getElementById('btnBookmark').textContent = added ? 'В закладках' : 'В закладки';
    };
    document.getElementById('btnWatch').onclick = () => {
      if(d.reason_id == null){ alert('Watchlist только для reason_id'); return; }
      const added = toggleWatchlist(d);
      document.getElementById('btnWatch').textContent = added ? 'В watchlist ✓' : 'В watchlist';
    };
    if (typeof revealRefresh === 'function') revealRefresh();
  }catch(e){
    meta.textContent = 'Ошибка: ' + (e.message||e);
  }
}
load();
</script>

</div>
</div>
</body>
</html>
"""


DIGEST_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Дайджест брака</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Дайджест</h1>
  <div class="subtitle" id="subtitle">Еженедельный срез KPI, алертов и сдвига ТОП-20</div>
</div>
<nav class="topnav no-print">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest" class="active">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">
<section class="hero-band no-print" data-reveal="fade-down">
  <div class="hero-kicker">Digest · weekly</div>
  <h2>Дайджест <span class="text-accent">брака</span></h2>
  <p class="hero-lead">Еженедельный срез KPI, алертов и сдвига ТОП-20 — смысл отчёта прежний.</p>
  <a class="scrolldown" href="#digestBody">К срезу <span class="chev" aria-hidden="true"></span></a>
</section>
<section class="mod-section no-print" id="digestFilters">
  <h2 class="section-title" data-reveal="fade-right">Параметры</h2>
  <div class="toolbar no-print" data-reveal="fade-up">
  <label>Год <input id="year" type="number" value="2026"></label>
  <label>Пред. нед. <input id="week_prev" type="number"></label>
  <label>Посл. нед. <input id="week_last" type="number"></label>
  <label>WH ids <input id="wh_ids" placeholder="пусто = каталог"></label>
  <div class="actions">
    <button class="btn primary" id="btnLoad" type="button">Собрать</button>
    <button class="btn" id="btnSaveSnap" type="button">Сохранить снимок</button>
    <button class="btn" id="btnPrint" type="button">Печать / PDF</button>
    <button class="btn" id="btnDownload" type="button">Скачать HTML</button>
  </div>
  </div>
</section>
<section class="mod-section no-rule" id="digestBody">
  <h2 class="section-title" data-reveal="fade-right">Срез <span class="text-accent">недели</span></h2>
  <div class="kpis stagger" id="kpis"></div>
  <div class="grid stagger" id="content">
  <section class="card" data-reveal="flip-left"><h3>Алерты роста</h3><div class="body" id="alerts"><div class="empty">Загрузка…</div></div></section>
  <section class="card" data-reveal="flip-left"><h3>Сдвиг ТОП-20</h3><div class="body" id="churn"><div class="empty">Загрузка…</div></div></section>
  <section class="card" data-reveal="fade-left"><h3>Корпуса</h3><div class="body" id="corpus"><div class="empty">Загрузка…</div></div></section>
  <section class="card" data-reveal="fade-left"><h3>Крупнейшие дельты</h3><div class="body" id="compare"><div class="empty">Загрузка…</div></div></section>
  <section class="card" style="grid-column:1 / -1" data-reveal="zoom-in"><h3>Снимок: тогда vs сейчас</h3><div class="body" id="snapshotBox"><div class="empty">Загрузка…</div></div></section>
  </div>
</section>
<script>
function fmt(n){ return Number(n||0).toLocaleString('ru-RU',{maximumFractionDigits:0}); }
function pct(n){ if(n==null) return '—'; return Number(n).toLocaleString('ru-RU',{maximumFractionDigits:1})+'%'; }
function dyn(n){ if(n==null) return '—'; const s=n>0?'+':''; return s+pct(n); }
function dynClass(n){
  if(n==null || Number.isNaN(Number(n))) return '';
  return Number(n) > 0 ? 'up' : (Number(n) < 0 ? 'down' : '');
}
function row(name, meta, val, valCls=''){
  return `<div class="row">
    <div><div class="name" title="${String(name).replaceAll('"','&quot;')}">${name}</div>
    ${meta ? `<div class="hint">${meta}</div>` : ''}</div>
    <div class="val ${valCls}">${val}</div>
  </div>`;
}
function hydrate(){
  const q = new URLSearchParams(location.search);
  if(q.get('year')) document.getElementById('year').value = q.get('year');
  if(q.get('week_prev')) document.getElementById('week_prev').value = q.get('week_prev');
  if(q.get('week_last')) document.getElementById('week_last').value = q.get('week_last');
  if(q.get('wh_ids')) document.getElementById('wh_ids').value = q.get('wh_ids');
}
async function load(){
  const q = new URLSearchParams();
  const year = document.getElementById('year').value;
  const wp = document.getElementById('week_prev').value;
  const wl = document.getElementById('week_last').value;
  const wh = (document.getElementById('wh_ids').value||'').trim();
  if(year) q.set('year', year);
  if(wp) q.set('week_prev', wp);
  if(wl) q.set('week_last', wl);
  if(wh) q.set('wh_ids', wh);
  const r = await fetch('/api/digest?' + q);
  const raw = await r.text();
  if(!r.ok) throw new Error(raw||r.statusText);
  const d = JSON.parse(raw);
  history.replaceState(null,'','/digest?'+q);
  document.getElementById('subtitle').textContent =
    `Год ${d.year} · недели ${d.week_prev}→${d.week_last}` +
    (d.freshness && d.freshness.max_date ? ` · данные на ${d.freshness.max_date}` : '');
  const k = d.kpis||{}, y=d.yoy||{};
  const yoyVal = y.yoy_pct==null ? 'нет данных' : dyn(y.yoy_pct);
  document.getElementById('kpis').innerHTML = [
    ['Всего брак', fmt(k.total_last), ''],
    ['ТОП-20', fmt(k.top20_last), ''],
    ['Покрытие', pct(k.top20_cover), ''],
    ['ORG0 %', pct(k.org0_share), ''],
    ['vs ср.4', dyn(k.top20_vs_avg4), dynClass(k.top20_vs_avg4)],
    ['YoY', yoyVal, y.yoy_pct==null ? '' : dynClass(y.yoy_pct)],
  ].map(([a,b,c])=>`<div class="kpi" data-reveal="flip-left"><div class="k">${a}</div><div class="v ${c}">${b}</div></div>`).join('');
  if (typeof revealRefresh === 'function') revealRefresh(document.getElementById('kpis'));

  const alerts = d.growth_alerts||[];
  document.getElementById('alerts').innerHTML = alerts.length
    ? `<div class="list">${alerts.slice(0,8).map(a => row(
        `${a.label}: ${a.name}`,
        `${fmt(a.w_prev)} → ${fmt(a.w_last)}`,
        dyn(a.vs_avg4!=null?a.vs_avg4:a.dynamics),
        'dyn'
      )).join('')}</div>`
    : '<div class="empty">Сильного роста нет</div>';

  const ch = (d.top20_churn&&d.top20_churn.defects)||{};
  const changed = Boolean(ch.membership_changed);
  const left = (changed ? (ch.entered||[]) : (ch.rank_up||[])).slice(0,6);
  const mid = (changed ? (ch.exited||[]) : (ch.rank_down||[])).slice(0,6);
  const leftTitle = changed ? 'Вошли' : '↑ в рейтинге';
  const midTitle = changed ? 'Вышли' : '↓ в рейтинге';
  const col = (title, arr, mode) => {
    if(!arr.length) return `<div class="churn-col"><h4>${title}</h4><div class="empty" style="padding:8px 0;text-align:left">нет</div></div>`;
    return `<div class="churn-col"><h4>${title}</h4><div class="list">${arr.map(x=>{
      let val = fmt(x.amount_last||x.delta);
      let cls = '';
      if(mode==='exited') val = fmt(x.amount_prev);
      if(mode==='rank'){ val = `#${x.rank_prev}→#${x.rank_last}`; cls = title.includes('↑') ? 'ok' : 'dyn'; }
      return row(x.name, mode==='rank' ? fmt(x.delta) : '', val, cls);
    }).join('')}</div></div>`;
  };
  document.getElementById('churn').innerHTML =
    (changed ? '' : '<div class="note">Состав ТОП-20 не изменился — показаны сдвиги ранга</div>') +
    `<div class="churn-grid">${col(leftTitle, left, changed?'entered':'rank')}${col(midTitle, mid, changed?'exited':'rank')}</div>`;

  const corpus = d.corpus_compare||[];
  document.getElementById('corpus').innerHTML = corpus.length
    ? `<div class="list">${corpus.map(c => row(
        c.name,
        `доля ${pct(c.share_of_total)} · ORG0 ${pct(c.org0_share)}`,
        `${fmt(c.amount_last)} · ${dyn(c.dynamics)}`,
        dynClass(c.dynamics)
      )).join('')}</div>`
    : '<div class="empty">Нет данных</div>';

  const cmp = [...((d.compare&&d.compare.defects)||[]), ...((d.compare&&d.compare.categories)||[])].slice(0,8);
  document.getElementById('compare').innerHTML = cmp.length
    ? `<div class="list">${cmp.map(r => row(
        r.name,
        `${fmt(r.amount_a)} → ${fmt(r.amount_b)}`,
        dyn(r.pct),
        dynClass(r.pct)
      )).join('')}</div>`
    : '<div class="empty">Нет данных</div>';
  renderSnapshotCompare(d);
}
function snapshotKey(){ return 'kurkuma_digest_snapshot'; }
function saveSnapshot(d){
  const snap = {
    saved_at: new Date().toISOString(),
    year: d.year, week_prev: d.week_prev, week_last: d.week_last,
    kpis: d.kpis || {},
    yoy: d.yoy || {},
    alerts: (d.growth_alerts||[]).slice(0,8).map(a=>({name:a.name, kind:a.kind, row_id:a.row_id, w_last:a.w_last, dynamics:a.dynamics, vs_avg4:a.vs_avg4})),
  };
  localStorage.setItem(snapshotKey(), JSON.stringify(snap));
  renderSnapshotCompare(d);
}
function loadSnapshot(){
  try { return JSON.parse(localStorage.getItem(snapshotKey())||'null'); } catch(_){ return null; }
}
function renderSnapshotCompare(current){
  const box = document.getElementById('snapshotBox');
  if(!box) return;
  const snap = loadSnapshot();
  if(!snap){
    box.innerHTML = '<div class="empty">Снимка ещё нет — нажмите «Сохранить снимок»</div>';
    return;
  }
  const ck = current.kpis||{}, sk = snap.kpis||{};
  const rows = [
    ['Всего брак', sk.total_last, ck.total_last],
    ['ТОП-20', sk.top20_last, ck.top20_last],
    ['Покрытие', sk.top20_cover, ck.top20_cover, true],
    ['ORG0 %', sk.org0_share, ck.org0_share, true],
  ];
  const when = (snap.saved_at||'').slice(0,16).replace('T',' ');
  box.innerHTML = `<div class="note">Снимок: ${when} · W${snap.week_prev}→${snap.week_last} (${snap.year})</div>
    <div class="list">${rows.map(([name,a,b,isPct])=>{
      const av = Number(a||0), bv = Number(b||0), delta = bv-av;
      const left = isPct ? pct(av) : fmt(av);
      const right = isPct ? pct(bv) : fmt(bv);
      const dtxt = isPct ? dyn(delta) : ((delta>0?'+':'')+fmt(delta));
      return row(name, `${left} → ${right}`, dtxt, dynClass(delta));
    }).join('')}</div>`;
}
document.getElementById('btnLoad').onclick = () => load().catch(e => alert(e.message||e));
document.getElementById('btnPrint').onclick = () => window.print();
document.getElementById('btnDownload').onclick = () => {
  const html = '<!DOCTYPE html>\\n' + document.documentElement.outerHTML;
  const blob = new Blob([html], {type:'text/html;charset=utf-8'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = 'brak_digest.html';
  a.click();
  URL.revokeObjectURL(a.href);
};
document.getElementById('btnSaveSnap').onclick = async () => {
  try {
    const q = new URLSearchParams(location.search);
    const r = await fetch('/api/digest?' + q);
    const d = await r.json();
    if(!r.ok) throw new Error(d.error||r.statusText);
    saveSnapshot(d);
  } catch(e){ alert(e.message||e); }
};
hydrate();
load().catch(e => { document.getElementById('alerts').innerHTML = '<div class="empty">Ошибка: '+(e.message||e)+'</div>'; });
</script>
</div>
</div>
</body>
</html>
"""


ACTIONS_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Доска действий</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Действия</h1>
  <div class="subtitle">Алерты роста со статусами из localStorage</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions" class="active">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">
<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Actions · board</div>
  <h2>Доска <span class="text-accent">действий</span></h2>
  <p class="hero-lead">Алерты роста со статусами из localStorage — колонки появляются каскадом.</p>
</section>
<section class="mod-section no-rule">
  <div class="toolbar" data-reveal="fade-up">
  <label>Год <input id="year" type="number" value="2026"></label>
  <label>Пред. нед. <input id="week_prev" type="number"></label>
  <label>Посл. нед. <input id="week_last" type="number"></label>
  <label>WH ids <input id="wh_ids" placeholder="пусто = каталог"></label>
  <button class="btn primary" id="btnLoad" type="button">Обновить</button>
  </div>
  <div class="board stagger" id="board"></div>
</section>
<script>
const STATUSES = ['new','watching','escalated','closed'];
function fmt(n){ return Number(n||0).toLocaleString('ru-RU',{maximumFractionDigits:0}); }
function dyn(n){ if(n==null) return '—'; const s=n>0?'+':''; return s+Number(n).toLocaleString('ru-RU',{maximumFractionDigits:1})+'%'; }
function notesKey(){
  const year=document.getElementById('year').value;
  const wl=document.getElementById('week_last').value||'na';
  const wh=(document.getElementById('wh_ids').value||'').trim()||'all';
  return `kurkuma_alert_notes:${year}:${wl}:${wh}`;
}
function loadNotes(){ try{return JSON.parse(localStorage.getItem(notesKey())||'{}')||{};}catch(_){return{};} }
function saveNotes(n){ localStorage.setItem(notesKey(), JSON.stringify(n||{})); }
function alertKey(a){ return a.alert_key || ((a.kind==='reason'?'r:':'c:')+(a.row_id??a.name)); }
function openCard(a){
  const q=new URLSearchParams({year:document.getElementById('year').value});
  const wl=document.getElementById('week_last').value; if(wl) q.set('week_last', wl);
  const wh=(document.getElementById('wh_ids').value||'').trim(); if(wh) q.set('wh_ids', wh);
  if(a.kind==='reason' && a.row_id!=null) q.set('reason_id', a.row_id);
  if(a.kind==='category') q.set('parent_name', a.name);
  location.href='/reason?'+q;
}
function render(alerts){
  const notes=loadNotes();
  const groups={new:[],watching:[],escalated:[],closed:[]};
  (alerts||[]).forEach(a=>{
    const key=alertKey(a);
    const st=(notes[key]&&notes[key].status)||'new';
    (groups[st]||groups.new).push({...a, _key:key, _note:notes[key]||{status:st,comment:''}});
  });
  const titles={new:'New',watching:'Watching',escalated:'Escalated',closed:'Closed'};
  document.getElementById('board').innerHTML = STATUSES.map(st=>{
    const list=groups[st]||[];
    return `<section class="col" data-reveal="flip-left"><h3>${titles[st]} <span class="n">${list.length}</span></h3>
      ${list.length?list.map(a=>`<div class="card" data-key="${a._key}">
        <div class="t">${a.label}: ${a.name}</div>
        <div class="m"><span class="dyn">${dyn(a.vs_avg4!=null?a.vs_avg4:a.dynamics)}</span> · ${fmt(a.w_prev)} → ${fmt(a.w_last)}</div>
        <select data-role="status">${STATUSES.map(s=>`<option value="${s}"${s===st?' selected':''}>${s}</option>`).join('')}</select>
        <textarea data-role="comment" placeholder="комментарий">${a._note.comment||''}</textarea>
      </div>`).join(''):'<div class="empty">Пусто</div>'}
    </section>`;
  }).join('');
  if (typeof revealRefresh === 'function') revealRefresh(document.getElementById('board'));
  document.querySelectorAll('.card').forEach(card=>{
    const key=card.dataset.key;
    const persist=()=>{
      const n=loadNotes();
      n[key]={status:card.querySelector('[data-role=status]').value, comment:card.querySelector('[data-role=comment]').value||''};
      saveNotes(n); render(alerts);
    };
    card.querySelector('[data-role=status]').onchange=persist;
    card.querySelector('[data-role=comment]').onchange=persist;
    card.addEventListener('click', (e)=>{
      if(e.target.closest('select,textarea')) return;
      const a=(alerts||[]).find(x=>alertKey(x)===key);
      if(a) openCard(a);
    });
  });
}
async function load(){
  const q=new URLSearchParams();
  const year=document.getElementById('year').value;
  const wp=document.getElementById('week_prev').value;
  const wl=document.getElementById('week_last').value;
  const wh=(document.getElementById('wh_ids').value||'').trim();
  if(year) q.set('year', year);
  if(wp) q.set('week_prev', wp);
  if(wl) q.set('week_last', wl);
  if(wh) q.set('wh_ids', wh);
  const r=await fetch('/api/digest?'+q);
  const d=await r.json();
  if(!r.ok) throw new Error(d.error||r.statusText);
  if(!document.getElementById('week_prev').value) document.getElementById('week_prev').value=d.week_prev;
  if(!document.getElementById('week_last').value) document.getElementById('week_last').value=d.week_last;
  history.replaceState(null,'','/actions?'+q);
  render(d.growth_alerts||[]);
}
document.getElementById('btnLoad').onclick=()=>load().catch(e=>alert(e.message||e));
const q0=new URLSearchParams(location.search);
if(q0.get('year')) document.getElementById('year').value=q0.get('year');
if(q0.get('week_prev')) document.getElementById('week_prev').value=q0.get('week_prev');
if(q0.get('week_last')) document.getElementById('week_last').value=q0.get('week_last');
if(q0.get('wh_ids')) document.getElementById('wh_ids').value=q0.get('wh_ids');
load().catch(e=>{ document.getElementById('board').innerHTML='<div class="empty">Ошибка: '+(e.message||e)+'</div>'; });
</script>
</div>
</div>
</body>
</html>
"""


STATUS_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Статус системы</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Статус</h1>
  <div class="subtitle">Окружение, база, matview и кэш</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status" class="active">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">System · status</div>
  <h2>Статус <span class="text-accent">системы</span></h2>
  <p class="hero-lead">Окружение, база, matview и кэш — без изменения смысла проверок.</p>
</section>
<div class="meta" id="meta" data-reveal="fade-up">Загрузка…</div>
<button class="btn secondary" id="btnReload" type="button" data-reveal="fade-up">Обновить</button>
<div class="grid stagger" id="grid"></div>
<script>
function cls(v) {
  if (v === 'ok' || v === 'set' || v === true) return 'ok';
  if (v === 'missing' || v === 'empty' || v === 'error' || v === false) return 'err';
  if (v === 'degraded') return 'warn';
  return '';
}
function row(k, v, c='') {
  return `<div class="row"><span class="k">${k}</span><span class="v ${c}">${v ?? '—'}</span></div>`;
}
function card(title, body) {
  return `<section class="card" data-reveal="flip-left"><h2>${title}</h2><div class="body">${body}</div></section>`;
}
async function loadStatus() {
  const meta = document.getElementById('meta');
  const grid = document.getElementById('grid');
  meta.textContent = 'Загрузка…';
  try {
    const r = await fetch('/api/status');
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const d = JSON.parse(raw);
    meta.innerHTML = `Общий статус: <b class="${cls(d.status)}">${d.status}</b> · env: ${d.vercel_env || 'local'}`;
    const envRows = Object.entries(d.env || {}).map(([k,v]) => row(k, v, cls(v))).join('');
    const db = d.database || {};
    const mv = d.matview || {};
    const cache = d.cache || {};
    const admin = d.admin || {};
    grid.innerHTML = [
      card('Environment', envRows + row('db_env_error', d.db_env_error || 'нет', d.db_env_error ? 'err' : 'ok')),
      card('Database', row('status', db.status, cls(db.status)) + row('detail', db.detail) + row('row_count', db.row_count) + row('max_date', db.max_date) + row('total_amount', db.total_amount)),
      card('Matview', row('enabled', String(mv.enabled), cls(mv.enabled)) + row('available', String(mv.available), cls(mv.available)) + row('name', mv.name) + row('row_count', mv.row_count) + row('max_year/week', `${mv.max_year ?? '—'} / ${mv.max_week ?? '—'}`) + row('last_refresh_age_sec', mv.last_refresh_age_sec) + row('last_refresh_note', mv.last_refresh_note || '—') + row('bootstrap_ok_age_sec', mv.bootstrap_ok_age_sec) + row('bootstrap_fail', mv.bootstrap_fail_msg || 'нет', mv.bootstrap_fail_msg ? 'err' : 'ok')),
      card('Cache / Admin', row('cache_entries', cache.entries) + row('report_ttl_sec', cache.report_ttl_sec) + row('weeks_ttl_sec', cache.weeks_ttl_sec) + row('refresh_token_required', String(admin.refresh_token_required)) + row('active_sessions', admin.active_sessions)),
    ].join('');
    if (typeof revealRefresh === 'function') revealRefresh(grid);
  } catch (e) {
    meta.textContent = 'Ошибка: ' + (e.message || e);
  }
}
document.getElementById('btnReload').onclick = loadStatus;
loadStatus();
</script>

</div>
</div>
</body>
</html>
"""


NOMENCLATURE_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Кол-во брака по номенклатуре</title>
<style>
__SHARED_CSS__
</style>
</head>
<body>
<div class="app-shell">
<div class="app-top"><div class="app-top-inner">
<div class="brand-row">
  <h1>Номенклатура</h1>
  <div class="subtitle">Количество брака по nm_id × корпуса за последнюю неделю</div>
</div>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature" class="active">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/reason">Карточка</a>
  <a href="/digest">Дайджест</a>
  <a href="/actions">Действия</a>
  <a href="/status">Статус</a>
</nav>
</div></div>
<div class="page-body">

<section class="hero-band" data-reveal="fade-down">
  <div class="hero-kicker">Nomenclature · week</div>
  <h2>Номенклатура <span class="text-accent">× корпуса</span></h2>
  <p class="hero-lead">Количество брака по nm_id за последнюю неделю — таблица та же.</p>
</section>

  <section class="panel" data-reveal="zoom-in">
    <div class="toolbar">
      <a class="btn secondary" href="/">На главную</a>
      <button type="button" class="btn primary" id="btnNmExport">Экспорт XLSX</button>
    </div>
    <h2>Последняя неделя: Номенклатура × Корпуса</h2>
    <div class="meta" id="meta">Загрузка…</div>
    <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>Номенклатура</th>
          <th class="num">1 корпус</th>
          <th class="num">2 корпус</th>
          <th class="num">3 корпус</th>
          <th class="num">Итог</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    </div>
    <div class="meta muted">Метрика: количество записей брака по `nm_id` за последнюю доступную ISO-неделю.</div>
  </section>
<script>
function fmtInt(v) { return Number(v || 0).toLocaleString('ru-RU'); }
function fmtPct(v) { return Number(v || 0).toLocaleString('ru-RU', { minimumFractionDigits: 2, maximumFractionDigits: 2 }); }

async function loadData() {
  const meta = document.getElementById('meta');
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '';
  try {
    const r = await fetch('/api/nomenclature/latest');
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    const rows = data.rows || [];
    rows.forEach((it) => {
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td>${it.nomenclature ?? '—'}</td>
        <td class="num">${fmtInt(it.corpus_1)}</td>
        <td class="num">${fmtInt(it.corpus_2)}</td>
        <td class="num">${fmtInt(it.corpus_3)}</td>
        <td class="num">${fmtInt(it.total)}</td>
      `;
      tbody.appendChild(tr);
    });
    const totalTr = document.createElement('tr');
    totalTr.className = 'total';
    const t = data.totals || {};
    totalTr.innerHTML = `
      <td>Итого</td>
      <td class="num">${fmtInt(t.corpus_1)}</td>
      <td class="num">${fmtInt(t.corpus_2)}</td>
      <td class="num">${fmtInt(t.corpus_3)}</td>
      <td class="num">${fmtInt(t.total)}</td>
    `;
    tbody.appendChild(totalTr);
    const y = data.latest_year ?? '—';
    const w = data.latest_week ?? '—';
    meta.textContent = `Период: ${y}-я неделя ${w} · номенклатур: ${fmtInt(rows.length)} · записей: ${fmtInt(t.total)}`;
  } catch (e) {
    meta.textContent = 'Ошибка загрузки: ' + (e.message || e);
  }
}

async function exportNomenclatureXlsx() {
  const meta = document.getElementById('meta');
  try {
    meta.textContent = 'Готовим XLSX...';
    const r = await fetch('/api/nomenclature/export/xlsx');
    if (!r.ok) throw new Error(await r.text());
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'nomenclature_latest_week.xlsx';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    meta.textContent = 'XLSX сформирован';
  } catch (e) {
    meta.textContent = 'Ошибка экспорта: ' + (e.message || e);
  }
}

document.getElementById('btnNmExport').addEventListener('click', exportNomenclatureXlsx);
loadData();
</script>

</div>
</div>
</body>
</html>
"""



def _materialize_shared_assets() -> None:
    global DASHBOARD_HTML, DETAILS_HTML, WEEKLY_HTML, REASON_HTML
    global DIGEST_HTML, ACTIONS_HTML, STATUS_HTML, NOMENCLATURE_HTML
    pages = [
        "DASHBOARD_HTML",
        "DETAILS_HTML",
        "WEEKLY_HTML",
        "REASON_HTML",
        "DIGEST_HTML",
        "ACTIONS_HTML",
        "STATUS_HTML",
        "NOMENCLATURE_HTML",
    ]
    g = globals()
    for name in pages:
        html = g[name].replace("__SHARED_CSS__", SHARED_CSS)
        if "__SHARED_JS__" in html:
            html = html.replace("__SHARED_JS__", SHARED_JS)
        elif "</body>" in html:
            html = html.replace("</body>", SHARED_JS + "\n</body>", 1)
        g[name] = html


_materialize_shared_assets()

def register_routes(application) -> None:
    from datetime import datetime, timezone

    from flask import jsonify, request, send_file

    if getattr(application, "_brak_routes_registered", False):
        return
    application._brak_routes_registered = True

    def _cfg() -> dict:
        return load_config()

    def _resolve_wh_ids(cfg: dict) -> list[int] | None:
        wh_raw = request.args.get("wh_ids", "")
        catalog_ids = catalog_wh_ids(cfg)
        if wh_raw.strip():
            wh_ids = parse_wh_ids(wh_raw)
            if not wh_ids:
                raise QueryParamError("Некорректный wh_ids: список пуст")
            return wh_ids
        if catalog_ids:
            return catalog_ids
        return None

    def _year_arg(default: int | None = None) -> int | None:
        try:
            return parse_year_value(request.args.get("year", ""), default)
        except ValueError as exc:
            raise QueryParamError(str(exc)) from exc

    def _float_arg(name: str, default: float) -> float:
        raw = request.args.get(name, "")
        if raw is None or str(raw).strip() == "":
            return default
        try:
            return float(str(raw).strip().replace(",", "."))
        except (TypeError, ValueError) as exc:
            raise QueryParamError(f"Некорректный {name}: ожидается число") from exc

    def _int_arg(name: str, default: int | None = None) -> int | None:
        raw = request.args.get(name, "")
        if raw is None or str(raw).strip() == "":
            return default
        try:
            return int(str(raw).strip())
        except (TypeError, ValueError) as exc:
            raise QueryParamError(f"Некорректный {name}: ожидается целое число") from exc

    def _optional_wh_ids() -> list[int] | None:
        wh_raw = str(request.args.get("wh_ids", "") or "").strip()
        if not wh_raw:
            return None
        wh_ids = parse_wh_ids(wh_raw)
        if not wh_ids:
            raise QueryParamError("Некорректный wh_ids: список пуст")
        return wh_ids

    @application.route("/")
    def index():
        try:
            cfg = _cfg()
            embed = {
                "wh_catalog": cfg.get("wh_catalog", []),
                "buildings": cfg.get("buildings", []),
                "week_year": cfg.get("week_year", 2026),
                "week_prev": cfg.get("week_prev", 20),
                "week_last": cfg.get("week_last", 21),
                "is_admin": False,
                "refresh_token_required": True,
                "show_all_weeks_default": False,
            }
            page = DASHBOARD_HTML.replace(
                "__CONFIG_JSON__", json.dumps(embed, ensure_ascii=False)
            )
            return page
        except Exception as exc:
            return f"<pre>Index error: {exc}</pre>", 500

    @application.route("/nomenclature")
    def nomenclature_page():
        return NOMENCLATURE_HTML

    @application.route("/details")
    def details_page():
        return DETAILS_HTML

    @application.route("/weekly")
    def weekly_page():
        return WEEKLY_HTML

    @application.route("/status")
    def status_page():
        return STATUS_HTML

    @application.route("/reason")
    def reason_page():
        return REASON_HTML

    @application.route("/digest")
    def digest_page():
        return DIGEST_HTML

    @application.route("/actions")
    def actions_page():
        return ACTIONS_HTML

    @application.route("/api/admin/login", methods=["POST"])
    def api_admin_login():
        _, expected_password = _admin_login_password()
        if not expected_password:
            sid = _admin_session_create()
            return jsonify({"ok": True, "session_id": sid, "note": "auth disabled"})
        try:
            payload = request.get_json(silent=True) or {}
            got_login = str(payload.get("login", "")).strip()
            got_password = str(payload.get("password", "")).strip()
        except Exception:
            got_login = ""
            got_password = ""
        expected_login, expected_password = _admin_login_password()
        if got_login != expected_login or got_password != expected_password:
            return jsonify({"error": "Неверный логин или пароль"}), 403
        sid = _admin_session_create()
        return jsonify({"ok": True, "session_id": sid})

    @application.route("/api/admin/logout", methods=["POST"])
    def api_admin_logout():
        sid = request.headers.get("X-Admin-Session", "").strip() or request.args.get(
            "admin_session", ""
        ).strip()
        _admin_session_delete(sid)
        return jsonify({"ok": True})

    @application.route("/api/refresh", methods=["POST"])
    def api_refresh():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        access_err = check_admin_access()
        if access_err:
            return jsonify({"error": access_err}), 403

        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")

            refresh_note = run_db_refresh()
            _cache_clear_all()
            stats = fetch_db_stats(office_id, wh_ids)
            weeks = fetch_available_weeks(year, office_id, wh_ids)
            week_prev = cfg.get("week_prev", 20)
            week_last = cfg.get("week_last", 21)
            if len(weeks) >= 2:
                week_prev, week_last = weeks[-2], weeks[-1]

            warm_report_cache_async(wh_ids, office_id, year, week_prev, week_last)
            return jsonify(
                {
                    "ok": True,
                    "refreshed_at": datetime.now(timezone.utc).isoformat(),
                    "refresh_note": refresh_note,
                    "row_count": stats["row_count"],
                    "max_date": stats["max_date"],
                    "total_amount": stats["total_amount"],
                    "year": year,
                    "weeks": weeks,
                    "week_prev": week_prev,
                    "week_last": week_last,
                }
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/matview/refresh", methods=["POST"])
    def api_matview_refresh():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        access_err = check_admin_access()
        if access_err:
            return jsonify({"error": access_err}), 403
        try:
            note = _refresh_report_matview()
            _cache_clear_all()
            return jsonify({"ok": True, "note": note or "Матпредставление обновлено"})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/report")
    def api_report():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503

        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))

            def _week_arg(name: str, default: int) -> int:
                raw = request.args.get(name, "")
                if raw is None or str(raw).strip() == "":
                    return default
                try:
                    return int(raw)
                except (TypeError, ValueError):
                    return default

            week_prev = _week_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _week_arg("week_last", cfg.get("week_last", 21))
            compare_a = _int_arg("compare_a", week_prev)
            compare_b = _int_arg("compare_b", week_last)
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")
            show_all_weeks = (request.args.get("show_all_weeks", "0") == "1")
            alert_wow = _float_arg("alert_wow", 15.0)
            alert_vs_avg4 = _float_arg("alert_vs_avg4", 20.0)
            alert_min_amount = _float_arg("alert_min_amount", 50000.0)
            data = build_report_data(
                wh_ids, office_id, year, week_prev, week_last, show_all_weeks=show_all_weeks
            )
            corpus_compare = fetch_corpus_comparison(
                office_id=office_id,
                year=year,
                week_prev=week_prev,
                week_last=week_last,
                cfg=cfg,
            )
            dim_breakdowns = fetch_dimension_breakdowns(
                wh_ids=wh_ids,
                office_id=office_id,
                year=year,
                week_prev=week_prev,
                week_last=week_last,
                limit=10,
            )
            growth_alerts = build_growth_alerts(
                data,
                min_dynamics=alert_wow,
                min_vs_avg4=alert_vs_avg4,
                min_amount=alert_min_amount,
            )
            extras = build_report_compare_and_churn(
                data,
                week_prev=week_prev,
                week_last=week_last,
                week_a=compare_a,
                week_b=compare_b,
            )
            yoy = fetch_yoy_totals(
                year=year,
                week_last=week_last,
                office_id=office_id,
                wh_ids=wh_ids,
            )
            org0_series = fetch_org0_series(
                year=year, office_id=office_id, wh_ids=wh_ids, limit_weeks=12
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

        weeks = data["weeks"]
        html_grid = (
            render_table(
                "Дефект ТОП-20, рубли",
                data["defects"],
                data["defects_total"],
                show_id=True,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Дефект",
                all_totals=data["all_totals"]["defects"],
                drill_kind="reason",
                org0_only=False,
            )
            + render_table(
                "Дефект ТОП-20, ORG 0, рубли",
                data["defects_org0"],
                data["defects_org0_total"],
                show_id=True,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Дефект",
                all_totals=data["all_totals"]["defects_org0"],
                drill_kind="reason",
                org0_only=True,
            )
            + render_table(
                "ТОП-20 категорий, рубли",
                data["categories"],
                data["categories_total"],
                show_id=False,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Категория",
                all_totals=data["all_totals"]["categories"],
                drill_kind="category",
                org0_only=False,
            )
            + render_table(
                "ТОП-20 категорий, ORG 0, рубли",
                data["categories_org0"],
                data["categories_org0_total"],
                show_id=False,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Категория",
                all_totals=data["all_totals"]["categories_org0"],
                drill_kind="category",
                org0_only=True,
            )
        )

        payload = {
            "html": html_grid,
            "year": year,
            "weeks": weeks,
            "week_prev": week_prev,
            "week_last": week_last,
            "kpis": build_kpi_payload(data),
            "corpus_compare": corpus_compare,
            "dim_breakdowns": dim_breakdowns,
            "growth_alerts": growth_alerts,
            "alert_thresholds": {
                "wow": alert_wow,
                "vs_avg4": alert_vs_avg4,
                "min_amount": alert_min_amount,
            },
            "compare": extras["compare"],
            "top20_churn": extras["top20_churn"],
            "yoy": yoy,
            "org0_series": org0_series,
            "freshness": fetch_freshness_payload(),
        }
        etag = '"' + stable_etag_payload(
            {
                "year": year,
                "week_prev": week_prev,
                "week_last": week_last,
                "compare_a": compare_a,
                "compare_b": compare_b,
                "show_all_weeks": show_all_weeks,
                "wh_ids": sorted(wh_ids) if wh_ids else [],
                "alert_wow": alert_wow,
                "alert_vs_avg4": alert_vs_avg4,
                "alert_min_amount": alert_min_amount,
                "kpis": payload["kpis"],
                "growth_alerts": growth_alerts,
                "dim_breakdowns": dim_breakdowns,
                "compare": extras["compare"],
                "top20_churn": extras["top20_churn"],
                "yoy": yoy,
            }
        ) + '"'
        if request.headers.get("If-None-Match") == etag:
            return ("", 304, {"ETag": etag})
        resp = jsonify(payload)
        resp.headers["ETag"] = etag
        resp.headers["Cache-Control"] = "private, max-age=30"
        return resp

    @application.route("/api/export/xlsx")
    def api_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))

            def _week_arg(name: str, default: int) -> int:
                raw = request.args.get(name, "")
                if raw is None or str(raw).strip() == "":
                    return default
                try:
                    return int(raw)
                except (TypeError, ValueError):
                    return default

            week_prev = _week_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _week_arg("week_last", cfg.get("week_last", 21))
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")
            show_all_weeks = (request.args.get("show_all_weeks", "0") == "1")
            report = build_report_data(
                wh_ids, office_id, year, week_prev, week_last, show_all_weeks=show_all_weeks
            )
            blob = export_report_xlsx(report, year, week_prev, week_last)
            scope = "all"
            if wh_ids:
                scope = f"wh{len(wh_ids)}"
            filename = f"write_offs_{scope}_{year}_w{week_prev}-{week_last}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/weeks")
    def api_weeks():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            wh_ids = _resolve_wh_ids(cfg)
            weeks = fetch_available_weeks(year, cfg.get("office_id"), wh_ids)
            week_prev = cfg.get("week_prev", 20)
            week_last = cfg.get("week_last", 21)
            if len(weeks) >= 2:
                week_prev, week_last = weeks[-2], weeks[-1]
            elif len(weeks) == 1:
                week_prev = week_last = weeks[0]
            return jsonify(
                {
                    "year": year,
                    "weeks": weeks,
                    "week_prev": week_prev,
                    "week_last": week_last,
                }
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/wh_ids")
    def api_wh_ids():
        return jsonify(fetch_wh_list(_cfg().get("office_id")))

    @application.route("/api/details")
    def api_details():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            page = _detail_positive_int_arg(
                request.args, "page", 1, min_value=1, max_value=100000
            )
            per_page = _detail_positive_int_arg(
                request.args, "per_page", 50, min_value=5, max_value=500
            )
            payload = fetch_detail_page(request.args, page=page, per_page=per_page)
            return jsonify(payload)
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/details/export/xlsx")
    def api_details_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            limit = _detail_positive_int_arg(
                request.args, "limit", 5000, min_value=1, max_value=20000
            )
            blob = export_detail_xlsx(request.args, limit=limit)
            reason = request.args.get("reason_id") or "all"
            parent = re.sub(r"[^\w\-]+", "_", str(request.args.get("parent_name") or "all"))[:24]
            date_from = request.args.get("date_from") or "na"
            date_to = request.args.get("date_to") or "na"
            filename = f"write_offs_details_r{reason}_{parent}_{date_from}_{date_to}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/weekly")
    def api_weekly():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 5, min_value=1, max_value=20
            )
            wh_ids = _optional_wh_ids()
            payload = fetch_weekly_dynamics(
                year=year,
                office_id=cfg.get("office_id"),
                wh_ids=wh_ids,
                top_n=top_n,
            )
            return jsonify(payload)
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/status")
    def api_status():
        try:
            return jsonify(fetch_status_payload())
        except Exception as exc:
            return jsonify({"status": "error", "detail": str(exc)}), 500

    @application.route("/api/freshness")
    def api_freshness():
        try:
            return jsonify(fetch_freshness_payload())
        except Exception as exc:
            return jsonify({"ok": False, "error": str(exc)}), 500

    @application.route("/api/reason")
    def api_reason():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            week_last = request.args.get("week_last", type=int)
            reason_id = _detail_int_arg(request.args, "reason_id")
            parent_name = str(request.args.get("parent_name", "") or "").strip() or None
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 5, min_value=1, max_value=20
            )
            wh_ids = _optional_wh_ids()
            payload = fetch_reason_card(
                reason_id=reason_id,
                parent_name=parent_name,
                office_id=cfg.get("office_id"),
                wh_ids=wh_ids,
                year=year,
                week_last=week_last,
                top_n=top_n,
            )
            return jsonify(payload)
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/reason/export/xlsx")
    def api_reason_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            week_last = request.args.get("week_last", type=int)
            reason_id = _detail_int_arg(request.args, "reason_id")
            parent_name = str(request.args.get("parent_name", "") or "").strip() or None
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 5, min_value=1, max_value=20
            )
            wh_ids = _optional_wh_ids()
            payload = fetch_reason_card(
                reason_id=reason_id,
                parent_name=parent_name,
                office_id=cfg.get("office_id"),
                wh_ids=wh_ids,
                year=year,
                week_last=week_last,
                top_n=top_n,
            )
            blob = export_reason_xlsx(payload)
            rid = reason_id if reason_id is not None else "cat"
            filename = f"reason_{rid}_{year}_w{payload.get('week_last')}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/weekly/export/xlsx")
    def api_weekly_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 5, min_value=1, max_value=20
            )
            wh_ids = _optional_wh_ids()
            payload = fetch_weekly_dynamics(
                year=year,
                office_id=cfg.get("office_id"),
                wh_ids=wh_ids,
                top_n=top_n,
            )
            blob = export_weekly_xlsx(payload)
            filename = f"weekly_{year}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/search")
    def api_search():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            wh_ids = _optional_wh_ids()
            q = str(request.args.get("q", "") or "")
            limit = _detail_positive_int_arg(
                request.args, "limit", 15, min_value=1, max_value=30
            )
            return jsonify(
                fetch_search(
                    q=q,
                    year=year,
                    office_id=cfg.get("office_id"),
                    wh_ids=wh_ids,
                    limit=limit,
                )
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/digest")
    def api_digest():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))

            def _week_arg(name: str, default: int) -> int:
                raw = request.args.get(name, "")
                if raw is None or str(raw).strip() == "":
                    return default
                try:
                    return int(raw)
                except (TypeError, ValueError):
                    return default

            week_prev = _week_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _week_arg("week_last", cfg.get("week_last", 21))
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")
            alert_wow = _float_arg("alert_wow", 15.0)
            alert_vs_avg4 = _float_arg("alert_vs_avg4", 20.0)
            alert_min_amount = _float_arg("alert_min_amount", 50000.0)
            data = build_report_data(
                wh_ids, office_id, year, week_prev, week_last, show_all_weeks=False
            )
            extras = build_report_compare_and_churn(
                data, week_prev=week_prev, week_last=week_last
            )
            return jsonify(
                {
                    "year": year,
                    "week_prev": week_prev,
                    "week_last": week_last,
                    "kpis": build_kpi_payload(data),
                    "yoy": fetch_yoy_totals(
                        year=year,
                        week_last=week_last,
                        office_id=office_id,
                        wh_ids=wh_ids,
                    ),
                    "growth_alerts": build_growth_alerts(
                        data,
                        min_dynamics=alert_wow,
                        min_vs_avg4=alert_vs_avg4,
                        min_amount=alert_min_amount,
                        limit=10,
                    ),
                    "top20_churn": extras["top20_churn"],
                    "compare": extras["compare"],
                    "corpus_compare": fetch_corpus_comparison(
                        office_id=office_id,
                        year=year,
                        week_prev=week_prev,
                        week_last=week_last,
                        cfg=cfg,
                    ),
                    "freshness": fetch_freshness_payload(),
                }
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/heatmap")
    def api_heatmap():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            wh_ids = _optional_wh_ids()
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 12, min_value=3, max_value=20
            )
            week_limit = _detail_positive_int_arg(
                request.args, "week_limit", 12, min_value=4, max_value=26
            )
            return jsonify(
                fetch_reason_heatmap(
                    year=year,
                    office_id=cfg.get("office_id"),
                    wh_ids=wh_ids,
                    top_n=top_n,
                    week_limit=week_limit,
                )
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/corpus/reasons")
    def api_corpus_reasons():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            corpus = _int_arg("corpus", None)
            if corpus is None or corpus not in (1, 2, 3):
                raise QueryParamError("Укажите corpus=1|2|3")
            week_prev = _int_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _int_arg("week_last", cfg.get("week_last", 21))
            if week_prev is None or week_last is None:
                raise QueryParamError("Укажите week_prev и week_last")
            limit = _detail_positive_int_arg(
                request.args, "limit", 10, min_value=3, max_value=20
            )
            return jsonify(
                fetch_corpus_reasons(
                    corpus=corpus,
                    year=year,
                    week_prev=week_prev,
                    week_last=week_last,
                    office_id=cfg.get("office_id"),
                    cfg=cfg,
                    limit=limit,
                )
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/watchlist")
    def api_watchlist():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(cfg.get("week_year", 2026))
            week_prev = _int_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _int_arg("week_last", cfg.get("week_last", 21))
            if week_prev is None or week_last is None:
                raise QueryParamError("Укажите week_prev и week_last")
            raw_ids = str(request.args.get("reason_ids", "") or "").strip()
            reason_ids = parse_wh_ids(raw_ids) if raw_ids else []
            wh_ids = _optional_wh_ids()
            return jsonify(
                fetch_watchlist_status(
                    reason_ids=reason_ids,
                    year=year,
                    week_prev=week_prev,
                    week_last=week_last,
                    office_id=cfg.get("office_id"),
                    wh_ids=wh_ids,
                )
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/nomenclature/latest")
    def api_nomenclature_latest():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(None)
            data = fetch_nomenclature_counts_latest_week(
                office_id=cfg.get("office_id"),
                year=year,
            )
            return jsonify(data)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/nomenclature/export/xlsx")
    def api_nomenclature_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = _year_arg(None)
            payload = fetch_nomenclature_counts_latest_week(
                office_id=cfg.get("office_id"),
                year=year,
            )
            limit = request.args.get("limit", type=int)
            if limit and limit > 0:
                rows = (payload.get("rows") or [])[:limit]
                payload = {
                    "latest_year": payload.get("latest_year"),
                    "latest_week": payload.get("latest_week"),
                    "rows": rows,
                    "totals": {
                        "corpus_1": sum(int(r.get("corpus_1", 0)) for r in rows),
                        "corpus_2": sum(int(r.get("corpus_2", 0)) for r in rows),
                        "corpus_3": sum(int(r.get("corpus_3", 0)) for r in rows),
                        "total": sum(int(r.get("total", 0)) for r in rows),
                    },
                }
            blob = export_nomenclature_xlsx(payload)
            y = payload.get("latest_year") or "na"
            w = payload.get("latest_week") or "na"
            filename = f"nomenclature_{y}_{w}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/health")
    def health():
        return jsonify({"status": "ok", "config": str(CONFIG_PATH.name)})

    @application.route("/health/env")
    def health_env():
        def _state(name: str) -> str:
            if name not in os.environ:
                return "missing"
            if os.environ.get(name, "").strip() == "":
                return "empty"
            return "set"

        return jsonify(
            {
                "status": "ok",
                "vercel_env": os.environ.get("VERCEL_ENV", "local"),
                "db_env_error": check_db_env(),
                "env": {
                    "DATABASE_URL": _state("DATABASE_URL"),
                    "DB_HOST": _state("DB_HOST"),
                    "DB_PORT": _state("DB_PORT"),
                    "DB_NAME": _state("DB_NAME"),
                    "DB_USER": _state("DB_USER"),
                    "DB_PASSWORD": _state("DB_PASSWORD"),
                    "DB_SSLMODE": _state("DB_SSLMODE"),
                },
            }
        )

    @application.route("/health/db")
    def health_db():
        env_err = check_db_env()
        if env_err:
            return jsonify({"status": "error", "detail": env_err}), 503
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
            return jsonify({"status": "ok", "database": "connected"})
        except Exception as exc:
            return jsonify({"status": "error", "detail": str(exc)}), 503


def create_app():
    from flask import Flask

    application = Flask(__name__)
    register_routes(application)
    return application


def run_server() -> None:
    host = os.environ.get("HTML_HOST", "127.0.0.1")
    port = int(os.environ.get("HTML_PORT", "8080"))
    print(f"Дашборд: http://{host}:{port}/")
    print(f"Корпуса: {CONFIG_PATH}")
    create_app().run(host=host, port=port, debug=False)


def main() -> int:
    if not os.environ.get("DB_PASSWORD"):
        print("Создайте .env с DB_PASSWORD", file=sys.stderr)
        return 1
    print(f"Кэш отчета: {CACHE_TTL_SEC}s, кэш недель: {WEEKS_CACHE_TTL_SEC}s")
    try:
        run_server()
    except Exception as e:
        print(f"Ошибка: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
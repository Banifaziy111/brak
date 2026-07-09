#!/usr/bin/env python3
"""
Дашборд brak_team.write_offs — 4 таблицы ТОП-20 как в отчёте.

  python write_offs_dashboard.py
  → http://127.0.0.1:8080/

Фильтр: все WH, корпус (несколько wh_id из wh_buildings.json) или свой набор галочками.
Настройка корпусов: wh_buildings.json
"""

from __future__ import annotations

import json
import os
import re
import secrets
import sys
from io import BytesIO
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from threading import RLock
from time import monotonic
from typing import Any

from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "wh_buildings.json"

DETAIL_COLUMNS = (
    ("shk_id", "ШК"),
    ("date", "Дата"),
    ("type", "Тип"),
    ("total_cost", "Стоимость"),
    ("amount", "Сумма"),
    ("share", "Доля"),
    ("office_id", "Офис"),
    ("wh_id", "Блок"),
    ("nm_id", "nm_id"),
    ("subject_name", "Предмет"),
    ("parent_name", "Родитель"),
    ("title", "Наименование"),
    ("brand_name", "Бренд"),
    ("state_id", "state_id"),
    ("reason_id", "reason_id"),
    ("reason_descr", "Причина"),
    ("seller_id", "seller_id"),
    ("supplier_id", "supplier_id"),
    ("owner_product", "Владелец"),
    ("cnt_org", "cnt_org"),
    ("cnt_ors", "cnt_ors"),
    ("cnt_ocr", "cnt_ocr"),
)
DETAIL_COL_NAMES = [c[0] for c in DETAIL_COLUMNS]
DETAIL_SORTABLE = {
    "date",
    "amount",
    "total_cost",
    "share",
    "office_id",
    "wh_id",
    "nm_id",
    "shk_id",
    "reason_id",
    "cnt_org",
    "cnt_ors",
    "cnt_ocr",
    "type",
    "parent_name",
    "reason_descr",
    "title",
    "brand_name",
    "subject_name",
}
DETAIL_CLICK_FILTERS = {
    "wh_id",
    "office_id",
    "nm_id",
    "shk_id",
    "reason_id",
    "parent_name",
    "type",
    "cnt_org",
    "brand_name",
}

_env_file = ROOT / ".env"
if _env_file.exists():
    load_dotenv(_env_file)

CACHE_TTL_SEC = int(os.environ.get("REPORT_CACHE_TTL_SEC", "120"))
WEEKS_CACHE_TTL_SEC = int(os.environ.get("WEEKS_CACHE_TTL_SEC", "300"))
NM_CACHE_TTL_SEC = int(os.environ.get("NM_CACHE_TTL_SEC", "300"))
MATVIEW_BOOTSTRAP_TTL_SEC = int(os.environ.get("MATVIEW_BOOTSTRAP_TTL_SEC", "600"))
_CACHE_LOCK = RLock()
_CACHE: dict[str, tuple[float, Any]] = {}
_MV_LOCK = RLock()
_MV_BOOTSTRAP_OK_AT = 0.0
_MV_BOOTSTRAP_FAIL_AT = 0.0
_MV_BOOTSTRAP_FAIL_MSG = ""
_MV_TEMP_SUFFIX = "__new"
_ADMIN_SESSIONS_LOCK = RLock()
_ADMIN_SESSIONS: dict[str, float] = {}
ADMIN_SESSION_TTL_SEC = int(os.environ.get("ADMIN_SESSION_TTL_SEC", "28800"))
_PG_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _cache_get(key: str) -> Any | None:
    now = monotonic()
    with _CACHE_LOCK:
        item = _CACHE.get(key)
        if not item:
            return None
        expires_at, payload = item
        if expires_at < now:
            _CACHE.pop(key, None)
            return None
        return payload


def _cache_set(key: str, payload: Any, ttl_sec: int) -> None:
    with _CACHE_LOCK:
        _CACHE[key] = (monotonic() + max(1, ttl_sec), payload)


def _cache_clear_all() -> None:
    with _CACHE_LOCK:
        _CACHE.clear()


def _use_report_matview() -> bool:
    return os.environ.get("USE_WRITE_OFFS_MATVIEW", "1").strip().lower() not in (
        "",
        "0",
        "false",
        "no",
    )


def _matview_name() -> str:
    return os.environ.get(
        "WRITE_OFFS_MATVIEW_NAME", "brak_team.write_offs_weekly_mv"
    ).strip() or "brak_team.write_offs_weekly_mv"


def _nm_matview_name() -> str:
    return os.environ.get(
        "WRITE_OFFS_NM_MATVIEW_NAME", "brak_team.write_offs_nm_weekly_mv"
    ).strip() or "brak_team.write_offs_nm_weekly_mv"


def _quote_pg_relation_name(name: str) -> str:
    parts = [p.strip() for p in name.split(".")]
    if len(parts) not in (1, 2) or any(not p for p in parts):
        raise ValueError(f"Некорректное имя PostgreSQL relation: {name!r}")
    for part in parts:
        if not _PG_IDENTIFIER_RE.fullmatch(part):
            raise ValueError(f"Некорректное имя PostgreSQL identifier: {part!r}")
    return ".".join(f'"{part}"' for part in parts)


def _relation_basename(name: str) -> str:
    return name.split(".")[-1].strip()


def _ensure_report_matview() -> None:
    """
    Ensures weekly aggregated materialized view exists.
    Expensive DDL check is throttled by MATVIEW_BOOTSTRAP_TTL_SEC.
    """
    global _MV_BOOTSTRAP_OK_AT, _MV_BOOTSTRAP_FAIL_AT, _MV_BOOTSTRAP_FAIL_MSG
    if not _use_report_matview():
        return
    now = monotonic()
    with _MV_LOCK:
        if _MV_BOOTSTRAP_OK_AT and (now - _MV_BOOTSTRAP_OK_AT) < MATVIEW_BOOTSTRAP_TTL_SEC:
            return
        if _MV_BOOTSTRAP_FAIL_AT and (now - _MV_BOOTSTRAP_FAIL_AT) < MATVIEW_BOOTSTRAP_TTL_SEC:
            raise RuntimeError(_MV_BOOTSTRAP_FAIL_MSG or "matview bootstrap failed")

    mv = _matview_name()
    mv_tmp = f"{mv}{_MV_TEMP_SUFFIX}"
    mv_sql = _quote_pg_relation_name(mv)
    mv_tmp_sql = _quote_pg_relation_name(mv_tmp)
    mv_basename_sql = _quote_pg_relation_name(_relation_basename(mv))
    ddl_tmp = f"""
        CREATE MATERIALIZED VIEW IF NOT EXISTS {mv_tmp_sql} AS
        SELECT
            EXTRACT(ISOYEAR FROM date)::int AS iso_year,
            EXTRACT(WEEK FROM date)::int AS week_no,
            office_id,
            wh_id,
            cnt_org,
            reason_id,
            COALESCE(reason_descr, '—') AS reason_descr,
            COALESCE(parent_name, '—') AS parent_name,
            SUM(amount)::numeric AS amount_sum,
            COUNT(*)::bigint AS rows_cnt,
            MAX(date) AS max_date
        FROM brak_team.write_offs
        WHERE date IS NOT NULL
        GROUP BY 1, 2, 3, 4, 5, 6, 7, 8
    """
    idx = [
        f"CREATE UNIQUE INDEX IF NOT EXISTS write_offs_weekly_mv_uq ON {mv_sql} (iso_year, week_no, office_id, wh_id, cnt_org, reason_id, reason_descr, parent_name)",
        f"CREATE INDEX IF NOT EXISTS write_offs_weekly_mv_filter_idx ON {mv_sql} (iso_year, office_id, wh_id, week_no, cnt_org)",
        f"CREATE INDEX IF NOT EXISTS write_offs_weekly_mv_week_idx ON {mv_sql} (iso_year, week_no)",
    ]
    try:
        with get_conn() as conn:
            conn.autocommit = True
            cur = conn.cursor()
            cur.execute(f"DROP MATERIALIZED VIEW IF EXISTS {mv_tmp_sql} CASCADE")
            cur.execute("SELECT to_regclass(%s)", (mv,))
            mv_regclass = cur.fetchone()
            if mv_regclass and mv_regclass[0]:
                cur.execute(
                    """
                    SELECT attname
                    FROM pg_attribute
                    WHERE attrelid = to_regclass(%s)
                      AND attnum > 0
                      AND NOT attisdropped
                    """,
                    (mv,),
                )
                cols = {str(r[0]) for r in cur.fetchall()}
                required = {
                    "iso_year",
                    "week_no",
                    "office_id",
                    "wh_id",
                    "cnt_org",
                    "reason_id",
                    "reason_descr",
                    "parent_name",
                    "amount_sum",
                    "rows_cnt",
                    "max_date",
                }
                if ("nm_id" in cols) or (not required.issubset(cols)):
                    cur.execute(f"DROP MATERIALIZED VIEW IF EXISTS {mv_sql} CASCADE")
            cur.execute("SELECT to_regclass(%s)", (mv,))
            mv_regclass = cur.fetchone()
            if not mv_regclass or not mv_regclass[0]:
                cur.execute(ddl_tmp)
                cur.execute(
                    f"ALTER MATERIALIZED VIEW {mv_tmp_sql} RENAME TO {mv_basename_sql}"
                )
            for stmt in idx:
                cur.execute(stmt)
        with _MV_LOCK:
            _MV_BOOTSTRAP_OK_AT = monotonic()
            _MV_BOOTSTRAP_FAIL_AT = 0.0
            _MV_BOOTSTRAP_FAIL_MSG = ""
    except Exception as exc:
        with _MV_LOCK:
            _MV_BOOTSTRAP_FAIL_AT = monotonic()
            _MV_BOOTSTRAP_FAIL_MSG = f"matview bootstrap failed: {exc}"
        raise


def _ensure_nm_matview() -> None:
    if not _use_report_matview():
        return
    mv = _nm_matview_name()
    mv_sql = _quote_pg_relation_name(mv)
    ddl = f"""
        CREATE MATERIALIZED VIEW IF NOT EXISTS {mv_sql} AS
        SELECT
            EXTRACT(ISOYEAR FROM date)::int AS iso_year,
            EXTRACT(WEEK FROM date)::int AS week_no,
            office_id,
            wh_id,
            nm_id,
            COUNT(*)::bigint AS rows_cnt,
            MAX(date) AS max_date
        FROM brak_team.write_offs
        WHERE date IS NOT NULL
          AND nm_id IS NOT NULL
        GROUP BY 1, 2, 3, 4, 5
    """
    idx = [
        f"CREATE UNIQUE INDEX IF NOT EXISTS write_offs_nm_weekly_mv_uq ON {mv_sql} (iso_year, week_no, office_id, wh_id, nm_id)",
        f"CREATE INDEX IF NOT EXISTS write_offs_nm_weekly_mv_filter_idx ON {mv_sql} (iso_year, office_id, week_no, wh_id)",
    ]
    with get_conn() as conn:
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute(ddl)
        for stmt in idx:
            cur.execute(stmt)


def _refresh_report_matview() -> str | None:
    if not _use_report_matview():
        return None
    _ensure_report_matview()
    mv = _matview_name()
    mv_sql = _quote_pg_relation_name(mv)
    nm_mv = _nm_matview_name()
    nm_mv_sql = _quote_pg_relation_name(nm_mv)
    _ensure_nm_matview()
    use_concurrently = os.environ.get(
        "WRITE_OFFS_MATVIEW_REFRESH_CONCURRENTLY", "1"
    ).strip().lower() not in ("0", "false", "no")
    with get_conn() as conn:
        conn.autocommit = True
        cur = conn.cursor()
        if use_concurrently:
            try:
                cur.execute(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {mv_sql}")
                try:
                    cur.execute(f"REFRESH MATERIALIZED VIEW CONCURRENTLY {nm_mv_sql}")
                    return "Матпредставления обновлены (CONCURRENTLY)"
                except Exception:
                    cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
                    return "Матпредставления обновлены"
            except Exception:
                cur.execute(f"REFRESH MATERIALIZED VIEW {mv_sql}")
                cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
                return "Матпредставления обновлены"
        cur.execute(f"REFRESH MATERIALIZED VIEW {mv_sql}")
        cur.execute(f"REFRESH MATERIALIZED VIEW {nm_mv_sql}")
        return "Матпредставления обновлены"


def _report_source() -> dict[str, str]:
    if _use_report_matview():
        try:
            _ensure_report_matview()
            return {
                "table": _matview_name(),
                "week_expr": "week_no",
                "isoyear_expr": "iso_year",
                "amount_expr": "amount_sum",
                "year_clause": "iso_year = %s",
                "base_not_null_clause": "",
            }
        except Exception as exc:
            print(
                f"[write_offs] matview unavailable, fallback to base table: {exc}",
                file=sys.stderr,
            )
    return {
        "table": "brak_team.write_offs",
        "week_expr": "EXTRACT(WEEK FROM date)::int",
        "isoyear_expr": "EXTRACT(ISOYEAR FROM date)::int",
        "amount_expr": "amount",
        "year_clause": "date >= %s AND date < %s",
        "base_not_null_clause": "date IS NOT NULL",
    }


@dataclass
class Row:
    row_id: int | None
    name: str
    amounts: dict[int, float]

    def amount(self, week: int) -> float:
        return self.amounts.get(week, 0.0)

    def dynamics(self, week_prev: int, week_last: int) -> float | None:
        w_prev, w_last = self.amount(week_prev), self.amount(week_last)
        if w_prev == 0:
            return None  # Исправлено: рост с нуля — это None, а не 100%
        return (w_last - w_prev) / w_prev * 100

    def average(self, week_prev: int, week_last: int) -> float:
        return (self.amount(week_prev) + self.amount(week_last)) / 2

    def pct_vs_avg(self, week_prev: int, week_last: int) -> float | None:
        avg = self.average(week_prev, week_last)
        if avg == 0:
            return None
        return self.amount(week_last) / avg * 100


class QueryParamError(ValueError):
    pass


def parse_wh_ids(raw: str) -> list[int]:
    wh_ids: list[int] = []
    seen: set[int] = set()
    for item in raw.split(","):
        part = item.strip()
        if not part:
            continue
        try:
            wh_id = int(part)
        except (TypeError, ValueError) as exc:
            raise QueryParamError(
                f"Некорректный wh_ids: значение {part!r} не является числом"
            ) from exc
        if wh_id not in seen:
            seen.add(wh_id)
            wh_ids.append(wh_id)
    return wh_ids


def normalize_database_url(url: str) -> str:
    from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

    url = url.strip()
    if "://" in url:
        scheme, rest = url.split("://", 1)
        if "+" in scheme:
            scheme = scheme.split("+", 1)[0]
        url = f"{scheme}://{rest}"

    parsed = urlparse(url)
    if not parsed.scheme:
        return url

    query = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k and v]
    if not query and parsed.query:
        pass
    fixed = parsed._replace(query=urlencode(query))
    return urlunparse(fixed)


def db_config() -> dict[str, Any]:
    timeout = int(os.environ.get("DB_CONNECT_TIMEOUT", "15"))
    host = os.environ.get("DB_HOST", "").strip()
    user = os.environ.get("DB_USER", "").strip()
    password = os.environ.get("DB_PASSWORD", "")

    if host and user and password:
        cfg: dict[str, Any] = {
            "host": host,
            "port": int(os.environ.get("DB_PORT", "5432")),
            "dbname": os.environ.get("DB_NAME", "botdb"),
            "user": user,
            "password": password,
            "connect_timeout": timeout,
        }
        sslmode = os.environ.get("DB_SSLMODE")
        if sslmode:
            cfg["sslmode"] = sslmode
        return cfg

    database_url = os.environ.get("DATABASE_URL", "").strip()
    if database_url:
        return {"dsn": normalize_database_url(database_url), "connect_timeout": timeout}

    return {
        "host": host or "localhost",
        "port": int(os.environ.get("DB_PORT", "5432")),
        "dbname": os.environ.get("DB_NAME", "botdb"),
        "user": user,
        "password": password,
        "connect_timeout": timeout,
    }


def check_db_env() -> str | None:
    if os.environ.get("DATABASE_URL", "").strip():
        return None
    missing = [k for k in ("DB_HOST", "DB_USER", "DB_PASSWORD") if not os.environ.get(k)]
    if missing:
        return "Не заданы: DATABASE_URL или " + ", ".join(missing)
    return None


@contextmanager
def get_conn():
    import psycopg2

    conn = psycopg2.connect(**db_config())
    try:
        yield conn
    finally:
        conn.close()


def load_config() -> dict:
    if CONFIG_PATH.exists():
        return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    return {
        "office_id": None,
        "week_year": 2026,
        "week_prev": 20,
        "week_last": 21,
        "wh_catalog": [],
        "buildings": [{"id": "all", "name": "Все WH", "wh_ids": []}],
    }


def catalog_wh_ids(cfg: dict) -> list[int]:
    catalog = cfg.get("wh_catalog") or []
    return [int(w["wh_id"]) for w in catalog]


def to_float(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, Decimal):
        return float(v)
    return float(v)


def fetch_wh_list(office_id: int | None) -> list[dict]:
    clauses = []
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"""
        SELECT wh_id,
               COUNT(*) AS cnt,
               ROUND(SUM(amount)::numeric, 0) AS total_amount
        FROM brak_team.write_offs
        {where}
        GROUP BY wh_id
        ORDER BY wh_id
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
    return [
        {"wh_id": r[0], "cnt": r[1], "total_amount": to_float(r[2])}
        for r in rows
    ]


def fetch_db_stats(
    office_id: int | None,
    wh_ids: list[int] | None,
) -> dict[str, Any]:
    clauses: list[str] = []
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = (" WHERE " + " AND ".join(clauses)) if clauses else ""
    sql = f"""
        SELECT COUNT(*)::bigint,
               MAX(date)::text,
               COALESCE(ROUND(SUM(amount)::numeric, 0), 0)
        FROM brak_team.write_offs
        {where}
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        row = cur.fetchone()
    return {
        "row_count": int(row[0]) if row else 0,
        "max_date": row[1] if row else None,
        "total_amount": to_float(row[2]) if row else 0.0,
    }


def _detail_int_arg(args: Any, name: str) -> int | None:
    raw = args.get(name, "")
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return int(str(raw).strip())
    except (TypeError, ValueError) as exc:
        raise QueryParamError(f"Некорректный {name}: значение должно быть числом") from exc


def _detail_positive_int_arg(
    args: Any,
    name: str,
    default: int,
    *,
    min_value: int,
    max_value: int,
) -> int:
    value = _detail_int_arg(args, name)
    if value is None:
        value = default
    return max(min_value, min(max_value, value))


def build_detail_where(args: Any) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []

    office_id = _detail_int_arg(args, "office_id")
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)

    wh_id = _detail_int_arg(args, "wh_id")
    wh_raw = str(args.get("wh_ids", "") or "").strip()
    if wh_id is not None and wh_raw:
        raise QueryParamError("Передайте только один фильтр: wh_id или wh_ids")
    if wh_id is not None:
        clauses.append("wh_id = %s")
        params.append(wh_id)
    elif wh_raw:
        wh_ids = parse_wh_ids(wh_raw)
        if wh_ids:
            clauses.append("wh_id = ANY(%s)")
            params.append(wh_ids)

    row_type = str(args.get("type", "") or "").strip()
    if row_type:
        clauses.append("type = %s")
        params.append(row_type)

    date_from = str(args.get("date_from", "") or "").strip()
    if date_from:
        clauses.append("date >= %s")
        params.append(date_from)

    date_to = str(args.get("date_to", "") or "").strip()
    if date_to:
        clauses.append("date < %s::date + interval '1 day'")
        params.append(date_to)

    reason_id = _detail_int_arg(args, "reason_id")
    if reason_id is not None:
        clauses.append("reason_id = %s")
        params.append(reason_id)

    parent_name = str(args.get("parent_name", "") or "").strip()
    if parent_name:
        clauses.append("COALESCE(parent_name, '—') = %s")
        params.append(parent_name)

    nm_id = _detail_int_arg(args, "nm_id")
    if nm_id is not None:
        clauses.append("nm_id = %s")
        params.append(nm_id)

    shk_id = _detail_int_arg(args, "shk_id")
    if shk_id is not None:
        clauses.append("shk_id = %s")
        params.append(shk_id)

    cnt_org = _detail_int_arg(args, "cnt_org")
    if cnt_org is not None:
        clauses.append("cnt_org = %s")
        params.append(cnt_org)

    brand_name = str(args.get("brand_name", "") or "").strip()
    if brand_name:
        clauses.append("COALESCE(brand_name, '') = %s")
        params.append(brand_name)

    search = str(args.get("search", "") or "").strip()
    if search:
        like = f"%{search}%"
        clauses.append(
            "("
            "title ILIKE %s OR reason_descr ILIKE %s OR brand_name ILIKE %s OR "
            "subject_name ILIKE %s OR parent_name ILIKE %s OR "
            "nm_id::text ILIKE %s OR shk_id::text ILIKE %s"
            ")"
        )
        params.extend([like, like, like, like, like, like, like])

    if not clauses:
        return "", params
    return " WHERE " + " AND ".join(clauses), params


def _detail_row_to_dict(row: tuple[Any, ...]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for idx, col in enumerate(DETAIL_COL_NAMES):
        value = row[idx]
        if isinstance(value, Decimal):
            value = float(value)
        elif hasattr(value, "isoformat"):
            value = value.isoformat()
        out[col] = value
    return out


def _detail_order_sql(args: Any) -> str:
    sort_by = str(args.get("sort_by", "") or "").strip()
    sort_dir = str(args.get("sort_dir", "") or "").strip().lower()
    if sort_by not in DETAIL_SORTABLE:
        sort_by = "date"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    nulls = "NULLS LAST" if sort_dir == "desc" else "NULLS FIRST"
    if sort_by == "date":
        return f"ORDER BY date {sort_dir} {nulls}, shk_id"
    return f"ORDER BY {sort_by} {sort_dir} {nulls}, date DESC NULLS LAST, shk_id"


def fetch_detail_page(
    args: Any,
    *,
    page: int,
    per_page: int,
) -> dict[str, Any]:
    where_sql, params = build_detail_where(args)
    cols = ", ".join(DETAIL_COL_NAMES)
    offset = (page - 1) * per_page
    order_sql = _detail_order_sql(args)
    sort_by = str(args.get("sort_by", "") or "").strip() or "date"
    sort_dir = str(args.get("sort_dir", "") or "").strip().lower() or "desc"
    if sort_by not in DETAIL_SORTABLE:
        sort_by = "date"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(f"SELECT COUNT(*)::bigint FROM brak_team.write_offs{where_sql}", params)
        total_row = cur.fetchone()
        total = int(total_row[0]) if total_row else 0
        cur.execute(
            f"""
            SELECT {cols}
            FROM brak_team.write_offs
            {where_sql}
            {order_sql}
            LIMIT %s OFFSET %s
            """,
            [*params, per_page, offset],
        )
        rows = cur.fetchall()

    pages = max(1, (total + per_page - 1) // per_page)
    return {
        "columns": [{"key": key, "label": label} for key, label in DETAIL_COLUMNS],
        "rows": [_detail_row_to_dict(tuple(r)) for r in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": pages,
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "click_filters": sorted(DETAIL_CLICK_FILTERS),
    }


def run_db_refresh() -> str | None:
    import urllib.error
    import urllib.request

    notes: list[str] = []
    sql = os.environ.get("DB_REFRESH_SQL", "").strip()
    if sql:
        with get_conn() as conn:
            conn.autocommit = True
            cur = conn.cursor()
            cur.execute(sql)
            try:
                row = cur.fetchone()
                if row and row[0] is not None:
                    notes.append(str(row[0]))
                else:
                    notes.append("SQL обновления выполнен")
            except Exception:
                notes.append("SQL обновления выполнен")

    url = os.environ.get("DB_REFRESH_URL", "").strip()
    if url:
        method = os.environ.get("DB_REFRESH_METHOD", "POST").upper()
        req = urllib.request.Request(url, method=method)
        token = os.environ.get("DB_REFRESH_TOKEN", "").strip()
        if token:
            req.add_header("Authorization", f"Bearer {token}")
        timeout = int(os.environ.get("DB_REFRESH_TIMEOUT", "120"))
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                body = resp.read().decode("utf-8", errors="replace")[:300]
                notes.append(body or f"HTTP {resp.status}")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:300]
            raise RuntimeError(f"DB_REFRESH_URL: HTTP {exc.code} {detail}") from exc

    try:
        mv_note = _refresh_report_matview()
        if mv_note:
            notes.append(mv_note)
    except Exception as exc:
        notes.append(f"matview refresh error: {exc}")
    if notes:
        return " | ".join(notes)
    return None


def check_refresh_access() -> str | None:
    token = os.environ.get("REFRESH_API_TOKEN", "").strip()
    if not token:
        return None

    from flask import request

    got = request.headers.get("X-Refresh-Token", "").strip() or request.args.get(
        "refresh_token", ""
    ).strip()
    if got != token:
        return "Недостаточно прав для обновления данных"
    return None


def _admin_login_password() -> tuple[str, str]:
    login = os.environ.get("ADMIN_LOGIN", "admin").strip() or "admin"
    password = os.environ.get("ADMIN_PASSWORD", "").strip()
    if not password:
        password = os.environ.get("REFRESH_API_TOKEN", "").strip()
    return login, password


def _admin_session_create() -> str:
    sid = secrets.token_urlsafe(32)
    with _ADMIN_SESSIONS_LOCK:
        _ADMIN_SESSIONS[sid] = monotonic() + max(60, ADMIN_SESSION_TTL_SEC)
    return sid


def _admin_session_valid(sid: str) -> bool:
    if not sid:
        return False
    now = monotonic()
    with _ADMIN_SESSIONS_LOCK:
        exp = _ADMIN_SESSIONS.get(sid)
        if exp is None:
            return False
        if exp < now:
            _ADMIN_SESSIONS.pop(sid, None)
            return False
        _ADMIN_SESSIONS[sid] = now + max(60, ADMIN_SESSION_TTL_SEC)
        return True


def _admin_session_delete(sid: str) -> None:
    if not sid:
        return
    with _ADMIN_SESSIONS_LOCK:
        _ADMIN_SESSIONS.pop(sid, None)


def check_admin_session_access() -> str | None:
    from flask import request

    sid = request.headers.get("X-Admin-Session", "").strip() or request.args.get(
        "admin_session", ""
    ).strip()
    if not _admin_session_valid(sid):
        return "Требуется вход администратора"
    return None


def check_admin_access() -> str | None:
    session_err = check_admin_session_access()
    if session_err is None:
        return None
    if not os.environ.get("REFRESH_API_TOKEN", "").strip():
        return session_err
    legacy_err = check_refresh_access()
    if legacy_err is None:
        return None
    return session_err


def fetch_totals_all(
    *,
    wh_ids: list[int] | None,
    office_id: int | None,
    org0_only: bool,
    year: int,
    week_prev: int,
    week_last: int,
) -> dict[str, float]:
    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != "brak_team.write_offs":
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    if org0_only:
        clauses.append("cnt_org = 0")
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    sql = f"""
        SELECT
            COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS total_prev,
            COALESCE(SUM({amount_expr}) FILTER (WHERE {week_expr} = %s), 0) AS total_last
        FROM {src["table"]}
        {where}
    """
    qparams = [week_prev, week_last, *params]
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, qparams)
        row = cur.fetchone()
    return {
        "w_prev": to_float(row[0] if row else 0),
        "w_last": to_float(row[1] if row else 0),
    }


def enrich_coverages(report: dict, all_totals: dict[str, dict[str, float]]) -> dict:
    for key, totals_key in (
        ("defects", "defects_total"),
        ("defects_org0", "defects_org0_total"),
        ("categories", "categories_total"),
        ("categories_org0", "categories_org0_total"),
    ):
        top = report[totals_key]
        full = all_totals[key]
        top_last = top["w_last"]
        all_last = full["w_last"]
        cover = (top_last / all_last * 100) if all_last else None
        top["all_w_prev"] = full["w_prev"]
        top["all_w_last"] = full["w_last"]
        top["top20_cover_last"] = cover
    return report


def build_kpi_payload(report: dict) -> dict[str, float]:
    d = report.get("defects_total", {}) or {}
    all_totals = report.get("all_totals") or {}
    all_defects = all_totals.get("defects") or {}
    all_org0 = all_totals.get("defects_org0") or {}
    total_last = to_float(all_defects.get("w_last"))
    org0_last = to_float(all_org0.get("w_last"))
    top20_last = to_float(d.get("w_last"))
    cover = d.get("top20_cover_last")
    if cover is None and total_last:
        cover = top20_last / total_last * 100
    return {
        "total_last": total_last,
        "top20_last": top20_last,
        "org0_last": org0_last,
        "org0_share": (org0_last / total_last * 100) if total_last else 0.0,
        "top20_cover": to_float(cover) if cover is not None else 0.0,
    }


def export_report_xlsx(report: dict, year: int, week_prev: int, week_last: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_prev = PatternFill("solid", fgColor="E2EFDA")
    fill_last = PatternFill("solid", fgColor="FFF2CC")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    fill_metric = PatternFill("solid", fgColor="F5F8FC")
    fill_heat_red = PatternFill("solid", fgColor="F8D7DA")
    fill_heat_green = PatternFill("solid", fgColor="D4EDDA")
    fill_heat_yellow = PatternFill("solid", fgColor="FFF3CD")
    fill_heat_share = PatternFill("solid", fgColor="FAD9DE")
    fill_pct_green = PatternFill("solid", fgColor="C6EFCE")
    fill_pct_red = PatternFill("solid", fgColor="FFC7CE")
    fill_pct_yellow = PatternFill("solid", fgColor="FFEB9C")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_bold = Font(bold=True, size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    def apply_heat(cell, value: float | None, mode: str) -> None:
        if value is None:
            return
        if mode == "dynamics":
            if value < -2:
                cell.fill = fill_heat_green
            elif value > 2:
                cell.fill = fill_heat_red
            else:
                cell.fill = fill_heat_yellow
        elif mode == "share":
            cell.fill = fill_heat_share
        else:
            v = value - 100
            if v < -2:
                cell.fill = fill_pct_green
            elif v > 2:
                cell.fill = fill_pct_red
            else:
                cell.fill = fill_pct_yellow

    def write_section(
        ws,
        title: str,
        rows: list[dict],
        total: dict,
        all_totals: dict[str, float],
        show_id: bool,
        name_header: str,
        weeks: list[int],
        week_prev: int,
        week_last: int,
    ) -> None:
        ws.append([title])
        row_title = ws.max_row
        ws.merge_cells(start_row=row_title, start_column=1, end_row=row_title, end_column=2 + len(weeks) + 4)
        c = ws.cell(row_title, 1)
        c.fill = fill_header
        c.font = font_header
        c.alignment = align_center
        c.border = border

        # Исправлено: заголовок "Доля в ТОП-20" соответствует логике расчёта
        headers = [("ИД" if show_id else "№"), name_header, *[str(w) for w in weeks], f"Динамика {week_last} к {week_prev}", "Доля в ТОП-20", "Среднее 2 нед.", "% посл. к средней"]
        ws.append(headers)
        hrow = ws.max_row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(hrow, col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_center
            cell.border = border
            if col >= 3 + len(weeks):
                cell.fill = fill_header
            if col == 3 + weeks.index(week_prev):
                cell.fill = fill_prev
            if col == 3 + weeks.index(week_last):
                cell.fill = fill_last

        def write_data_row(r: dict, total_style: bool = False, label: str | None = None):
            row_idx = ws.max_row + 1
            c1 = ws.cell(row_idx, 1, label if label is not None else (r.get("row_id") if r.get("row_id") is not None else r.get("num")))
            c2 = ws.cell(row_idx, 2, r.get("name", ""))
            c1.alignment = align_center
            c2.alignment = align_left
            c1.border = c2.border = border
            if total_style:
                c1.font = c2.font = font_bold
                c1.fill = c2.fill = fill_total

            for i, w in enumerate(weeks, start=3):
                val = r.get("amounts", {}).get(w, 0.0)
                cw = ws.cell(row_idx, i, val)
                cw.number_format = "# ##0"
                cw.alignment = align_right
                cw.border = border
                if total_style:
                    cw.font = font_bold
                    cw.fill = fill_total
                elif w == week_prev:
                    cw.fill = fill_prev
                elif w == week_last:
                    cw.fill = fill_last

            base = 3 + len(weeks)
            c_dyn = ws.cell(row_idx, base, r.get("dynamics"))
            c_share = ws.cell(row_idx, base + 1, r.get("share"))
            c_avg = ws.cell(row_idx, base + 2, r.get("average"))
            c_pct = ws.cell(row_idx, base + 3, r.get("pct_vs_avg"))
            for c in (c_dyn, c_share, c_avg, c_pct):
                c.border = border
                c.alignment = align_right
                c.fill = fill_metric
            c_dyn.number_format = "0.00%"
            c_share.number_format = "0.00%"
            c_avg.number_format = "# ##0"
            c_pct.number_format = "0.00%"
            if c_dyn.value is not None:
                c_dyn.value = c_dyn.value / 100
            if c_share.value is not None:
                c_share.value = c_share.value / 100
            if c_pct.value is not None:
                c_pct.value = c_pct.value / 100

            apply_heat(c_dyn, r.get("dynamics"), "dynamics")
            apply_heat(c_share, r.get("share"), "share")
            apply_heat(c_pct, r.get("pct_vs_avg"), "pct_vs_avg")

            if total_style:
                for c in (c_dyn, c_share, c_avg, c_pct):
                    c.font = font_bold
                    c.fill = fill_total

        for r in rows:
            write_data_row(r)

        write_data_row(
            {
                "name": "Итого ТОП-20",
                "amounts": total.get("amounts", {}),
                "dynamics": total.get("dynamics"),
                "share": total.get("share"),
                "average": total.get("average"),
                "pct_vs_avg": total.get("pct_vs_avg"),
            },
            total_style=True,
            label="Итого",
        )

        all_prev = to_float(all_totals.get("w_prev"))
        all_last = to_float(all_totals.get("w_last"))
        all_amounts = {w: 0.0 for w in weeks}
        all_amounts[week_prev] = all_prev
        all_amounts[week_last] = all_last
        all_avg = (all_prev + all_last) / 2 if (all_prev or all_last) else 0.0
        all_dyn = ((all_last - all_prev) / all_prev * 100) if all_prev else None
        all_pct = (all_last / all_avg * 100) if all_avg else None
        cover = (total.get("w_last", 0) / all_last * 100) if all_last else None
        write_data_row(
            {
                "name": f"Итого по всем (покрытие ТОП-20: {fmt_pct(cover)})",
                "amounts": all_amounts,
                "dynamics": all_dyn,
                "share": None,
                "average": all_avg,
                "pct_vs_avg": all_pct,
            },
            total_style=True,
            label="Итого по всем",
        )
        ws.append([])

    weeks = report["weeks"]
    sections = [
        ("Дефект ТОП-20, рубли", report["defects"], report["defects_total"], report["all_totals"]["defects"], True, "Дефект"),
        ("Дефект ТОП-20, ORG 0, рубли", report["defects_org0"], report["defects_org0_total"], report["all_totals"]["defects_org0"], True, "Дефект"),
        ("ТОП-20 категорий, рубли", report["categories"], report["categories_total"], report["all_totals"]["categories"], False, "Категория"),
        ("ТОП-20 категорий, ORG 0, рубли", report["categories_org0"], report["categories_org0_total"], report["all_totals"]["categories_org0"], False, "Категория"),
    ]

    ws = wb.create_sheet("Дашборд")
    ws.append(["Отчет по браку", "", "Год", year, "Пред. неделя", week_prev, "Посл. неделя", week_last])
    for c in range(1, 9):
        cell = ws.cell(1, c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.alignment = align_center
        cell.border = border
    ws.append([])

    for sec in sections:
        write_section(ws, *sec, weeks=weeks, week_prev=week_prev, week_last=week_last)

    widths = {
        1: 12,
        2: 42,
    }
    for i in range(len(weeks)):
        widths[3 + i] = 12
    widths[3 + len(weeks)] = 16
    widths[4 + len(weeks)] = 12
    widths[5 + len(weeks)] = 14
    widths[6 + len(weeks)] = 18
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_nomenclature_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"

    fill_header = PatternFill("solid", fgColor="1F4E79")
    fill_total = PatternFill("solid", fgColor="D9E2F3")
    border = Border(
        left=Side(style="thin", color="B4B4B4"),
        right=Side(style="thin", color="B4B4B4"),
        top=Side(style="thin", color="B4B4B4"),
        bottom=Side(style="thin", color="B4B4B4"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    font_bold = Font(bold=True, size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    latest_year = payload.get("latest_year")
    latest_week = payload.get("latest_week")
    ws.append(["Кол-во брака по номенклатуре", "", "Год", latest_year, "Неделя", latest_week])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    for c in range(1, 7):
        cell = ws.cell(1, c)
        cell.fill = fill_total
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border

    headers = ["Номенклатура", "1 корпус", "2 корпус", "3 корпус", "Итог"]
    ws.append(headers)
    hrow = ws.max_row
    for i in range(1, len(headers) + 1):
        cell = ws.cell(hrow, i)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border

    rows = payload.get("rows") or []
    for r in rows:
        ws.append(
            [
                r.get("nomenclature"),
                r.get("corpus_1", 0),
                r.get("corpus_2", 0),
                r.get("corpus_3", 0),
                r.get("total", 0),
            ]
        )

    totals = payload.get("totals") or {}
    ws.append(
        [
            "Итого",
            totals.get("corpus_1", 0),
            totals.get("corpus_2", 0),
            totals.get("corpus_3", 0),
            totals.get("total", 0),
        ]
    )
    tr = ws.max_row
    for c in range(1, 6):
        cell = ws.cell(tr, c)
        cell.fill = fill_total
        cell.font = font_bold
        cell.border = border
        if c == 1:
            cell.alignment = align_center
        else:
            cell.number_format = "# ##0"
            cell.alignment = align_right

    ws.column_dimensions[get_column_letter(1)].width = 18
    ws.column_dimensions[get_column_letter(2)].width = 14
    ws.column_dimensions[get_column_letter(3)].width = 14
    ws.column_dimensions[get_column_letter(4)].width = 14
    ws.column_dimensions[get_column_letter(5)].width = 14

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def export_detail_xlsx(args: Any, *, limit: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    where_sql, params = build_detail_where(args)
    cols = ", ".join(DETAIL_COL_NAMES)
    order_sql = _detail_order_sql(args)
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT {cols}
            FROM brak_team.write_offs
            {where_sql}
            {order_sql}
            LIMIT %s
            """,
            [*params, limit],
        )
        rows = cur.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Детализация"

    fill_header = PatternFill("solid", fgColor="1F4E79")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    font_header = Font(color="FFFFFF", bold=True, size=10)
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    ws.append([label for _, label in DETAIL_COLUMNS])
    for col in range(1, len(DETAIL_COLUMNS) + 1):
        cell = ws.cell(1, col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = border

    numeric_cols = {
        "shk_id",
        "total_cost",
        "amount",
        "share",
        "office_id",
        "wh_id",
        "nm_id",
        "state_id",
        "reason_id",
        "seller_id",
        "supplier_id",
        "cnt_org",
        "cnt_ors",
        "cnt_ocr",
    }
    for row in rows:
        ws.append(list(row))
        row_idx = ws.max_row
        for col_idx, col_name in enumerate(DETAIL_COL_NAMES, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            if col_name in numeric_cols:
                cell.alignment = align_right
                cell.number_format = "# ##0.00" if col_name in ("amount", "total_cost", "share") else "# ##0"
            else:
                cell.alignment = align_left

    widths = {
        "shk_id": 14,
        "date": 16,
        "type": 12,
        "title": 42,
        "reason_descr": 34,
        "subject_name": 24,
        "parent_name": 24,
        "brand_name": 22,
        "owner_product": 24,
    }
    for idx, col_name in enumerate(DETAIL_COL_NAMES, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 14)

    ws.freeze_panes = "A2"
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def fetch_weekly_dynamics(
    *,
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
    top_n: int = 5,
) -> dict[str, Any]:
    sorted_wh = sorted(wh_ids) if wh_ids else []
    cache_key = f"weekly:{year}:{office_id}:{','.join(map(str, sorted_wh))}:{top_n}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != "brak_team.write_offs":
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    # Base table: raw rows. Matview: aggregated group rows (approx. volume).
    row_count_expr = "COUNT(*)::bigint"

    sql_weeks = f"""
        SELECT {week_expr}::int AS week_no,
               COALESCE(SUM({amount_expr}), 0) AS amount_all,
               COALESCE(SUM({amount_expr}) FILTER (WHERE cnt_org = 0), 0) AS amount_org0,
               {row_count_expr} AS row_count
        FROM {src["table"]}
        {where}
        GROUP BY 1
        ORDER BY 1
    """
    sql_top = f"""
        SELECT week_no, reason_id, reason_descr, amount_sum
        FROM (
            SELECT {week_expr}::int AS week_no,
                   reason_id,
                   MAX(COALESCE(reason_descr, '—')) AS reason_descr,
                   COALESCE(SUM({amount_expr}), 0) AS amount_sum,
                   ROW_NUMBER() OVER (
                       PARTITION BY {week_expr}::int
                       ORDER BY COALESCE(SUM({amount_expr}), 0) DESC, reason_id
                   ) AS rn
            FROM {src["table"]}
            {where}
            GROUP BY {week_expr}::int, reason_id
        ) t
        WHERE rn <= %s
        ORDER BY week_no, rn
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql_weeks, params)
        week_rows = cur.fetchall()
        cur.execute(sql_top, [*params, max(1, min(20, top_n))])
        top_rows = cur.fetchall()

    weeks = []
    for r in week_rows:
        amount_all = to_float(r[1])
        amount_org0 = to_float(r[2])
        weeks.append(
            {
                "week": int(r[0]),
                "amount_all": amount_all,
                "amount_org0": amount_org0,
                "org0_share": (amount_org0 / amount_all * 100) if amount_all else 0.0,
                "row_count": int(r[3] or 0),
            }
        )
    top_by_week: dict[int, list[dict[str, Any]]] = {}
    for r in top_rows:
        week = int(r[0])
        top_by_week.setdefault(week, []).append(
            {
                "reason_id": int(r[1]) if r[1] is not None else None,
                "reason_descr": r[2] or "—",
                "amount": to_float(r[3]),
            }
        )
    for w in weeks:
        w["top_reasons"] = top_by_week.get(w["week"], [])

    payload = {
        "year": year,
        "source": src["table"],
        "weeks": weeks,
        "totals": {
            "amount_all": sum(w["amount_all"] for w in weeks),
            "amount_org0": sum(w["amount_org0"] for w in weeks),
            "row_count": sum(w["row_count"] for w in weeks),
        },
    }
    _cache_set(cache_key, payload, WEEKS_CACHE_TTL_SEC)
    return payload


def fetch_status_payload() -> dict[str, Any]:
    env_err = check_db_env()
    env_states = {}
    for name in (
        "DATABASE_URL",
        "DB_HOST",
        "DB_PORT",
        "DB_NAME",
        "DB_USER",
        "DB_PASSWORD",
        "DB_SSLMODE",
    ):
        if name not in os.environ:
            env_states[name] = "missing"
        elif os.environ.get(name, "").strip() == "":
            env_states[name] = "empty"
        else:
            env_states[name] = "set"

    db_status = "error"
    db_detail = env_err or ""
    stats: dict[str, Any] = {}
    matview: dict[str, Any] = {
        "enabled": _use_report_matview(),
        "name": _matview_name(),
        "available": False,
        "bootstrap_ok_age_sec": None,
        "bootstrap_fail_age_sec": None,
        "bootstrap_fail_msg": "",
    }
    now = monotonic()
    with _MV_LOCK:
        if _MV_BOOTSTRAP_OK_AT:
            matview["bootstrap_ok_age_sec"] = round(now - _MV_BOOTSTRAP_OK_AT, 1)
        if _MV_BOOTSTRAP_FAIL_AT:
            matview["bootstrap_fail_age_sec"] = round(now - _MV_BOOTSTRAP_FAIL_AT, 1)
            matview["bootstrap_fail_msg"] = _MV_BOOTSTRAP_FAIL_MSG

    if not env_err:
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
                db_status = "ok"
                db_detail = "connected"
                cfg = load_config()
                stats = fetch_db_stats(cfg.get("office_id"), None)
                if _use_report_matview():
                    try:
                        _ensure_report_matview()
                        cur.execute(
                            f"""
                            SELECT COUNT(*)::bigint,
                                   MAX(iso_year)::int,
                                   MAX(week_no)::int
                            FROM {_matview_name()}
                            """
                        )
                        mv_row = cur.fetchone()
                        matview["available"] = True
                        matview["row_count"] = int(mv_row[0]) if mv_row else 0
                        matview["max_year"] = int(mv_row[1]) if mv_row and mv_row[1] is not None else None
                        matview["max_week"] = int(mv_row[2]) if mv_row and mv_row[2] is not None else None
                    except Exception as exc:
                        matview["available"] = False
                        matview["error"] = str(exc)
        except Exception as exc:
            db_status = "error"
            db_detail = str(exc)

    with _CACHE_LOCK:
        cache_entries = len(_CACHE)

    return {
        "status": "ok" if db_status == "ok" and not env_err else "degraded",
        "vercel_env": os.environ.get("VERCEL_ENV", "local"),
        "db_env_error": env_err,
        "env": env_states,
        "database": {"status": db_status, "detail": db_detail, **stats},
        "matview": matview,
        "cache": {
            "entries": cache_entries,
            "report_ttl_sec": CACHE_TTL_SEC,
            "weeks_ttl_sec": WEEKS_CACHE_TTL_SEC,
        },
        "admin": {
            "refresh_token_required": bool(
                (
                    os.environ.get("ADMIN_PASSWORD")
                    or os.environ.get("DB_REFRESH_TOKEN")
                    or ""
                ).strip()
            ),
            "active_sessions": len(_ADMIN_SESSIONS),
        },
    }


def fetch_available_weeks(
    year: int,
    office_id: int | None,
    wh_ids: list[int] | None,
) -> list[int]:
    # Исправлено: сортируем wh_ids для корректного кэширования
    sorted_wh = sorted(wh_ids) if wh_ids else []
    key = f"weeks:{year}:{office_id}:{','.join(map(str, sorted_wh))}"
    cached = _cache_get(key)
    if cached is not None:
        return cached

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != "brak_team.write_offs":
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)
    week_expr = src["week_expr"]
    sql = f"""
        SELECT DISTINCT {week_expr}::int AS w
        FROM {src["table"]}
        {where}
        ORDER BY w
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
    weeks = [int(r[0]) for r in rows]
    _cache_set(key, weeks, WEEKS_CACHE_TTL_SEC)
    return weeks


def fetch_nomenclature_counts_latest_week(
    *,
    office_id: int | None,
    year: int | None,
) -> dict[str, Any]:
    cache_key = f"nm_latest:{office_id}:{year if year is not None else 'all'}"
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    cfg = load_config()
    wh_catalog = cfg.get("wh_catalog") or []
    wh_to_corpus: dict[int, int] = {}
    for w in wh_catalog:
        try:
            wid = int(w.get("wh_id"))
        except Exception:
            continue
        corpus_raw = w.get("corpus")
        try:
            corpus = int(corpus_raw) if corpus_raw is not None else 0
        except Exception:
            corpus = 0
        if corpus in (1, 2, 3):
            wh_to_corpus[wid] = corpus

    use_nm_mv = _use_report_matview()
    nm_src = {
        "table": "brak_team.write_offs",
        "week_expr": "EXTRACT(WEEK FROM date)::int",
        "isoyear_expr": "EXTRACT(ISOYEAR FROM date)::int",
        "year_clause": "date >= %s AND date < %s",
        "base_not_null_clause": "date IS NOT NULL",
        "rows_expr": "1",
    }
    if use_nm_mv:
        try:
            _ensure_nm_matview()
            nm_src = {
                "table": _nm_matview_name(),
                "week_expr": "week_no",
                "isoyear_expr": "iso_year",
                "year_clause": "iso_year = %s",
                "base_not_null_clause": "",
                "rows_expr": "rows_cnt",
            }
        except Exception as exc:
            print(
                f"[write_offs] nm matview unavailable, fallback to base table: {exc}",
                file=sys.stderr,
            )
    clauses: list[str] = []
    if nm_src["base_not_null_clause"]:
        clauses.append(nm_src["base_not_null_clause"])
    clauses.append("nm_id IS NOT NULL")
    params: list[Any] = []
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if year is not None:
        if use_nm_mv and nm_src["table"] != "brak_team.write_offs":
            clauses.append(nm_src["year_clause"])
            params.append(year)
        else:
            y_start = date.fromisocalendar(year, 1, 1)
            y_end = date.fromisocalendar(year + 1, 1, 1)
            clauses.append(nm_src["year_clause"])
            params.extend([y_start, y_end])

    where = " WHERE " + " AND ".join(clauses)
    week_expr = nm_src["week_expr"]
    year_expr = nm_src["isoyear_expr"]
    rows_expr = nm_src["rows_expr"]
    sql = f"""
        WITH base AS (
            SELECT
                {year_expr} AS iso_year,
                {week_expr} AS week_no,
                wh_id,
                nm_id,
                {rows_expr}::bigint AS rows_cnt
            FROM {nm_src["table"]}
            {where}
        ),
        lw AS (
            SELECT iso_year, week_no
            FROM base
            GROUP BY iso_year, week_no
            ORDER BY iso_year DESC, week_no DESC
            LIMIT 1
        )
        SELECT b.nm_id, b.wh_id, COALESCE(SUM(b.rows_cnt), 0)::bigint AS cnt
        FROM base b
        JOIN lw ON lw.iso_year = b.iso_year AND lw.week_no = b.week_no
        GROUP BY b.nm_id, b.wh_id
        ORDER BY b.nm_id, b.wh_id
    """
    sql_week = f"""
        WITH base AS (
            SELECT
                {year_expr} AS iso_year,
                {week_expr} AS week_no
            FROM {nm_src["table"]}
            {where}
        )
        SELECT iso_year, week_no
        FROM base
        GROUP BY iso_year, week_no
        ORDER BY iso_year DESC, week_no DESC
        LIMIT 1
    """
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql_week, params)
        wk = cur.fetchone()
        cur.execute(sql, params)
        rows = cur.fetchall()

    by_nm: dict[int, dict[str, int]] = {}
    for nm_id, wh_id, cnt in rows:
        nm = int(nm_id)
        corpus = wh_to_corpus.get(int(wh_id), 0)
        if nm not in by_nm:
            by_nm[nm] = {"c1": 0, "c2": 0, "c3": 0, "total": 0}
        if corpus == 1:
            by_nm[nm]["c1"] += int(cnt)
        elif corpus == 2:
            by_nm[nm]["c2"] += int(cnt)
        elif corpus == 3:
            by_nm[nm]["c3"] += int(cnt)
        by_nm[nm]["total"] += int(cnt)

    matrix_rows = [
        {
            "nomenclature": nm,
            "corpus_1": vals["c1"],
            "corpus_2": vals["c2"],
            "corpus_3": vals["c3"],
            "total": vals["total"],
        }
        for nm, vals in by_nm.items()
        if vals["total"] > 0
    ]
    matrix_rows.sort(key=lambda x: (-x["total"], x["nomenclature"]))

    totals = {
        "corpus_1": sum(r["corpus_1"] for r in matrix_rows),
        "corpus_2": sum(r["corpus_2"] for r in matrix_rows),
        "corpus_3": sum(r["corpus_3"] for r in matrix_rows),
        "total": sum(r["total"] for r in matrix_rows),
    }

    latest_year = int(wk[0]) if wk else None
    latest_week = int(wk[1]) if wk else None
    payload = {
        "latest_year": latest_year,
        "latest_week": latest_week,
        "rows": matrix_rows,
        "totals": totals,
    }
    _cache_set(cache_key, payload, NM_CACHE_TTL_SEC)
    return payload


def fetch_top_bundle(
    *,
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    weeks: list[int],
    week_last: int,
    limit: int = 20,
) -> dict[str, list[Row]]:
    if not weeks:
        weeks = [week_last]

    src = _report_source()
    clauses: list[str] = []
    params: list[Any] = []
    if src["base_not_null_clause"]:
        clauses.append(src["base_not_null_clause"])
    if _use_report_matview() and src["table"] != "brak_team.write_offs":
        clauses.append(src["year_clause"])
        params.append(year)
    else:
        year_start = date.fromisocalendar(year, 1, 1)
        year_end = date.fromisocalendar(year + 1, 1, 1)
        clauses.append(src["year_clause"])
        params.extend([year_start, year_end])
    if office_id is not None:
        clauses.append("office_id = %s")
        params.append(office_id)
    if wh_ids:
        clauses.append("wh_id = ANY(%s)")
        params.append(wh_ids)
    where = " WHERE " + " AND ".join(clauses)

    week_expr = src["week_expr"]
    amount_expr = src["amount_expr"]
    week_cols = ",\n               ".join(
        f"COALESCE(SUM(amount) FILTER (WHERE week_no = %s), 0) AS w_{w}"
        for w in weeks
    )

    base_week_select = "week_no" if week_expr == "week_no" else f"{week_expr} AS week_no"
    sql = f"""
WITH base AS (
    SELECT reason_id,
           COALESCE(reason_descr, '—') AS reason_descr,
           COALESCE(parent_name, '—') AS parent_name,
           {base_week_select},
           cnt_org,
           {amount_expr} AS amount
    FROM {src["table"]}
    {where}
),
defects_all AS (
    SELECT reason_id AS row_id, MAX(reason_descr) AS name, {week_cols}
    FROM base
    GROUP BY reason_id
),
defects_org0 AS (
    SELECT reason_id AS row_id, MAX(reason_descr) AS name, {week_cols}
    FROM base
    WHERE cnt_org = 0
    GROUP BY reason_id
),
cats_all AS (
    SELECT NULL::int AS row_id, parent_name AS name, {week_cols}
    FROM base
    GROUP BY parent_name
),
cats_org0 AS (
    SELECT NULL::int AS row_id, parent_name AS name, {week_cols}
    FROM base
    WHERE cnt_org = 0
    GROUP BY parent_name
),
u AS (
    SELECT 'defects'::text AS bucket, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM defects_all
    UNION ALL
    SELECT 'defects_org0'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM defects_org0
    UNION ALL
    SELECT 'categories'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM cats_all
    UNION ALL
    SELECT 'categories_org0'::text, row_id, name, {", ".join(f"w_{w}" for w in weeks)} FROM cats_org0
),
r AS (
    SELECT *, ROW_NUMBER() OVER (PARTITION BY bucket ORDER BY w_{week_last} DESC NULLS LAST) AS rn
    FROM u
)
SELECT bucket, row_id, name, {", ".join(f"w_{w}" for w in weeks)}
FROM r
WHERE rn <= %s
ORDER BY bucket, rn
"""
    qparams = [*params, *weeks, *weeks, *weeks, *weeks, limit]

    buckets: dict[str, list[Row]] = {
        "defects": [],
        "defects_org0": [],
        "categories": [],
        "categories_org0": [],
    }
    with get_conn() as conn:
        cur = conn.cursor()
        cur.execute(sql, qparams)
        raw = cur.fetchall()

    for rec in raw:
        bucket = str(rec[0])
        row_id = int(rec[1]) if rec[1] is not None else None
        name = str(rec[2])
        amounts = {w: to_float(rec[3 + i]) for i, w in enumerate(weeks)}
        buckets[bucket].append(Row(row_id=row_id, name=name, amounts=amounts))
    return buckets


def warm_report_cache_async(
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
) -> None:
    import threading

    def _worker() -> None:
        try:
            build_report_data(wh_ids, office_id, year, week_prev, week_last)
        except Exception:
            pass

    threading.Thread(target=_worker, daemon=True).start()


def add_shares(rows: list[Row], week_prev: int, week_last: int) -> list[dict]:
    """
    Доля считается от суммы ТОП-20. Сумма всех долей в таблице = 100%.
    """
    total_last = sum(r.amount(week_last) for r in rows)
    out: list[dict] = []
    for i, r in enumerate(rows, start=1):
        share = (r.amount(week_last) / total_last * 100) if total_last else 0
        out.append(
            {
                "num": i,
                "row_id": r.row_id,
                "name": r.name,
                "amounts": dict(r.amounts),
                "w_prev": r.amount(week_prev),
                "w_last": r.amount(week_last),
                "dynamics": r.dynamics(week_prev, week_last),
                "share": share,
                "average": r.average(week_prev, week_last),
                "pct_vs_avg": r.pct_vs_avg(week_prev, week_last),
            }
        )
    return out


def totals(rows: list[dict], weeks: list[int], week_prev: int, week_last: int) -> dict:
    """
    Для итоговой строки ТОП-20 доля всегда 100%.
    """
    amounts = {w: sum(x["amounts"].get(w, 0) for x in rows) for w in weeks}
    w_prev = amounts.get(week_prev, 0)
    w_last = amounts.get(week_last, 0)
    avg = (w_prev + w_last) / 2 if rows else 0
    dyn = ((w_last - w_prev) / w_prev * 100) if w_prev else None
    pct_avg = (w_last / avg * 100) if avg else None
    return {
        "amounts": amounts,
        "w_prev": w_prev,
        "w_last": w_last,
        "dynamics": dyn,
        "share": 100.0,  # Всегда 100% для итога ТОП-20
        "average": avg,
        "pct_vs_avg": pct_avg,
    }


def fmt_num(n: float) -> str:
    return f"{n:,.0f}".replace(",", " ")


def fmt_pct(n: float | None) -> str:
    if n is None:
        return "—"
    return f"{n:.2f}%"


def heat_style(value: float | None, mode: str) -> str:
    if value is None:
        return ""
    if mode == "dynamics":
        v = max(-50, min(50, value))
        if v <= 0:
            return "background:#d4edda;color:#155724" if v < -2 else "background:#fff3cd"
        return "background:#f8d7da;color:#721c24" if v > 2 else "background:#fff3cd"
    if mode == "share":
        v = max(0, min(20, value))
        alpha = v / 20
        return f"background:rgba(255,199,206,{0.15 + alpha * 0.5})"
    v = value - 100
    v = max(-30, min(30, v))
    if v < -2:
        return "background:#c6efce"
    if v > 2:
        return "background:#ffc7ce"
    return "background:#ffeb9c"


def _week_cell_style(week: int, week_prev: int, week_last: int) -> str:
    if week == week_last:
        return "background:#fff2cc;font-weight:600"
    if week == week_prev:
        return "background:#e2efda"
    return ""


def render_table(
    title: str,
    rows: list[dict],
    total: dict,
    *,
    show_id: bool,
    weeks: list[int],
    week_prev: int,
    week_last: int,
    name_header: str = "Наименование",
    all_totals: dict[str, float] | None = None,
    drill_kind: str | None = None,
    org0_only: bool = False,
) -> str:
    id_hdr = (
        "<th class='sticky col-id'>ИД</th>"
        if show_id
        else "<th class='sticky col-id'>№</th>"
    )
    week_hdrs = "".join(
        f"<th class='n week-col'>{w}</th>" for w in weeks
    )
    body = []
    for r in rows:
        id_cell = (
            f"<td class='c sticky col-id'>{r['row_id']}</td>"
            if show_id
            else f"<td class='c sticky col-id'>{r['num']}</td>"
        )
        week_cells = "".join(
            f"<td class='n week-col' style='{_e(_week_cell_style(w, week_prev, week_last))}'>"
            f"{fmt_num(r['amounts'].get(w, 0))}</td>"
            for w in weeks
        )
        drill_attrs = ""
        if drill_kind == "reason" and r.get("row_id") is not None:
            drill_attrs = (
                f' class="drill-row" data-drill="reason" data-reason-id="{_e(r["row_id"])}"'
                f' data-org0="{1 if org0_only else 0}" title="Открыть детализацию"'
            )
        elif drill_kind == "category" and r.get("name"):
            drill_attrs = (
                f' class="drill-row" data-drill="category" data-parent-name="{_e(r["name"])}"'
                f' data-org0="{1 if org0_only else 0}" title="Открыть детализацию"'
            )
        body.append(
            f"<tr{drill_attrs}>{id_cell}"
            f"<td class='name sticky col-name' title='{_e(r['name'])}'>{_e(r['name'])}</td>"
            f"{week_cells}"
            f"<td class='n metric' style='{_e(heat_style(r['dynamics'], 'dynamics'))}'>{fmt_pct(r['dynamics'])}</td>"
            f"<td class='n metric' style='{_e(heat_style(r['share'], 'share'))}'>{fmt_pct(r['share'])}</td>"
            f"<td class='n metric'>{fmt_num(r['average'])}</td>"
            f"<td class='n metric' style='{_e(heat_style(r['pct_vs_avg'], 'pct_vs_avg'))}'>{fmt_pct(r['pct_vs_avg'])}</td>"
            "</tr>"
        )

    t = total
    week_total_cells = "".join(
        f"<td class='n week-col' style='{_e(_week_cell_style(w, week_prev, week_last))}'>"
        f"<b>{fmt_num(t['amounts'].get(w, 0))}</b></td>"
        for w in weeks
    )
    body.append(
        f"<tr class='total'>"
        f"<td colspan='2' class='sticky col-id'><b>Итого</b></td>"
        f"{week_total_cells}"
        f"<td class='n metric' style='{_e(heat_style(t['dynamics'], 'dynamics'))}'><b>{fmt_pct(t['dynamics'])}</b></td>"
        f"<td class='n metric'><b>{fmt_pct(t['share'])}</b></td>"
        f"<td class='n metric'><b>{fmt_num(t['average'])}</b></td>"
        f"<td class='n metric' style='{_e(heat_style(t['pct_vs_avg'], 'pct_vs_avg'))}'><b>{fmt_pct(t['pct_vs_avg'])}</b></td>"
        f"</tr>"
    )
    if all_totals is not None:
        all_prev = to_float(all_totals.get("w_prev"))
        all_last = to_float(all_totals.get("w_last"))
        cover = (t["w_last"] / all_last * 100) if all_last else None
        all_amounts = {w: 0.0 for w in weeks}
        all_amounts[week_prev] = all_prev
        all_amounts[week_last] = all_last
        all_week_cells = "".join(
            f"<td class='n week-col'><b>{fmt_num(all_amounts.get(w, 0))}</b></td>"
            for w in weeks
        )
        all_avg = (all_prev + all_last) / 2 if (all_prev or all_last) else 0.0
        all_dyn = ((all_last - all_prev) / all_prev * 100) if all_prev else None
        all_pct = (all_last / all_avg * 100) if all_avg else None
        body.append(
            f"<tr class='total'>"
            f"<td colspan='2' class='sticky col-id'><b>Итого по всем</b></td>"
            f"{all_week_cells}"
            f"<td class='n metric'><b>{fmt_pct(all_dyn)}</b></td>"
            f"<td class='n metric'><b>—</b></td>"
            f"<td class='n metric'><b>{fmt_num(all_avg)}</b></td>"
            f"<td class='n metric'><b>{fmt_pct(all_pct)} · покрытие ТОП-20: {fmt_pct(cover)}</b></td>"
            f"</tr>"
        )

    # Исправлено: заголовок "Доля в ТОП-20" соответствует логике расчёта
    return f"""
<section class="panel">
  <h2>{_e(title)}</h2>
  <div class="table-scroll">
  <table>
    <colgroup>
      <col class="col-id">
      <col class="col-name">
    </colgroup>
    <thead>
      <tr>
        {id_hdr}
        <th class="sticky col-name">{_e(name_header)}</th>
        {week_hdrs}
        <th class="metric">Динамика {week_last} к {week_prev}</th>
        <th class="metric">Доля в ТОП-20</th>
        <th class="metric">Среднее за 2 нед.</th>
        <th class="metric">% посл. к средней</th>
      </tr>
    </thead>
    <tbody>{''.join(body)}</tbody>
  </table>
  </div>
</section>
"""


def _e(s: Any) -> str:
    import html

    return html.escape(str(s)) if s is not None else ""


def build_report_data(
    wh_ids: list[int] | None,
    office_id: int | None,
    year: int,
    week_prev: int,
    week_last: int,
    weeks: list[int] | None = None,
    show_all_weeks: bool = True,
) -> dict:
    # Исправлено: сортируем wh_ids и weeks для корректного кэширования
    sorted_wh = sorted(wh_ids) if wh_ids else []
    sorted_weeks = sorted(weeks) if weeks else []
    key = (
        f"report:{year}:{week_prev}:{week_last}:{office_id}:{int(show_all_weeks)}:"
        f"{','.join(map(str, sorted_wh))}:{','.join(map(str, sorted_weeks))}"
    )
    cached = _cache_get(key)
    if cached is not None:
        return cached

    if weeks is None:
        weeks = fetch_available_weeks(year, office_id, wh_ids)
    if not weeks:
        weeks = [week_prev, week_last]
    if week_prev not in weeks:
        weeks = sorted(set(weeks) | {week_prev})
    if week_last not in weeks:
        weeks = sorted(set(weeks) | {week_last})
    if not show_all_weeks:
        weeks = sorted(set([week_prev, week_last]))

    bundle = fetch_top_bundle(
        wh_ids=wh_ids,
        office_id=office_id,
        year=year,
        weeks=weeks,
        week_last=week_last,
        limit=20,
    )
    defects = bundle["defects"]
    defects_org0 = bundle["defects_org0"]
    cats = bundle["categories"]
    cats_org0 = bundle["categories_org0"]

    d_rows = add_shares(defects, week_prev, week_last)
    d0_rows = add_shares(defects_org0, week_prev, week_last)
    c_rows = add_shares(cats, week_prev, week_last)
    c0_rows = add_shares(cats_org0, week_prev, week_last)

    report = {
        "weeks": weeks,
        "defects": d_rows,
        "defects_total": totals(d_rows, weeks, week_prev, week_last),
        "defects_org0": d0_rows,
        "defects_org0_total": totals(d0_rows, weeks, week_prev, week_last),
        "categories": c_rows,
        "categories_total": totals(c_rows, weeks, week_prev, week_last),
        "categories_org0": c0_rows,
        "categories_org0_total": totals(c0_rows, weeks, week_prev, week_last),
    }
    total_all = fetch_totals_all(
        wh_ids=wh_ids,
        office_id=office_id,
        org0_only=False,
        year=year,
        week_prev=week_prev,
        week_last=week_last,
    )
    total_org0 = fetch_totals_all(
        wh_ids=wh_ids,
        office_id=office_id,
        org0_only=True,
        year=year,
        week_prev=week_prev,
        week_last=week_last,
    )
    all_totals = {
        "defects": total_all,
        "defects_org0": total_org0,
        "categories": total_all,
        "categories_org0": total_org0,
    }
    report["all_totals"] = all_totals
    final_report = enrich_coverages(report, all_totals)
    _cache_set(key, final_report, CACHE_TTL_SEC)
    return final_report


DASHBOARD_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Брак — ТОП-20</title>
<style>
* { box-sizing: border-box; }
body { margin: 0; font: 13px/1.35 Calibri, "Segoe UI", sans-serif; background: #e8eaed; color: #000; }
header { background: #1f4e79; color: #fff; padding: 12px 16px; }
header h1 { margin: 0; font-size: 18px; font-weight: 600; }
.toolbar { background: #fff; padding: 12px 16px; border-bottom: 1px solid #ccc;
           display: flex; flex-wrap: wrap; gap: 16px; align-items: flex-start; }
.group { border: 1px solid #ddd; border-radius: 6px; padding: 8px 12px; background: #fafafa; }
.group legend { font-weight: 600; font-size: 12px; padding: 0 4px; }
.building-btns { display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; }
.building-btns button { padding: 6px 12px; border: 1px solid #1f4e79; background: #fff;
  border-radius: 4px; cursor: pointer; font-size: 12px; }
.building-btns button.active { background: #1f4e79; color: #fff; }
.wh-grid { display: flex; flex-direction: column; gap: 2px;
           max-height: 280px; overflow: auto; margin-top: 6px; min-width: 340px; }
.wh-grid label { font-size: 11px; display: flex; align-items: center; gap: 6px; cursor: pointer; }
.wh-grid .corpus-hdr { font-weight: 700; font-size: 11px; color: #1f4e79; margin-top: 8px;
                      padding: 4px 0 2px; border-bottom: 1px solid #ccc; }
.wh-grid .corpus-hdr:first-child { margin-top: 0; }
.weeks input, .weeks select { padding: 4px; margin-right: 8px; font-size: 12px; }
.weeks select { min-width: 64px; }
.weeks .hint { display: block; font-size: 11px; color: #666; margin-top: 6px; max-width: 320px; }
.panel .table-scroll { overflow-x: auto; max-width: 100%; }
.panel table { min-width: max-content; }
th.week-col, td.week-col { min-width: 72px; }
th.metric, td.metric { min-width: 110px; }
th.metric { background: #1f4e79; color: #fff; }
td.metric { background: #f5f8fc; color: #000; }
th.sticky, td.sticky { position: sticky; z-index: 2; }
th.sticky { background: #1f4e79; z-index: 4; }
td.sticky { background: #fff; }
.col-id { width: 52px; min-width: 52px; }
.col-name { width: 260px; min-width: 200px; }
th.col-id, td.col-id { left: 0; }
th.col-name, td.col-name { left: 52px; box-shadow: 2px 0 4px rgba(0,0,0,.08); }
td.col-name { white-space: normal; line-height: 1.25; vertical-align: top; }
tr.total td.sticky { background: #d9e2f3; }
tr.total th.sticky { background: #1f4e79; }
.actions { display: flex; align-items: flex-end; gap: 8px; }
.actions button { padding: 8px 16px; border: none; border-radius: 4px; cursor: pointer;
  font-weight: 600; font-size: 12px; }
.actions button.primary { background: #1f4e79; color: #fff; }
.actions button.primary:hover { background: #163a5c; }
.actions button.primary:disabled { opacity: 0.6; cursor: wait; }
.actions button.secondary { background: #217346; color: #fff; }
.actions button.secondary:hover { background: #1a5c38; }
.actions button.export { background: #6b7280; color: #fff; }
.actions button.export:hover { background: #4b5563; }
.actions .hidden { display: none; }
#status { padding: 8px 16px; color: #555; font-size: 12px; }
.modal-overlay { position: fixed; inset: 0; background: rgba(15,23,42,.45); display: none; align-items: center; justify-content: center; z-index: 50; }
.modal-overlay.show { display: flex; }
.modal { width: 360px; max-width: calc(100vw - 24px); background: #fff; border-radius: 8px; box-shadow: 0 10px 40px rgba(0,0,0,.25); border: 1px solid #cbd5e1; }
.modal h3 { margin: 0; padding: 10px 12px; background: #1f4e79; color: #fff; font-size: 14px; }
.modal .body { padding: 12px; display: grid; gap: 10px; }
.modal label { display: grid; gap: 4px; font-size: 12px; color: #334155; }
.modal input { padding: 8px; border: 1px solid #cbd5e1; border-radius: 4px; font-size: 13px; }
.modal .row { display: flex; gap: 8px; justify-content: flex-end; }
.modal .hint { color: #b91c1c; min-height: 16px; font-size: 12px; }
.grid { display: grid; grid-template-columns: 1fr 1fr; gap: 12px; padding: 12px; }
@media (max-width: 1200px) { .grid { grid-template-columns: 1fr; } }
.panel { background: #fff; border: 1px solid #999; overflow: auto; }
.panel h2 { margin: 0; padding: 8px 10px; font-size: 13px; font-weight: 600;
            background: #1f4e79; color: #fff; text-align: center; }
table { width: 100%; border-collapse: collapse; }
th, td { border: 1px solid #b4b4b4; padding: 3px 6px; }
th { background: #1f4e79; color: #fff; font-weight: 600; text-align: center; font-size: 11px; }
td.name { text-align: left; }
td.c, td.n { text-align: right; white-space: nowrap; }
tr.total td { background: #d9e2f3; font-weight: 600; }
.loading { opacity: 0.5; pointer-events: none; }

:root {
  --bg: #f3f6fb;
  --surface: #ffffff;
  --surface-2: #f8fafc;
  --text: #0f172a;
  --muted: #64748b;
  --line: #dbe4f0;
  --primary: #2563eb;
  --primary-2: #1d4ed8;
  --accent: #0f766e;
  --danger: #b91c1c;
  --shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
  --radius: 12px;
}
body {
  font: 14px/1.45 Inter, "Segoe UI", Roboto, Arial, sans-serif;
  background:
    radial-gradient(1200px 520px at 0% -10%, rgba(37, 99, 235, 0.08), transparent 60%),
    radial-gradient(900px 420px at 100% -20%, rgba(14, 116, 144, 0.08), transparent 55%),
    var(--bg);
  color: var(--text);
}
header {
  background: linear-gradient(100deg, #0f4c81 0%, #2563eb 55%, #1d4ed8 100%);
  border-bottom: 1px solid rgba(255, 255, 255, 0.25);
  box-shadow: 0 8px 24px rgba(29, 78, 216, 0.22);
  padding: 18px 20px 20px;
}
header h1 { font-size: 24px; font-weight: 750; letter-spacing: .2px; }
.header-subtitle {
  margin-top: 5px;
  color: rgba(255, 255, 255, 0.82);
  font-size: 13px;
}
.toolbar {
  margin: 16px 14px 12px;
  border: 1px solid var(--line);
  border-radius: 16px;
  background: rgba(255, 255, 255, 0.92);
  box-shadow: var(--shadow);
  padding: 16px;
  backdrop-filter: blur(8px);
}
.group {
  border: 1px solid var(--line);
  border-radius: 14px;
  background: var(--surface-2);
  padding: 12px 14px;
}
.group legend {
  color: #1e3a8a;
  font-size: 11px;
  font-weight: 800;
  text-transform: uppercase;
  letter-spacing: .6px;
  padding: 0 6px;
}
.building-btns button {
  border: 1px solid #bfd2ff;
  background: #eef4ff;
  color: #1e40af;
  border-radius: 999px;
  transition: all .16s ease;
}
.building-btns button:hover { background: #dbeafe; border-color: #93c5fd; }
.building-btns button.active {
  background: linear-gradient(180deg, #2563eb, #1d4ed8);
  border-color: #1d4ed8;
  color: #fff;
}
.wh-grid {
  border: 1px solid var(--line);
  border-radius: 12px;
  background: #fff;
  padding: 9px;
  box-shadow: inset 0 1px 2px rgba(15,23,42,.04);
}
.wh-grid label {
  font-size: 12px;
  color: #1f2937;
  padding: 3px 4px;
  border-radius: 7px;
}
.wh-grid label:hover { background: #f8fafc; }
.wh-grid .corpus-hdr { color: #1d4ed8; border-bottom: 1px dashed #c7d2fe; }
.weeks input, .weeks select {
  border: 1px solid #cbd5e1;
  border-radius: 8px;
  background: #fff;
  padding: 6px 8px;
}
.weeks .hint { color: var(--muted); }
.actions { gap: 10px; }
.actions button {
  border-radius: 10px;
  font-size: 12px;
  font-weight: 650;
  border: 1px solid transparent;
  box-shadow: 0 1px 0 rgba(15,23,42,.03);
  transition: all .16s ease;
}
.actions button:hover { transform: translateY(-1px); box-shadow: 0 8px 18px rgba(15,23,42,.12); }
.actions button:disabled { opacity: .55; cursor: not-allowed; transform: none; box-shadow: none; }
.actions button.primary { background: linear-gradient(180deg, var(--primary), var(--primary-2)); color: #fff; }
.actions button.primary:hover { background: linear-gradient(180deg, #1d4ed8, #1e40af); }
.actions button.secondary { background: #f0f9ff; color: #075985; border-color: #bae6fd; }
.actions button.secondary:hover { background: #e0f2fe; }
.actions button.export { background: #f8fafc; color: #334155; border-color: #dbe4f0; }
.actions button.export:hover { background: #f1f5f9; }
#status {
  margin: 0 14px 12px;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 12px;
  padding: 10px 12px;
  color: #334155;
  box-shadow: 0 3px 10px rgba(15, 23, 42, 0.06);
}
.grid { gap: 14px; padding: 0 14px 14px; }
.panel {
  border: 1px solid var(--line);
  border-radius: 16px;
  background: var(--surface);
  box-shadow: var(--shadow);
  overflow: hidden;
}
.panel h2 {
  background: linear-gradient(90deg, #eff6ff 0%, #f8fafc 100%);
  color: #0f172a;
  border-bottom: 1px solid var(--line);
  font-size: 13px;
  font-weight: 700;
  text-align: left;
  padding: 10px 12px;
}
.panel .table-scroll {
  background: #fff;
  border-top: 1px solid #eef2f7;
}
th, td { border-color: #e2e8f0; padding: 5px 8px; }
th {
  background: #f8fafc;
  color: #334155;
  font-size: 11px;
  position: relative;
}
th.metric {
  background: #eff6ff;
  color: #1e3a8a;
}
td.metric { background: #f8fafc; color: #0f172a; }
th.sticky { background: #f8fafc; z-index: 4; }
td.sticky { background: #fff; }
tbody tr:not(.total):nth-child(even) td { background-color: #fbfdff; }
tbody tr:not(.total):hover td { background-color: #eef6ff; }
tbody tr:not(.total):hover td.sticky { background-color: #eef6ff; }
tr.drill-row { cursor: pointer; }
tr.drill-row:hover td { background-color: #dbeafe !important; }
tr.total td { background: #eef2ff; }
.presets {
  margin: 0 14px 10px;
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
  align-items: center;
}
.presets .label { color: #64748b; font-size: 12px; font-weight: 700; margin-right: 4px; }
.presets button {
  border: 1px solid #bfd2ff;
  background: #eef4ff;
  color: #1e40af;
  border-radius: 999px;
  padding: 6px 10px;
  font-size: 12px;
  font-weight: 700;
  cursor: pointer;
}
.presets button:hover { background: #dbeafe; }
.modal h3 {
  background: linear-gradient(100deg, #1d4ed8, #2563eb);
}
.modal input:focus {
  outline: none;
  border-color: #60a5fa;
  box-shadow: 0 0 0 3px rgba(59,130,246,.16);
}
.topnav {
  margin: 12px 14px 10px;
  display: flex;
  gap: 8px;
}
.topnav a {
  text-decoration: none;
  padding: 8px 12px;
  border: 1px solid #bfd2ff;
  background: #eef4ff;
  color: #1e40af;
  border-radius: 999px;
  font-weight: 700;
  font-size: 12px;
  box-shadow: 0 1px 0 rgba(15,23,42,.04);
}
.topnav a:hover { background: #dbeafe; border-color: #93c5fd; }
.topnav a.active {
  background: linear-gradient(180deg, #2563eb, #1d4ed8);
  color: #fff;
  border-color: #1d4ed8;
}
.kpis {
  margin: 0 14px 12px;
  display: grid;
  grid-template-columns: repeat(4, minmax(180px, 1fr));
  gap: 12px;
}
.kpi {
  background: linear-gradient(180deg, #ffffff, #f8fbff);
  border: 1px solid var(--line);
  border-radius: 14px;
  padding: 12px 14px;
  box-shadow: 0 6px 18px rgba(15,23,42,.07);
}
.kpi .k { color: #64748b; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .35px; margin-bottom: 5px; }
.kpi .v { color: #0f172a; font-size: 22px; font-weight: 800; font-variant-numeric: tabular-nums; }
.table-tools {
  margin: 0 14px 10px;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 12px;
  padding: 10px 12px;
  box-shadow: 0 3px 10px rgba(15,23,42,.05);
}
.table-tools input {
  border: 1px solid #cbd5e1;
  border-radius: 8px;
  padding: 6px 8px;
  min-width: 260px;
}
.table-tools input:focus,
.weeks input:focus,
.weeks select:focus {
  outline: none;
  border-color: #60a5fa;
  box-shadow: 0 0 0 3px rgba(59,130,246,.14);
}

@media (max-width: 900px) {
  .toolbar { margin: 10px; padding: 10px; }
  .topnav { margin: 10px 10px 8px; }
  .kpis { margin: 0 10px 10px; grid-template-columns: 1fr 1fr; }
  .table-tools { margin: 0 10px 8px; }
  #status { margin: 0 10px 10px; }
  .grid { padding: 0 10px 10px; }
  .actions { width: 100%; }
}
</style>
</head>
<body>
<header>
  <h1>Отчёт по браку — write_offs</h1>
  <div class="header-subtitle">ТОП-20 дефектов и категорий по корпусам, неделям и динамике</div>
</header>
<nav class="topnav">
  <a href="/" class="active">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/status">Статус</a>
</nav>
<div class="toolbar">
  <fieldset class="group">
    <legend>Корпус / WH</legend>
    <div class="building-btns" id="buildingBtns"></div>
    <div class="wh-grid" id="whGrid"></div>
  </fieldset>
  <fieldset class="group weeks">
    <legend>Недели (ISO)</legend>
    <label>Год <input type="number" id="year" value="2026"></label>
    <label>Расчёт: пред. <select id="weekPrev"></select></label>
    <label>посл. <select id="weekLast"></select></label>
    <span class="hint">Все недели года — в таблице (прокрутка). Динамика, доля и среднее — только по двум выбранным неделям.</span>
  </fieldset>
  <div class="actions">
    <button type="button" id="btnRefreshData" class="primary">Обновить данные</button>
    <button type="button" id="btnAdminLogin" class="secondary">Вход админа</button>
    <button type="button" id="btnAdminLogout" class="export">Выход админа</button>
    <button type="button" id="btnToggleWeeks" class="secondary">Показать все недели</button>
    <button type="button" id="btnNomenclature" class="secondary">Номенклатура</button>
    <button type="button" id="btnDetails" class="secondary">Детализация</button>
    <button type="button" id="btnClearWh" class="export">Сбросить фильтр</button>
    <button type="button" id="btnApply" class="secondary">Применить</button>
    <button type="button" id="btnAllWh" class="secondary">Все WH</button>
    <button type="button" id="btnExportXlsx" class="export">Экспорт XLSX</button>
  </div>
</div>
<div class="kpis" id="kpis">
  <div class="kpi"><div class="k">Всего брак (посл. нед.)</div><div class="v" id="kpiTotal">—</div></div>
  <div class="kpi"><div class="k">ТОП-20 (посл. нед.)</div><div class="v" id="kpiTop20">—</div></div>
  <div class="kpi"><div class="k">ORG0 от общего</div><div class="v" id="kpiOrg0Share">—</div></div>
  <div class="kpi"><div class="k">Покрытие ТОП-20</div><div class="v" id="kpiCover">—</div></div>
</div>
<div class="presets" id="presetsBar">
  <span class="label">Пресеты:</span>
  <button type="button" data-preset="all">Все корпуса</button>
  <button type="button" data-preset="korpus_1">1 корпус</button>
  <button type="button" data-preset="korpus_2">2 корпус</button>
  <button type="button" data-preset="korpus_3">3 корпус</button>
  <button type="button" data-preset="latest">Последние 2 недели</button>
  <button type="button" id="btnSavePreset">Сохранить текущий</button>
</div>
<div class="table-tools">
  <label>Поиск: <input type="text" id="tableSearch" placeholder="Дефект / категория / ID"></label>
</div>
<div id="status">Загрузка…</div>
<div id="adminModal" class="modal-overlay" aria-hidden="true">
  <div class="modal" role="dialog" aria-modal="true" aria-labelledby="adminModalTitle">
    <h3 id="adminModalTitle">Вход администратора</h3>
    <div class="body">
      <label>Логин
        <input id="adminLoginInput" type="text" autocomplete="username" placeholder="admin">
      </label>
      <label>Пароль
        <input id="adminPasswordInput" type="password" autocomplete="current-password" placeholder="Введите пароль">
      </label>
      <div id="adminAuthError" class="hint"></div>
      <div class="row">
        <button type="button" id="btnAdminCancel" class="export">Отмена</button>
        <button type="button" id="btnAdminSubmit" class="primary">Войти</button>
      </div>
    </div>
  </div>
</div>
<div class="grid" id="reportGrid"></div>
<script>
const CONFIG = __CONFIG_JSON__;
const CATALOG = CONFIG.wh_catalog && CONFIG.wh_catalog.length
  ? CONFIG.wh_catalog
  : (CONFIG.wh_list || []).map(w => ({ wh_id: w.wh_id, name: String(w.wh_id), corpus: 0 }));
const ALL_WH_IDS = CATALOG.map(w => w.wh_id);

let selectedWh = new Set();
let activeBuilding = 'custom';
let availableWeeks = [];
let showAllWeeks = false;
let adminSessionId = '';
let isAdminSession = false;

function whLabel(id) {
  const w = CATALOG.find(x => x.wh_id === id);
  return w ? (id + ' — ' + w.name) : String(id);
}

function fmtInt(v) {
  return Number(v || 0).toLocaleString('ru-RU');
}

function fmtPct(v) {
  return Number(v || 0).toLocaleString('ru-RU', { minimumFractionDigits: 2, maximumFractionDigits: 2 }) + '%';
}

function updateKpis(kpis) {
  const d = kpis || {};
  document.getElementById('kpiTotal').textContent = fmtInt(d.total_last || 0);
  document.getElementById('kpiTop20').textContent = fmtInt(d.top20_last || 0);
  document.getElementById('kpiOrg0Share').textContent = fmtPct(d.org0_share || 0);
  document.getElementById('kpiCover').textContent = fmtPct(d.top20_cover || 0);
}

function currentFilterState() {
  const { wp, wl } = selectedWeeks();
  return {
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
    show_all_weeks: showAllWeeks,
    building: activeBuilding,
    wh_ids: Array.from(selectedWh),
  };
}

function applyFilterState(state) {
  if (!state) return;
  if (state.year) document.getElementById('year').value = state.year;
  if (Array.isArray(state.wh_ids)) {
    selectedWh = new Set(state.wh_ids.map(Number));
    document.querySelectorAll('#whGrid input').forEach(cb => {
      cb.checked = selectedWh.has(parseInt(cb.value, 10));
    });
  }
  if (state.building) {
    activeBuilding = state.building;
    syncBuildingButtons();
  }
  if (typeof state.show_all_weeks === 'boolean') {
    showAllWeeks = state.show_all_weeks;
    updateWeeksToggleLabel();
  }
  fillWeekSelects(
    availableWeeks.length ? availableWeeks : [Number(state.week_prev), Number(state.week_last)].filter(Boolean),
    Number(state.week_prev),
    Number(state.week_last)
  );
}

function loadSavedPresets() {
  try {
    return JSON.parse(localStorage.getItem('dashboardPresets') || '[]');
  } catch (_) {
    return [];
  }
}

function renderSavedPresets() {
  const bar = document.getElementById('presetsBar');
  bar.querySelectorAll('[data-saved-preset]').forEach(el => el.remove());
  loadSavedPresets().forEach((p, idx) => {
    const btn = document.createElement('button');
    btn.type = 'button';
    btn.dataset.savedPreset = String(idx);
    btn.textContent = p.name || ('Пресет ' + (idx + 1));
    btn.addEventListener('click', async () => {
      applyFilterState(p.state);
      await refreshWeeks();
      applyFilterState(p.state);
      loadReport();
    });
    bar.appendChild(btn);
  });
}

function openDetailsDrill(params) {
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const year = Number(document.getElementById('year').value) || new Date().getFullYear();
  const q = new URLSearchParams(params || {});
  if (wh) q.set('wh_ids', wh);
  // Approximate ISO week range for selected last week.
  try {
    const start = new Date(Date.UTC(year, 0, 1 + (Number(wl) - 1) * 7));
    const day = start.getUTCDay() || 7;
    start.setUTCDate(start.getUTCDate() - day + 1);
    const end = new Date(start);
    end.setUTCDate(end.getUTCDate() + 6);
    q.set('date_from', start.toISOString().slice(0, 10));
    q.set('date_to', end.toISOString().slice(0, 10));
  } catch (_) {}
  q.set('week_prev', wp);
  q.set('week_last', wl);
  q.set('year', String(year));
  window.location.href = '/details?' + q.toString();
}

function applyTableSearch() {
  const q = (document.getElementById('tableSearch').value || '').trim().toLowerCase();
  document.querySelectorAll('#reportGrid tbody tr').forEach(tr => {
    if (tr.classList.contains('total')) {
      tr.style.display = '';
      return;
    }
    if (!q) {
      tr.style.display = '';
      return;
    }
    tr.style.display = tr.textContent.toLowerCase().includes(q) ? '' : 'none';
  });
}

function init() {
  if (CONFIG.week_year) document.getElementById('year').value = CONFIG.week_year;
  adminSessionId = (localStorage.getItem('adminSessionId') || '').trim();
  isAdminSession = Boolean(adminSessionId);
  applyAdminUi();
  showAllWeeks = Boolean(CONFIG.show_all_weeks_default);
  updateWeeksToggleLabel();
  const grid = document.getElementById('whGrid');
  let lastCorpus = null;
  CATALOG.forEach(w => {
    if (w.corpus && w.corpus !== lastCorpus) {
      const hdr = document.createElement('div');
      hdr.className = 'corpus-hdr';
      hdr.textContent = w.corpus + ' корпус';
      grid.appendChild(hdr);
      lastCorpus = w.corpus;
    }
    const lab = document.createElement('label');
    const cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.value = w.wh_id;
    cb.checked = true;
    cb.addEventListener('change', () => {
      activeBuilding = 'custom';
      syncBuildingButtons();
      const id = parseInt(cb.value, 10);
      if (cb.checked) selectedWh.add(id); else selectedWh.delete(id);
    });
    selectedWh.add(w.wh_id);
    lab.append(cb, document.createTextNode(' ' + whLabel(w.wh_id)));
    grid.appendChild(lab);
  });

  const btns = document.getElementById('buildingBtns');
  CONFIG.buildings.forEach(b => {
    const btn = document.createElement('button');
    btn.type = 'button';
    btn.textContent = b.name;
    btn.dataset.id = b.id;
    btn.addEventListener('click', () => selectBuilding(b));
    btns.appendChild(btn);
  });

  document.getElementById('year').addEventListener('change', () => refreshWeeks().then(loadReport));
  document.getElementById('weekPrev').addEventListener('change', loadReport);
  document.getElementById('weekLast').addEventListener('change', loadReport);
  document.getElementById('btnAdminLogin').onclick = adminLogin;
  document.getElementById('btnAdminLogout').onclick = adminLogout;
  document.getElementById('btnAdminCancel').onclick = closeAdminModal;
  document.getElementById('btnAdminSubmit').onclick = submitAdminLogin;
  document.getElementById('adminModal').addEventListener('click', (e) => {
    if (e.target && e.target.id === 'adminModal') closeAdminModal();
  });
  document.getElementById('adminPasswordInput').addEventListener('keydown', (e) => {
    if (e.key === 'Enter') submitAdminLogin();
  });
  document.getElementById('btnToggleWeeks').onclick = () => {
    showAllWeeks = !showAllWeeks;
    updateWeeksToggleLabel();
    loadReport();
  };
  document.getElementById('btnNomenclature').onclick = () => {
    window.location.href = '/nomenclature';
  };
  document.getElementById('btnDetails').onclick = () => {
    openDetailsDrill({});
  };
  document.getElementById('btnClearWh').onclick = () => {
    selectedWh = new Set();
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = false);
    activeBuilding = 'custom';
    syncBuildingButtons();
    document.getElementById('status').textContent = 'Фильтр WH сброшен. Выберите нужный блок/склады и нажмите "Применить".';
  };
  document.getElementById('tableSearch').addEventListener('input', applyTableSearch);
  document.getElementById('btnApply').onclick = loadReport;
  document.getElementById('btnRefreshData').onclick = refreshData;
  document.getElementById('btnExportXlsx').onclick = exportXlsx;
  document.getElementById('btnAllWh').onclick = async () => {
    selectedWh = new Set(ALL_WH_IDS);
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = true);
    activeBuilding = 'all';
    syncBuildingButtons();
    await refreshWeeks();
    loadReport();
  };
  document.querySelectorAll('#presetsBar [data-preset]').forEach(btn => {
    btn.addEventListener('click', async () => {
      const key = btn.dataset.preset;
      if (key === 'all') {
        selectedWh = new Set(ALL_WH_IDS);
        activeBuilding = 'all';
      } else if (key.startsWith('korpus_')) {
        const b = (CONFIG.buildings || []).find(x => x.id === key);
        if (b) {
          selectedWh = new Set(b.wh_ids || []);
          activeBuilding = b.id;
        }
      } else if (key === 'latest') {
        await refreshWeeks(true);
      }
      document.querySelectorAll('#whGrid input').forEach(cb => {
        cb.checked = selectedWh.has(parseInt(cb.value, 10));
      });
      syncBuildingButtons();
      await refreshWeeks(key === 'latest');
      loadReport();
    });
  });
  document.getElementById('btnSavePreset').onclick = () => {
    const name = prompt('Название пресета', 'Мой фильтр');
    if (!name) return;
    const list = loadSavedPresets();
    list.push({ name: name.trim(), state: currentFilterState() });
    localStorage.setItem('dashboardPresets', JSON.stringify(list.slice(-8)));
    renderSavedPresets();
  };
  renderSavedPresets();

  refreshWeeks(true).then(() => {
    if (CONFIG.buildings.length) selectBuilding(CONFIG.buildings[0]);
    else loadReport();
  });
}

function applyAdminUi() {
  const btnRefresh = document.getElementById('btnRefreshData');
  const btnLogin = document.getElementById('btnAdminLogin');
  const btnLogout = document.getElementById('btnAdminLogout');
  const needAuth = Boolean(CONFIG.refresh_token_required);
  btnRefresh.classList.toggle('hidden', !isAdminSession);
  btnLogin.disabled = isAdminSession;
  btnLogout.disabled = !isAdminSession;
}

function adminLogin() {
  openAdminModal();
}

function adminLogout() {
  const status = document.getElementById('status');
  const sid = adminSessionId;
  adminSessionId = '';
  localStorage.removeItem('adminSessionId');
  isAdminSession = false;
  applyAdminUi();
  fetch('/api/admin/logout', {
    method: 'POST',
    headers: sid ? { 'X-Admin-Session': sid } : {},
  }).catch(() => {});
  status.textContent = 'Режим администратора выключен.';
}

function openAdminModal() {
  const modal = document.getElementById('adminModal');
  document.getElementById('adminAuthError').textContent = '';
  document.getElementById('adminLoginInput').value = '';
  document.getElementById('adminPasswordInput').value = '';
  modal.classList.add('show');
  modal.setAttribute('aria-hidden', 'false');
  setTimeout(() => document.getElementById('adminLoginInput').focus(), 0);
}

function closeAdminModal() {
  const modal = document.getElementById('adminModal');
  modal.classList.remove('show');
  modal.setAttribute('aria-hidden', 'true');
}

async function submitAdminLogin() {
  const login = document.getElementById('adminLoginInput').value.trim();
  const password = document.getElementById('adminPasswordInput').value;
  const err = document.getElementById('adminAuthError');
  if (!login || !password) {
    err.textContent = 'Введите логин и пароль.';
    return;
  }
  try {
    const r = await fetch('/api/admin/login', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ login, password }),
    });
    const data = await parseApiResponse(r);
    adminSessionId = String(data.session_id || '');
    if (!adminSessionId) throw new Error('Сессия не создана');
    localStorage.setItem('adminSessionId', adminSessionId);
    isAdminSession = true;
    applyAdminUi();
    closeAdminModal();
    document.getElementById('status').textContent = 'Режим администратора включён.';
  } catch (e) {
    err.textContent = e.message || 'Ошибка авторизации';
  }
}

function updateWeeksToggleLabel() {
  const btn = document.getElementById('btnToggleWeeks');
  if (!btn) return;
  btn.textContent = showAllWeeks ? 'Скрыть лишние недели' : 'Показать все недели';
}

function fillWeekSelects(weeks, prev, last) {
  const selPrev = document.getElementById('weekPrev');
  const selLast = document.getElementById('weekLast');
  selPrev.innerHTML = '';
  selLast.innerHTML = '';
  weeks.forEach(w => {
    const o1 = document.createElement('option');
    o1.value = w; o1.textContent = w;
    const o2 = o1.cloneNode(true);
    selPrev.appendChild(o1);
    selLast.appendChild(o2);
  });
  if (weeks.length >= 2) {
    selPrev.value = String(prev ?? weeks[weeks.length - 2]);
    selLast.value = String(last ?? weeks[weeks.length - 1]);
  } else if (weeks.length === 1) {
    selPrev.value = selLast.value = String(weeks[0]);
  }
}

async function refreshWeeks(forceLatest = false) {
  const year = document.getElementById('year').value;
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const q = new URLSearchParams({ year });
  if (wh) q.set('wh_ids', wh);
  try {
    const r = await fetch('/api/weeks?' + q);
    if (!r.ok) return;
    const data = await r.json();
    availableWeeks = data.weeks || [];
    const curPrev = parseInt(document.getElementById('weekPrev').value, 10);
    const curLast = parseInt(document.getElementById('weekLast').value, 10);
    const hasCurPrev = Number.isFinite(curPrev) && availableWeeks.includes(curPrev);
    const hasCurLast = Number.isFinite(curLast) && availableWeeks.includes(curLast);
    if (forceLatest || !hasCurPrev || !hasCurLast) {
      fillWeekSelects(availableWeeks, data.week_prev, data.week_last);
    } else {
      fillWeekSelects(availableWeeks, curPrev, curLast);
    }
  } catch (_) {}
}

async function selectBuilding(b) {
  activeBuilding = b.id;
  syncBuildingButtons();
  if (!b.wh_ids || b.wh_ids.length === 0) {
    selectedWh = new Set(ALL_WH_IDS);
    document.querySelectorAll('#whGrid input').forEach(cb => cb.checked = true);
  } else {
    selectedWh = new Set(b.wh_ids);
    document.querySelectorAll('#whGrid input').forEach(cb => {
      cb.checked = selectedWh.has(parseInt(cb.value, 10));
    });
  }
  await refreshWeeks();
  loadReport();
}

function syncBuildingButtons() {
  document.querySelectorAll('#buildingBtns button').forEach(btn => {
    btn.classList.toggle('active', btn.dataset.id === activeBuilding);
  });
}

function selectedWeeks() {
  let wp = document.getElementById('weekPrev').value;
  let wl = document.getElementById('weekLast').value;
  if (!wp || !wl) {
    if (availableWeeks.length >= 2) {
      wp = String(availableWeeks[availableWeeks.length - 2]);
      wl = String(availableWeeks[availableWeeks.length - 1]);
    } else {
      wp = String(CONFIG.week_prev || 20);
      wl = String(CONFIG.week_last || 21);
    }
    fillWeekSelects(availableWeeks.length ? availableWeeks : [parseInt(wp,10)], parseInt(wp,10), parseInt(wl,10));
  }
  return { wp, wl };
}

async function parseApiResponse(r) {
  const raw = await r.text();
  if (!r.ok) {
    try {
      const err = JSON.parse(raw);
      throw new Error(err.error || raw);
    } catch (e) {
      if (e instanceof Error && e.message !== raw) throw e;
      throw new Error(raw || r.statusText);
    }
  }
  return JSON.parse(raw);
}

async function refreshData() {
  const status = document.getElementById('status');
  const grid = document.getElementById('reportGrid');
  const btn = document.getElementById('btnRefreshData');
  if (CONFIG.refresh_token_required && !isAdminSession) {
    status.textContent = 'Сначала выполните вход администратора.';
    return;
  }
  btn.disabled = true;
  grid.classList.add('loading');
  status.textContent = 'Обновление базы и загрузка отчёта…';
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const year = document.getElementById('year').value;
  const q = new URLSearchParams({ year });
  if (wh) q.set('wh_ids', wh);
  const headers = adminSessionId ? { 'X-Admin-Session': adminSessionId } : {};
  try {
    const data = await parseApiResponse(await fetch('/api/refresh?' + q, { method: 'POST', headers }));
    if (data.weeks && data.weeks.length) {
      availableWeeks = data.weeks;
      const wp = data.week_prev ?? availableWeeks[availableWeeks.length - 2];
      const wl = data.week_last ?? availableWeeks[availableWeeks.length - 1];
      fillWeekSelects(availableWeeks, wp, wl);
    } else {
      await refreshWeeks();
    }
    await loadReport();
    const parts = [
      'Данные обновлены',
      data.row_count != null ? data.row_count.toLocaleString('ru') + ' строк' : '',
      data.max_date ? 'посл. дата ' + data.max_date : '',
    ].filter(Boolean);
    if (data.refresh_note) parts.push(data.refresh_note);
    status.textContent = parts.join(' · ');
  } catch (e) {
    if (String(e.message || '').includes('Недостаточно прав')) {
      isAdminSession = false;
      adminSessionId = '';
      localStorage.removeItem('adminSessionId');
      applyAdminUi();
      status.textContent = 'Недостаточно прав: выполните вход администратора с корректным токеном.';
    } else {
    status.textContent = 'Ошибка обновления: ' + e.message;
    }
  } finally {
    btn.disabled = false;
    grid.classList.remove('loading');
  }
}

async function exportXlsx() {
  const status = document.getElementById('status');
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
  });
  q.set('show_all_weeks', showAllWeeks ? '1' : '0');
  if (wh) q.set('wh_ids', wh);
  try {
    status.textContent = 'Готовим XLSX...';
    const r = await fetch('/api/export/xlsx?' + q);
    if (!r.ok) throw new Error(await r.text());
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    const building = activeBuilding && activeBuilding !== 'custom' ? activeBuilding : 'wh';
    a.download = `write_offs_${building}_${q.get('year')}_w${wp}-${wl}.xlsx`;
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    status.textContent = 'XLSX сформирован';
  } catch (e) {
    status.textContent = 'Ошибка экспорта: ' + e.message;
  }
}

async function loadReport() {
  const status = document.getElementById('status');
  const grid = document.getElementById('reportGrid');
  grid.classList.add('loading');
  const wh = selectedWh.size ? Array.from(selectedWh).join(',') : '';
  const { wp, wl } = selectedWeeks();
  const q = new URLSearchParams({
    year: document.getElementById('year').value,
    week_prev: wp,
    week_last: wl,
  });
  q.set('show_all_weeks', showAllWeeks ? '1' : '0');
  if (wh) q.set('wh_ids', wh);
  status.textContent = 'Загрузка…';
  try {
    const data = await parseApiResponse(await fetch('/api/report?' + q));
    grid.innerHTML = data.html;
    updateKpis(data.kpis || null);
    applyTableSearch();
    grid.querySelectorAll('tr.drill-row').forEach(tr => {
      tr.addEventListener('click', () => {
        const kind = tr.dataset.drill;
        const params = {};
        if (kind === 'reason' && tr.dataset.reasonId) params.reason_id = tr.dataset.reasonId;
        if (kind === 'category' && tr.dataset.parentName) params.parent_name = tr.dataset.parentName;
        if (tr.dataset.org0 === '1') params.cnt_org = '0';
        openDetailsDrill(params);
      });
    });
    const sorted = Array.from(selectedWh).sort((a,b)=>a-b);
    const label = selectedWh.size === ALL_WH_IDS.length
      ? 'все корпуса (' + ALL_WH_IDS.length + ' WH)'
      : sorted.map(whLabel).join('; ');
    status.textContent = `WH: ${label} · расчёт нед. ${data.week_prev}→${data.week_last}, в таблице: ${(data.weeks || []).join(', ')} (${data.year}). Клик по строке → детализация.`;
  } catch (e) {
    status.textContent = 'Ошибка: ' + e.message;
  } finally {
    grid.classList.remove('loading');
  }
}

init();
</script>
</body>
</html>
"""


DETAILS_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Детализация write_offs</title>
<style>
* { box-sizing: border-box; }
:root {
  --bg: #f3f6fb;
  --surface: #ffffff;
  --text: #0f172a;
  --muted: #64748b;
  --line: #dbe4f0;
  --primary: #2563eb;
  --primary-2: #1d4ed8;
  --shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
}
body {
  margin: 0;
  font: 14px/1.45 Inter, "Segoe UI", Roboto, Arial, sans-serif;
  background:
    radial-gradient(1200px 520px at 0% -10%, rgba(37, 99, 235, 0.08), transparent 60%),
    var(--bg);
  color: var(--text);
}
header {
  background: linear-gradient(100deg, #0f4c81 0%, #2563eb 55%, #1d4ed8 100%);
  color: #fff;
  padding: 18px 20px 20px;
  box-shadow: 0 8px 24px rgba(29, 78, 216, 0.22);
}
header h1 { margin: 0; font-size: 24px; font-weight: 750; }
.subtitle { margin-top: 5px; color: rgba(255,255,255,.82); font-size: 13px; }
.topnav { margin: 12px 14px 10px; display: flex; gap: 8px; flex-wrap: wrap; }
.topnav a {
  text-decoration: none;
  padding: 8px 12px;
  border: 1px solid #bfd2ff;
  background: #eef4ff;
  color: #1e40af;
  border-radius: 999px;
  font-weight: 700;
  font-size: 12px;
}
.topnav a:hover { background: #dbeafe; border-color: #93c5fd; }
.topnav a.active { background: linear-gradient(180deg, #2563eb, #1d4ed8); color: #fff; border-color: #1d4ed8; }
.filters, .meta, .panel, .chips {
  margin: 0 14px 12px;
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: 16px;
  box-shadow: var(--shadow);
}
.filters {
  padding: 14px;
  display: flex;
  gap: 12px;
  align-items: end;
  flex-wrap: wrap;
}
.filters label {
  display: grid;
  gap: 4px;
  color: #334155;
  font-size: 12px;
  font-weight: 700;
}
.filters input, .filters select {
  min-width: 110px;
  border: 1px solid #cbd5e1;
  border-radius: 9px;
  background: #fff;
  padding: 7px 9px;
  font: inherit;
}
.filters input[name="search"] { min-width: 240px; }
.filters input[name="wh_ids"], .filters input[name="parent_name"] { min-width: 160px; }
.filters input:focus, .filters select:focus {
  outline: none;
  border-color: #60a5fa;
  box-shadow: 0 0 0 3px rgba(59,130,246,.14);
}
.btn {
  border: 1px solid transparent;
  border-radius: 10px;
  padding: 8px 13px;
  font-weight: 700;
  cursor: pointer;
  transition: all .16s ease;
}
.btn:hover { transform: translateY(-1px); box-shadow: 0 8px 18px rgba(15,23,42,.12); }
.btn.primary { background: linear-gradient(180deg, var(--primary), var(--primary-2)); color: #fff; }
.btn.secondary { background: #f0f9ff; color: #075985; border-color: #bae6fd; }
.btn.export { background: #f8fafc; color: #334155; border-color: #dbe4f0; }
.meta { padding: 10px 12px; color: #334155; }
.chips { padding: 10px 12px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
.chip {
  display: inline-flex;
  gap: 6px;
  align-items: center;
  background: #eff6ff;
  color: #1e40af;
  border: 1px solid #bfdbfe;
  border-radius: 999px;
  padding: 4px 10px;
  font-size: 12px;
  font-weight: 700;
}
.chip button {
  border: 0;
  background: transparent;
  color: #1d4ed8;
  cursor: pointer;
  font-weight: 800;
}
.panel { overflow: hidden; }
.table-wrap {
  overflow: auto;
  max-height: calc(100vh - 340px);
  background: #fff;
}
table { width: max-content; min-width: 100%; border-collapse: collapse; }
th, td {
  border-bottom: 1px solid #e2e8f0;
  padding: 6px 10px;
  text-align: left;
  max-width: 320px;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}
th {
  position: sticky;
  top: 0;
  z-index: 2;
  background: #f8fafc;
  color: #334155;
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: .25px;
}
th.sortable { cursor: pointer; user-select: none; }
th.sortable:hover { background: #eef6ff; color: #1d4ed8; }
th.active-sort { color: #1d4ed8; background: #e0efff; }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
td.clickable { cursor: pointer; color: #1d4ed8; text-decoration: underline; text-underline-offset: 2px; }
td.clickable:hover { background: #dbeafe !important; }
tbody tr:nth-child(even) td { background: #fbfdff; }
tbody tr:hover td { background: #eef6ff; }
.pager {
  padding: 10px 12px;
  display: flex;
  gap: 8px;
  align-items: center;
  flex-wrap: wrap;
  border-top: 1px solid #e2e8f0;
  background: #f8fafc;
}
.pager .spacer { flex: 1; }
.muted { color: var(--muted); }
@media (max-width: 900px) {
  .filters, .meta, .panel, .chips { margin-left: 10px; margin-right: 10px; }
  .filters input[name="search"] { min-width: 180px; }
}
</style>
</head>
<body>
<header>
  <h1>Полная детализация</h1>
  <div class="subtitle">Сырые строки `brak_team.write_offs` · клик по ячейке фильтрует · клик по заголовку сортирует</div>
</header>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details" class="active">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/status">Статус</a>
</nav>
<form class="filters" id="filters">
  <label>Дата от <input name="date_from" type="date"></label>
  <label>Дата до <input name="date_to" type="date"></label>
  <label>Офис <input name="office_id" type="number" placeholder="office_id"></label>
  <label>WH <input name="wh_id" type="number" placeholder="wh_id"></label>
  <label>WH list <input name="wh_ids" placeholder="1,2,3"></label>
  <label>Тип <input name="type" placeholder="type"></label>
  <label>reason_id <input name="reason_id" type="number"></label>
  <label>Категория <input name="parent_name" placeholder="parent_name"></label>
  <label>nm_id <input name="nm_id" type="number"></label>
  <label>shk_id <input name="shk_id" type="number"></label>
  <label>cnt_org <input name="cnt_org" type="number"></label>
  <label>Бренд <input name="brand_name" placeholder="brand_name"></label>
  <label>Поиск <input name="search" placeholder="title, причина, бренд, nm_id, shk_id"></label>
  <label>На стр.
    <select name="per_page">
      <option>25</option>
      <option selected>50</option>
      <option>100</option>
      <option>250</option>
      <option>500</option>
    </select>
  </label>
  <input type="hidden" name="sort_by" value="date">
  <input type="hidden" name="sort_dir" value="desc">
  <button class="btn primary" type="submit">Применить</button>
  <button class="btn export" type="button" id="btnReset">Сбросить</button>
  <button class="btn secondary" type="button" id="btnExport">Экспорт XLSX</button>
</form>
<div class="chips" id="activeFilters"></div>
<div class="meta" id="meta">Загрузка…</div>
<section class="panel">
  <div class="table-wrap">
    <table>
      <thead><tr id="thead"></tr></thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
  <div class="pager">
    <button class="btn export" type="button" id="btnPrev">Назад</button>
    <button class="btn export" type="button" id="btnNext">Вперёд</button>
    <span class="muted" id="pageInfo"></span>
    <span class="spacer"></span>
    <span class="muted">Клик по ячейке = фильтр. Экспорт по умолчанию 5000 строк.</span>
  </div>
</section>
<script>
let currentPage = 1;
let columns = [];
let clickFilters = new Set(['wh_id','office_id','nm_id','shk_id','reason_id','parent_name','type','cnt_org','brand_name']);
const SORTABLE = new Set(['date','amount','total_cost','share','office_id','wh_id','nm_id','shk_id','reason_id','cnt_org','cnt_ors','cnt_ocr','type','parent_name','reason_descr','title','brand_name','subject_name']);

function fmtValue(v) {
  if (v === null || v === undefined || v === '') return '—';
  if (typeof v === 'number') return v.toLocaleString('ru-RU');
  return String(v);
}

function formEl(name) {
  return document.getElementById('filters').elements[name];
}

function paramsFor(page) {
  const fd = new FormData(document.getElementById('filters'));
  const q = new URLSearchParams();
  for (const [k, v] of fd.entries()) {
    const value = String(v || '').trim();
    if (value) q.set(k, value);
  }
  q.set('page', String(page));
  return q;
}

function hydrateFiltersFromUrl() {
  const q = new URLSearchParams(window.location.search);
  const form = document.getElementById('filters');
  for (const [k, v] of q.entries()) {
    const el = form.elements[k];
    if (el) el.value = v;
  }
  if (!form.elements.sort_by.value) form.elements.sort_by.value = 'date';
  if (!form.elements.sort_dir.value) form.elements.sort_dir.value = 'desc';
  currentPage = parseInt(q.get('page') || '1', 10) || 1;
}

function renderActiveFilters() {
  const box = document.getElementById('activeFilters');
  const keys = ['date_from','date_to','office_id','wh_id','wh_ids','type','reason_id','parent_name','nm_id','shk_id','cnt_org','brand_name','search'];
  const chips = [];
  keys.forEach(k => {
    const el = formEl(k);
    const v = (el && el.value || '').trim();
    if (!v) return;
    chips.push(`<span class="chip">${k}=${v}<button type="button" data-clear="${k}" title="Сбросить">×</button></span>`);
  });
  const sortBy = formEl('sort_by').value || 'date';
  const sortDir = formEl('sort_dir').value || 'desc';
  chips.push(`<span class="chip">sort=${sortBy} ${sortDir}</span>`);
  box.innerHTML = chips.length ? chips.join('') : '<span class="muted">Активных фильтров нет. Кликните по ячейке таблицы, чтобы отфильтровать.</span>';
  box.querySelectorAll('[data-clear]').forEach(btn => {
    btn.addEventListener('click', () => {
      const el = formEl(btn.dataset.clear);
      if (el) el.value = '';
      loadDetails(1);
    });
  });
}

function setSort(key) {
  const by = formEl('sort_by');
  const dir = formEl('sort_dir');
  if (by.value === key) {
    dir.value = dir.value === 'asc' ? 'desc' : 'asc';
  } else {
    by.value = key;
    dir.value = (key === 'date' || key === 'amount' || key === 'total_cost') ? 'desc' : 'asc';
  }
  loadDetails(1);
}

function applyCellFilter(key, value) {
  if (value === null || value === undefined || value === '') return;
  const el = formEl(key);
  if (!el) return;
  el.value = String(value);
  if (key === 'wh_id') {
    const list = formEl('wh_ids');
    if (list) list.value = '';
  }
  loadDetails(1);
}

async function loadDetails(page = 1, pushState = true) {
  const meta = document.getElementById('meta');
  const tbody = document.getElementById('tbody');
  const thead = document.getElementById('thead');
  meta.textContent = 'Загрузка…';
  tbody.innerHTML = '';
  const q = paramsFor(page);
  try {
    const r = await fetch('/api/details?' + q);
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    columns = data.columns || [];
    if (Array.isArray(data.click_filters)) clickFilters = new Set(data.click_filters);
    currentPage = data.page || page;
    if (data.sort_by) formEl('sort_by').value = data.sort_by;
    if (data.sort_dir) formEl('sort_dir').value = data.sort_dir;
    if (pushState) history.replaceState(null, '', '/details?' + q);
    const sortBy = formEl('sort_by').value;
    const sortDir = formEl('sort_dir').value;
    thead.innerHTML = columns.map(c => {
      const sortable = SORTABLE.has(c.key);
      const active = sortable && c.key === sortBy;
      const mark = active ? (sortDir === 'asc' ? ' ▲' : ' ▼') : '';
      const cls = sortable ? ` class="sortable${active ? ' active-sort' : ''}" data-sort="${c.key}"` : '';
      return `<th${cls}>${c.label}${mark}</th>`;
    }).join('');
    thead.querySelectorAll('[data-sort]').forEach(th => {
      th.addEventListener('click', () => setSort(th.dataset.sort));
    });
    tbody.innerHTML = (data.rows || []).map(row => {
      return '<tr>' + columns.map(c => {
        const rawVal = row[c.key];
        const value = fmtValue(rawVal);
        const isNum = typeof rawVal === 'number';
        const canClick = clickFilters.has(c.key) && rawVal !== null && rawVal !== undefined && rawVal !== '';
        const cls = [isNum ? 'num' : '', canClick ? 'clickable' : ''].filter(Boolean).join(' ');
        const attrs = canClick
          ? ` class="${cls}" data-filter-key="${c.key}" data-filter-value="${String(rawVal).replaceAll('"', '&quot;')}" title="Фильтр: ${c.key}=${String(value).replaceAll('"', '&quot;')}"`
          : ` class="${cls}" title="${String(value).replaceAll('"', '&quot;')}"`;
        return `<td${attrs}>${value}</td>`;
      }).join('') + '</tr>';
    }).join('');
    tbody.querySelectorAll('[data-filter-key]').forEach(td => {
      td.addEventListener('click', () => applyCellFilter(td.dataset.filterKey, td.dataset.filterValue));
    });
    meta.textContent = `Всего по фильтру: ${Number(data.total || 0).toLocaleString('ru-RU')} · показано: ${(data.rows || []).length}`;
    document.getElementById('pageInfo').textContent = `Страница ${data.page} / ${data.pages}`;
    document.getElementById('btnPrev').disabled = data.page <= 1;
    document.getElementById('btnNext').disabled = data.page >= data.pages;
    renderActiveFilters();
  } catch (e) {
    meta.textContent = 'Ошибка загрузки: ' + (e.message || e);
  }
}

function exportXlsx() {
  const q = paramsFor(currentPage);
  q.delete('page');
  q.set('limit', '5000');
  const reason = q.get('reason_id') || 'all';
  const parent = (q.get('parent_name') || 'all').replaceAll(/[^\\w\\-]+/g, '_').slice(0, 24);
  window.location.href = '/api/details/export/xlsx?' + q;
  // filename is set by server; keep query rich for filters
  void reason; void parent;
}

document.getElementById('filters').addEventListener('submit', (e) => {
  e.preventDefault();
  loadDetails(1);
});
document.getElementById('btnReset').addEventListener('click', () => {
  document.getElementById('filters').reset();
  formEl('sort_by').value = 'date';
  formEl('sort_dir').value = 'desc';
  loadDetails(1);
});
document.getElementById('btnExport').addEventListener('click', exportXlsx);
document.getElementById('btnPrev').addEventListener('click', () => loadDetails(Math.max(1, currentPage - 1)));
document.getElementById('btnNext').addEventListener('click', () => loadDetails(currentPage + 1));

hydrateFiltersFromUrl();
loadDetails(currentPage, false);
</script>
</body>
</html>
"""


WEEKLY_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Динамика по неделям</title>
<style>
* { box-sizing: border-box; }
body { margin: 0; font: 14px/1.45 Inter, "Segoe UI", Roboto, Arial, sans-serif; background: #f3f6fb; color: #0f172a; }
header { background: linear-gradient(100deg, #0f4c81 0%, #2563eb 55%, #1d4ed8 100%); color: #fff; padding: 18px 20px; }
header h1 { margin: 0; font-size: 22px; }
.subtitle { margin-top: 4px; opacity: .85; font-size: 13px; }
.topnav { margin: 12px 14px 10px; display: flex; gap: 8px; flex-wrap: wrap; }
.topnav a { text-decoration: none; padding: 8px 12px; border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; border-radius: 999px; font-weight: 700; font-size: 12px; }
.topnav a.active { background: #1d4ed8; color: #fff; border-color: #1d4ed8; }
.toolbar, .meta, .panel { margin: 0 14px 12px; background: #fff; border: 1px solid #dbe4f0; border-radius: 14px; box-shadow: 0 10px 28px rgba(15,23,42,.08); }
.toolbar { padding: 12px; display: flex; gap: 10px; flex-wrap: wrap; align-items: end; }
.toolbar label { display: grid; gap: 4px; font-size: 12px; font-weight: 700; color: #334155; }
.toolbar input, .toolbar select { border: 1px solid #cbd5e1; border-radius: 8px; padding: 7px 9px; font: inherit; min-width: 120px; }
.btn { border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; border-radius: 10px; padding: 8px 12px; font-weight: 700; cursor: pointer; }
.btn.primary { background: linear-gradient(180deg, #2563eb, #1d4ed8); color: #fff; border-color: #1d4ed8; }
.meta { padding: 10px 12px; color: #334155; }
.panel { overflow: hidden; }
.chart { padding: 14px 16px 8px; }
.bars { display: grid; grid-auto-flow: column; grid-auto-columns: minmax(28px, 1fr); gap: 6px; align-items: end; height: 220px; }
.bar-wrap { display: flex; flex-direction: column; align-items: center; height: 100%; justify-content: end; gap: 4px; }
.bar { width: 100%; max-width: 36px; border-radius: 8px 8px 4px 4px; background: linear-gradient(180deg, #60a5fa, #2563eb); min-height: 2px; }
.bar.org0 { background: linear-gradient(180deg, #fbbf24, #d97706); opacity: .9; position: absolute; left: 0; right: 0; bottom: 0; border-radius: 8px 8px 4px 4px; }
.bar-stack { position: relative; width: 100%; max-width: 36px; height: var(--h); }
.wlabel { font-size: 10px; color: #64748b; }
table { width: 100%; border-collapse: collapse; }
th, td { border-top: 1px solid #e2e8f0; padding: 8px 10px; text-align: left; }
th { background: #f8fafc; font-size: 12px; color: #334155; }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
.top { font-size: 12px; color: #475569; max-width: 420px; }
.muted { color: #64748b; }
</style>
</head>
<body>
<header>
  <h1>Динамика по неделям</h1>
  <div class="subtitle">Сумма брака, ORG0 и топ причин по ISO-неделям</div>
</header>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly" class="active">Динамика</a>
  <a href="/status">Статус</a>
</nav>
<div class="toolbar">
  <label>Год <input id="year" type="number" value="2026"></label>
  <label>WH ids <input id="wh_ids" placeholder="пусто = все"></label>
  <label>Топ причин
    <select id="top_n">
      <option>3</option>
      <option selected>5</option>
      <option>10</option>
    </select>
  </label>
  <button class="btn primary" id="btnLoad" type="button">Загрузить</button>
</div>
<div class="meta" id="meta">Загрузка…</div>
<section class="panel">
  <div class="chart">
    <div class="bars" id="bars"></div>
  </div>
  <div style="overflow:auto">
    <table>
      <thead>
        <tr>
          <th>Неделя</th>
          <th class="num">Всего, ₽</th>
          <th class="num">ORG0, ₽</th>
          <th class="num">ORG0 %</th>
          <th class="num">Строк</th>
          <th>Топ причин</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
  </div>
</section>
<script>
function fmt(n) { return Number(n || 0).toLocaleString('ru-RU', { maximumFractionDigits: 0 }); }
function pct(n) { return (Number(n || 0)).toLocaleString('ru-RU', { maximumFractionDigits: 1 }) + '%'; }

function hydrate() {
  const q = new URLSearchParams(location.search);
  if (q.get('year')) document.getElementById('year').value = q.get('year');
  if (q.get('wh_ids')) document.getElementById('wh_ids').value = q.get('wh_ids');
  if (q.get('top_n')) document.getElementById('top_n').value = q.get('top_n');
}

async function loadWeekly() {
  const meta = document.getElementById('meta');
  const year = document.getElementById('year').value || new Date().getFullYear();
  const wh = (document.getElementById('wh_ids').value || '').trim();
  const topN = document.getElementById('top_n').value || '5';
  const q = new URLSearchParams({ year, top_n: topN });
  if (wh) q.set('wh_ids', wh);
  meta.textContent = 'Загрузка…';
  try {
    const r = await fetch('/api/weekly?' + q);
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    history.replaceState(null, '', '/weekly?' + q);
    const weeks = data.weeks || [];
    const maxAmt = Math.max(1, ...weeks.map(w => w.amount_all || 0));
    document.getElementById('bars').innerHTML = weeks.map(w => {
      const h = Math.max(2, Math.round((w.amount_all || 0) / maxAmt * 180));
      const h0 = Math.max(0, Math.round((w.amount_org0 || 0) / maxAmt * 180));
      return `<div class="bar-wrap" title="W${w.week}: ${fmt(w.amount_all)}">
        <div class="bar-stack" style="--h:${h}px">
          <div class="bar" style="height:${h}px"></div>
          <div class="bar org0" style="height:${h0}px"></div>
        </div>
        <div class="wlabel">${w.week}</div>
      </div>`;
    }).join('');
    document.getElementById('tbody').innerHTML = weeks.map(w => {
      const top = (w.top_reasons || []).map(t => `${t.reason_descr} (${fmt(t.amount)})`).join('; ');
      return `<tr>
        <td>${w.week}</td>
        <td class="num">${fmt(w.amount_all)}</td>
        <td class="num">${fmt(w.amount_org0)}</td>
        <td class="num">${pct(w.org0_share)}</td>
        <td class="num">${fmt(w.row_count)}</td>
        <td class="top">${top || '—'}</td>
      </tr>`;
    }).join('');
    const t = data.totals || {};
    meta.textContent = `Год ${data.year} · источник ${data.source} · всего ${fmt(t.amount_all)} ₽ · ORG0 ${fmt(t.amount_org0)} ₽ · недель ${weeks.length}`;
  } catch (e) {
    meta.textContent = 'Ошибка: ' + (e.message || e);
  }
}

document.getElementById('btnLoad').onclick = loadWeekly;
hydrate();
loadWeekly();
</script>
</body>
</html>
"""


STATUS_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Статус системы</title>
<style>
* { box-sizing: border-box; }
body { margin: 0; font: 14px/1.45 Inter, "Segoe UI", Roboto, Arial, sans-serif; background: #f3f6fb; color: #0f172a; }
header { background: linear-gradient(100deg, #0f4c81 0%, #2563eb 55%, #1d4ed8 100%); color: #fff; padding: 18px 20px; }
header h1 { margin: 0; font-size: 22px; }
.topnav { margin: 12px 14px 10px; display: flex; gap: 8px; flex-wrap: wrap; }
.topnav a { text-decoration: none; padding: 8px 12px; border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; border-radius: 999px; font-weight: 700; font-size: 12px; }
.topnav a.active { background: #1d4ed8; color: #fff; border-color: #1d4ed8; }
.grid { margin: 0 14px 14px; display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 12px; }
.card { background: #fff; border: 1px solid #dbe4f0; border-radius: 14px; box-shadow: 0 10px 28px rgba(15,23,42,.08); padding: 14px; }
.card h2 { margin: 0 0 10px; font-size: 14px; color: #1e3a8a; }
.row { display: flex; justify-content: space-between; gap: 10px; padding: 6px 0; border-bottom: 1px solid #f1f5f9; font-size: 13px; }
.row:last-child { border-bottom: 0; }
.k { color: #64748b; }
.v { font-weight: 700; text-align: right; word-break: break-all; }
.ok { color: #15803d; }
.warn { color: #b45309; }
.err { color: #b91c1c; }
.meta { margin: 0 14px 12px; color: #334155; }
.btn { margin-left: 14px; border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; border-radius: 10px; padding: 8px 12px; font-weight: 700; cursor: pointer; }
</style>
</head>
<body>
<header>
  <h1>Статус системы</h1>
</header>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/status" class="active">Статус</a>
</nav>
<div class="meta" id="meta">Загрузка…</div>
<button class="btn" id="btnReload" type="button">Обновить</button>
<div class="grid" id="grid"></div>
<script>
function cls(v) {
  if (v === 'ok' || v === 'set' || v === true) return 'ok';
  if (v === 'missing' || v === 'empty' || v === 'error' || v === false) return 'err';
  if (v === 'degraded') return 'warn';
  return '';
}
function row(k, v, c='') {
  return `<div class="row"><span class="k">${k}</span><span class="v ${c}">${v ?? '—'}</span></div>`;
}
function card(title, body) {
  return `<section class="card"><h2>${title}</h2>${body}</section>`;
}
async function loadStatus() {
  const meta = document.getElementById('meta');
  const grid = document.getElementById('grid');
  meta.textContent = 'Загрузка…';
  try {
    const r = await fetch('/api/status');
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const d = JSON.parse(raw);
    meta.innerHTML = `Общий статус: <b class="${cls(d.status)}">${d.status}</b> · env: ${d.vercel_env || 'local'}`;
    const envRows = Object.entries(d.env || {}).map(([k,v]) => row(k, v, cls(v))).join('');
    const db = d.database || {};
    const mv = d.matview || {};
    const cache = d.cache || {};
    const admin = d.admin || {};
    grid.innerHTML = [
      card('Environment', envRows + row('db_env_error', d.db_env_error || 'нет', d.db_env_error ? 'err' : 'ok')),
      card('Database', row('status', db.status, cls(db.status)) + row('detail', db.detail) + row('row_count', db.row_count) + row('max_date', db.max_date) + row('total_amount', db.total_amount)),
      card('Matview', row('enabled', String(mv.enabled), cls(mv.enabled)) + row('available', String(mv.available), cls(mv.available)) + row('name', mv.name) + row('row_count', mv.row_count) + row('max_year/week', `${mv.max_year ?? '—'} / ${mv.max_week ?? '—'}`) + row('bootstrap_ok_age_sec', mv.bootstrap_ok_age_sec) + row('bootstrap_fail', mv.bootstrap_fail_msg || 'нет', mv.bootstrap_fail_msg ? 'err' : 'ok')),
      card('Cache / Admin', row('cache_entries', cache.entries) + row('report_ttl_sec', cache.report_ttl_sec) + row('weeks_ttl_sec', cache.weeks_ttl_sec) + row('refresh_token_required', String(admin.refresh_token_required)) + row('active_sessions', admin.active_sessions)),
    ].join('');
  } catch (e) {
    meta.textContent = 'Ошибка: ' + (e.message || e);
  }
}
document.getElementById('btnReload').onclick = loadStatus;
loadStatus();
</script>
</body>
</html>
"""


NOMENCLATURE_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Кол-во брака по номенклатуре</title>
<style>
* { box-sizing: border-box; }
body { margin: 0; font: 14px/1.45 Inter, "Segoe UI", Roboto, Arial, sans-serif; background: #f3f6fb; color: #0f172a; }
header { background: linear-gradient(100deg, #0f4c81 0%, #2563eb 55%, #1d4ed8 100%); color: #fff; padding: 12px 16px; }
header h1 { margin: 0; font-size: 20px; font-weight: 700; }
.topnav { margin: 10px 12px 8px; display: flex; gap: 8px; }
.topnav a { text-decoration: none; padding: 8px 12px; border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; border-radius: 999px; font-weight: 600; font-size: 12px; }
.topnav a.active { background: #1d4ed8; color: #fff; border-color: #1d4ed8; }
.wrap { padding: 14px 12px; }
.panel { background: #fff; border: 1px solid #dbe4f0; border-radius: 12px; box-shadow: 0 10px 28px rgba(15,23,42,.08); overflow: hidden; }
.panel h2 { margin: 0; padding: 10px 12px; font-size: 14px; border-bottom: 1px solid #e2e8f0; background: #eff6ff; color: #1e3a8a; }
.toolbar { padding: 10px 12px; border-bottom: 1px solid #e2e8f0; background: #fff; }
.btn { display: inline-block; padding: 8px 12px; border-radius: 8px; border: 1px solid #bfd2ff; background: #eef4ff; color: #1e40af; text-decoration: none; font-size: 12px; font-weight: 600; }
.btn:hover { background: #dbeafe; border-color: #93c5fd; }
.meta { padding: 10px 12px; font-size: 13px; color: #334155; border-bottom: 1px solid #e2e8f0; }
table { width: 100%; border-collapse: collapse; }
th, td { border: 1px solid #e2e8f0; padding: 8px 10px; }
th { background: #f8fafc; color: #334155; text-align: left; font-size: 12px; }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
tr.total td { background: #eef2ff; font-weight: 700; }
.muted { color: #64748b; font-size: 12px; }
</style>
</head>
<body>
<header><h1>Кол-во брака по номенклатуре</h1></header>
<nav class="topnav">
  <a href="/">Дашборд</a>
  <a href="/nomenclature" class="active">Номенклатура</a>
  <a href="/details">Детализация</a>
  <a href="/weekly">Динамика</a>
  <a href="/status">Статус</a>
</nav>
<div class="wrap">
  <section class="panel">
    <div class="toolbar">
      <a class="btn" href="/">На главную</a>
      <button type="button" class="btn" id="btnNmExport">Экспорт XLSX</button>
    </div>
    <h2>Последняя неделя: Номенклатура × Корпуса</h2>
    <div class="meta" id="meta">Загрузка…</div>
    <table>
      <thead>
        <tr>
          <th>Номенклатура</th>
          <th class="num">1 корпус</th>
          <th class="num">2 корпус</th>
          <th class="num">3 корпус</th>
          <th class="num">Итог</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="meta muted">Метрика: количество записей брака по `nm_id` за последнюю доступную ISO-неделю.</div>
  </section>
</div>
<script>
function fmtInt(v) { return Number(v || 0).toLocaleString('ru-RU'); }
function fmtPct(v) { return Number(v || 0).toLocaleString('ru-RU', { minimumFractionDigits: 2, maximumFractionDigits: 2 }); }

async function loadData() {
  const meta = document.getElementById('meta');
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '';
  try {
    const r = await fetch('/api/nomenclature/latest');
    const raw = await r.text();
    if (!r.ok) throw new Error(raw || r.statusText);
    const data = JSON.parse(raw);
    const rows = data.rows || [];
    rows.forEach((it) => {
      const tr = document.createElement('tr');
      tr.innerHTML = `
        <td>${it.nomenclature ?? '—'}</td>
        <td class="num">${fmtInt(it.corpus_1)}</td>
        <td class="num">${fmtInt(it.corpus_2)}</td>
        <td class="num">${fmtInt(it.corpus_3)}</td>
        <td class="num">${fmtInt(it.total)}</td>
      `;
      tbody.appendChild(tr);
    });
    const totalTr = document.createElement('tr');
    totalTr.className = 'total';
    const t = data.totals || {};
    totalTr.innerHTML = `
      <td>Итого</td>
      <td class="num">${fmtInt(t.corpus_1)}</td>
      <td class="num">${fmtInt(t.corpus_2)}</td>
      <td class="num">${fmtInt(t.corpus_3)}</td>
      <td class="num">${fmtInt(t.total)}</td>
    `;
    tbody.appendChild(totalTr);
    const y = data.latest_year ?? '—';
    const w = data.latest_week ?? '—';
    meta.textContent = `Период: ${y}-я неделя ${w} · номенклатур: ${fmtInt(rows.length)} · записей: ${fmtInt(t.total)}`;
  } catch (e) {
    meta.textContent = 'Ошибка загрузки: ' + (e.message || e);
  }
}

async function exportNomenclatureXlsx() {
  const meta = document.getElementById('meta');
  try {
    meta.textContent = 'Готовим XLSX...';
    const r = await fetch('/api/nomenclature/export/xlsx');
    if (!r.ok) throw new Error(await r.text());
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'nomenclature_latest_week.xlsx';
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    meta.textContent = 'XLSX сформирован';
  } catch (e) {
    meta.textContent = 'Ошибка экспорта: ' + (e.message || e);
  }
}

document.getElementById('btnNmExport').addEventListener('click', exportNomenclatureXlsx);
loadData();
</script>
</body>
</html>
"""


def register_routes(application) -> None:
    from datetime import datetime, timezone

    from flask import jsonify, request, send_file

    if getattr(application, "_brak_routes_registered", False):
        return
    application._brak_routes_registered = True

    def _cfg() -> dict:
        return load_config()

    def _resolve_wh_ids(cfg: dict) -> list[int] | None:
        wh_raw = request.args.get("wh_ids", "")
        catalog_ids = catalog_wh_ids(cfg)
        if wh_raw.strip():
            return parse_wh_ids(wh_raw)
        if catalog_ids:
            return catalog_ids
        return None

    @application.route("/")
    def index():
        try:
            cfg = _cfg()
            embed = {
                "wh_catalog": cfg.get("wh_catalog", []),
                "buildings": cfg.get("buildings", []),
                "week_year": cfg.get("week_year", 2026),
                "week_prev": cfg.get("week_prev", 20),
                "week_last": cfg.get("week_last", 21),
                "is_admin": False,
                "refresh_token_required": True,
                "show_all_weeks_default": False,
            }
            page = DASHBOARD_HTML.replace(
                "__CONFIG_JSON__", json.dumps(embed, ensure_ascii=False)
            )
            return page
        except Exception as exc:
            return f"<pre>Index error: {exc}</pre>", 500

    @application.route("/nomenclature")
    def nomenclature_page():
        return NOMENCLATURE_HTML

    @application.route("/details")
    def details_page():
        return DETAILS_HTML

    @application.route("/weekly")
    def weekly_page():
        return WEEKLY_HTML

    @application.route("/status")
    def status_page():
        return STATUS_HTML

    @application.route("/api/admin/login", methods=["POST"])
    def api_admin_login():
        _, expected_password = _admin_login_password()
        if not expected_password:
            sid = _admin_session_create()
            return jsonify({"ok": True, "session_id": sid, "note": "auth disabled"})
        try:
            payload = request.get_json(silent=True) or {}
            got_login = str(payload.get("login", "")).strip()
            got_password = str(payload.get("password", "")).strip()
        except Exception:
            got_login = ""
            got_password = ""
        expected_login, expected_password = _admin_login_password()
        if got_login != expected_login or got_password != expected_password:
            return jsonify({"error": "Неверный логин или пароль"}), 403
        sid = _admin_session_create()
        return jsonify({"ok": True, "session_id": sid})

    @application.route("/api/admin/logout", methods=["POST"])
    def api_admin_logout():
        sid = request.headers.get("X-Admin-Session", "").strip() or request.args.get(
            "admin_session", ""
        ).strip()
        _admin_session_delete(sid)
        return jsonify({"ok": True})

    @application.route("/api/refresh", methods=["POST"])
    def api_refresh():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        access_err = check_admin_access()
        if access_err:
            return jsonify({"error": access_err}), 403

        try:
            cfg = _cfg()
            year = request.args.get("year", cfg.get("week_year", 2026), type=int)
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")

            refresh_note = run_db_refresh()
            _cache_clear_all()
            stats = fetch_db_stats(office_id, wh_ids)
            weeks = fetch_available_weeks(year, office_id, wh_ids)
            week_prev = cfg.get("week_prev", 20)
            week_last = cfg.get("week_last", 21)
            if len(weeks) >= 2:
                week_prev, week_last = weeks[-2], weeks[-1]

            warm_report_cache_async(wh_ids, office_id, year, week_prev, week_last)
            return jsonify(
                {
                    "ok": True,
                    "refreshed_at": datetime.now(timezone.utc).isoformat(),
                    "refresh_note": refresh_note,
                    "row_count": stats["row_count"],
                    "max_date": stats["max_date"],
                    "total_amount": stats["total_amount"],
                    "year": year,
                    "weeks": weeks,
                    "week_prev": week_prev,
                    "week_last": week_last,
                }
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/matview/refresh", methods=["POST"])
    def api_matview_refresh():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        access_err = check_admin_access()
        if access_err:
            return jsonify({"error": access_err}), 403
        try:
            note = _refresh_report_matview()
            _cache_clear_all()
            return jsonify({"ok": True, "note": note or "Матпредставление обновлено"})
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/report")
    def api_report():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503

        try:
            cfg = _cfg()
            year = request.args.get("year", cfg.get("week_year", 2026), type=int)

            def _week_arg(name: str, default: int) -> int:
                raw = request.args.get(name, "")
                if raw is None or str(raw).strip() == "":
                    return default
                try:
                    return int(raw)
                except (TypeError, ValueError):
                    return default

            week_prev = _week_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _week_arg("week_last", cfg.get("week_last", 21))
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")
            show_all_weeks = (request.args.get("show_all_weeks", "0") == "1")
            data = build_report_data(
                wh_ids, office_id, year, week_prev, week_last, show_all_weeks=show_all_weeks
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

        weeks = data["weeks"]
        html_grid = (
            render_table(
                "Дефект ТОП-20, рубли",
                data["defects"],
                data["defects_total"],
                show_id=True,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Дефект",
                all_totals=data["all_totals"]["defects"],
                drill_kind="reason",
                org0_only=False,
            )
            + render_table(
                "Дефект ТОП-20, ORG 0, рубли",
                data["defects_org0"],
                data["defects_org0_total"],
                show_id=True,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Дефект",
                all_totals=data["all_totals"]["defects_org0"],
                drill_kind="reason",
                org0_only=True,
            )
            + render_table(
                "ТОП-20 категорий, рубли",
                data["categories"],
                data["categories_total"],
                show_id=False,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Категория",
                all_totals=data["all_totals"]["categories"],
                drill_kind="category",
                org0_only=False,
            )
            + render_table(
                "ТОП-20 категорий, ORG 0, рубли",
                data["categories_org0"],
                data["categories_org0_total"],
                show_id=False,
                weeks=weeks,
                week_prev=week_prev,
                week_last=week_last,
                name_header="Категория",
                all_totals=data["all_totals"]["categories_org0"],
                drill_kind="category",
                org0_only=True,
            )
        )

        return jsonify(
            {
                "html": html_grid,
                "year": year,
                "weeks": weeks,
                "week_prev": week_prev,
                "week_last": week_last,
                "kpis": build_kpi_payload(data),
            }
        )

    @application.route("/api/export/xlsx")
    def api_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = request.args.get("year", cfg.get("week_year", 2026), type=int)

            def _week_arg(name: str, default: int) -> int:
                raw = request.args.get(name, "")
                if raw is None or str(raw).strip() == "":
                    return default
                try:
                    return int(raw)
                except (TypeError, ValueError):
                    return default

            week_prev = _week_arg("week_prev", cfg.get("week_prev", 20))
            week_last = _week_arg("week_last", cfg.get("week_last", 21))
            wh_ids = _resolve_wh_ids(cfg)
            office_id = cfg.get("office_id")
            show_all_weeks = (request.args.get("show_all_weeks", "0") == "1")
            report = build_report_data(
                wh_ids, office_id, year, week_prev, week_last, show_all_weeks=show_all_weeks
            )
            blob = export_report_xlsx(report, year, week_prev, week_last)
            scope = "all"
            if wh_ids:
                scope = f"wh{len(wh_ids)}"
            filename = f"write_offs_{scope}_{year}_w{week_prev}-{week_last}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/weeks")
    def api_weeks():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = request.args.get("year", cfg.get("week_year", 2026), type=int)
            wh_ids = _resolve_wh_ids(cfg)
            weeks = fetch_available_weeks(year, cfg.get("office_id"), wh_ids)
            week_prev = cfg.get("week_prev", 20)
            week_last = cfg.get("week_last", 21)
            if len(weeks) >= 2:
                week_prev, week_last = weeks[-2], weeks[-1]
            elif len(weeks) == 1:
                week_prev = week_last = weeks[0]
            return jsonify(
                {
                    "year": year,
                    "weeks": weeks,
                    "week_prev": week_prev,
                    "week_last": week_last,
                }
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/wh_ids")
    def api_wh_ids():
        return jsonify(fetch_wh_list(_cfg().get("office_id")))

    @application.route("/api/details")
    def api_details():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            page = _detail_positive_int_arg(
                request.args, "page", 1, min_value=1, max_value=100000
            )
            per_page = _detail_positive_int_arg(
                request.args, "per_page", 50, min_value=5, max_value=500
            )
            payload = fetch_detail_page(request.args, page=page, per_page=per_page)
            return jsonify(payload)
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/details/export/xlsx")
    def api_details_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            limit = _detail_positive_int_arg(
                request.args, "limit", 5000, min_value=1, max_value=20000
            )
            blob = export_detail_xlsx(request.args, limit=limit)
            reason = request.args.get("reason_id") or "all"
            parent = re.sub(r"[^\w\-]+", "_", str(request.args.get("parent_name") or "all"))[:24]
            date_from = request.args.get("date_from") or "na"
            date_to = request.args.get("date_to") or "na"
            filename = f"write_offs_details_r{reason}_{parent}_{date_from}_{date_to}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/weekly")
    def api_weekly():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = request.args.get("year", cfg.get("week_year", 2026), type=int)
            top_n = _detail_positive_int_arg(
                request.args, "top_n", 5, min_value=1, max_value=20
            )
            wh_raw = str(request.args.get("wh_ids", "") or "").strip()
            wh_ids = parse_wh_ids(wh_raw) if wh_raw else None
            payload = fetch_weekly_dynamics(
                year=year,
                office_id=cfg.get("office_id"),
                wh_ids=wh_ids,
                top_n=top_n,
            )
            return jsonify(payload)
        except QueryParamError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/status")
    def api_status():
        try:
            return jsonify(fetch_status_payload())
        except Exception as exc:
            return jsonify({"status": "error", "detail": str(exc)}), 500

    @application.route("/api/nomenclature/latest")
    def api_nomenclature_latest():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = request.args.get("year", type=int)
            data = fetch_nomenclature_counts_latest_week(
                office_id=cfg.get("office_id"),
                year=year,
            )
            return jsonify(data)
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/api/nomenclature/export/xlsx")
    def api_nomenclature_export_xlsx():
        env_err = check_db_env()
        if env_err:
            return jsonify({"error": env_err}), 503
        try:
            cfg = _cfg()
            year = request.args.get("year", type=int)
            payload = fetch_nomenclature_counts_latest_week(
                office_id=cfg.get("office_id"),
                year=year,
            )
            limit = request.args.get("limit", type=int)
            if limit and limit > 0:
                rows = (payload.get("rows") or [])[:limit]
                payload = {
                    "latest_year": payload.get("latest_year"),
                    "latest_week": payload.get("latest_week"),
                    "rows": rows,
                    "totals": {
                        "corpus_1": sum(int(r.get("corpus_1", 0)) for r in rows),
                        "corpus_2": sum(int(r.get("corpus_2", 0)) for r in rows),
                        "corpus_3": sum(int(r.get("corpus_3", 0)) for r in rows),
                        "total": sum(int(r.get("total", 0)) for r in rows),
                    },
                }
            blob = export_nomenclature_xlsx(payload)
            y = payload.get("latest_year") or "na"
            w = payload.get("latest_week") or "na"
            filename = f"nomenclature_{y}_{w}.xlsx"
            return send_file(
                BytesIO(blob),
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as exc:
            return jsonify({"error": str(exc)}), 500

    @application.route("/health")
    def health():
        return jsonify({"status": "ok", "config": str(CONFIG_PATH.name)})

    @application.route("/health/env")
    def health_env():
        def _state(name: str) -> str:
            if name not in os.environ:
                return "missing"
            if os.environ.get(name, "").strip() == "":
                return "empty"
            return "set"

        return jsonify(
            {
                "status": "ok",
                "vercel_env": os.environ.get("VERCEL_ENV", "local"),
                "db_env_error": check_db_env(),
                "env": {
                    "DATABASE_URL": _state("DATABASE_URL"),
                    "DB_HOST": _state("DB_HOST"),
                    "DB_PORT": _state("DB_PORT"),
                    "DB_NAME": _state("DB_NAME"),
                    "DB_USER": _state("DB_USER"),
                    "DB_PASSWORD": _state("DB_PASSWORD"),
                    "DB_SSLMODE": _state("DB_SSLMODE"),
                },
            }
        )

    @application.route("/health/db")
    def health_db():
        env_err = check_db_env()
        if env_err:
            return jsonify({"status": "error", "detail": env_err}), 503
        try:
            with get_conn() as conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
            return jsonify({"status": "ok", "database": "connected"})
        except Exception as exc:
            return jsonify({"status": "error", "detail": str(exc)}), 503


def create_app():
    from flask import Flask

    application = Flask(__name__)
    register_routes(application)
    return application


def run_server() -> None:
    host = os.environ.get("HTML_HOST", "127.0.0.1")
    port = int(os.environ.get("HTML_PORT", "8080"))
    print(f"Дашборд: http://{host}:{port}/")
    print(f"Корпуса: {CONFIG_PATH}")
    create_app().run(host=host, port=port, debug=False)


def main() -> int:
    if not os.environ.get("DB_PASSWORD"):
        print("Создайте .env с DB_PASSWORD", file=sys.stderr)
        return 1
    print(f"Кэш отчета: {CACHE_TTL_SEC}s, кэш недель: {WEEKS_CACHE_TTL_SEC}s")
    try:
        run_server()
    except Exception as e:
        print(f"Ошибка: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())